"""Tests para emision masiva de comprobantes."""

from copy import deepcopy
from datetime import date, datetime, timedelta
from decimal import Decimal
import hashlib
from io import BytesIO
import json
from types import SimpleNamespace

import pytest
from httpx import AsyncClient
from openpyxl import Workbook, load_workbook
from openpyxl.utils.datetime import to_excel
from sqlalchemy import JSON, func, select, update
from sqlalchemy.exc import OperationalError
from sqlalchemy.exc import TimeoutError as SQLAlchemyTimeoutError
from sqlalchemy.ext.asyncio import AsyncSession

from app.arca.exceptions import (
    ArcaErrorGlobalEstructurado,
    ArcaServiceError,
    CabeceraRespuestaFecae,
    MensajeArcaEstructurado,
)
from app.arca.models import CAEResponse
from app.arca.models import ComprobanteResponse as ArcaComprobanteResponse
from app.core.config import settings
from app.models.certificado import Certificado
from app.models.comprobante import Comprobante
from app.models.comprobante_item import ComprobanteItem
from app.models.empresa import Empresa
from app.models.elegibilidad_rece import (
    PuntoVentaElegibilidadReceActual,
    PuntoVentaElegibilidadReceRevision,
    PuntoVentaGuardaEmisionRece,
)
from app.models.lote_comprobante import (
    LoteComprobante,
    LoteComprobanteEvento,
    LoteComprobanteFila,
    LoteComprobanteGrupo,
)
from app.models.idempotencia_fiscal import IntentoEmisionFiscal, OperacionIdempotente
from app.models.punto_venta import PuntoVenta
from app.schemas.comprobante import EmitirComprobanteRequest, EmitirComprobanteResponse
from app.schemas.lote_comprobante import (
    LoteComprobanteResponse,
    LoteComprobanteSeguimientoResponse,
    LoteProcesamientoResponse,
    LoteReconciliacionExternaItem,
)
from app.services.facturacion_service import FacturacionService
from app.services.lote_comprobantes_service import (
    LoteComprobanteConflictoError,
    LoteComprobanteError,
    LoteComprobantesService,
)
from app.services.idempotencia_fiscal_service import IdempotenciaFiscalService
from app.services.elegibilidad_rece_service import (
    ContextoElegibilidadRece,
    ElegibilidadReceService,
)
from app.services.lote_worker import LoteWorker, get_lote_worker_status


def _crear_error_db_temporal(
    error_type: type[Exception],
) -> SQLAlchemyTimeoutError | OperationalError:
    """Construye errores transitorios de SQLAlchemy sin una base externa."""
    if error_type is SQLAlchemyTimeoutError:
        return SQLAlchemyTimeoutError()
    return OperationalError(
        "UPDATE lotes_comprobantes SET estado = :estado",
        {"estado": "procesando"},
        RuntimeError("base temporalmente no disponible"),
    )


# Identificadores sintéticos de fixtures. Se construyen en partes para evitar
# versionar por accidente datos fiscales reales o emitidos.
CUIT_RECEPTOR_TEST_NO_REAL = "".join(("20", "40937847", "2"))
CUIT_RECEPTOR_TEST_NO_REAL_INT = int(CUIT_RECEPTOR_TEST_NO_REAL)
CAE_TEST_NO_REAL_SERIE = "".join(("1234567", "89012"))
CAE_TEST_NO_REAL_PREFIX = f"{CAE_TEST_NO_REAL_SERIE}3"
CAE_TEST_NO_REAL = f"{CAE_TEST_NO_REAL_SERIE}34"
CAE_TEST_NO_REAL_ALT = f"{CAE_TEST_NO_REAL_SERIE}35"
CAE_TEST_NO_REAL_36 = f"{CAE_TEST_NO_REAL_SERIE}36"
CAE_TEST_NO_REAL_37 = f"{CAE_TEST_NO_REAL_SERIE}37"
CAE_TEST_NO_REAL_38 = f"{CAE_TEST_NO_REAL_SERIE}38"
CAE_TEST_NO_REAL_39 = f"{CAE_TEST_NO_REAL_SERIE}39"
CAE_TEST_NO_REAL_40 = f"{CAE_TEST_NO_REAL_SERIE}40"
FECHA_FISCAL_PF02B2 = date(2026, 7, 29)
FECHA_FISCAL_CONTROLADA_PF19B = date(2026, 8, 9)
FECHA_DOCUMENTO_RECE_TEST = date(2026, 8, 1)
FECHA_VIGENCIA_RECE_TEST = date(2099, 12, 31)
INSTANTE_RECE_TEST = datetime(2026, 8, 1, 12, 0, 0)


class _FechaFiscalControlada(date):
    """Reloj estable para conservar activa la ventana fiscal de facturación."""

    @classmethod
    def today(cls) -> date:
        """Devuelve la fecha fiscal explícita compartida por estos tests."""
        return cls(
            FECHA_FISCAL_CONTROLADA_PF19B.year,
            FECHA_FISCAL_CONTROLADA_PF19B.month,
            FECHA_FISCAL_CONTROLADA_PF19B.day,
        )


@pytest.fixture(autouse=True)
def _controlar_reloj_fecha_fiscal(monkeypatch: pytest.MonkeyPatch) -> None:
    """Fija el reloj de facturación sin omitir su validación de ventana ARCA."""
    monkeypatch.setattr(
        "app.services.facturacion_service.date",
        _FechaFiscalControlada,
    )


@pytest.fixture(autouse=True)
def _configurar_ambiente_rece_productivo(monkeypatch: pytest.MonkeyPatch) -> None:
    """Fija el ambiente RECE requerido sin depender del entorno del runner."""
    monkeypatch.setattr(settings, "arca_env", "produccion")


@pytest.fixture(autouse=True)
def _desactivar_batch_arca_por_defecto(monkeypatch: pytest.MonkeyPatch):
    """Evita consultas WSAA/WSFE reales en tests que no prueban batching ARCA."""
    monkeypatch.setattr(settings, "arca_fecaesolicitar_batch_enabled", False)


def _build_lote_excel(
    empresa_cuit: str,
    punto_venta_numero: int | str = 1,
    concepto: int | str = 1,
    tipo_comprobante: int = 6,
    iva: int | float = 21,
    cliente_tipo_documento: str = "CUIT",
    cliente_numero_documento: str = CUIT_RECEPTOR_TEST_NO_REAL,
    cliente_razon_social: str = "Cliente Lote SA",
    cliente_condicion_iva: str = "Responsable Inscripto",
    item_precio_unitario: int | float = 1000,
    fecha_servicio_desde: date | str = "",
    fecha_servicio_hasta: date | str = "",
    fecha_vto_pago: date | str = "",
    asociado_tipo_comprobante: int | str = "",
    asociado_punto_venta: int | str = "",
    asociado_numero: int | str = "",
    asociado_fecha: date | str = "",
    asociado_cuit: str = "",
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Comprobantes"
    sheet.append(
        [
            "comprobante_ref",
            "empresa_cuit",
            "punto_venta_numero",
            "tipo_comprobante",
            "concepto",
            "fecha_emision",
            "cliente_tipo_documento",
            "cliente_numero_documento",
            "cliente_razon_social",
            "cliente_condicion_iva",
            "cliente_domicilio",
            "fecha_servicio_desde",
            "fecha_servicio_hasta",
            "fecha_vto_pago",
            "item_codigo",
            "item_descripcion",
            "item_cantidad",
            "item_unidad",
            "item_precio_unitario",
            "item_descuento_porcentaje",
            "item_iva_porcentaje",
            "observaciones",
            "asociado_tipo_comprobante",
            "asociado_punto_venta",
            "asociado_numero",
            "asociado_fecha",
            "asociado_cuit",
        ]
    )
    asociado_fecha_valor = (
        asociado_fecha.isoformat()
        if isinstance(asociado_fecha, date)
        else asociado_fecha
    )
    fecha_servicio_desde_valor = (
        fecha_servicio_desde.isoformat()
        if isinstance(fecha_servicio_desde, date)
        else fecha_servicio_desde
    )
    fecha_servicio_hasta_valor = (
        fecha_servicio_hasta.isoformat()
        if isinstance(fecha_servicio_hasta, date)
        else fecha_servicio_hasta
    )
    fecha_vto_pago_valor = (
        fecha_vto_pago.isoformat()
        if isinstance(fecha_vto_pago, date)
        else fecha_vto_pago
    )
    sheet.append(
        [
            "LOTE-001",
            empresa_cuit,
            punto_venta_numero,
            tipo_comprobante,
            concepto,
            FECHA_FISCAL_CONTROLADA_PF19B.isoformat(),
            cliente_tipo_documento,
            cliente_numero_documento,
            cliente_razon_social,
            cliente_condicion_iva,
            "Av. Siempre Viva 123",
            fecha_servicio_desde_valor,
            fecha_servicio_hasta_valor,
            fecha_vto_pago_valor,
            "ITEM-001",
            "Servicio mensual",
            1,
            "unidad",
            item_precio_unitario,
            0,
            iva,
            "Factura de prueba",
            asociado_tipo_comprobante,
            asociado_punto_venta,
            asociado_numero,
            asociado_fecha_valor,
            asociado_cuit,
        ]
    )

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


async def _confirmacion_fecha_fiscal_header_lote(
    db_session: AsyncSession,
    *,
    lote_id: int,
    estados: set[str],
    grupo_ids: list[int] | None = None,
    idempotency_key: str = "idem-lote-test",
) -> dict[str, str]:
    """Obtiene el token fiscal RECE exacto del lote usado por el test."""
    confirmacion = await LoteComprobantesService(
        db_session
    ).obtener_confirmacion_fiscal_grupos(
        lote_id=lote_id,
        empresa_id=int(
            await db_session.scalar(
                select(LoteComprobante.empresa_id).where(LoteComprobante.id == lote_id)
            )
        ),
        estados=estados,
        grupo_ids=grupo_ids,
    )
    return {
        "X-Confirmacion-Fecha-Fiscal": str(confirmacion["confirmacion_fecha_fiscal"]),
        "X-Idempotency-Key": idempotency_key,
    }


@pytest.mark.asyncio
async def test_archivo_observado_escapa_textos_con_formulas(
    db_session: AsyncSession,
    test_empresa,
):
    """El Excel observado no debe abrir formulas tomadas del archivo original."""
    datos = {column: "" for column in LoteComprobantesService.TEMPLATE_COLUMNS}
    datos.update(
        {
            "comprobante_ref": "=SUM(1,1)",
            "empresa_cuit": test_empresa.cuit,
            "item_descripcion": "+cmd",
            "observaciones": " @malicioso",
        }
    )
    lote = LoteComprobante(
        nombre_archivo="observado.xlsx",
        archivo_hash="hash-observado-formulas",
        estado="con_errores",
        total_filas=1,
        total_grupos=1,
        grupos_con_error=1,
        empresa_id=test_empresa.id,
    )
    grupo = LoteComprobanteGrupo(
        lote=lote,
        comprobante_ref="=SUM(1,1)",
        orden=1,
        estado="con_error",
        mensajes_json=["=HYPERLINK('http://malicioso')"],
    )
    fila = LoteComprobanteFila(
        lote=lote,
        grupo=grupo,
        fila_excel=2,
        comprobante_ref="=SUM(1,1)",
        estado="con_error",
        datos_json=datos,
        mensajes_json=["=HYPERLINK('http://malicioso')"],
    )
    db_session.add_all([lote, grupo, fila])
    await db_session.commit()
    await db_session.refresh(lote)

    service = LoteComprobantesService(db_session)
    contenido = await service.generar_archivo_observado(lote.id, test_empresa.id)

    workbook = load_workbook(BytesIO(contenido), data_only=False)
    sheet = workbook["Resultados"]
    headers = [cell.value for cell in sheet[1]]
    row = {
        header: sheet.cell(row=2, column=index + 1).value
        for index, header in enumerate(headers)
    }

    assert row["comprobante_ref"] == "'=SUM(1,1)"
    assert row["item_descripcion"] == "'+cmd"
    assert row["observaciones"] == "' @malicioso"
    assert row["resultado_mensajes"] == "'=HYPERLINK('http://malicioso')"
    assert sheet["A2"].data_type == "s"


def _build_lote_excel_multi_grupo(empresa_cuit: str, total_grupos: int = 2) -> bytes:
    """Construye un Excel de prueba con varios comprobantes independientes."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Comprobantes"
    sheet.append(
        [
            "comprobante_ref",
            "empresa_cuit",
            "punto_venta_numero",
            "tipo_comprobante",
            "concepto",
            "fecha_emision",
            "cliente_tipo_documento",
            "cliente_numero_documento",
            "cliente_razon_social",
            "cliente_condicion_iva",
            "cliente_domicilio",
            "fecha_servicio_desde",
            "fecha_servicio_hasta",
            "fecha_vto_pago",
            "item_codigo",
            "item_descripcion",
            "item_cantidad",
            "item_unidad",
            "item_precio_unitario",
            "item_descuento_porcentaje",
            "item_iva_porcentaje",
            "observaciones",
            "asociado_tipo_comprobante",
            "asociado_punto_venta",
            "asociado_numero",
            "asociado_fecha",
            "asociado_cuit",
        ]
    )
    for index in range(1, total_grupos + 1):
        sheet.append(
            [
                f"LOTE-{index:03d}",
                empresa_cuit,
                1,
                6,
                1,
                FECHA_FISCAL_CONTROLADA_PF19B.isoformat(),
                "CUIT",
                CUIT_RECEPTOR_TEST_NO_REAL,
                f"Cliente Lote {index}",
                "Responsable Inscripto",
                "Av. Siempre Viva 123",
                "",
                "",
                "",
                f"ITEM-{index:03d}",
                "Servicio mensual",
                1,
                "unidad",
                1000,
                0,
                21,
                "Factura de prueba",
                "",
                "",
                "",
                "",
                "",
            ]
        )

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _build_extracto_bancario_excel(
    empresa_cuit: str,
    fecha_movimiento: date | None = None,
    fecha_como_serial: bool = False,
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Extracto"
    sheet.append(
        [
            "Fecha",
            "Créditos",
            "Leyendas Adicionales1",
            "Leyendas Adicionales2",
            "Pto Vta",
        ]
    )
    fecha_base = fecha_movimiento or FECHA_FISCAL_CONTROLADA_PF19B
    fecha = (
        to_excel(fecha_base) if fecha_como_serial else fecha_base.strftime("%d/%m/%Y")
    )
    sheet.append([fecha, "59.500,00", "CLIENTE UNO", CUIT_RECEPTOR_TEST_NO_REAL, 1])
    sheet.append([fecha, "70.500,00", "CLIENTE DOS", CUIT_RECEPTOR_TEST_NO_REAL, 10])
    sheet.append([fecha, "140.000,00", "CLIENTE TRES", CUIT_RECEPTOR_TEST_NO_REAL, 13])

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _build_cano_factura_b_excel(fecha_movimiento: date | None = None) -> bytes:
    """Genera una muestra del formato Cano con Factura B e IVA discriminado."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Hoja1"
    sheet.append(
        [
            "Fecha",
            "Tipo",
            "Punto de Venta",
            "Número Desde",
            "Número Hasta",
            "Cód. Autorización",
            "Tipo Doc. Receptor",
            "Nro. Doc. Receptor",
            "Denominación Receptor",
            "Tipo Cambio",
            "Moneda",
            "Imp. Neto Gravado",
            "Imp. Neto No Gravado",
            "Imp. Op. Exentas",
            "Otros Tributos",
            "IVA",
            "Imp. Total",
        ]
    )
    fecha = fecha_movimiento or FECHA_FISCAL_CONTROLADA_PF19B
    sheet.append(
        [
            fecha,
            "6 - Factura B",
            2,
            1,
            "",
            "",
            "DNI",
            "HEBER YOEL ASANCHEZ CA -",
            "",
            1,
            "$",
            74380.1652892562,
            0,
            0,
            "",
            15619.8347107438,
            90000,
        ]
    )

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _config_formato_cano_factura_b() -> dict:
    """Devuelve la configuración del formato Cano Factura B con IVA 21%."""
    return {
        "tipo": "cano_factura_b_iva_21",
        "header_row": 1,
        "modo_agrupacion": "fila",
        "campos": {
            "fecha_origen": {
                "origen": "header",
                "encabezados": ["Fecha"],
                "transformacion": "fecha",
                "requerido": True,
            },
            "importe_total": {
                "origen": "header",
                "encabezados": ["Imp. Total"],
                "transformacion": "decimal",
                "requerido": False,
            },
            "item_precio_unitario": {
                "origen": "header",
                "encabezados": ["Imp. Neto Gravado"],
                "transformacion": "decimal",
                "requerido": True,
            },
            "cliente_razon_social": {
                "origen": "header",
                "encabezados": ["Nro. Doc. Receptor", "Denominación Receptor"],
                "transformacion": "texto",
                "requerido": False,
                "default": "",
            },
            "punto_venta_numero": {
                "origen": "header",
                "encabezados": ["Punto de Venta"],
                "transformacion": "entero",
                "requerido": True,
            },
            "tipo_comprobante": {"origen": "constante", "valor": 6},
            "cliente_condicion_iva": {
                "origen": "constante",
                "valor": "Consumidor Final",
            },
            "item_cantidad": {"origen": "constante", "valor": 1},
            "item_unidad": {"origen": "constante", "valor": "unidad"},
            "item_iva_porcentaje": {"origen": "constante", "valor": 21},
            "item_descuento_porcentaje": {"origen": "constante", "valor": 0},
            "guardar_cliente": {"origen": "constante", "valor": False},
        },
    }


def _fecha_argentina(value: date | str) -> str:
    """Formatea una fecha de test en DD/MM/AAAA."""
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    return datetime.strptime(str(value), "%Y-%m-%d").strftime("%d/%m/%Y")


def test_reconciliacion_externa_item_rechaza_fecha_calendario_invalida():
    """El schema no debe aceptar fechas externas imposibles."""
    with pytest.raises(ValueError):
        LoteReconciliacionExternaItem(
            grupo_id=1,
            tipo_comprobante=6,
            punto_venta_numero=1,
            numero=123,
            fecha_emision="31/02/2026",
            total=Decimal("1210.00"),
            motivo="Emitido manualmente por ARCA Web",
        )


def _hashes_fiscales_request(
    request: EmitirComprobanteRequest,
    punto_venta_numero: int,
    total: Decimal,
) -> tuple[str, str]:
    """Calcula los hashes fiscales del request igual que producción."""
    payload = request.model_dump(mode="json")
    payload_hash = IdempotenciaFiscalService.calcular_payload_hash(
        IdempotenciaFiscalService.payload_sin_confirmacion_duplicado(payload)
    )
    huella = IdempotenciaFiscalService.calcular_huella_logica(
        request=request,
        punto_venta_numero=punto_venta_numero,
        total=total,
    )
    return payload_hash, huella


def _payload_lote_basico(
    empresa_id: int,
    punto_venta_id: int,
    fecha_fiscal: date,
    razon_social: str = "Cliente Lote SA",
) -> dict[str, object]:
    """Construye un payload fiscal sintético y estable para tests de lotes."""
    return {
        "empresa_id": empresa_id,
        "punto_venta_id": punto_venta_id,
        "tipo_comprobante": 6,
        "concepto": 1,
        "fecha_emision": fecha_fiscal.isoformat(),
        "confirmacion_fecha_fiscal": True,
        "tipo_documento": 80,
        "numero_documento": CUIT_RECEPTOR_TEST_NO_REAL,
        "razon_social": razon_social,
        "condicion_iva": "RI",
        "domicilio": "Av. Siempre Viva 123",
        "moneda": "PES",
        "cotizacion": "1",
        "guardar_cliente": False,
        "items": [
            {
                "descripcion": "Servicio mensual",
                "cantidad": "1",
                "unidad": "unidad",
                "precio_unitario": "1000",
                "iva_porcentaje": "21",
            }
        ],
    }


def _opciones_fechas(
    fecha_emision_modo: str = "archivo",
    fecha_emision_fija: date | str | None = None,
    concepto_modo: str = "productos",
    descripcion_item_modo: str = "archivo",
    descripcion_item_fija: str | None = None,
    punto_venta_modo: str = "archivo",
    punto_venta_numero: int | None = None,
    fecha_servicio_desde_fija: date | str | None = None,
    fecha_servicio_hasta_fija: date | str | None = None,
    fecha_vto_pago_fija: date | str | None = None,
) -> dict[str, str]:
    """Devuelve opciones explícitas para validar lotes."""
    data = {
        "concepto_modo": concepto_modo,
        "descripcion_item_modo": descripcion_item_modo,
        "punto_venta_modo": punto_venta_modo,
        "fecha_emision_modo": fecha_emision_modo,
        "fecha_servicio_desde_modo": "archivo",
        "fecha_servicio_hasta_modo": "archivo",
        "fecha_vto_pago_modo": "archivo",
    }
    for key, value in {
        "fecha_emision_fija": fecha_emision_fija,
        "fecha_servicio_desde_fija": fecha_servicio_desde_fija,
        "fecha_servicio_hasta_fija": fecha_servicio_hasta_fija,
        "fecha_vto_pago_fija": fecha_vto_pago_fija,
    }.items():
        if value:
            data[key] = value.isoformat() if isinstance(value, date) else value
    if descripcion_item_fija:
        data["descripcion_item_fija"] = descripcion_item_fija
    if punto_venta_numero:
        data["punto_venta_numero"] = str(punto_venta_numero)
    return data


async def _crear_punto_venta_rece_verificado(
    db_session: AsyncSession,
    empresa: Empresa,
    *,
    usuario_id: int,
    numero: int,
    nombre: str,
    documento_emitido_en: date,
    vigente_hasta: date,
    observado_en: datetime,
) -> PuntoVenta:
    """Crea un punto sintético con ledger y cabeza RECE positivos explícitos."""
    assert settings.arca_env == "produccion"
    punto = PuntoVenta(
        numero=numero,
        nombre=nombre,
        activo=True,
        es_webservice=True,
        empresa_id=empresa.id,
        revision_fiscal=1,
    )
    db_session.add(punto)
    await db_session.flush()
    elegibilidad = ElegibilidadReceService(db_session)
    await elegibilidad.crear_contextos_iniciales_no_verificados(
        punto,
        creado_por_usuario_id=usuario_id,
    )
    revision = PuntoVentaElegibilidadReceRevision(
        empresa_id=empresa.id,
        punto_venta_id=punto.id,
        ambiente="produccion",
        revision=2,
        estado="verificado_rece",
        fuente="constancia_arca_atestada",
        evidencia_tipo="rece_aplicativo_web_services_v1",
        evidencia_sha256=f"{numero:064x}",
        clasificador_version="rece-v1-test",
        empresa_cuit_snapshot=empresa.cuit,
        punto_venta_numero_snapshot=numero,
        punto_revision_fiscal=1,
        documento_emitido_en=documento_emitido_en,
        vigente_hasta=vigente_hasta,
        observado_en=observado_en,
        verificado_en=observado_en,
        creado_por_usuario_id=usuario_id,
        actor_usuario_id_snapshot=usuario_id,
        created_at=observado_en,
    )
    db_session.add(revision)
    await db_session.flush()
    head = await db_session.scalar(
        select(PuntoVentaElegibilidadReceActual).where(
            PuntoVentaElegibilidadReceActual.punto_venta_id == punto.id,
            PuntoVentaElegibilidadReceActual.ambiente == "produccion",
        )
    )
    assert head is not None
    head.revision_actual_id = revision.id
    await db_session.flush()
    return punto


@pytest.fixture
async def test_punto_venta(
    db_session: AsyncSession,
    test_empresa,
    test_user,
) -> PuntoVenta:
    """Crea un punto con evidencia RECE positiva únicamente para tests felices."""
    punto = await _crear_punto_venta_rece_verificado(
        db_session,
        test_empresa,
        usuario_id=int(test_user.id),
        numero=1,
        nombre="Principal",
        documento_emitido_en=FECHA_DOCUMENTO_RECE_TEST,
        vigente_hasta=FECHA_VIGENCIA_RECE_TEST,
        observado_en=INSTANTE_RECE_TEST,
    )
    await db_session.commit()
    return punto


@pytest.fixture
async def test_certificado(db_session: AsyncSession, test_empresa) -> Certificado:
    certificado = Certificado(
        nombre="Certificado homologacion",
        cuit=test_empresa.cuit,
        fecha_emision=date(2026, 1, 1),
        fecha_vencimiento=date(2026, 12, 31),
        archivo_crt="empresa-test.crt",
        archivo_key="empresa-test.key",
        activo=True,
        ambiente=settings.arca_env,
        empresa_id=test_empresa.id,
    )
    db_session.add(certificado)
    await db_session.commit()
    await db_session.refresh(certificado)
    return certificado


async def _persistir_comprobante_autorizado(
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta: PuntoVenta,
    *,
    tipo_comprobante: int,
    numero: int,
    fecha_emision: date,
    cae: str,
    cae_vencimiento: date,
    total: Decimal,
) -> int:
    """Crea un comprobante sintético para fakes que simulan CAE autorizado."""
    comprobante = Comprobante(
        tipo_comprobante=tipo_comprobante,
        concepto=1,
        numero=numero,
        fecha_emision=fecha_emision,
        subtotal=total,
        descuento=Decimal("0.00"),
        iva_21=Decimal("0.00"),
        iva_10_5=Decimal("0.00"),
        iva_27=Decimal("0.00"),
        otros_impuestos=Decimal("0.00"),
        total=total,
        cae=cae,
        cae_vencimiento=cae_vencimiento,
        estado="autorizado",
        empresa_id=test_empresa.id,
        punto_venta_id=test_punto_venta.id,
        receptor_tipo_documento=99,
        receptor_numero_documento="0",
        receptor_razon_social="A CONSUMIDOR FINAL",
        receptor_condicion_iva="Consumidor Final",
    )
    db_session.add(comprobante)
    await db_session.flush()
    return comprobante.id


async def _crear_lote_validado_por_api(
    client: AsyncClient,
    auth_headers: dict,
    empresa_cuit: str,
    nombre_archivo: str = "lote-resolucion.xlsx",
    total_grupos: int = 1,
) -> int:
    """Crea un lote validado usando el endpoint público de carga masiva."""
    archivo = (
        _build_lote_excel_multi_grupo(empresa_cuit, total_grupos)
        if total_grupos > 1
        else _build_lote_excel(empresa_cuit)
    )
    response = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data=_opciones_fechas(),
        files={
            "archivo": (
                nombre_archivo,
                archivo,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 200, response.text
    return int(response.json()["lote"]["id"])


async def _marcar_grupos_lote(
    db_session: AsyncSession,
    lote_id: int,
    estados: list[str],
) -> list[LoteComprobanteGrupo]:
    """Actualiza estados de grupos y recalcula el lote en pruebas."""
    grupos = list(
        (
            await db_session.execute(
                select(LoteComprobanteGrupo)
                .where(LoteComprobanteGrupo.lote_id == lote_id)
                .order_by(LoteComprobanteGrupo.orden)
            )
        )
        .scalars()
        .all()
    )
    assert len(grupos) >= len(estados)
    for grupo, estado in zip(grupos, estados, strict=False):
        grupo.estado = estado
        grupo.mensajes_json = [f"Estado de prueba: {estado}"]
        if estado in {"autorizado", "requiere_reconciliacion"}:
            grupo.cae = CAE_TEST_NO_REAL
            grupo.numero_asignado = 100 + grupo.orden
        filas = list(
            (
                await db_session.execute(
                    select(LoteComprobanteFila).where(
                        LoteComprobanteFila.grupo_id == grupo.id
                    )
                )
            )
            .scalars()
            .all()
        )
        for fila in filas:
            fila.estado = estado
            fila.mensajes_json = grupo.mensajes_json

    lote = await db_session.get(LoteComprobante, lote_id)
    assert lote is not None
    await db_session.flush()
    service = LoteComprobantesService(db_session)
    await service._actualizar_estado_lote(lote)
    await db_session.commit()
    return grupos


async def _preparar_reintento_manual_pf02b2(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta: PuntoVenta,
    wsfe_client_class: type,
    *,
    nombre_archivo: str,
    total_grupos: int = 1,
    ultimo_local: int | None = None,
) -> tuple[int, list[LoteComprobanteGrupo]]:
    """Prepara un reintento manual determinista con un doble WSFE controlado."""
    lote_id = await _crear_lote_validado_por_api(
        client,
        auth_headers,
        test_empresa.cuit,
        nombre_archivo=nombre_archivo,
        total_grupos=total_grupos,
    )
    grupos = await _marcar_grupos_lote(
        db_session,
        lote_id,
        ["fallido"] * total_grupos,
    )
    for grupo in grupos:
        payload = dict(grupo.payload_json or {})
        payload["fecha_emision"] = FECHA_FISCAL_PF02B2.isoformat()
        grupo.payload_json = payload

    if ultimo_local is not None:
        await _persistir_comprobante_autorizado(
            db_session,
            test_empresa,
            test_punto_venta,
            tipo_comprobante=6,
            numero=ultimo_local,
            fecha_emision=date(2026, 7, 28),
            cae=CAE_TEST_NO_REAL,
            cae_vencimiento=date(2026, 8, 31),
            total=Decimal("1210.00"),
        )
    await db_session.commit()

    async def fake_validar_datos(self, request):
        """Aísla la ventana temporal ARCA para usar una fecha fiscal fija."""

    async def fake_ticket(self, empresa, certificado):
        """Evita leer certificados o contactar WSAA."""
        return SimpleNamespace(token="token-test", sign="sign-test")

    async def fake_validar_punto(self, wsfe_client, punto_venta_numero):
        """Mantiene el punto sintético habilitado sin consultar ARCA."""

    monkeypatch.setattr(
        "app.services.facturacion_service.WSFEv1Client",
        wsfe_client_class,
    )
    monkeypatch.setattr(FacturacionService, "_validar_datos", fake_validar_datos)
    monkeypatch.setattr(FacturacionService, "_obtener_ticket_acceso", fake_ticket)
    monkeypatch.setattr(
        FacturacionService,
        "_validar_punto_venta_habilitado",
        fake_validar_punto,
    )
    return lote_id, grupos


async def _crear_lote_stale_moderno_intacto(
    db_session: AsyncSession,
    empresa: Empresa,
    *,
    grupos_payload: list[tuple[str, dict]],
    idempotency_key: str,
) -> tuple[LoteComprobante, tuple[LoteComprobanteGrupo, ...]]:
    """Persiste ownership worker moderno para grupos stale aún intactos."""
    empresa_id = int(empresa.id)
    elegibilidad = ElegibilidadReceService(db_session)
    contextos_por_punto: dict[int, ContextoElegibilidadRece] = {}
    for _, payload in grupos_payload:
        punto_venta_id = int(payload["punto_venta_id"])
        if punto_venta_id in contextos_por_punto:
            continue
        contextos_por_punto[
            punto_venta_id
        ] = await elegibilidad.exigir_contexto_preautorizacion(
            empresa_id=empresa_id,
            punto_venta_id=punto_venta_id,
            ambiente=settings.arca_env,
            tipo_comprobante=int(payload["tipo_comprobante"]),
        )

    lote = LoteComprobante(
        nombre_archivo=f"{idempotency_key}.xlsx",
        archivo_hash=f"hash-{idempotency_key}",
        estado="validado",
        total_filas=len(grupos_payload),
        total_grupos=len(grupos_payload),
        grupos_validos=len(grupos_payload),
        empresa_id=empresa_id,
        metadata_json={
            "opciones_concepto": {"concepto_modo": "archivo"},
            "opciones_descripcion_item": {"descripcion_item_modo": "archivo"},
        },
    )
    db_session.add(lote)
    await db_session.flush()
    lote_id = int(lote.id)

    grupos: list[LoteComprobanteGrupo] = []
    for orden, (comprobante_ref, payload) in enumerate(grupos_payload, start=1):
        contexto = contextos_por_punto[int(payload["punto_venta_id"])]
        grupo = LoteComprobanteGrupo(
            lote_id=lote_id,
            empresa_id=empresa_id,
            comprobante_ref=comprobante_ref,
            orden=orden,
            estado="validado",
            tipo_comprobante=int(payload["tipo_comprobante"]),
            punto_venta_id=contexto.punto_venta_id,
            punto_venta_numero=contexto.punto_venta_numero,
            ambiente=contexto.ambiente,
            punto_venta_elegibilidad_revision_id=(contexto.elegibilidad_revision_id),
            punto_venta_revision_fiscal=contexto.punto_venta_revision_fiscal,
            cliente_documento=str(payload.get("numero_documento") or ""),
            cliente_razon_social=str(payload.get("razon_social") or ""),
            total_estimado=Decimal("1210.00"),
            payload_json=deepcopy(payload),
            mensajes_json=["Validado correctamente. Listo para emitir."],
        )
        db_session.add(grupo)
        grupos.append(grupo)
    await db_session.flush()

    service = LoteComprobantesService(db_session)
    material_rece = await service.calcular_material_idempotente_grupos(
        lote_id=lote_id,
        empresa_id=empresa_id,
        estados={"validado"},
    )
    idempotencia = IdempotenciaFiscalService(db_session)
    payload_hash = idempotencia.calcular_payload_hash(
        {
            "lote_id": lote_id,
            "grupo_ids": material_rece["grupo_ids"],
            "grupos_hash": material_rece["grupos_hash"],
        }
    )
    operacion, creada = await idempotencia.obtener_o_crear_operacion(
        empresa_id=empresa_id,
        usuario_id=None,
        idempotency_key=idempotency_key,
        tipo_operacion="procesar_lote",
        payload_hash=payload_hash,
        lote_id=lote_id,
        contextos_rece=list(contextos_por_punto.values()),
    )
    assert creada is True
    operacion_id = int(operacion.id)

    lote = await db_session.get(LoteComprobante, lote_id)
    assert lote is not None
    lote = await service.encolar_lote(
        lote_id=lote_id,
        empresa_id=empresa_id,
        operacion_id=operacion_id,
        material_rece=material_rece,
        commit=False,
    )
    respuesta_encolada = LoteProcesamientoResponse(
        lote=LoteComprobanteResponse.model_validate(lote),
        mensaje="El lote quedó en cola y se está procesando en segundo plano.",
        en_progreso=True,
    )
    publicada = await idempotencia.guardar_respuesta_operacion_cas(
        operacion_id=operacion_id,
        response_json=respuesta_encolada,
        estado="en_proceso",
        estado_esperado="en_proceso",
        respuesta_esperada_nula=True,
        commit=False,
    )
    assert publicada is True
    await db_session.commit()

    await service._tomar_lote_para_procesamiento(
        lote_id=lote_id,
        empresa_id=empresa_id,
        procesamiento_async=True,
        modo_procesamiento="background",
    )
    await db_session.flush()
    lote = await db_session.get(LoteComprobante, lote_id)
    assert lote is not None
    await db_session.refresh(lote)
    lote.updated_at = datetime.utcnow() - timedelta(
        minutes=settings.batch_processing_stale_minutes + 1
    )
    await service._guardar_respuesta_operacion_background(lote, operacion_id)
    await db_session.commit()

    db_session.expire_all()
    lote = await db_session.get(LoteComprobante, lote_id)
    operacion = await db_session.get(OperacionIdempotente, operacion_id)
    grupos_actuales = tuple(
        (
            await db_session.scalars(
                select(LoteComprobanteGrupo)
                .where(LoteComprobanteGrupo.lote_id == lote_id)
                .order_by(LoteComprobanteGrupo.orden)
            )
        ).all()
    )
    assert lote is not None
    assert operacion is not None
    assert lote.estado == "procesando"
    assert lote.metadata_json["operacion_idempotente_id"] == operacion_id
    assert lote.metadata_json["pf19b_rece_material"] == material_rece
    assert operacion.estado == "en_proceso"
    assert operacion.response_json["en_progreso"] is True
    assert operacion.response_json["lote"]["estado"] == "procesando"
    assert (
        operacion.response_json["lote"]["metadata_json"]["pf19b_rece_material"]
        == material_rece
    )
    return lote, grupos_actuales


def _instalar_oraculos_stale_sin_arca(
    service: LoteComprobantesService,
) -> dict[str, int]:
    """Hace observables WSAA, FEComp y FECAE en casos stale fail-closed."""
    llamadas = {"wsaa": 0, "fecomp": 0, "fecae": 0}

    async def fail_wsaa(*args, **kwargs):
        llamadas["wsaa"] += 1
        raise AssertionError("El caso stale no debe solicitar WSAA")

    async def fail_fecomp(*args, **kwargs):
        llamadas["fecomp"] += 1
        raise AssertionError("El caso stale no debe consultar FEComp")

    async def fail_fecae(*args, **kwargs):
        llamadas["fecae"] += 1
        raise AssertionError("El caso stale no debe solicitar FECAE")

    service.facturacion_service._obtener_ticket_acceso = fail_wsaa
    service.facturacion_service.verificar_numeracion_segura_para_emision = fail_fecomp
    service.facturacion_service.emitir_comprobante = fail_fecae
    service.facturacion_service._emitir_comprobante_locked = fail_fecae
    return llamadas


@pytest.mark.asyncio
async def test_descargar_plantilla_lote(
    client: AsyncClient,
    auth_headers: dict,
):
    response = await client.get(
        "/api/lotes-comprobantes/plantilla", headers=auth_headers
    )

    assert response.status_code == 200
    assert (
        response.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment;" in response.headers["content-disposition"]


@pytest.mark.asyncio
async def test_obtener_resumen_y_grupos_paginados_lote(
    client: AsyncClient,
    auth_headers: dict,
    db_session: AsyncSession,
    test_empresa,
):
    """El detalle paginado debe evitar traer todo el lote para abrir la UI."""
    lote = LoteComprobante(
        nombre_archivo="lote-grande.xlsx",
        archivo_hash="hash-lote-grande-paginado",
        estado="validado",
        total_filas=3,
        total_grupos=3,
        grupos_validos=2,
        grupos_con_error=1,
        empresa_id=test_empresa.id,
    )
    db_session.add(lote)
    await db_session.flush()

    for index, estado in enumerate(["validado", "validado", "con_error"], start=1):
        payload = {
            "fecha_emision": "2026-05-20",
            "concepto": 1,
            "items": [
                {
                    "cantidad": 1,
                    "precio_unitario": 1000,
                    "descuento_porcentaje": 0,
                    "iva_porcentaje": 21,
                }
            ],
        }
        grupo = LoteComprobanteGrupo(
            lote_id=lote.id,
            empresa_id=lote.empresa_id,
            comprobante_ref=f"LOTE-{index:03d}",
            orden=index,
            estado=estado,
            tipo_comprobante=6,
            punto_venta_numero=1,
            cliente_documento=CUIT_RECEPTOR_TEST_NO_REAL,
            cliente_razon_social=f"Cliente {index}",
            total_estimado=Decimal("1210"),
            payload_json=payload,
            mensajes_json=["Validado correctamente. Listo para emitir."]
            if estado == "validado"
            else ["Observado"],
        )
        db_session.add(grupo)
        await db_session.flush()
        db_session.add(
            LoteComprobanteFila(
                lote_id=lote.id,
                grupo_id=grupo.id,
                fila_excel=index + 1,
                comprobante_ref=grupo.comprobante_ref,
                estado=estado,
                datos_json={"item_descripcion": f"Servicio {index}"},
                mensajes_json=grupo.mensajes_json,
            )
        )
    await db_session.commit()
    material_resumen = await LoteComprobantesService(
        db_session
    ).calcular_material_idempotente_grupos(
        lote_id=lote.id,
        empresa_id=test_empresa.id,
        estados={"validado"},
    )

    resumen = await client.get(
        f"/api/lotes-comprobantes/{lote.id}/resumen",
        headers=auth_headers,
    )
    assert resumen.status_code == 200, resumen.text
    resumen_data = resumen.json()
    assert "grupos" not in resumen_data
    assert "filas" not in resumen_data
    assert resumen_data["confirmacion_fecha_fiscal"] == (
        "fechas=2026-05-20;puntos_venta=1;" f"rece={material_resumen['grupos_hash']}"
    )
    assert resumen_data["fechas_emision_validas"] == ["2026-05-20"]
    assert resumen_data["puntos_venta_validos"] == [1]
    assert resumen_data["totales_listos_para_emitir"] == {
        "comprobantes": 2,
        "neto": 2000,
        "iva21": 420,
        "iva105": 0,
        "total": 2420,
        "valores_invalidos": 0,
    }

    pagina = await client.get(
        f"/api/lotes-comprobantes/{lote.id}/grupos?page=1&per_page=2",
        headers=auth_headers,
    )
    assert pagina.status_code == 200, pagina.text
    pagina_data = pagina.json()
    assert pagina_data["total"] == 3
    assert pagina_data["total_pages"] == 2
    assert [item["comprobante_ref"] for item in pagina_data["items"]] == [
        "LOTE-001",
        "LOTE-002",
    ]
    assert pagina_data["items"][0]["descripcion_facturada"] == "Servicio 1"

    filtrada = await client.get(
        f"/api/lotes-comprobantes/{lote.id}/grupos?estado=validado&per_page=10",
        headers=auth_headers,
    )
    assert filtrada.status_code == 200, filtrada.text
    filtrada_data = filtrada.json()
    assert filtrada_data["total"] == 2
    assert {item["estado"] for item in filtrada_data["items"]} == {"validado"}


@pytest.mark.asyncio
async def test_seguimiento_lote_es_liviano_y_no_muta_updated_at(
    client: AsyncClient,
    auth_headers: dict,
    db_session: AsyncSession,
    test_empresa,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """El polling consulta una vez el resumen persistido y no refresca el lote."""
    lote = LoteComprobante(
        nombre_archivo="lote-seguimiento.xlsx",
        archivo_hash="hash-lote-seguimiento-liviano",
        estado="en_cola",
        modo_procesamiento="background",
        procesamiento_async=True,
        total_filas=200,
        total_grupos=100,
        grupos_validos=100,
        empresa_id=test_empresa.id,
    )
    db_session.add(lote)
    await db_session.commit()
    await db_session.refresh(lote)
    updated_at_antes = lote.updated_at

    original = LoteComprobantesService.obtener_seguimiento_lote
    consultas = 0

    async def contar_consulta(
        self: LoteComprobantesService,
        lote_id: int,
        empresa_id: int,
    ) -> LoteComprobanteSeguimientoResponse:
        nonlocal consultas
        consultas += 1
        return await original(self, lote_id, empresa_id)

    monkeypatch.setattr(
        LoteComprobantesService,
        "obtener_seguimiento_lote",
        contar_consulta,
    )

    response = await client.get(
        f"/api/lotes-comprobantes/{lote.id}/seguimiento",
        headers=auth_headers,
    )

    assert response.status_code == 200, response.text
    data = response.json()
    assert consultas == 1
    assert data["id"] == lote.id
    assert data["estado"] == "en_cola"
    assert data["modo_procesamiento"] == "background"
    assert set(data) == {
        "id",
        "estado",
        "modo_procesamiento",
        "procesamiento_async",
        "total_filas",
        "total_grupos",
        "grupos_validos",
        "grupos_con_error",
        "grupos_emitidos",
        "grupos_fallidos",
        "grupos_reconciliados_externos",
        "grupos_descartados",
        "mensaje_resumen",
        "started_at",
        "finished_at",
        "updated_at",
    }
    for forbidden in (
        "archivo_hash",
        "metadata_json",
        "mapeo_usado_json",
        "headers_detectados_json",
        "empresa_id",
        "usuario_id",
        "formato_importacion_id",
    ):
        assert forbidden not in data
    await db_session.refresh(lote)
    assert lote.updated_at == updated_at_antes


@pytest.mark.asyncio
async def test_seguimiento_lote_respeta_scope_del_emisor(
    client: AsyncClient,
    auth_headers: dict,
    db_session: AsyncSession,
    test_empresa,
) -> None:
    """Un usuario común no puede seguir lotes de otro emisor."""
    otra_empresa = Empresa(
        razon_social="Otra Empresa Test S.A.",
        cuit="20999999991",
        condicion_iva="RI",
        domicilio="Av. Prueba 456",
        localidad="Buenos Aires",
        provincia="Buenos Aires",
        codigo_postal="1000",
        inicio_actividades=date(2020, 1, 1),
    )
    db_session.add(otra_empresa)
    await db_session.flush()
    lote_ajeno = LoteComprobante(
        nombre_archivo="lote-ajeno-seguimiento.xlsx",
        archivo_hash="hash-lote-ajeno-seguimiento",
        estado="en_cola",
        total_filas=1,
        total_grupos=1,
        grupos_validos=1,
        empresa_id=otra_empresa.id,
    )
    db_session.add(lote_ajeno)
    await db_session.commit()

    response = await client.get(
        f"/api/lotes-comprobantes/{lote_ajeno.id}/seguimiento",
        headers={**auth_headers, "X-Empresa-Id": str(otra_empresa.id)},
    )

    assert response.status_code == 403
    assert "permiso" in response.json()["detail"]


@pytest.mark.asyncio
async def test_validar_lote_registra_grupos_y_filas(
    client: AsyncClient,
    auth_headers: dict,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    response = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data=_opciones_fechas(),
        files={
            "archivo": (
                "lote.xlsx",
                _build_lote_excel(test_empresa.cuit),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200, response.text
    data = response.json()
    assert data["puede_emitirse"] is True
    assert data["lote"]["estado"] == "validado"
    assert data["lote"]["grupos_validos"] == 1
    assert data["lote"]["grupos_con_error"] == 0

    detalle = await client.get(
        f"/api/lotes-comprobantes/{data['lote']['id']}",
        headers=auth_headers,
    )
    assert detalle.status_code == 200
    detalle_data = detalle.json()
    assert len(detalle_data["grupos"]) == 1
    assert len(detalle_data["filas"]) == 1
    assert detalle_data["grupos"][0]["estado"] == "validado"


@pytest.mark.asyncio
async def test_validar_lote_productos_acepta_fechas_servicio_omitidas(
    client: AsyncClient,
    auth_headers: dict,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """Un lote de productos no debe exigir campos de servicio en multipart."""
    data = _opciones_fechas(concepto_modo="productos")
    for key in [
        "fecha_servicio_desde_modo",
        "fecha_servicio_hasta_modo",
        "fecha_vto_pago_modo",
    ]:
        data.pop(key)

    response = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data=data,
        files={
            "archivo": (
                "lote-productos-sin-servicio.xlsx",
                _build_lote_excel(test_empresa.cuit),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200, response.text
    detalle = await client.get(
        f"/api/lotes-comprobantes/{response.json()['lote']['id']}",
        headers=auth_headers,
    )
    assert detalle.status_code == 200, detalle.text
    grupo = detalle.json()["grupos"][0]
    assert grupo["concepto"] == 1
    assert grupo["fecha_servicio_desde"] is None
    assert grupo["fecha_servicio_hasta"] is None
    assert grupo["fecha_vto_pago"] is None


@pytest.mark.asyncio
async def test_validar_lote_mixto_no_anuncia_emision_si_estado_no_procesable(
    client: AsyncClient,
    auth_headers: dict,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """Un lote con errores conserva contrato consistente: no puede emitirse."""
    workbook = load_workbook(BytesIO(_build_lote_excel_multi_grupo(test_empresa.cuit)))
    sheet = workbook["Comprobantes"]
    sheet.cell(row=3, column=4).value = 11
    sheet.cell(row=3, column=21).value = 21
    stream = BytesIO()
    workbook.save(stream)

    response = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data=_opciones_fechas(),
        files={
            "archivo": (
                "lote-mixto.xlsx",
                stream.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200, response.text
    data = response.json()
    assert data["puede_emitirse"] is False
    assert data["lote"]["estado"] == "con_errores"
    assert data["lote"]["grupos_validos"] == 1
    assert data["lote"]["grupos_con_error"] == 1


@pytest.mark.asyncio
async def test_validar_lote_rechaza_xlsx_malformado(
    client: AsyncClient,
    auth_headers: dict,
):
    """Un .xlsx corrupto debe devolver error funcional, no 500."""
    response = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data=_opciones_fechas(),
        files={
            "archivo": (
                "corrupto.xlsx",
                b"esto no es un zip",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 400
    assert "No se pudo leer el archivo Excel" in response.json()["detail"]


@pytest.mark.asyncio
async def test_validar_lote_rechaza_archivo_demasiado_grande(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
):
    """El límite de bytes debe aplicarse antes de parsear el Excel."""
    monkeypatch.setattr(settings, "batch_max_upload_bytes", 10)

    response = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data=_opciones_fechas(),
        files={
            "archivo": (
                "grande.xlsx",
                b"01234567890",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 400
    assert "tamaño máximo" in response.json()["detail"]


@pytest.mark.asyncio
async def test_validar_lote_punto_venta_fijo_sobrescribe_archivo(
    client: AsyncClient,
    auth_headers: dict,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """Debe permitir fijar un punto de venta habilitado para todo el lote."""
    punto_fijo = PuntoVenta(
        numero=13,
        nombre="Web Services 13",
        activo=True,
        es_webservice=True,
        empresa_id=test_empresa.id,
    )
    db_session.add(punto_fijo)
    await db_session.commit()

    response = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data=_opciones_fechas(
            punto_venta_modo="fijo",
            punto_venta_numero=13,
        ),
        files={
            "archivo": (
                "lote-pv-fijo.xlsx",
                _build_lote_excel(test_empresa.cuit, punto_venta_numero=1),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200, response.text
    detalle = await client.get(
        f"/api/lotes-comprobantes/{response.json()['lote']['id']}",
        headers=auth_headers,
    )
    detalle_data = detalle.json()
    assert detalle_data["grupos"][0]["punto_venta_numero"] == 13
    assert detalle_data["filas"][0]["datos_json"]["punto_venta_numero"] == 13
    assert (
        detalle_data["metadata_json"]["opciones_punto_venta"]["punto_venta_numero"]
        == 13
    )


@pytest.mark.asyncio
async def test_validar_lote_acepta_fecha_fija_argentina(
    client: AsyncClient,
    auth_headers: dict,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """Debe aceptar DD/MM/AAAA en fechas fijas del formulario."""
    response = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data=_opciones_fechas(
            fecha_emision_modo="fija",
            fecha_emision_fija="20/05/2026",
        ),
        files={
            "archivo": (
                "lote-fecha-fija-argentina.xlsx",
                _build_lote_excel(test_empresa.cuit),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200, response.text
    detalle = await client.get(
        f"/api/lotes-comprobantes/{response.json()['lote']['id']}",
        headers=auth_headers,
    )
    assert detalle.json()["grupos"][0]["fecha_emision"] == "2026-05-20"


@pytest.mark.asyncio
async def test_validar_lote_rechaza_fecha_fija_argentina_invalida(
    client: AsyncClient,
    auth_headers: dict,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """Debe rechazar fechas fijas con calendario imposible."""
    response = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data=_opciones_fechas(
            fecha_emision_modo="fija",
            fecha_emision_fija="31/02/2026",
        ),
        files={
            "archivo": (
                "lote-fecha-fija-invalida.xlsx",
                _build_lote_excel(test_empresa.cuit),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 400
    assert "fecha_emision_fija debe ser una fecha válida" in response.json()["detail"]


@pytest.mark.asyncio
async def test_validar_lote_rechaza_punto_venta_fijo_no_habilitado(
    client: AsyncClient,
    auth_headers: dict,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """No debe validar con un punto fijo no usable por el emisor activo."""
    response = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data=_opciones_fechas(
            punto_venta_modo="fijo",
            punto_venta_numero=99,
        ),
        files={
            "archivo": (
                "lote-pv-invalido.xlsx",
                _build_lote_excel(test_empresa.cuit),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 400
    assert "Puntos de venta" in response.json()["detail"]


@pytest.mark.asyncio
async def test_validar_lote_guarda_snapshot_perfil_carga_masiva(
    client: AsyncClient,
    auth_headers: dict,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """Debe conservar el perfil usado aunque luego se edite."""
    test_certificado.ambiente = settings.arca_env
    perfil = await client.post(
        "/api/perfiles-carga-masiva",
        headers=auth_headers,
        json={
            "nombre": "Servicios mensuales",
            "descripcion": "Perfil de prueba",
            "configuracion_json": {
                "version": 1,
                "formato_importacion_version_id": None,
                "concepto_modo": "productos",
                "descripcion_item_modo": "archivo",
                "fecha_emision": {"modo": "archivo"},
                "periodo_servicio": {"modo": "archivo"},
                "fecha_vto_pago": {"modo": "archivo"},
            },
            "es_predeterminado": True,
            "activo": True,
        },
    )
    assert perfil.status_code == 201, perfil.text

    response = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data={
            **_opciones_fechas(),
            "perfil_carga_masiva_id": str(perfil.json()["id"]),
        },
        files={
            "archivo": (
                "lote-con-perfil.xlsx",
                _build_lote_excel(test_empresa.cuit),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200, response.text
    detalle = await client.get(
        f"/api/lotes-comprobantes/{response.json()['lote']['id']}",
        headers=auth_headers,
    )
    metadata = detalle.json()["metadata_json"]
    assert metadata["perfil_carga_masiva"]["id"] == perfil.json()["id"]
    assert metadata["perfil_carga_masiva"]["nombre"] == "Servicios mensuales"
    assert (
        metadata["perfil_carga_masiva"]["configuracion_json"]["concepto_modo"]
        == "productos"
    )


@pytest.mark.asyncio
async def test_validar_lote_rechaza_descripcion_item_faltante(
    client: AsyncClient,
    auth_headers: dict,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """No debe validar un lote sin política de descripción facturada."""
    opciones = _opciones_fechas()
    opciones.pop("descripcion_item_modo")

    response = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data=opciones,
        files={
            "archivo": (
                "lote-sin-descripcion-modo.xlsx",
                _build_lote_excel(test_empresa.cuit),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 422


@pytest.mark.asyncio
async def test_validar_lote_descripcion_item_fija_sobrescribe_archivo(
    client: AsyncClient,
    auth_headers: dict,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """Debe aplicar la descripción fija elegida para todo el lote."""
    response = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data=_opciones_fechas(
            descripcion_item_modo="fija",
            descripcion_item_fija="Honorarios profesionales",
        ),
        files={
            "archivo": (
                "lote-descripcion-fija.xlsx",
                _build_lote_excel(test_empresa.cuit),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200, response.text
    data = response.json()
    detalle = await client.get(
        f"/api/lotes-comprobantes/{data['lote']['id']}",
        headers=auth_headers,
    )
    fila = detalle.json()["filas"][0]
    assert fila["datos_json"]["item_descripcion"] == "Honorarios profesionales"


@pytest.mark.asyncio
async def test_validar_lote_rechaza_empresa_distinta(
    client: AsyncClient,
    auth_headers: dict,
    test_punto_venta,
    test_certificado,
):
    response = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data=_opciones_fechas(),
        files={
            "archivo": (
                "lote-otra-empresa.xlsx",
                _build_lote_excel("30999999999"),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 400
    assert "empresa activa" in response.json()["detail"]


@pytest.mark.asyncio
async def test_validar_lote_rechaza_iva_invalido(
    client: AsyncClient,
    auth_headers: dict,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    response = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data=_opciones_fechas(),
        files={
            "archivo": (
                "lote-iva-invalido.xlsx",
                _build_lote_excel(test_empresa.cuit, iva=22),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200, response.text
    data = response.json()
    assert data["puede_emitirse"] is False
    assert data["lote"]["estado"] == "con_errores"
    assert data["lote"]["grupos_con_error"] == 1

    detalle = await client.get(
        f"/api/lotes-comprobantes/{data['lote']['id']}",
        headers=auth_headers,
    )
    mensajes = detalle.json()["grupos"][0]["mensajes_json"]
    assert any("alícuota de IVA" in mensaje for mensaje in mensajes)


@pytest.mark.asyncio
async def test_validar_lote_rechaza_factura_c_con_iva(
    client: AsyncClient,
    auth_headers: dict,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """Factura C no puede quedar lista si el archivo trae IVA."""
    test_empresa.condicion_iva = "Exento"
    await db_session.commit()

    response = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data=_opciones_fechas(),
        files={
            "archivo": (
                "lote-factura-c-con-iva.xlsx",
                _build_lote_excel(
                    test_empresa.cuit,
                    tipo_comprobante=11,
                    iva=21,
                    cliente_tipo_documento="",
                    cliente_numero_documento="",
                    cliente_razon_social="A CONSUMIDOR FINAL",
                    cliente_condicion_iva="Consumidor Final",
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200, response.text
    data = response.json()
    assert data["puede_emitirse"] is False
    assert data["lote"]["estado"] == "con_errores"

    detalle = await client.get(
        f"/api/lotes-comprobantes/{data['lote']['id']}",
        headers=auth_headers,
    )
    mensajes = detalle.json()["grupos"][0]["mensajes_json"]
    assert any("tipo C no pueden incluir IVA" in mensaje for mensaje in mensajes)


@pytest.mark.asyncio
async def test_validar_lote_consumidor_final_sin_documento_bajo_umbral(
    client: AsyncClient,
    auth_headers: dict,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    response = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data=_opciones_fechas(),
        files={
            "archivo": (
                "lote-cf-sin-documento.xlsx",
                _build_lote_excel(
                    test_empresa.cuit,
                    cliente_tipo_documento="",
                    cliente_numero_documento="",
                    cliente_razon_social="",
                    cliente_condicion_iva="Consumidor Final",
                    item_precio_unitario=1000,
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200, response.text
    data = response.json()
    assert data["puede_emitirse"] is True

    detalle = await client.get(
        f"/api/lotes-comprobantes/{data['lote']['id']}",
        headers=auth_headers,
    )
    grupo = detalle.json()["grupos"][0]
    assert grupo["estado"] == "validado"
    assert grupo["cliente_documento"] == "0"
    assert grupo["cliente_razon_social"] == "A CONSUMIDOR FINAL"


@pytest.mark.asyncio
async def test_validar_lote_nota_credito_requiere_comprobante_asociado(
    client: AsyncClient,
    auth_headers: dict,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """Una nota de crédito no puede quedar lista sin comprobante asociado."""
    test_empresa.condicion_iva = "Exento"
    await db_session.commit()

    response = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data=_opciones_fechas(concepto_modo="servicios"),
        files={
            "archivo": (
                "lote-nc-sin-asociado.xlsx",
                _build_lote_excel(
                    test_empresa.cuit,
                    tipo_comprobante=13,
                    concepto=2,
                    iva=0,
                    cliente_tipo_documento="",
                    cliente_numero_documento="",
                    cliente_razon_social="A CONSUMIDOR FINAL",
                    cliente_condicion_iva="Consumidor Final",
                    fecha_servicio_desde=FECHA_FISCAL_CONTROLADA_PF19B,
                    fecha_servicio_hasta=FECHA_FISCAL_CONTROLADA_PF19B,
                    fecha_vto_pago=FECHA_FISCAL_CONTROLADA_PF19B,
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200, response.text
    data = response.json()
    assert data["puede_emitirse"] is False
    assert data["lote"]["grupos_con_error"] == 1

    detalle = await client.get(
        f"/api/lotes-comprobantes/{data['lote']['id']}",
        headers=auth_headers,
    )
    mensajes = detalle.json()["grupos"][0]["mensajes_json"]
    assert any("requiere comprobante asociado" in mensaje for mensaje in mensajes)


@pytest.mark.asyncio
async def test_validar_lote_nota_debito_requiere_comprobante_asociado(
    client: AsyncClient,
    auth_headers: dict,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """Una nota de débito no puede quedar lista sin comprobante asociado."""
    test_empresa.condicion_iva = "Exento"
    await db_session.commit()

    response = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data=_opciones_fechas(concepto_modo="servicios"),
        files={
            "archivo": (
                "lote-nd-sin-asociado.xlsx",
                _build_lote_excel(
                    test_empresa.cuit,
                    tipo_comprobante=12,
                    concepto=2,
                    iva=0,
                    cliente_tipo_documento="",
                    cliente_numero_documento="",
                    cliente_razon_social="A CONSUMIDOR FINAL",
                    cliente_condicion_iva="Consumidor Final",
                    fecha_servicio_desde=date(2026, 5, 31),
                    fecha_servicio_hasta=date(2026, 5, 31),
                    fecha_vto_pago=date(2026, 5, 31),
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200, response.text
    data = response.json()
    assert data["puede_emitirse"] is False
    assert data["lote"]["grupos_con_error"] == 1

    detalle = await client.get(
        f"/api/lotes-comprobantes/{data['lote']['id']}",
        headers=auth_headers,
    )
    mensajes = detalle.json()["grupos"][0]["mensajes_json"]
    assert any("requiere comprobante asociado" in mensaje for mensaje in mensajes)


@pytest.mark.asyncio
async def test_validar_lote_nota_credito_guarda_comprobante_asociado_en_payload(
    client: AsyncClient,
    auth_headers: dict,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """El lote debe persistir el asociado que luego se informa como CbtesAsoc."""
    test_empresa.condicion_iva = "Exento"
    await db_session.commit()

    response = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data=_opciones_fechas(concepto_modo="servicios"),
        files={
            "archivo": (
                "lote-nc-con-asociado.xlsx",
                _build_lote_excel(
                    test_empresa.cuit,
                    tipo_comprobante=13,
                    concepto=2,
                    iva=0,
                    cliente_tipo_documento="",
                    cliente_numero_documento="",
                    cliente_razon_social="A CONSUMIDOR FINAL",
                    cliente_condicion_iva="Consumidor Final",
                    fecha_servicio_desde=FECHA_FISCAL_CONTROLADA_PF19B,
                    fecha_servicio_hasta=FECHA_FISCAL_CONTROLADA_PF19B,
                    fecha_vto_pago=FECHA_FISCAL_CONTROLADA_PF19B,
                    asociado_tipo_comprobante=11,
                    asociado_punto_venta=1,
                    asociado_numero=1234,
                    asociado_fecha=date(2026, 4, 30),
                    asociado_cuit=test_empresa.cuit,
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200, response.text
    data = response.json()
    assert data["puede_emitirse"] is True
    stmt = select(LoteComprobanteGrupo).where(
        LoteComprobanteGrupo.lote_id == data["lote"]["id"]
    )
    grupo = (await db_session.execute(stmt)).scalar_one()
    asociado = grupo.payload_json["comprobantes_asociados"][0]
    assert asociado["tipo_comprobante"] == 11
    assert asociado["punto_venta"] == 1
    assert asociado["numero"] == 1234
    assert asociado["fecha"] == "2026-04-30"
    assert asociado["cuit"] == test_empresa.cuit


@pytest.mark.asyncio
async def test_detectar_formato_extracto_bancario(
    client: AsyncClient,
    auth_headers: dict,
    test_empresa,
):
    response = await client.post(
        "/api/formatos-importacion/detectar",
        headers=auth_headers,
        files={
            "archivo": (
                "extracto-bancario.xlsx",
                _build_extracto_bancario_excel(test_empresa.cuit),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200, response.text
    data = response.json()
    assert data["headers_detectados"] == [
        "Fecha",
        "Créditos",
        "Leyendas Adicionales1",
        "Leyendas Adicionales2",
        "Pto Vta",
    ]
    assert data["formato_sugerido_version_id"] is not None
    assert data["candidatos"][0]["confianza"] == "alta"


@pytest.mark.asyncio
async def test_detectar_formato_rechaza_archivo_demasiado_grande(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
):
    """La detección de formatos debe aplicar el mismo límite de upload."""
    monkeypatch.setattr(settings, "batch_max_upload_bytes", 10)

    response = await client.post(
        "/api/formatos-importacion/detectar",
        headers=auth_headers,
        files={
            "archivo": (
                "grande.xlsx",
                b"01234567890",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 400
    assert "tamaño máximo" in response.json()["detail"]


@pytest.mark.asyncio
async def test_crear_formato_rechaza_configuracion_malformada(
    client: AsyncClient,
    auth_headers: dict,
):
    """La configuración de formato debe validarse antes de persistir."""
    response = await client.post(
        "/api/formatos-importacion",
        headers=auth_headers,
        json={
            "nombre": "Formato invalido",
            "descripcion": "No debe persistirse",
            "configuracion_json": {"campos": {"importe_total": None}},
        },
    )

    assert response.status_code == 400
    assert "debe ser un objeto" in response.json()["detail"]


@pytest.mark.asyncio
async def test_validar_lote_formato_cano_factura_b_iva_21(
    client: AsyncClient,
    auth_headers: dict,
    db_session: AsyncSession,
    test_empresa,
    test_user,
    test_punto_venta,
    test_certificado,
):
    """Debe validar el formato Cano como Factura B con IVA 21%."""
    test_empresa.condicion_iva = "RI"
    await _crear_punto_venta_rece_verificado(
        db_session,
        test_empresa,
        usuario_id=int(test_user.id),
        numero=2,
        nombre="Cano PV 2",
        documento_emitido_en=FECHA_DOCUMENTO_RECE_TEST,
        vigente_hasta=FECHA_VIGENCIA_RECE_TEST,
        observado_en=INSTANTE_RECE_TEST,
    )
    await db_session.commit()

    crear = await client.post(
        "/api/formatos-importacion",
        headers=auth_headers,
        json={
            "nombre": "Cano - Factura B IVA 21%",
            "descripcion": (
                "Formato particular para planillas Cano: neto gravado, IVA "
                "discriminado y total de Factura B a consumidor final."
            ),
            "configuracion_json": _config_formato_cano_factura_b(),
        },
    )
    assert crear.status_code == 201, crear.text

    contenido = _build_cano_factura_b_excel(FECHA_FISCAL_CONTROLADA_PF19B)
    detectar = await client.post(
        "/api/formatos-importacion/detectar",
        headers=auth_headers,
        files={
            "archivo": (
                "cano-factura-b.xlsx",
                contenido,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert detectar.status_code == 200, detectar.text
    formato_version_id = detectar.json()["formato_sugerido_version_id"]
    assert formato_version_id == crear.json()["version_vigente"]["id"]

    response = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data={
            **_opciones_fechas(
                concepto_modo="productos",
                descripcion_item_modo="fija",
                descripcion_item_fija="Venta mostrador",
            ),
            "formato_version_id": str(formato_version_id),
        },
        files={
            "archivo": (
                "cano-factura-b.xlsx",
                contenido,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200, response.text
    data = response.json()
    assert data["puede_emitirse"] is True
    assert data["lote"]["grupos_validos"] == 1

    detalle = await client.get(
        f"/api/lotes-comprobantes/{data['lote']['id']}",
        headers=auth_headers,
    )
    grupo = detalle.json()["grupos"][0]
    fila = detalle.json()["filas"][0]["datos_json"]
    assert grupo["tipo_comprobante"] == 6
    assert grupo["punto_venta_numero"] == 2
    assert grupo["cliente_documento"] == "0"
    assert grupo["total_estimado"] == "90000.00"
    assert fila["item_precio_unitario"] == "74380.1652892562"
    assert fila["item_iva_porcentaje"] == 21


@pytest.mark.asyncio
async def test_validar_lote_formato_cano_bloquea_total_usado_como_neto(
    client: AsyncClient,
    auth_headers: dict,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """Debe bloquear un formato que recalcula IVA sobre un total ya final."""
    test_empresa.condicion_iva = "RI"
    db_session.add(
        PuntoVenta(
            numero=2,
            nombre="Cano PV 2",
            activo=True,
            es_webservice=True,
            empresa_id=test_empresa.id,
        )
    )
    await db_session.commit()

    configuracion = _config_formato_cano_factura_b()
    configuracion["campos"]["item_precio_unitario"] = {
        "origen": "header",
        "encabezados": ["Imp. Total"],
        "transformacion": "decimal",
        "requerido": True,
    }
    crear = await client.post(
        "/api/formatos-importacion",
        headers=auth_headers,
        json={
            "nombre": "Cano - Formato erroneo total como neto",
            "descripcion": "Config de prueba que no debe quedar emitible.",
            "configuracion_json": configuracion,
        },
    )
    assert crear.status_code == 201, crear.text

    contenido = _build_cano_factura_b_excel(FECHA_FISCAL_CONTROLADA_PF19B)
    response = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data={
            **_opciones_fechas(
                concepto_modo="productos",
                descripcion_item_modo="fija",
                descripcion_item_fija="Venta mostrador",
            ),
            "formato_version_id": str(crear.json()["version_vigente"]["id"]),
        },
        files={
            "archivo": (
                "cano-total-como-neto.xlsx",
                contenido,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200, response.text
    data = response.json()
    assert data["puede_emitirse"] is False
    assert data["lote"]["grupos_con_error"] == 1

    detalle = await client.get(
        f"/api/lotes-comprobantes/{data['lote']['id']}",
        headers=auth_headers,
    )
    mensajes = detalle.json()["grupos"][0]["mensajes_json"]
    assert any("no coincide con el total informado" in mensaje for mensaje in mensajes)


@pytest.mark.asyncio
async def test_validar_lote_extracto_bancario_varios_puntos_venta(
    client: AsyncClient,
    auth_headers: dict,
    db_session: AsyncSession,
    test_empresa,
    test_user,
    test_punto_venta,
    test_certificado,
):
    test_empresa.condicion_iva = "Exento"
    for numero in [10, 13]:
        await _crear_punto_venta_rece_verificado(
            db_session,
            test_empresa,
            usuario_id=int(test_user.id),
            numero=numero,
            nombre=f"Punto {numero}",
            documento_emitido_en=FECHA_DOCUMENTO_RECE_TEST,
            vigente_hasta=FECHA_VIGENCIA_RECE_TEST,
            observado_en=INSTANTE_RECE_TEST,
        )
    await db_session.commit()
    contenido = _build_extracto_bancario_excel(
        test_empresa.cuit,
        fecha_movimiento=FECHA_FISCAL_CONTROLADA_PF19B,
    )
    detectar = await client.post(
        "/api/formatos-importacion/detectar",
        headers=auth_headers,
        files={
            "archivo": (
                "extracto-bancario-multi-pv.xlsx",
                contenido,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    formato_version_id = detectar.json()["formato_sugerido_version_id"]

    response = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data={
            **_opciones_fechas(
                concepto_modo="servicios",
                descripcion_item_modo="fija",
                descripcion_item_fija="Honorarios",
            ),
            "formato_version_id": str(formato_version_id),
        },
        files={
            "archivo": (
                "extracto-bancario-multi-pv.xlsx",
                contenido,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200, response.text
    data = response.json()
    assert data["puede_emitirse"] is True
    assert data["lote"]["grupos_validos"] == 3
    assert data["lote"]["formato_importacion_version_id"] is not None

    detalle = await client.get(
        f"/api/lotes-comprobantes/{data['lote']['id']}",
        headers=auth_headers,
    )
    grupos = detalle.json()["grupos"]
    assert [grupo["punto_venta_numero"] for grupo in grupos] == [1, 10, 13]
    assert [grupo["cliente_documento"] for grupo in grupos] == ["0", "0", "0"]
    assert [grupo["total_estimado"] for grupo in grupos] == [
        "59500.00",
        "70500.00",
        "140000.00",
    ]
    assert [grupo["concepto"] for grupo in grupos] == [2, 2, 2]
    assert grupos[0]["fecha_emision"] == FECHA_FISCAL_CONTROLADA_PF19B.isoformat()


@pytest.mark.asyncio
async def test_validar_lote_formato_con_header_blanco_preserva_indices(
    client: AsyncClient,
    auth_headers: dict,
    db_session: AsyncSession,
    test_empresa,
    test_user,
    test_punto_venta,
    test_certificado,
):
    """Los headers vacíos no deben desplazar los índices físicos del Excel."""
    test_empresa.condicion_iva = "Exento"
    for numero in [10, 13]:
        await _crear_punto_venta_rece_verificado(
            db_session,
            test_empresa,
            usuario_id=int(test_user.id),
            numero=numero,
            nombre=f"Punto {numero}",
            documento_emitido_en=FECHA_DOCUMENTO_RECE_TEST,
            vigente_hasta=FECHA_VIGENCIA_RECE_TEST,
            observado_en=INSTANTE_RECE_TEST,
        )
    await db_session.commit()
    workbook = load_workbook(
        BytesIO(
            _build_extracto_bancario_excel(
                test_empresa.cuit,
                fecha_movimiento=FECHA_FISCAL_CONTROLADA_PF19B,
            )
        )
    )
    workbook.active.insert_cols(1)
    stream = BytesIO()
    workbook.save(stream)
    contenido = stream.getvalue()

    detectar = await client.post(
        "/api/formatos-importacion/detectar",
        headers=auth_headers,
        files={
            "archivo": (
                "extracto-header-blanco.xlsx",
                contenido,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert detectar.status_code == 200, detectar.text
    assert detectar.json()["headers_detectados"][0] == ""
    formato_version_id = detectar.json()["formato_sugerido_version_id"]

    response = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data={
            **_opciones_fechas(
                concepto_modo="servicios",
                descripcion_item_modo="fija",
                descripcion_item_fija="Honorarios",
            ),
            "formato_version_id": str(formato_version_id),
        },
        files={
            "archivo": (
                "extracto-header-blanco.xlsx",
                contenido,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200, response.text
    data = response.json()
    assert data["puede_emitirse"] is True

    detalle = await client.get(
        f"/api/lotes-comprobantes/{data['lote']['id']}",
        headers=auth_headers,
    )
    grupos = detalle.json()["grupos"]
    assert [grupo["punto_venta_numero"] for grupo in grupos] == [1, 10, 13]
    assert [grupo["total_estimado"] for grupo in grupos] == [
        "59500.00",
        "70500.00",
        "140000.00",
    ]


@pytest.mark.asyncio
async def test_validar_lote_rechaza_fecha_emision_fuera_de_ventana_arca(
    client: AsyncClient,
    auth_headers: dict,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """Debe observar el lote si la fecha del archivo no puede usarse en ARCA."""
    test_empresa.condicion_iva = "Exento"
    for numero in [10, 13]:
        db_session.add(
            PuntoVenta(
                numero=numero,
                nombre=f"Punto {numero}",
                activo=True,
                es_webservice=True,
                empresa_id=test_empresa.id,
            )
        )
    await db_session.commit()
    contenido = _build_extracto_bancario_excel(
        test_empresa.cuit,
        fecha_movimiento=FECHA_FISCAL_CONTROLADA_PF19B - timedelta(days=20),
        fecha_como_serial=True,
    )
    detectar = await client.post(
        "/api/formatos-importacion/detectar",
        headers=auth_headers,
        files={
            "archivo": (
                "extracto-fecha-vieja.xlsx",
                contenido,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    formato_version_id = detectar.json()["formato_sugerido_version_id"]

    response = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data={
            **_opciones_fechas(
                concepto_modo="servicios",
                descripcion_item_modo="fija",
                descripcion_item_fija="Honorarios",
            ),
            "formato_version_id": str(formato_version_id),
        },
        files={
            "archivo": (
                "extracto-fecha-vieja.xlsx",
                contenido,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200, response.text
    data = response.json()
    assert data["puede_emitirse"] is False
    assert data["lote"]["grupos_con_error"] == 3

    detalle = await client.get(
        f"/api/lotes-comprobantes/{data['lote']['id']}",
        headers=auth_headers,
    )
    mensajes = detalle.json()["grupos"][0]["mensajes_json"]
    assert any("ventana ARCA" in mensaje for mensaje in mensajes)


@pytest.mark.asyncio
async def test_validar_lote_concepto_definido_por_archivo(
    client: AsyncClient,
    auth_headers: dict,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """Debe aceptar Producto/Servicio del Excel cuando se elige archivo."""
    response = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data=_opciones_fechas(concepto_modo="archivo"),
        files={
            "archivo": (
                "lote-concepto-archivo.xlsx",
                _build_lote_excel(test_empresa.cuit, concepto="Producto"),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200, response.text
    data = response.json()
    assert data["puede_emitirse"] is True

    detalle = await client.get(
        f"/api/lotes-comprobantes/{data['lote']['id']}",
        headers=auth_headers,
    )
    grupo = detalle.json()["grupos"][0]
    assert grupo["concepto"] == 1


@pytest.mark.asyncio
async def test_validar_lote_rechaza_concepto_archivo_sin_columna(
    client: AsyncClient,
    auth_headers: dict,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """Debe rechazar 'Definido por archivo' si el formato no trae columna."""
    test_empresa.condicion_iva = "Exento"
    contenido = _build_extracto_bancario_excel(test_empresa.cuit)
    detectar = await client.post(
        "/api/formatos-importacion/detectar",
        headers=auth_headers,
        files={
            "archivo": (
                "extracto-sin-concepto.xlsx",
                contenido,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    formato_version_id = detectar.json()["formato_sugerido_version_id"]

    response = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data={
            **_opciones_fechas(
                concepto_modo="archivo",
                descripcion_item_modo="fija",
                descripcion_item_fija="Honorarios",
            ),
            "formato_version_id": str(formato_version_id),
        },
        files={
            "archivo": (
                "extracto-sin-concepto.xlsx",
                contenido,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 400
    assert "columna de concepto fiscal" in response.json()["detail"]


@pytest.mark.asyncio
async def test_validar_lote_rechaza_descripcion_archivo_sin_columna(
    client: AsyncClient,
    auth_headers: dict,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """Debe rechazar descripción desde archivo si el formato no la mapea."""
    test_empresa.condicion_iva = "Exento"
    contenido = _build_extracto_bancario_excel(test_empresa.cuit)
    detectar = await client.post(
        "/api/formatos-importacion/detectar",
        headers=auth_headers,
        files={
            "archivo": (
                "extracto-sin-descripcion.xlsx",
                contenido,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    formato_version_id = detectar.json()["formato_sugerido_version_id"]

    response = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data={
            **_opciones_fechas(
                concepto_modo="servicios",
                descripcion_item_modo="archivo",
            ),
            "formato_version_id": str(formato_version_id),
        },
        files={
            "archivo": (
                "extracto-sin-descripcion.xlsx",
                contenido,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 400
    assert "descripción facturada" in response.json()["detail"]


@pytest.mark.asyncio
async def test_validar_lote_extracto_bancario_exige_confirmar_formato(
    client: AsyncClient,
    auth_headers: dict,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    response = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data=_opciones_fechas(),
        files={
            "archivo": (
                "extracto-sin-formato.xlsx",
                _build_extracto_bancario_excel(test_empresa.cuit),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 400
    assert "formato de importación" in response.json()["detail"]


@pytest.mark.asyncio
async def test_validar_lote_extracto_bancario_rechaza_factura_c_para_ri(
    client: AsyncClient,
    auth_headers: dict,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    for numero in [10, 13]:
        db_session.add(
            PuntoVenta(
                numero=numero,
                nombre=f"Punto {numero}",
                activo=True,
                es_webservice=True,
                empresa_id=test_empresa.id,
            )
        )
    await db_session.commit()
    contenido = _build_extracto_bancario_excel(test_empresa.cuit)
    detectar = await client.post(
        "/api/formatos-importacion/detectar",
        headers=auth_headers,
        files={
            "archivo": (
                "extracto-ri.xlsx",
                contenido,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    formato_version_id = detectar.json()["formato_sugerido_version_id"]

    response = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data={
            **_opciones_fechas(
                concepto_modo="servicios",
                descripcion_item_modo="fija",
                descripcion_item_fija="Honorarios",
            ),
            "formato_version_id": str(formato_version_id),
        },
        files={
            "archivo": (
                "extracto-ri.xlsx",
                contenido,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200, response.text
    data = response.json()
    assert data["puede_emitirse"] is False
    assert data["lote"]["grupos_con_error"] == 3

    detalle = await client.get(
        f"/api/lotes-comprobantes/{data['lote']['id']}",
        headers=auth_headers,
    )
    mensajes = detalle.json()["grupos"][0]["mensajes_json"]
    assert any("Responsable Inscripto" in mensaje for mensaje in mensajes)


@pytest.mark.asyncio
async def test_validar_lote_consumidor_final_sin_documento_sobre_umbral(
    client: AsyncClient,
    auth_headers: dict,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    response = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data=_opciones_fechas(),
        files={
            "archivo": (
                "lote-cf-sin-documento-alto.xlsx",
                _build_lote_excel(
                    test_empresa.cuit,
                    cliente_tipo_documento="",
                    cliente_numero_documento="",
                    cliente_razon_social="",
                    cliente_condicion_iva="Consumidor Final",
                    item_precio_unitario=10_000_000,
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200, response.text
    data = response.json()
    assert data["puede_emitirse"] is False

    detalle = await client.get(
        f"/api/lotes-comprobantes/{data['lote']['id']}",
        headers=auth_headers,
    )
    mensajes = detalle.json()["grupos"][0]["mensajes_json"]
    assert any("$10.000.000" in mensaje for mensaje in mensajes)


@pytest.mark.asyncio
async def test_validar_lote_rechaza_archivo_duplicado(
    client: AsyncClient,
    auth_headers: dict,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    contenido = _build_lote_excel(test_empresa.cuit)
    files = {
        "archivo": (
            "lote-duplicado.xlsx",
            contenido,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }

    primera = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data=_opciones_fechas(),
        files=files,
    )
    assert primera.status_code == 200, primera.text

    segunda = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data=_opciones_fechas(),
        files={
            "archivo": (
                "lote-duplicado.xlsx",
                contenido,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert segunda.status_code == 400
    assert "ya fue cargado" in segunda.json()["detail"]


@pytest.mark.asyncio
async def test_validar_lote_permite_reintentar_fallido_sin_emitidos(
    client: AsyncClient,
    auth_headers: dict,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """Un lote fallido sin CAE emitidos puede revalidarse con el mismo archivo."""
    contenido = _build_lote_excel(test_empresa.cuit)
    files = {
        "archivo": (
            "lote-reintento.xlsx",
            contenido,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }

    primera = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data=_opciones_fechas(),
        files=files,
    )
    assert primera.status_code == 200, primera.text
    lote_id = primera.json()["lote"]["id"]

    lote_previo = await db_session.get(LoteComprobante, lote_id)
    lote_previo.estado = "fallido"
    lote_previo.grupos_validos = 0
    lote_previo.grupos_fallidos = 1
    lote_previo.grupos_emitidos = 0
    await db_session.commit()

    segunda = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data=_opciones_fechas(),
        files={
            "archivo": (
                "lote-reintento.xlsx",
                contenido,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert segunda.status_code == 200, segunda.text
    assert segunda.json()["lote"]["id"] != lote_id
    await db_session.refresh(lote_previo)
    assert lote_previo.metadata_json["reemplazado_por_reintento"][
        "archivo_hash_original"
    ]


@pytest.mark.asyncio
async def test_procesar_lote_sync_actualiza_resultados(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    test_certificado.ambiente = settings.arca_env
    llamadas = 0

    async def fake_emitir(self, request, **kwargs):
        nonlocal llamadas
        llamadas += 1
        comprobante_id = await _persistir_comprobante_autorizado(
            db_session,
            test_empresa,
            test_punto_venta,
            tipo_comprobante=request.tipo_comprobante,
            numero=456,
            fecha_emision=request.fecha_emision,
            cae=CAE_TEST_NO_REAL,
            cae_vencimiento=date(2026, 3, 31),
            total=Decimal("1210.00"),
        )
        return EmitirComprobanteResponse(
            exito=True,
            comprobante_id=comprobante_id,
            tipo_comprobante=request.tipo_comprobante,
            punto_venta=1,
            numero=456,
            fecha=request.fecha_emision,
            cae=CAE_TEST_NO_REAL,
            cae_vencimiento=date(2026, 3, 31),
            total=Decimal("1210.00"),
            mensaje="Comprobante autorizado",
            errores=[],
        )

    monkeypatch.setattr(
        "app.services.facturacion_service.FacturacionService.emitir_comprobante",
        fake_emitir,
    )

    validar = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data=_opciones_fechas(),
        files={
            "archivo": (
                "lote-procesar.xlsx",
                _build_lote_excel(test_empresa.cuit),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert validar.status_code == 200, validar.text
    lote_id = validar.json()["lote"]["id"]
    headers_procesar = await _confirmacion_fecha_fiscal_header_lote(
        db_session,
        lote_id=lote_id,
        estados={"validado"},
    )

    procesar = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/procesar",
        headers={**auth_headers, **headers_procesar},
    )

    assert procesar.status_code == 200, procesar.text
    data = procesar.json()
    assert data["en_progreso"] is False
    assert data["lote"]["estado"] == "completado"
    assert data["lote"]["grupos_emitidos"] == 1
    assert data["lote"]["grupos_fallidos"] == 0
    assert llamadas == 1

    replay = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/procesar",
        headers={**auth_headers, **headers_procesar},
    )

    assert replay.status_code == 200, replay.text
    replay_data = replay.json()
    assert replay_data["en_progreso"] is False
    assert replay_data["lote"]["estado"] == "completado"
    assert replay_data["lote"]["grupos_emitidos"] == 1
    assert llamadas == 1

    detalle = await client.get(
        f"/api/lotes-comprobantes/{lote_id}/resultados",
        headers=auth_headers,
    )
    assert detalle.status_code == 200
    grupo = detalle.json()["grupos"][0]
    assert grupo["estado"] == "autorizado"
    assert grupo["cae"] == CAE_TEST_NO_REAL


@pytest.mark.asyncio
async def test_procesar_lote_sanitiza_payload_con_clave_desconocida(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
) -> None:
    """El procesamiento no debe exponer el valor de un payload no canónico."""
    lote_id = await _crear_lote_validado_por_api(
        client,
        auth_headers,
        test_empresa.cuit,
        nombre_archivo="lote-payload-no-canonico.xlsx",
    )
    grupos = await _marcar_grupos_lote(db_session, lote_id, ["validado"])
    grupo = grupos[0]
    valor_no_publicable = "dato-sintetico-no-publicable"
    payload = dict(grupo.payload_json or {})
    payload["instruccion_fiscal_desconocida"] = valor_no_publicable
    grupo.payload_json = payload
    await db_session.commit()
    llamadas_emision = 0

    async def fail_emitir(self, request, **kwargs):
        nonlocal llamadas_emision
        llamadas_emision += 1
        raise AssertionError("No debe emitir un payload fiscal no canónico")

    monkeypatch.setattr(FacturacionService, "emitir_comprobante", fail_emitir)

    headers_procesar = await _confirmacion_fecha_fiscal_header_lote(
        db_session,
        lote_id=lote_id,
        estados={"validado"},
        idempotency_key="idem-lote-payload-no-canonico",
    )

    response = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/procesar",
        headers={**auth_headers, **headers_procesar},
    )

    assert response.status_code == 200, response.text
    assert llamadas_emision == 0
    await db_session.refresh(grupo)
    assert grupo.estado == "fallido"
    assert grupo.mensajes_json == [
        "El payload fiscal guardado no cumple el contrato vigente. "
        "No se solicitó CAE; revisá el lote antes de reintentar."
    ]
    assert valor_no_publicable not in response.text
    assert valor_no_publicable not in str(grupo.mensajes_json)
    assert grupo.numero_asignado is None
    assert grupo.cae is None
    assert grupo.comprobante_id is None


@pytest.mark.asyncio
async def test_procesar_lote_background_encola_lote_chico(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """Permite iniciar un lote chico en segundo plano para observar progreso."""
    test_certificado.ambiente = settings.arca_env
    monkeypatch.setattr(
        "app.api.lotes_comprobantes.ensure_lote_worker_running",
        lambda app: True,
    )
    llamadas = 0

    async def fake_emitir(self, request, **kwargs):
        nonlocal llamadas
        llamadas += 1
        numero = 500 + llamadas
        comprobante_id = await _persistir_comprobante_autorizado(
            db_session,
            test_empresa,
            test_punto_venta,
            tipo_comprobante=request.tipo_comprobante,
            numero=numero,
            fecha_emision=request.fecha_emision,
            cae=CAE_TEST_NO_REAL,
            cae_vencimiento=date(2026, 3, 31),
            total=Decimal("1210.00"),
        )
        return EmitirComprobanteResponse(
            exito=True,
            comprobante_id=comprobante_id,
            tipo_comprobante=request.tipo_comprobante,
            punto_venta=1,
            numero=numero,
            fecha=request.fecha_emision,
            cae=CAE_TEST_NO_REAL,
            cae_vencimiento=date(2026, 3, 31),
            total=Decimal("1210.00"),
            mensaje="Comprobante autorizado",
            errores=[],
        )

    monkeypatch.setattr(
        "app.services.facturacion_service.FacturacionService.emitir_comprobante",
        fake_emitir,
    )
    validar = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data=_opciones_fechas(),
        files={
            "archivo": (
                "lote-background-chico.xlsx",
                _build_lote_excel(test_empresa.cuit),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert validar.status_code == 200, validar.text
    lote_id = validar.json()["lote"]["id"]
    confirmacion = await _confirmacion_fecha_fiscal_header_lote(
        db_session,
        lote_id=lote_id,
        estados={"validado"},
    )

    procesar = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/procesar?background=true",
        headers={**auth_headers, **confirmacion},
    )

    assert procesar.status_code == 200, procesar.text
    data = procesar.json()
    assert data["en_progreso"] is True
    assert data["lote"]["estado"] == "en_cola"
    assert data["lote"]["modo_procesamiento"] == "background"

    operacion = (
        (
            await db_session.execute(
                select(OperacionIdempotente).where(
                    OperacionIdempotente.idempotency_key == "idem-lote-test"
                )
            )
        )
        .scalars()
        .one()
    )
    assert operacion.estado == "en_proceso"
    assert operacion.response_json["en_progreso"] is True

    async with AsyncSession(bind=db_session.bind, expire_on_commit=False) as observador:
        lote_publicado = await observador.get(LoteComprobante, lote_id)
        operacion_publicada = await observador.get(
            OperacionIdempotente,
            operacion.id,
        )
        assert lote_publicado.estado == "en_cola"
        assert lote_publicado.metadata_json["operacion_idempotente_id"] == operacion.id
        assert operacion_publicada.estado == "en_proceso"
        assert operacion_publicada.response_json["en_progreso"] is True
        assert operacion_publicada.response_json["lote"]["id"] == lote_id

    service = LoteComprobantesService(db_session)
    lote = await service.procesar_lote(lote_id, test_empresa.id, reanudar=True)
    await db_session.refresh(operacion)

    assert lote.estado == "completado"
    assert lote.procesamiento_async is True
    assert lote.modo_procesamiento == "background"
    assert llamadas == 1
    assert operacion.estado == "finalizado"
    assert operacion.response_json["en_progreso"] is False
    assert operacion.response_json["lote"]["estado"] == "completado"


@pytest.mark.asyncio
async def test_procesar_background_encolado_durable_no_reabre_operacion(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
) -> None:
    """Una falla al publicar ownership revierte también el encolado."""
    monkeypatch.setattr(
        "app.api.lotes_comprobantes.ensure_lote_worker_running",
        lambda app: True,
    )

    async def fail_guardar_respuesta(self, **kwargs):
        raise SQLAlchemyTimeoutError()

    monkeypatch.setattr(
        IdempotenciaFiscalService,
        "guardar_respuesta_operacion_cas",
        fail_guardar_respuesta,
    )
    lote_id = await _crear_lote_validado_por_api(
        client,
        auth_headers,
        test_empresa.cuit,
        nombre_archivo="lote-background-respuesta-db.xlsx",
    )
    confirmacion = await _confirmacion_fecha_fiscal_header_lote(
        db_session,
        lote_id=lote_id,
        estados={"validado"},
    )

    response = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/procesar?background=true",
        headers={**auth_headers, **confirmacion},
    )

    assert response.status_code == 503, response.text
    assert response.headers["Retry-After"] == "2"
    async with AsyncSession(bind=db_session.bind, expire_on_commit=False) as observador:
        lote = await observador.get(LoteComprobante, lote_id)
        assert lote is not None
        assert lote.estado == "validado"
        assert lote.procesamiento_async is False
        assert lote.metadata_json.get("operacion_idempotente_id") is None
        operacion = await observador.scalar(
            select(OperacionIdempotente).where(
                OperacionIdempotente.idempotency_key == "idem-lote-test"
            )
        )
        assert operacion is not None
        assert operacion.estado == "interrumpida_pre_arca"
        assert operacion.response_json is None
        intentos = await observador.scalars(
            select(IntentoEmisionFiscal).where(
                IntentoEmisionFiscal.operacion_id == operacion.id
            )
        )
        assert intentos.all() == []
        publicable_al_worker = await observador.scalar(
            select(LoteComprobante.id).where(
                LoteComprobante.id == lote_id,
                LoteComprobante.estado == "en_cola",
            )
        )
        assert publicable_al_worker is None


@pytest.mark.asyncio
@pytest.mark.parametrize(
    "perdida_ownership",
    ["progreso_adulterado", "terminal", "cas_updated_at"],
)
async def test_publicacion_background_revierte_lote_si_pierde_ownership(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
    perdida_ownership: str,
) -> None:
    """La publicación worker es atómica ante ownership adulterado o perdido."""
    monkeypatch.setattr(
        "app.api.lotes_comprobantes.ensure_lote_worker_running",
        lambda app: True,
    )
    lote_id = await _crear_lote_validado_por_api(
        client,
        auth_headers,
        test_empresa.cuit,
        nombre_archivo=f"lote-publicacion-{perdida_ownership}.xlsx",
    )
    idempotency_key = f"idem-publicacion-{perdida_ownership}"
    headers = await _confirmacion_fecha_fiscal_header_lote(
        db_session,
        lote_id=lote_id,
        estados={"validado"},
        idempotency_key=idempotency_key,
    )
    encolado = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/procesar?background=true",
        headers={**auth_headers, **headers},
    )
    assert encolado.status_code == 200, encolado.text
    lote = await db_session.get(LoteComprobante, lote_id)
    operacion = await db_session.scalar(
        select(OperacionIdempotente).where(
            OperacionIdempotente.idempotency_key == idempotency_key
        )
    )
    assert lote is not None
    assert operacion is not None
    operacion_id = int(operacion.id)

    if perdida_ownership == "progreso_adulterado":
        respuesta_adulterada = deepcopy(operacion.response_json)
        respuesta_adulterada["lote"]["metadata_json"]["pf19b_rece_material"][
            "grupos_hash"
        ] = ("f" * 64)
        operacion.response_json = respuesta_adulterada
        await db_session.commit()
    elif perdida_ownership == "terminal":
        respuesta_terminal = deepcopy(operacion.response_json)
        respuesta_terminal["en_progreso"] = False
        respuesta_terminal["lote"]["estado"] = "completado"
        operacion.estado = "finalizado"
        operacion.response_json = respuesta_terminal
        await db_session.commit()

    lote.estado = "completado"
    lote.finished_at = datetime.utcnow()
    if perdida_ownership == "cas_updated_at":
        execute_original = db_session.execute

        async def perder_cas_updated_at(statement, *args, **kwargs):
            """Simula rowcount cero tras perder el CAS de publicación."""
            if (
                getattr(statement, "is_update", False)
                and getattr(getattr(statement, "table", None), "name", None)
                == "operaciones_idempotentes"
            ):
                return SimpleNamespace(rowcount=0)
            return await execute_original(statement, *args, **kwargs)

        monkeypatch.setattr(db_session, "execute", perder_cas_updated_at)

    with pytest.raises(LoteComprobanteConflictoError):
        await LoteComprobantesService(
            db_session
        )._guardar_respuesta_operacion_background(lote, operacion_id)

    async with AsyncSession(bind=db_session.bind, expire_on_commit=False) as observador:
        lote_visible = await observador.get(LoteComprobante, lote_id)
        operacion_visible = await observador.get(OperacionIdempotente, operacion_id)
    assert lote_visible is not None
    assert operacion_visible is not None
    assert lote_visible.estado == "en_cola"
    if perdida_ownership == "terminal":
        assert operacion_visible.estado == "finalizado"
        assert operacion_visible.response_json["en_progreso"] is False
    else:
        assert operacion_visible.estado == "en_proceso"
        if perdida_ownership == "progreso_adulterado":
            assert (
                operacion_visible.response_json["lote"]["metadata_json"][
                    "pf19b_rece_material"
                ]["grupos_hash"]
                == "f" * 64
            )
        else:
            assert operacion_visible.response_json["en_progreso"] is True


@pytest.mark.asyncio
async def test_publicacion_background_confirma_reconciliacion_sin_reabrir(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
) -> None:
    """El worker reemplaza su progress solo por el terminal recon del mismo lote."""
    monkeypatch.setattr(
        "app.api.lotes_comprobantes.ensure_lote_worker_running",
        lambda app: True,
    )
    lote_id = await _crear_lote_validado_por_api(
        client,
        auth_headers,
        test_empresa.cuit,
        nombre_archivo="lote-publicacion-reconciliacion.xlsx",
    )
    idempotency_key = "idem-publicacion-reconciliacion"
    headers = await _confirmacion_fecha_fiscal_header_lote(
        db_session,
        lote_id=lote_id,
        estados={"validado"},
        idempotency_key=idempotency_key,
    )
    encolado = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/procesar?background=true",
        headers={**auth_headers, **headers},
    )
    assert encolado.status_code == 200, encolado.text
    lote = await db_session.get(LoteComprobante, lote_id)
    operacion = await db_session.scalar(
        select(OperacionIdempotente).where(
            OperacionIdempotente.idempotency_key == idempotency_key
        )
    )
    assert lote is not None
    assert operacion is not None
    operacion_id = int(operacion.id)
    assert operacion.estado == "en_proceso"
    assert operacion.response_json["en_progreso"] is True

    operacion.estado = "requiere_reconciliacion"
    lote.estado = "requiere_reconciliacion"
    lote.finished_at = datetime.utcnow()
    lote.mensaje_resumen = "El lote requiere reconciliación fiscal."
    await db_session.commit()
    await db_session.refresh(lote)
    await LoteComprobantesService(db_session)._guardar_respuesta_operacion_background(
        lote, operacion_id
    )
    await db_session.commit()

    async with AsyncSession(bind=db_session.bind, expire_on_commit=False) as observador:
        lote_visible = await observador.get(LoteComprobante, lote_id)
        operacion_visible = await observador.get(
            OperacionIdempotente,
            operacion_id,
        )
    assert lote_visible is not None
    assert lote_visible.estado == "requiere_reconciliacion"
    assert operacion_visible is not None
    assert operacion_visible.estado == "requiere_reconciliacion"
    assert operacion_visible.response_json["en_progreso"] is False
    assert (
        operacion_visible.response_json["lote"]["estado"] == "requiere_reconciliacion"
    )


@pytest.mark.asyncio
async def test_procesar_lote_background_sin_worker_no_encola(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """Sin worker disponible no debe mutar el lote ni crear idempotencia."""
    test_certificado.ambiente = settings.arca_env
    monkeypatch.setattr(settings, "batch_worker_enabled", False)

    validar = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data=_opciones_fechas(),
        files={
            "archivo": (
                "lote-worker-no-disponible.xlsx",
                _build_lote_excel(test_empresa.cuit),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert validar.status_code == 200, validar.text
    lote_id = validar.json()["lote"]["id"]
    headers_procesar = await _confirmacion_fecha_fiscal_header_lote(
        db_session,
        lote_id=lote_id,
        estados={"validado"},
    )

    procesar = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/procesar?background=true",
        headers={**auth_headers, **headers_procesar},
    )

    assert procesar.status_code == 503, procesar.text
    detail = procesar.json()["detail"]
    assert detail["categoria_error"] == "worker_lotes_no_disponible"
    assert "No se encoló el lote" in detail["mensaje"]

    lote = await db_session.get(LoteComprobante, lote_id)
    await db_session.refresh(lote)
    assert lote.estado == "validado"
    assert lote.procesamiento_async is False
    operacion = (
        await db_session.execute(
            select(OperacionIdempotente).where(
                OperacionIdempotente.idempotency_key == "idem-lote-test"
            )
        )
    ).scalar_one_or_none()
    assert operacion is None


@pytest.mark.asyncio
@pytest.mark.parametrize(
    ("estado_operacion", "respuesta_es_error"),
    [
        pytest.param("finalizado", False, id="finalizado"),
        pytest.param("fallido", True, id="fallido-legacy"),
    ],
)
async def test_replay_terminal_background_no_depende_del_worker(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
    estado_operacion: str,
    respuesta_es_error: bool,
) -> None:
    """Un resultado terminal legacy se reproduce antes del gate del worker."""
    lote_id = await _crear_lote_validado_por_api(
        client,
        auth_headers,
        test_empresa.cuit,
        nombre_archivo="lote-replay-terminal-worker-caido.xlsx",
    )
    idempotency_key = "idem-replay-terminal-worker-caido"
    headers = await _confirmacion_fecha_fiscal_header_lote(
        db_session,
        lote_id=lote_id,
        estados={"validado"},
        idempotency_key=idempotency_key,
    )
    service = LoteComprobantesService(db_session)
    material = await service.calcular_material_idempotente_grupos(
        lote_id=lote_id,
        empresa_id=test_empresa.id,
        estados={
            "validado",
            "procesando",
            "autorizado",
            "fallido",
            "requiere_reconciliacion",
        },
    )
    payload = {
        "lote_id": lote_id,
        "background": True,
        "confirmacion_fecha_fiscal": headers["X-Confirmacion-Fecha-Fiscal"],
        "grupo_ids": material["grupo_ids"],
        "grupos_hash": material["grupos_hash"],
    }
    lote = await db_session.get(LoteComprobante, lote_id)
    assert lote is not None
    if respuesta_es_error:
        respuesta_terminal = {
            "mensaje": "Fallo terminal legacy ya confirmado.",
            "errores": ["El procesamiento ya terminó con un error conocido."],
            "categoria_error": "lote_fallido_legacy",
            "status_code": 409,
        }
    else:
        respuesta_terminal = {
            "lote": LoteComprobanteResponse.model_validate(lote).model_dump(
                mode="json"
            ),
            "mensaje": "Resultado terminal sintético ya confirmado.",
            "en_progreso": False,
        }
    operacion = OperacionIdempotente(
        empresa_id=test_empresa.id,
        idempotency_key=idempotency_key,
        tipo_operacion="procesar_lote",
        payload_hash=IdempotenciaFiscalService.calcular_payload_hash(payload),
        lote_id=lote_id,
        estado=estado_operacion,
        response_json=respuesta_terminal,
    )
    db_session.add(operacion)
    await db_session.commit()

    async def fail_resolver(*args, **kwargs):
        """El replay terminal no debe volver a resolver RECE ni crear operación."""
        raise AssertionError("No debe resolver una operación terminal nuevamente")

    async def fail_procesar(*args, **kwargs):
        """El replay terminal no debe entrar al servicio de emisión."""
        raise AssertionError("No debe procesar un lote con respuesta terminal")

    monkeypatch.setattr(
        "app.api.lotes_comprobantes.ensure_lote_worker_running",
        lambda app: False,
    )
    monkeypatch.setattr(
        "app.api.lotes_comprobantes._resolver_operacion_lote",
        fail_resolver,
    )
    monkeypatch.setattr(LoteComprobantesService, "procesar_lote", fail_procesar)

    response = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/procesar?background=true",
        headers={**auth_headers, **headers},
    )

    if respuesta_es_error:
        assert response.status_code == 409, response.text
        assert response.json()["detail"] == respuesta_terminal
    else:
        assert response.status_code == 200, response.text
        assert response.json() == LoteProcesamientoResponse.model_validate(
            respuesta_terminal
        ).model_dump(mode="json")
    assert await db_session.scalar(select(func.count(IntentoEmisionFiscal.id))) == 0
    assert (
        await db_session.scalar(select(func.count(PuntoVentaGuardaEmisionRece.id))) == 0
    )
    await db_session.refresh(operacion)
    assert operacion.estado == estado_operacion
    assert operacion.response_json == respuesta_terminal


@pytest.mark.asyncio
async def test_procesar_lote_actualiza_contadores_parciales(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """Debe persistir avance real entre grupos durante la emisión."""
    test_certificado.ambiente = settings.arca_env
    validar = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data=_opciones_fechas(),
        files={
            "archivo": (
                "lote-progreso-parcial.xlsx",
                _build_lote_excel_multi_grupo(test_empresa.cuit),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert validar.status_code == 200, validar.text
    lote_id = validar.json()["lote"]["id"]
    llamadas = 0
    avance_observado = None

    async def fake_emitir(self, request, **kwargs):
        nonlocal llamadas, avance_observado
        llamadas += 1
        if llamadas == 2:
            result = await db_session.execute(
                select(LoteComprobante).where(LoteComprobante.id == lote_id)
            )
            lote = result.scalar_one()
            avance_observado = (
                lote.grupos_emitidos,
                lote.grupos_fallidos,
                lote.grupos_validos,
                lote.mensaje_resumen,
            )
        numero = 200 + llamadas
        cae = f"{CAE_TEST_NO_REAL_PREFIX}{llamadas}"
        comprobante_id = await _persistir_comprobante_autorizado(
            db_session,
            test_empresa,
            test_punto_venta,
            tipo_comprobante=request.tipo_comprobante,
            numero=numero,
            fecha_emision=request.fecha_emision,
            cae=cae,
            cae_vencimiento=date(2026, 3, 31),
            total=Decimal("1210.00"),
        )
        return EmitirComprobanteResponse(
            exito=True,
            comprobante_id=comprobante_id,
            tipo_comprobante=request.tipo_comprobante,
            punto_venta=1,
            numero=numero,
            fecha=request.fecha_emision,
            cae=cae,
            cae_vencimiento=date(2026, 3, 31),
            total=Decimal("1210.00"),
            mensaje="Comprobante autorizado",
            errores=[],
        )

    monkeypatch.setattr(
        "app.services.facturacion_service.FacturacionService.emitir_comprobante",
        fake_emitir,
    )
    headers_procesar = await _confirmacion_fecha_fiscal_header_lote(
        db_session,
        lote_id=lote_id,
        estados={"validado"},
    )

    procesar = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/procesar",
        headers={**auth_headers, **headers_procesar},
    )

    assert procesar.status_code == 200, procesar.text
    assert avance_observado == (
        1,
        0,
        1,
        "Procesando comprobante 1 de 2...",
    )
    data = procesar.json()
    assert data["lote"]["estado"] == "completado"
    assert data["lote"]["grupos_emitidos"] == 2


@pytest.mark.asyncio
async def test_procesar_lote_usa_sublotes_arca_segun_regxreq(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """Un lote elegible se divide en sublotes según RegXReq."""
    test_certificado.ambiente = settings.arca_env
    monkeypatch.setattr(settings, "arca_fecaesolicitar_batch_enabled", True)
    llamadas_batch: list[int] = []
    numero = 0

    async def fake_regxreq(self, empresa_id):
        return 2

    async def fake_emitir_lote(
        self,
        requests,
        max_registros=None,
        contextos=None,
        fase_solicitud_arca=None,
        commit_rechazo_global=True,
    ):
        nonlocal numero
        assert fase_solicitud_arca.iniciada is False
        llamadas_batch.append(len(requests))
        respuestas = []
        for request in requests:
            numero += 1
            cae = f"{CAE_TEST_NO_REAL_PREFIX}{numero}"
            comprobante_id = await _persistir_comprobante_autorizado(
                db_session,
                test_empresa,
                test_punto_venta,
                tipo_comprobante=request.tipo_comprobante,
                numero=numero,
                fecha_emision=request.fecha_emision,
                cae=cae,
                cae_vencimiento=date(2026, 3, 31),
                total=Decimal("1210.00"),
            )
            respuestas.append(
                EmitirComprobanteResponse(
                    exito=True,
                    comprobante_id=comprobante_id,
                    tipo_comprobante=request.tipo_comprobante,
                    punto_venta=1,
                    numero=numero,
                    fecha=request.fecha_emision,
                    cae=cae,
                    cae_vencimiento=date(2026, 3, 31),
                    total=Decimal("1210.00"),
                    mensaje="Comprobante autorizado",
                    errores=[],
                )
            )
        return respuestas

    async def fail_emitir_unitario(self, request, **kwargs):
        raise AssertionError("No debe usar emisión unitaria en sublotes de tamaño 2")

    monkeypatch.setattr(
        "app.services.facturacion_service.FacturacionService.obtener_registros_maximos_por_request",
        fake_regxreq,
    )
    monkeypatch.setattr(
        "app.services.facturacion_service.FacturacionService.emitir_comprobantes_lote",
        fake_emitir_lote,
    )
    monkeypatch.setattr(
        "app.services.facturacion_service.FacturacionService.emitir_comprobante",
        fail_emitir_unitario,
    )

    validar = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data=_opciones_fechas(),
        files={
            "archivo": (
                "lote-batch-regxreq.xlsx",
                _build_lote_excel_multi_grupo(test_empresa.cuit, total_grupos=4),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert validar.status_code == 200, validar.text
    lote_id = validar.json()["lote"]["id"]
    headers_procesar = await _confirmacion_fecha_fiscal_header_lote(
        db_session,
        lote_id=lote_id,
        estados={"validado"},
    )

    procesar = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/procesar",
        headers={**auth_headers, **headers_procesar},
    )

    assert procesar.status_code == 200, procesar.text
    data = procesar.json()
    assert llamadas_batch == [2, 2]
    assert data["lote"]["estado"] == "completado"
    assert data["lote"]["metadata_json"]["arca_batch"]["reg_x_req"] == 2
    assert data["lote"]["metadata_json"]["arca_batch"]["chunk_size"] == 2
    assert data["lote"]["metadata_json"]["arca_batch"]["modo"] == "batch"


@pytest.mark.asyncio
async def test_procesar_lote_10005_cierra_sublote_y_aborta_remanentes_sin_replay_arca(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
) -> None:
    """Todos los enviados conservan 10005 y los demás quedan no enviados."""
    test_certificado.ambiente = settings.arca_env
    monkeypatch.setattr(settings, "arca_fecaesolicitar_batch_enabled", True)
    llamadas_fecae = 0

    class FakeWSFEClient:
        """Expone capacidad dos y rechaza globalmente el primer sublote."""

        def __init__(self, *args, **kwargs) -> None:
            """Acepta la firma productiva sin abrir red."""

        async def fe_comp_tot_x_request(self):
            return 2

        async def fe_comp_ultimo_autorizado(self, punto, tipo):
            return 0

        async def fe_cae_solicitar_lote(self, arca_requests):
            nonlocal llamadas_fecae
            llamadas_fecae += 1
            primero = arca_requests[0]
            raise ArcaErrorGlobalEstructurado(
                cabecera=CabeceraRespuestaFecae(
                    cuit=int(test_empresa.cuit),
                    punto_venta=primero.punto_venta,
                    tipo_comprobante=primero.tipo_cbte,
                    cantidad=len(arca_requests),
                    resultado="R",
                ),
                errores=(MensajeArcaEstructurado(10005, "mensaje privado ARCA"),),
                eventos=(),
                detalles_presentes=False,
                senales_cae_presentes=False,
                request_cuit=int(test_empresa.cuit),
                request_punto_venta=primero.punto_venta,
                request_tipo_comprobante=primero.tipo_cbte,
                request_cantidad=len(arca_requests),
                request_rangos=tuple(
                    (request.cbte_desde, request.cbte_hasta)
                    for request in arca_requests
                ),
            )

    async def fake_ticket(self, empresa, certificado):
        return SimpleNamespace(token="token", sign="sign")

    async def fake_validar_punto(self, wsfe_client, numero):
        return None

    monkeypatch.setattr(
        "app.services.facturacion_service.WSFEv1Client",
        FakeWSFEClient,
    )
    monkeypatch.setattr(FacturacionService, "_obtener_ticket_acceso", fake_ticket)
    monkeypatch.setattr(
        FacturacionService,
        "_validar_punto_venta_habilitado",
        fake_validar_punto,
    )

    lote_id = await _crear_lote_validado_por_api(
        client,
        auth_headers,
        test_empresa.cuit,
        nombre_archivo="lote-pf19c-10005.xlsx",
        total_grupos=4,
    )
    headers = await _confirmacion_fecha_fiscal_header_lote(
        db_session,
        lote_id=lote_id,
        estados={"validado"},
        idempotency_key="idem-lote-pf19c-10005",
    )
    primera = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/procesar",
        headers={**auth_headers, **headers},
    )
    assert primera.status_code == 200, primera.text
    assert llamadas_fecae == 1
    errores_arca_esperados = [
        {
            "codigo": 10005,
            "alcance": "global",
            "mensaje": "El punto de venta no está dado de alta como RECE en ARCA.",
        }
    ]
    assert primera.json()["errores_arca"] == errores_arca_esperados
    assert (
        primera.json()["lote"]["metadata_json"]["pf19c_rechazo_global"]["errores_arca"]
        == errores_arca_esperados
    )

    db_session.expire_all()
    grupos = list(
        (
            await db_session.scalars(
                select(LoteComprobanteGrupo)
                .where(LoteComprobanteGrupo.lote_id == lote_id)
                .order_by(LoteComprobanteGrupo.orden)
            )
        ).all()
    )
    intentos = list(
        (
            await db_session.scalars(
                select(IntentoEmisionFiscal)
                .where(IntentoEmisionFiscal.lote_id == lote_id)
                .order_by(IntentoEmisionFiscal.id)
            )
        ).all()
    )
    assert len(intentos) == 2
    assert {intento.grupo_id for intento in intentos} == {
        grupos[0].id,
        grupos[1].id,
    }
    assert {intento.estado for intento in intentos} == {"rechazado_arca"}
    assert all(
        intento.errores_arca_json == errores_arca_esperados for intento in intentos
    )
    assert all(grupo.estado == "fallido" for grupo in grupos), [
        grupo.estado for grupo in grupos
    ]
    assert all(
        any(
            "no_enviado_por_rechazo_global" in mensaje
            for mensaje in grupos[indice].mensajes_json
        )
        for indice in (2, 3)
    )
    operacion = await db_session.scalar(
        select(OperacionIdempotente).where(
            OperacionIdempotente.idempotency_key == "idem-lote-pf19c-10005"
        )
    )
    assert operacion is not None
    assert operacion.estado == "rechazado_arca"
    assert (
        primera.json()["lote"]["metadata_json"]["pf19c_rechazo_global"]["operacion_id"]
        == operacion.id
    )
    assert operacion.response_json == primera.json()

    segunda = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/procesar",
        headers={**auth_headers, **headers},
    )
    assert segunda.status_code == 200, segunda.text
    assert segunda.json() == primera.json()
    assert llamadas_fecae == 1

    respuesta_original = deepcopy(operacion.response_json)
    respuesta_adulterada = deepcopy(respuesta_original)
    respuesta_adulterada["errores_arca"][0]["codigo"] = 10006
    operacion.response_json = respuesta_adulterada
    await db_session.commit()
    desconocida = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/procesar",
        headers={**auth_headers, **headers},
    )
    assert desconocida.status_code == 409, desconocida.text
    assert llamadas_fecae == 1

    respuesta_sin_evidencia = deepcopy(respuesta_original)
    respuesta_sin_evidencia["errores_arca"] = []
    respuesta_sin_evidencia["lote"]["metadata_json"].pop("pf19c_rechazo_global")
    operacion.response_json = respuesta_sin_evidencia
    await db_session.commit()
    reatestiguada = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/procesar",
        headers={**auth_headers, **headers},
    )
    assert reatestiguada.status_code == 409, reatestiguada.text
    assert llamadas_fecae == 1


@pytest.mark.asyncio
async def test_reintento_exitoso_no_publica_10005_historico_y_replay_a_es_exacto(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
) -> None:
    """El retry B revalida RECE sin heredar el rechazo durable de A."""
    test_certificado.ambiente = settings.arca_env
    monkeypatch.setattr(settings, "arca_fecaesolicitar_batch_enabled", False)

    class FakeWSFEClient:
        """Rechaza A con 10005 y autoriza B sin abrir red."""

        solicitudes_fecae = 0

        def __init__(self, *args, **kwargs) -> None:
            """Acepta la firma productiva."""

        async def fe_comp_ultimo_autorizado(self, punto, tipo):
            """Mantiene disponible el mismo número tras el rechazo."""
            return 0

        async def fe_cae_solicitar(self, arca_request):
            """Expone el rechazo inicial y la autorización posterior."""
            FakeWSFEClient.solicitudes_fecae += 1
            if FakeWSFEClient.solicitudes_fecae == 1:
                raise ArcaErrorGlobalEstructurado(
                    cabecera=CabeceraRespuestaFecae(
                        cuit=int(test_empresa.cuit),
                        punto_venta=arca_request.punto_venta,
                        tipo_comprobante=arca_request.tipo_cbte,
                        cantidad=1,
                        resultado="R",
                    ),
                    errores=(MensajeArcaEstructurado(10005, "mensaje privado ARCA"),),
                    eventos=(),
                    detalles_presentes=False,
                    senales_cae_presentes=False,
                    request_cuit=int(test_empresa.cuit),
                    request_punto_venta=arca_request.punto_venta,
                    request_tipo_comprobante=arca_request.tipo_cbte,
                    request_cantidad=1,
                    request_rangos=(
                        (arca_request.cbte_desde, arca_request.cbte_hasta),
                    ),
                )
            return CAEResponse(
                cae=CAE_TEST_NO_REAL_ALT,
                cae_vencimiento="20260831",
                numero_comprobante=arca_request.cbte_desde,
                tipo_cbte=arca_request.tipo_cbte,
                punto_venta=arca_request.punto_venta,
                resultado="A",
            )

    async def fake_ticket(self, empresa, certificado):
        return SimpleNamespace(token="token", sign="sign")

    async def fake_validar_punto(self, wsfe_client, numero):
        return None

    monkeypatch.setattr(
        "app.services.facturacion_service.WSFEv1Client",
        FakeWSFEClient,
    )
    monkeypatch.setattr(FacturacionService, "_obtener_ticket_acceso", fake_ticket)
    monkeypatch.setattr(
        FacturacionService,
        "_validar_punto_venta_habilitado",
        fake_validar_punto,
    )

    lote_id = await _crear_lote_validado_por_api(
        client,
        auth_headers,
        test_empresa.cuit,
        nombre_archivo="lote-pf19c-owner-historico.xlsx",
    )
    headers_a = await _confirmacion_fecha_fiscal_header_lote(
        db_session,
        lote_id=lote_id,
        estados={"validado"},
        idempotency_key="idem-pf19c-owner-a",
    )
    respuesta_a = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/procesar",
        headers={**auth_headers, **headers_a},
    )
    assert respuesta_a.status_code == 200, respuesta_a.text
    assert respuesta_a.json()["errores_arca"][0]["codigo"] == 10005

    db_session.expire_all()
    [grupo] = list(
        (
            await db_session.scalars(
                select(LoteComprobanteGrupo).where(
                    LoteComprobanteGrupo.lote_id == lote_id
                )
            )
        ).all()
    )
    operacion_a = await db_session.scalar(
        select(OperacionIdempotente).where(
            OperacionIdempotente.idempotency_key == "idem-pf19c-owner-a"
        )
    )
    assert operacion_a is not None
    assert operacion_a.estado == "rechazado_arca"
    assert grupo.estado == "fallido", grupo.estado
    operacion_a_id = int(operacion_a.id)
    assert (
        respuesta_a.json()["lote"]["metadata_json"]["pf19c_rechazo_global"][
            "operacion_id"
        ]
        == operacion_a_id
    )

    headers_b = await _confirmacion_fecha_fiscal_header_lote(
        db_session,
        lote_id=lote_id,
        estados={"fallido"},
        grupo_ids=[int(grupo.id)],
        idempotency_key="idem-pf19c-owner-b",
    )
    respuesta_b = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/reintentar-fallidos",
        headers={**auth_headers, **headers_b},
        json={"grupo_ids": [int(grupo.id)]},
    )
    assert respuesta_b.status_code == 200, respuesta_b.text
    assert respuesta_b.json()["errores_arca"] == []
    assert respuesta_b.json()["lote"]["estado"] == "completado"
    assert (
        respuesta_b.json()["lote"]["metadata_json"]["pf19c_rechazo_global"][
            "operacion_id"
        ]
        == operacion_a_id
    )

    operacion_b = await db_session.scalar(
        select(OperacionIdempotente).where(
            OperacionIdempotente.idempotency_key == "idem-pf19c-owner-b"
        )
    )
    assert operacion_b is not None
    assert operacion_b.estado == "finalizado"
    assert operacion_b.response_json == respuesta_b.json()
    intentos = list(
        (
            await db_session.scalars(
                select(IntentoEmisionFiscal)
                .where(IntentoEmisionFiscal.lote_id == lote_id)
                .order_by(IntentoEmisionFiscal.id)
            )
        ).all()
    )
    assert [(intento.operacion_id, intento.estado) for intento in intentos] == [
        (operacion_a_id, "rechazado_arca"),
        (int(operacion_b.id), "autorizado"),
    ]
    assert FakeWSFEClient.solicitudes_fecae == 2

    replay_a = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/procesar",
        headers={**auth_headers, **headers_a},
    )
    assert replay_a.status_code == 200, replay_a.text
    assert replay_a.json() == respuesta_a.json()
    assert FakeWSFEClient.solicitudes_fecae == 2

    operacion_a_sin_respuesta = await db_session.get(
        OperacionIdempotente,
        operacion_a_id,
        populate_existing=True,
    )
    assert operacion_a_sin_respuesta is not None
    operacion_a_sin_respuesta.response_json = None
    await db_session.commit()
    reconstruccion_cruzada = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/procesar",
        headers={**auth_headers, **headers_a},
    )
    assert reconstruccion_cruzada.status_code == 409, reconstruccion_cruzada.text
    assert FakeWSFEClient.solicitudes_fecae == 2


@pytest.mark.parametrize(
    "error_raw",
    [
        {
            "codigo": 10006,
            "alcance": "global",
            "mensaje": "ARCA informó un error global para el requerimiento.",
        },
        {
            "codigo": "10005",
            "alcance": "global",
            "mensaje": "El punto de venta no está dado de alta como RECE en ARCA.",
        },
        {
            "codigo": 10005,
            "alcance": "global",
            "mensaje": "mensaje privado ARCA",
        },
    ],
    ids=["desconocido", "coaccionable", "mensaje-no-canonico"],
)
def test_metadata_lote_desconocida_no_publica_rechazo_global_10005(
    error_raw: dict,
) -> None:
    """Metadata no canónica nunca se reatestigua como evidencia terminal 10005."""
    metadata = {
        "pf19c_rechazo_global": {
            "operacion_id": 17,
            "categoria": "arca_rechazo_global_excluyente",
            "grupos_rechazo_ids": [1],
            "grupos_no_enviados_ids": [2],
            "errores_arca": [error_raw],
        }
    }

    assert (
        LoteComprobantesService.errores_arca_publicables_desde_metadata(
            metadata,
            operacion_id=17,
        )
        == []
    )


@pytest.mark.parametrize(
    ("owner_marker", "owner_actual"),
    [
        pytest.param(17, 18, id="operacion-distinta"),
        pytest.param(True, 1, id="marker-bool"),
        pytest.param(17, True, id="owner-actual-bool"),
        pytest.param("17", 17, id="marker-string"),
    ],
)
def test_metadata_lote_10005_exige_owner_entero_exacto(
    owner_marker: object,
    owner_actual: object,
) -> None:
    """Un marker canónico no prueba el rechazo de otra operación."""
    metadata = {
        "pf19c_rechazo_global": {
            "operacion_id": owner_marker,
            "categoria": "arca_rechazo_global_excluyente",
            "grupos_rechazo_ids": [1],
            "grupos_no_enviados_ids": [],
            "errores_arca": [
                {
                    "codigo": 10005,
                    "alcance": "global",
                    "mensaje": (
                        "El punto de venta no está dado de alta como RECE en ARCA."
                    ),
                }
            ],
        }
    }

    assert (
        LoteComprobantesService.errores_arca_publicables_desde_metadata(
            metadata,
            operacion_id=owner_actual,
        )
        == []
    )


@pytest.mark.parametrize("estado", ["finalizado", "fallido", "fallido_verificado"])
def test_replay_lote_sin_10005_conserva_terminal_historico(estado: str) -> None:
    """Terminales históricos sin evidencia global siguen siendo publicables."""
    respuesta = SimpleNamespace(
        lote=SimpleNamespace(
            id=9,
            empresa_id=1,
            estado="fallido",
            metadata_json={},
        ),
        errores_arca=[],
    )

    assert LoteComprobantesService.respuesta_lote_coincide_operacion(
        respuesta,
        estado_operacion=estado,
        operacion_id=17,
        lote_id=9,
        empresa_id=1,
    )


def test_ownership_worker_acepta_legacy_o_errores_vacios_pero_no_evidencia() -> None:
    """El progress worker admite el default nuevo sin aceptar un 10005 terminal."""
    grupo = {
        "grupo_id": 1,
        "empresa_id": 1,
        "punto_venta_id": 1,
        "punto_venta_numero": 41,
        "ambiente": "produccion",
        "elegibilidad_revision_id": 1,
        "punto_venta_revision_fiscal": 1,
        "tipo_comprobante": 6,
        "payload_hash": "a" * 64,
    }
    material = {
        "grupo_ids": [1],
        "grupos_hash": hashlib.sha256(
            json.dumps([grupo], sort_keys=True, separators=(",", ":")).encode("utf-8")
        ).hexdigest(),
        "grupos": [grupo],
    }
    respuesta_legacy = {
        "lote": {
            "id": 9,
            "empresa_id": 1,
            "estado": "en_cola",
            "modo_procesamiento": "background",
            "procesamiento_async": True,
            "metadata_json": {
                "operacion_idempotente_id": 7,
                "pf19b_rece_material": material,
            },
        },
        "mensaje": "El lote quedó en cola.",
        "en_progreso": True,
    }
    respuesta_nueva = {**respuesta_legacy, "errores_arca": []}
    respuesta_adulterada = {
        **respuesta_legacy,
        "errores_arca": [
            {
                "codigo": 10005,
                "alcance": "global",
                "mensaje": (
                    "El punto de venta no está dado de alta como RECE en ARCA."
                ),
            }
        ],
    }

    for respuesta in (respuesta_legacy, respuesta_nueva):
        assert IdempotenciaFiscalService.respuesta_worker_en_progreso_valida(
            respuesta,
            lote_id=9,
            empresa_id=1,
            operacion_id=7,
            material_rece=material,
        )
    assert not IdempotenciaFiscalService.respuesta_worker_en_progreso_valida(
        respuesta_adulterada,
        lote_id=9,
        empresa_id=1,
        operacion_id=7,
        material_rece=material,
    )


@pytest.mark.asyncio
async def test_procesar_lote_error_inesperado_post_arca_inmoviliza_todo_y_detiene(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
) -> None:
    """Un fallo post-FECAE no publica datos crudos ni inicia otro sublote."""
    test_certificado.ambiente = settings.arca_env
    monkeypatch.setattr(settings, "arca_fecaesolicitar_batch_enabled", True)
    llamadas_batch = 0

    async def fake_regxreq(self, empresa_id):
        return 2

    async def fake_emitir_lote(
        self,
        requests,
        max_registros=None,
        contextos=None,
        fase_solicitud_arca=None,
        commit_rechazo_global=True,
    ):
        nonlocal llamadas_batch
        llamadas_batch += 1
        fase_solicitud_arca.marcar_iniciada()
        raise RuntimeError("detalle privado post ARCA")

    async def fail_emitir_unitario(self, request, **kwargs):
        raise AssertionError("No debe degradar a unitario")

    monkeypatch.setattr(
        FacturacionService,
        "obtener_registros_maximos_por_request",
        fake_regxreq,
    )
    monkeypatch.setattr(
        FacturacionService,
        "emitir_comprobantes_lote",
        fake_emitir_lote,
    )
    monkeypatch.setattr(
        FacturacionService,
        "emitir_comprobante",
        fail_emitir_unitario,
    )
    lote_id = await _crear_lote_validado_por_api(
        client,
        auth_headers,
        test_empresa.cuit,
        nombre_archivo="lote-pf19c-incierto.xlsx",
        total_grupos=4,
    )
    headers = await _confirmacion_fecha_fiscal_header_lote(
        db_session,
        lote_id=lote_id,
        estados={"validado"},
        idempotency_key="idem-lote-pf19c-incierto",
    )
    response = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/procesar",
        headers={**auth_headers, **headers},
    )
    assert response.status_code == 200, response.text
    assert llamadas_batch == 1
    assert "detalle privado" not in response.text
    db_session.expire_all()
    grupos = list(
        (
            await db_session.scalars(
                select(LoteComprobanteGrupo).where(
                    LoteComprobanteGrupo.lote_id == lote_id
                )
            )
        ).all()
    )
    assert len(grupos) == 4
    assert {grupo.estado for grupo in grupos} == {"requiere_reconciliacion"}
    operacion = await db_session.scalar(
        select(OperacionIdempotente).where(
            OperacionIdempotente.idempotency_key == "idem-lote-pf19c-incierto"
        )
    )
    assert operacion is not None
    assert operacion.estado == "requiere_reconciliacion"


@pytest.mark.asyncio
async def test_procesar_lote_fallback_regxreq_degrada_a_unitario_con_aviso(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """Si RegXReq no está disponible, el lote usa modo unitario y avisa."""
    test_certificado.ambiente = settings.arca_env
    monkeypatch.setattr(settings, "arca_fecaesolicitar_batch_enabled", True)

    async def fake_regxreq(self, empresa_id):
        raise RuntimeError("RegXReq no disponible")

    async def fake_emitir(self, request, **kwargs):
        comprobante_id = await _persistir_comprobante_autorizado(
            db_session,
            test_empresa,
            test_punto_venta,
            tipo_comprobante=request.tipo_comprobante,
            numero=456,
            fecha_emision=request.fecha_emision,
            cae=CAE_TEST_NO_REAL,
            cae_vencimiento=date(2026, 3, 31),
            total=Decimal("1210.00"),
        )
        return EmitirComprobanteResponse(
            exito=True,
            comprobante_id=comprobante_id,
            tipo_comprobante=request.tipo_comprobante,
            punto_venta=1,
            numero=456,
            fecha=request.fecha_emision,
            cae=CAE_TEST_NO_REAL,
            cae_vencimiento=date(2026, 3, 31),
            total=Decimal("1210.00"),
            mensaje="Comprobante autorizado",
            errores=[],
        )

    monkeypatch.setattr(
        "app.services.facturacion_service.FacturacionService.obtener_registros_maximos_por_request",
        fake_regxreq,
    )
    monkeypatch.setattr(
        "app.services.facturacion_service.FacturacionService.emitir_comprobante",
        fake_emitir,
    )

    validar = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data=_opciones_fechas(),
        files={
            "archivo": (
                "lote-fallback-regxreq.xlsx",
                _build_lote_excel(test_empresa.cuit),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert validar.status_code == 200, validar.text
    lote_id = validar.json()["lote"]["id"]
    headers_procesar = await _confirmacion_fecha_fiscal_header_lote(
        db_session,
        lote_id=lote_id,
        estados={"validado"},
    )

    procesar = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/procesar",
        headers={**auth_headers, **headers_procesar},
    )

    assert procesar.status_code == 200, procesar.text
    data = procesar.json()
    arca_batch = data["lote"]["metadata_json"]["arca_batch"]
    assert arca_batch["modo"] == "unitario_fallback"
    assert arca_batch["fallback_unitario"] is True
    assert "RegXReq no disponible" in arca_batch["fallback_motivo"]
    assert "modo unitario" in data["lote"]["mensaje_resumen"]


@pytest.mark.asyncio
@pytest.mark.parametrize(
    "error_type",
    [SQLAlchemyTimeoutError, OperationalError],
    ids=["timeout", "operational"],
)
async def test_procesar_lote_db_temporal_pre_arca_devuelve_503_sin_fallar_grupo(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
    error_type: type[Exception],
) -> None:
    """Un lote pre-ARCA queda intacto y la API delega el 503 sanitizado."""
    test_certificado.ambiente = settings.arca_env

    async def fail_emitir(self, request, **kwargs):
        raise _crear_error_db_temporal(error_type)

    monkeypatch.setattr(
        "app.services.facturacion_service.FacturacionService.emitir_comprobante",
        fail_emitir,
    )

    lote_id = await _crear_lote_validado_por_api(
        client,
        auth_headers,
        test_empresa.cuit,
        nombre_archivo="lote-db-temporal-pre-arca.xlsx",
    )
    headers_procesar = await _confirmacion_fecha_fiscal_header_lote(
        db_session,
        lote_id=lote_id,
        estados={"validado"},
    )
    response = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/procesar",
        headers={**auth_headers, **headers_procesar},
    )

    assert response.status_code == 503, response.text
    assert response.headers["Retry-After"] == "2"
    assert "UPDATE lotes_comprobantes" not in response.text
    assert "base temporalmente no disponible" not in response.text
    grupo = await db_session.scalar(
        select(LoteComprobanteGrupo).where(LoteComprobanteGrupo.lote_id == lote_id)
    )
    assert grupo is not None
    assert grupo.estado == "validado"
    assert not any(
        "UPDATE lotes_comprobantes" in mensaje
        for mensaje in (grupo.mensajes_json or [])
    )
    lote = await db_session.get(LoteComprobante, lote_id)
    assert lote is not None
    assert lote.estado == "validado"
    operacion = await db_session.scalar(
        select(OperacionIdempotente).where(
            OperacionIdempotente.idempotency_key == "idem-lote-test"
        )
    )
    assert operacion is not None
    assert operacion.estado == "interrumpida_pre_arca"
    assert operacion.response_json is None
    intentos = (
        (
            await db_session.execute(
                select(IntentoEmisionFiscal).where(
                    IntentoEmisionFiscal.operacion_id == operacion.id
                )
            )
        )
        .scalars()
        .all()
    )
    assert intentos == []


@pytest.mark.asyncio
async def test_procesar_lote_post_arca_db_temporal_devuelve_409_sanitizado(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
) -> None:
    """Una caída DB con FECAE iniciado nunca se presenta como reintentable."""
    test_certificado.ambiente = settings.arca_env

    async def fail_post_arca(self, lote_id, empresa_id, **kwargs):
        kwargs["fase_solicitud_arca"].marcar_iniciada()
        raise _crear_error_db_temporal(OperationalError)

    monkeypatch.setattr(
        LoteComprobantesService,
        "procesar_lote",
        fail_post_arca,
    )
    lote_id = await _crear_lote_validado_por_api(
        client,
        auth_headers,
        test_empresa.cuit,
        nombre_archivo="lote-db-temporal-post-arca.xlsx",
    )
    headers_procesar = await _confirmacion_fecha_fiscal_header_lote(
        db_session,
        lote_id=lote_id,
        estados={"validado"},
    )

    response = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/procesar",
        headers={**auth_headers, **headers_procesar},
    )

    assert response.status_code == 409
    assert "Retry-After" not in response.headers
    detail = response.json()["detail"]
    assert detail["requiere_reconciliacion"] is True
    assert "UPDATE lotes_comprobantes" not in response.text
    assert "base temporalmente no disponible" not in response.text


@pytest.mark.asyncio
async def test_procesar_lote_post_arca_requiere_reconciliacion(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """Un fallo post-ARCA en lote no debe quedar como reintentable."""
    test_certificado.ambiente = settings.arca_env

    async def fake_emitir(self, request, **kwargs):
        operacion_id = int(kwargs["operacion_id"])
        transicion = await self.db.execute(
            update(OperacionIdempotente)
            .where(
                OperacionIdempotente.id == operacion_id,
                OperacionIdempotente.estado == "en_proceso",
                OperacionIdempotente.response_json.is_(None),
            )
            .values(estado="requiere_reconciliacion")
        )
        assert transicion.rowcount == 1
        await self.db.flush()
        return EmitirComprobanteResponse(
            exito=False,
            tipo_comprobante=request.tipo_comprobante,
            punto_venta=1,
            numero=654,
            fecha=request.fecha_emision,
            cae=CAE_TEST_NO_REAL_ALT,
            cae_vencimiento=date(2026, 3, 31),
            total=Decimal("1210.00"),
            mensaje="ARCA autorizó el comprobante, pero FactuFlow no pudo guardarlo",
            errores=["No reintentes esta emisión"],
            requiere_reconciliacion=True,
            categoria_error="post_arca_persistencia",
        )

    monkeypatch.setattr(
        "app.services.facturacion_service.FacturacionService.emitir_comprobante",
        fake_emitir,
    )

    validar = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data=_opciones_fechas(),
        files={
            "archivo": (
                "lote-reconciliacion.xlsx",
                _build_lote_excel(test_empresa.cuit),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert validar.status_code == 200, validar.text
    lote_id = validar.json()["lote"]["id"]
    headers_procesar = await _confirmacion_fecha_fiscal_header_lote(
        db_session,
        lote_id=lote_id,
        estados={"validado"},
    )

    procesar = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/procesar",
        headers={**auth_headers, **headers_procesar},
    )

    assert procesar.status_code == 200, procesar.text
    data = procesar.json()
    assert data["lote"]["estado"] == "requiere_reconciliacion"
    assert data["lote"]["grupos_emitidos"] == 0
    assert data["lote"]["grupos_fallidos"] == 0

    detalle = await client.get(
        f"/api/lotes-comprobantes/{lote_id}/resultados",
        headers=auth_headers,
    )
    grupo = detalle.json()["grupos"][0]
    assert grupo["estado"] == "requiere_reconciliacion"
    assert grupo["cae"] == CAE_TEST_NO_REAL_ALT
    assert grupo["numero_asignado"] == 654

    service = LoteComprobantesService(db_session)
    lote = await service.obtener_lote(lote_id, test_empresa.id)
    assert service._lote_permite_reintento(lote) is False
    operacion = (
        await db_session.execute(
            select(OperacionIdempotente).where(
                OperacionIdempotente.lote_id == lote_id,
                OperacionIdempotente.tipo_operacion == "procesar_lote",
            )
        )
    ).scalar_one()
    assert operacion.estado == "requiere_reconciliacion"
    assert operacion.response_json is not None


@pytest.mark.asyncio
async def test_descartar_grupos_cierra_lote_con_descartes(
    client: AsyncClient,
    auth_headers: dict,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """Descartar pendientes no emitidos debe cerrar un lote parcial."""
    lote_id = await _crear_lote_validado_por_api(
        client,
        auth_headers,
        test_empresa.cuit,
        nombre_archivo="lote-descartar-pendientes.xlsx",
        total_grupos=2,
    )
    grupos = await _marcar_grupos_lote(
        db_session,
        lote_id,
        ["autorizado", "fallido"],
    )

    response = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/descartar-grupos",
        headers=auth_headers,
        json={
            "grupo_ids": [grupos[1].id],
            "motivo": "Emitido manualmente en otro flujo operativo",
        },
    )

    assert response.status_code == 200, response.text
    data = response.json()["lote"]
    assert data["estado"] == "cerrado_con_descartes"
    assert data["grupos_emitidos"] == 1
    assert data["grupos_descartados"] == 1

    grupo_descartado = await db_session.get(LoteComprobanteGrupo, grupos[1].id)
    assert grupo_descartado.estado == "descartado"


@pytest.mark.asyncio
async def test_reintentar_fallidos_exige_confirmacion_fecha_fiscal(
    client: AsyncClient,
    auth_headers: dict,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """El reintento de fallidos también debe confirmar fecha fiscal exacta."""
    lote_id = await _crear_lote_validado_por_api(
        client,
        auth_headers,
        test_empresa.cuit,
        nombre_archivo="lote-reintento-sin-confirmacion.xlsx",
    )
    await _marcar_grupos_lote(db_session, lote_id, ["fallido"])

    response = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/reintentar-fallidos",
        headers={
            **auth_headers,
            "X-Idempotency-Key": "idem-reintento-sin-confirmacion",
        },
        json={"grupo_ids": []},
    )

    assert response.status_code == 400
    detail = response.json()["detail"]
    assert "confirmar la fecha fiscal exacta" in detail["mensaje"]
    assert "0001" in detail["mensaje"]


@pytest.mark.asyncio
async def test_reintentar_fallidos_reclama_grupo_antes_de_emitir(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """El reintento debe sacar el grupo de fallido antes de pedir CAE."""
    lote_id = await _crear_lote_validado_por_api(
        client,
        auth_headers,
        test_empresa.cuit,
        nombre_archivo="lote-reintento-claim.xlsx",
    )
    grupos = await _marcar_grupos_lote(db_session, lote_id, ["fallido"])
    grupo = grupos[0]
    estados_vistos: list[str] = []

    async def fake_emitir_locked(self, request, commit=True, **kwargs):
        estado = await db_session.scalar(
            select(LoteComprobanteGrupo.estado).where(
                LoteComprobanteGrupo.id == grupo.id
            )
        )
        estados_vistos.append(str(estado))
        assert commit is False
        return EmitirComprobanteResponse(
            exito=False,
            tipo_comprobante=request.tipo_comprobante,
            punto_venta=grupo.punto_venta_numero,
            numero=0,
            fecha=request.fecha_emision,
            total=Decimal("1210.00"),
            mensaje="Error controlado",
            errores=["Error controlado"],
        )

    monkeypatch.setattr(
        "app.services.facturacion_service.FacturacionService._emitir_comprobante_locked",
        fake_emitir_locked,
    )
    headers_reintento = await _confirmacion_fecha_fiscal_header_lote(
        db_session,
        lote_id=lote_id,
        estados={"fallido"},
        grupo_ids=[grupo.id],
    )

    response = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/reintentar-fallidos",
        headers={**auth_headers, **headers_reintento},
        json={"grupo_ids": [grupo.id]},
    )

    assert response.status_code == 200, response.text
    assert estados_vistos == ["reintentando"]
    await db_session.refresh(grupo)
    assert grupo.estado == "fallido"


@pytest.mark.asyncio
@pytest.mark.parametrize("adulteracion", ["owner", "material"])
async def test_reintentar_fallidos_revalida_ownership_post_claim_sin_arca(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
    adulteracion: str,
) -> None:
    """Una mutación post-claim bloquea todo I/O fiscal y no habilita recovery ajeno."""
    lote_id = await _crear_lote_validado_por_api(
        client,
        auth_headers,
        test_empresa.cuit,
        nombre_archivo=f"lote-reintento-post-claim-{adulteracion}.xlsx",
    )
    [grupo] = await _marcar_grupos_lote(db_session, lote_id, ["fallido"])
    grupo_id = int(grupo.id)
    headers_reintento = await _confirmacion_fecha_fiscal_header_lote(
        db_session,
        lote_id=lote_id,
        estados={"fallido"},
        grupo_ids=[grupo_id],
        idempotency_key=f"idem-reintento-post-claim-{adulteracion}",
    )
    original_claim = LoteComprobantesService._reclamar_grupo_para_reintento
    operacion_perdedora_id: int | None = None
    owner_adulterado_id: int | None = None

    async def claim_con_adulteracion(self, **kwargs):
        nonlocal operacion_perdedora_id, owner_adulterado_id
        resultado = await original_claim(self, **kwargs)
        if kwargs.get("solo_revalidar") is True or resultado[0] is None:
            return resultado
        operacion_perdedora_id = int(kwargs["operacion_id"])
        lote = await self.db.get(LoteComprobante, lote_id)
        operacion = await self.db.get(
            OperacionIdempotente,
            operacion_perdedora_id,
        )
        assert lote is not None
        assert operacion is not None
        metadata = deepcopy(lote.metadata_json or {})
        if adulteracion == "owner":
            owner_adulterado = OperacionIdempotente(
                empresa_id=operacion.empresa_id,
                usuario_id=operacion.usuario_id,
                idempotency_key="idem-reintento-owner-concurrente",
                tipo_operacion="reintentar_fallidos_lote",
                payload_hash="f" * 64,
                estado="en_proceso",
                lote_id=lote_id,
                rece_snapshot_hash=operacion.rece_snapshot_hash,
            )
            self.db.add(owner_adulterado)
            await self.db.flush()
            owner_adulterado_id = int(owner_adulterado.id)
            metadata["operacion_idempotente_id"] = owner_adulterado_id
        else:
            material = deepcopy(metadata["pf19b_rece_material"])
            material["grupos_hash"] = "0" * 64
            metadata["pf19b_rece_material"] = material
        lote.metadata_json = metadata
        await self.db.commit()
        return resultado

    llamadas = {"ticket": 0, "fecomp": 0, "fecae": 0}

    async def fail_ticket(*args, **kwargs):
        llamadas["ticket"] += 1
        raise AssertionError("No debe solicitar ticket con ownership adulterado")

    async def fail_fecomp(*args, **kwargs):
        llamadas["fecomp"] += 1
        raise AssertionError("No debe consultar numeración con ownership adulterado")

    async def fail_fecae(*args, **kwargs):
        llamadas["fecae"] += 1
        raise AssertionError("No debe solicitar CAE con ownership adulterado")

    monkeypatch.setattr(
        LoteComprobantesService,
        "_reclamar_grupo_para_reintento",
        claim_con_adulteracion,
    )
    monkeypatch.setattr(
        FacturacionService,
        "_obtener_ticket_acceso",
        fail_ticket,
    )
    monkeypatch.setattr(
        FacturacionService,
        "_obtener_diagnostico_numeracion",
        fail_fecomp,
    )
    monkeypatch.setattr(
        "app.arca.wsfev1.WSFEv1Client.fe_cae_solicitar",
        fail_fecae,
    )

    response = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/reintentar-fallidos",
        headers={**auth_headers, **headers_reintento},
        json={"grupo_ids": [grupo_id]},
    )

    assert response.status_code == 409, response.text
    assert llamadas == {"ticket": 0, "fecomp": 0, "fecae": 0}
    assert operacion_perdedora_id is not None
    db_session.expire_all()
    grupo_durable = await db_session.get(LoteComprobanteGrupo, grupo_id)
    lote_durable = await db_session.get(LoteComprobante, lote_id)
    assert grupo_durable is not None
    assert lote_durable is not None
    assert grupo_durable.estado == "reintentando"
    if adulteracion == "owner":
        assert owner_adulterado_id is not None
        assert (
            lote_durable.metadata_json["operacion_idempotente_id"]
            == owner_adulterado_id
        )

    recovery = await LoteComprobantesService(
        db_session
    ).recuperar_reintento_interrumpido_pre_arca(
        lote_id=lote_id,
        grupo_id=grupo_id,
        operacion_id=operacion_perdedora_id,
        mensajes_previos=["Fallo previo sintético."],
    )

    assert recovery == "no_recuperable"
    db_session.expire_all()
    assert (await db_session.get(LoteComprobanteGrupo, grupo_id)).estado == (
        "reintentando"
    )


@pytest.mark.asyncio
@pytest.mark.parametrize(
    "estado_intento_previo", ["en_proceso", "requiere_reconciliacion"]
)
async def test_reintentar_fallidos_no_transfiere_owner_con_intento_previo_activo(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
    estado_intento_previo: str,
) -> None:
    """Un owner terminal con intento activo o incierto no puede transferirse."""
    lote_id = await _crear_lote_validado_por_api(
        client,
        auth_headers,
        test_empresa.cuit,
        nombre_archivo=f"lote-owner-previo-{estado_intento_previo}.xlsx",
    )
    [grupo] = await _marcar_grupos_lote(db_session, lote_id, ["fallido"])
    grupo_id = int(grupo.id)
    original_emitir = FacturacionService._emitir_comprobante_locked

    async def fake_emitir(self, request, **kwargs):
        return EmitirComprobanteResponse(
            exito=False,
            tipo_comprobante=request.tipo_comprobante,
            punto_venta=grupo.punto_venta_numero,
            numero=0,
            fecha=request.fecha_emision,
            total=Decimal("1210.00"),
            mensaje="Fallo verificado sintético.",
            errores=["Fallo verificado sintético."],
        )

    monkeypatch.setattr(
        FacturacionService,
        "_emitir_comprobante_locked",
        fake_emitir,
    )
    headers_previos = await _confirmacion_fecha_fiscal_header_lote(
        db_session,
        lote_id=lote_id,
        estados={"fallido"},
        grupo_ids=[grupo_id],
        idempotency_key=f"idem-owner-previo-{estado_intento_previo}",
    )
    primera = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/reintentar-fallidos",
        headers={**auth_headers, **headers_previos},
        json={"grupo_ids": [grupo_id]},
    )
    assert primera.status_code == 200, primera.text

    operacion_previa = await db_session.scalar(
        select(OperacionIdempotente).where(
            OperacionIdempotente.idempotency_key
            == f"idem-owner-previo-{estado_intento_previo}"
        )
    )
    grupo_previo = await db_session.get(LoteComprobanteGrupo, grupo_id)
    assert operacion_previa is not None
    assert grupo_previo is not None
    assert grupo_previo.estado == "fallido"
    owner_previo_id = int(operacion_previa.id)
    owner_previo_usuario_id = operacion_previa.usuario_id
    request = EmitirComprobanteRequest.model_validate(grupo_previo.payload_json or {})
    contexto = ContextoElegibilidadRece(
        empresa_id=int(grupo_previo.empresa_id),
        punto_venta_id=int(grupo_previo.punto_venta_id),
        punto_venta_numero=int(grupo_previo.punto_venta_numero),
        ambiente=str(grupo_previo.ambiente),
        elegibilidad_revision_id=int(grupo_previo.punto_venta_elegibilidad_revision_id),
        punto_venta_revision_fiscal=int(grupo_previo.punto_venta_revision_fiscal),
    )
    guarda_cerrada = PuntoVentaGuardaEmisionRece(
        token=("a" if estado_intento_previo == "en_proceso" else "b") * 64,
        fase="cerrada_pre_arca",
        operacion_id=owner_previo_id,
        empresa_id=grupo_previo.empresa_id,
        punto_venta_id=grupo_previo.punto_venta_id,
        ambiente=grupo_previo.ambiente,
        elegibilidad_revision_id=grupo_previo.punto_venta_elegibilidad_revision_id,
        punto_venta_revision_fiscal=grupo_previo.punto_venta_revision_fiscal,
        cerrada_en=datetime.utcnow(),
    )
    db_session.add(guarda_cerrada)
    await db_session.flush()
    await db_session.refresh(test_punto_venta)
    intento_previo = await IdempotenciaFiscalService(db_session).crear_intento_emision(
        request=request,
        punto_venta=test_punto_venta,
        numero_planificado=1,
        total=FacturacionService(db_session)._calcular_totales(request.items)["total"],
        operacion_id=owner_previo_id,
        usuario_id=owner_previo_usuario_id,
        lote_id=lote_id,
        grupo_id=grupo_id,
        contexto_rece=contexto,
        guarda_rece_id=int(guarda_cerrada.id),
        commit=False,
    )
    intento_previo.estado = estado_intento_previo
    await db_session.commit()

    monkeypatch.setattr(
        FacturacionService,
        "_emitir_comprobante_locked",
        original_emitir,
    )
    llamadas = {"ticket": 0, "fecomp": 0, "fecae": 0}

    async def fail_ticket(*args, **kwargs):
        llamadas["ticket"] += 1
        raise AssertionError("No debe solicitar ticket con intento previo bloqueante")

    async def fail_fecomp(*args, **kwargs):
        llamadas["fecomp"] += 1
        raise AssertionError(
            "No debe consultar numeración con intento previo bloqueante"
        )

    async def fail_fecae(*args, **kwargs):
        llamadas["fecae"] += 1
        raise AssertionError("No debe solicitar CAE con intento previo bloqueante")

    monkeypatch.setattr(FacturacionService, "_obtener_ticket_acceso", fail_ticket)
    monkeypatch.setattr(
        FacturacionService,
        "_obtener_diagnostico_numeracion",
        fail_fecomp,
    )
    monkeypatch.setattr(
        "app.arca.wsfev1.WSFEv1Client.fe_cae_solicitar",
        fail_fecae,
    )
    headers_nuevos = await _confirmacion_fecha_fiscal_header_lote(
        db_session,
        lote_id=lote_id,
        estados={"fallido"},
        grupo_ids=[grupo_id],
        idempotency_key=f"idem-owner-nuevo-{estado_intento_previo}",
    )

    segunda = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/reintentar-fallidos",
        headers={**auth_headers, **headers_nuevos},
        json={"grupo_ids": [grupo_id]},
    )

    assert segunda.status_code == 409, segunda.text
    assert llamadas == {"ticket": 0, "fecomp": 0, "fecae": 0}
    db_session.expire_all()
    lote_durable = await db_session.get(LoteComprobante, lote_id)
    grupo_durable = await db_session.get(LoteComprobanteGrupo, grupo_id)
    assert lote_durable is not None
    assert grupo_durable is not None
    assert lote_durable.metadata_json["operacion_idempotente_id"] == owner_previo_id
    assert grupo_durable.estado == "fallido"


@pytest.mark.asyncio
async def test_reintentar_fallidos_segunda_key_pierde_cas_y_conserva_owner_previo(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
) -> None:
    """La segunda key que pierde el CAS no publica metadata ni devuelve 200 vacío."""
    lote_id = await _crear_lote_validado_por_api(
        client,
        auth_headers,
        test_empresa.cuit,
        nombre_archivo="lote-dos-keys-cas.xlsx",
    )
    [grupo] = await _marcar_grupos_lote(db_session, lote_id, ["fallido"])
    grupo_id = int(grupo.id)

    async def fake_emitir(self, request, **kwargs):
        return EmitirComprobanteResponse(
            exito=False,
            tipo_comprobante=request.tipo_comprobante,
            punto_venta=grupo.punto_venta_numero,
            numero=0,
            fecha=request.fecha_emision,
            total=Decimal("1210.00"),
            mensaje="Fallo verificado sintético.",
            errores=["Fallo verificado sintético."],
        )

    monkeypatch.setattr(
        FacturacionService,
        "_emitir_comprobante_locked",
        fake_emitir,
    )
    headers_owner = await _confirmacion_fecha_fiscal_header_lote(
        db_session,
        lote_id=lote_id,
        estados={"fallido"},
        grupo_ids=[grupo_id],
        idempotency_key="idem-dos-keys-owner",
    )
    primera = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/reintentar-fallidos",
        headers={**auth_headers, **headers_owner},
        json={"grupo_ids": [grupo_id]},
    )
    assert primera.status_code == 200, primera.text
    operacion_owner = await db_session.scalar(
        select(OperacionIdempotente).where(
            OperacionIdempotente.idempotency_key == "idem-dos-keys-owner"
        )
    )
    assert operacion_owner is not None
    owner_id = int(operacion_owner.id)

    headers_perdedora = await _confirmacion_fecha_fiscal_header_lote(
        db_session,
        lote_id=lote_id,
        estados={"fallido"},
        grupo_ids=[grupo_id],
        idempotency_key="idem-dos-keys-perdedora",
    )
    original_execute = db_session.execute
    cas_interceptado = False

    async def execute_con_cas_perdido(statement, *args, **kwargs):
        nonlocal cas_interceptado
        tabla = getattr(statement, "table", None)
        if (
            not cas_interceptado
            and getattr(statement, "is_update", False)
            and getattr(tabla, "name", None) == "lotes_comprobantes_grupos"
        ):
            cas_interceptado = True
            return SimpleNamespace(rowcount=0)
        return await original_execute(statement, *args, **kwargs)

    monkeypatch.setattr(db_session, "execute", execute_con_cas_perdido)

    segunda = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/reintentar-fallidos",
        headers={**auth_headers, **headers_perdedora},
        json={"grupo_ids": [grupo_id]},
    )

    assert segunda.status_code == 409, segunda.text
    assert cas_interceptado is True
    db_session.expire_all()
    lote_durable = await db_session.get(LoteComprobante, lote_id)
    grupo_durable = await db_session.get(LoteComprobanteGrupo, grupo_id)
    assert lote_durable is not None
    assert grupo_durable is not None
    assert lote_durable.metadata_json["operacion_idempotente_id"] == owner_id
    assert grupo_durable.estado == "fallido"


@pytest.mark.asyncio
async def test_reintentar_fallidos_bloquea_payload_con_clave_desconocida(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
) -> None:
    """Un reintento no debe emitir si el snapshot fiscal no es canónico."""
    lote_id = await _crear_lote_validado_por_api(
        client,
        auth_headers,
        test_empresa.cuit,
        nombre_archivo="lote-reintento-payload-no-canonico.xlsx",
    )
    grupos = await _marcar_grupos_lote(db_session, lote_id, ["fallido"])
    grupo = grupos[0]
    payload = dict(grupo.payload_json or {})
    payload["cotizaccion"] = "2"
    grupo.payload_json = payload
    await db_session.commit()
    llamadas_emision = 0

    async def fail_emitir_locked(self, request, commit=True, **kwargs):
        nonlocal llamadas_emision
        llamadas_emision += 1
        raise AssertionError("No debe emitir un payload fiscal no canónico")

    monkeypatch.setattr(
        "app.services.facturacion_service.FacturacionService._emitir_comprobante_locked",
        fail_emitir_locked,
    )
    headers_reintento = await _confirmacion_fecha_fiscal_header_lote(
        db_session,
        lote_id=lote_id,
        estados={"fallido"},
        grupo_ids=[grupo.id],
    )

    response = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/reintentar-fallidos",
        headers={**auth_headers, **headers_reintento},
        json={"grupo_ids": [grupo.id]},
    )

    assert response.status_code == 200, response.text
    assert llamadas_emision == 0
    await db_session.refresh(grupo)
    assert grupo.estado == "fallido"
    assert grupo.mensajes_json == [
        "No se pudo completar el reintento antes de solicitar CAE. "
        "El detalle técnico quedó registrado en logs privados."
    ]
    assert "cotizaccion" not in str(grupo.mensajes_json)
    assert grupo.numero_asignado is None
    assert grupo.cae is None
    assert grupo.comprobante_id is None


@pytest.mark.asyncio
async def test_reintentar_fallidos_usa_historia_externa_y_replay_no_reemite(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
) -> None:
    """El reintento usa el siguiente ARCA y el replay no vuelve a solicitar CAE."""

    class FakeWSFEClient:
        """Simula historia externa estable y una autorización verificable."""

        consultas_numeracion = 0
        numeros_solicitados: list[int] = []

        def __init__(self, *args, **kwargs) -> None:
            """Acepta la firma del cliente real sin usar red."""

        async def fe_comp_ultimo_autorizado(self, punto_venta_numero, tipo):
            """Informa un comprobante externo posterior a la historia local."""
            FakeWSFEClient.consultas_numeracion += 1
            return 77

        async def fe_cae_solicitar(self, arca_request):
            """Autoriza únicamente el número confirmado por ambos preflights."""
            FakeWSFEClient.numeros_solicitados.append(arca_request.cbte_desde)
            return CAEResponse(
                cae=CAE_TEST_NO_REAL_ALT,
                cae_vencimiento="20260831",
                numero_comprobante=arca_request.cbte_desde,
                tipo_cbte=arca_request.tipo_cbte,
                punto_venta=arca_request.punto_venta,
                resultado="A",
            )

    lote_id, grupos = await _preparar_reintento_manual_pf02b2(
        client,
        auth_headers,
        monkeypatch,
        db_session,
        test_empresa,
        test_punto_venta,
        FakeWSFEClient,
        nombre_archivo="lote-reintento-historia-externa.xlsx",
        total_grupos=2,
        ultimo_local=76,
    )
    grupo_id = grupos[0].id
    otro_grupo_id = grupos[1].id
    headers_reintento = await _confirmacion_fecha_fiscal_header_lote(
        db_session,
        lote_id=lote_id,
        estados={"fallido"},
        grupo_ids=[grupo_id],
        idempotency_key="idem-reintento-historia-externa",
    )
    headers = {**auth_headers, **headers_reintento}
    body = {"grupo_ids": [grupo_id]}

    primera = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/reintentar-fallidos",
        headers=headers,
        json=body,
    )
    segunda = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/reintentar-fallidos",
        headers=headers,
        json=body,
    )
    conflicto = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/reintentar-fallidos",
        headers=headers,
        json={"grupo_ids": [otro_grupo_id]},
    )

    assert primera.status_code == 200, primera.text
    assert segunda.status_code == 200, segunda.text
    assert conflicto.status_code == 409, conflicto.text
    assert segunda.json() == primera.json()
    assert FakeWSFEClient.consultas_numeracion == 2
    assert FakeWSFEClient.numeros_solicitados == [78]
    db_session.expire_all()
    grupo = await db_session.get(LoteComprobanteGrupo, grupo_id)
    assert grupo is not None
    assert grupo.estado == "autorizado"
    assert grupo.numero_asignado == 78
    assert grupo.cae == CAE_TEST_NO_REAL_ALT
    assert grupo.comprobante_id is not None
    intentos = (
        (
            await db_session.execute(
                select(IntentoEmisionFiscal).where(
                    IntentoEmisionFiscal.grupo_id == grupo.id
                )
            )
        )
        .scalars()
        .all()
    )
    assert len(intentos) == 1
    assert intentos[0].estado == "autorizado"
    assert intentos[0].numero_planificado == 78
    operacion = await db_session.scalar(
        select(OperacionIdempotente).where(
            OperacionIdempotente.idempotency_key == "idem-reintento-historia-externa"
        )
    )
    assert operacion is not None
    assert operacion.estado == "finalizado"


@pytest.mark.asyncio
async def test_reintentar_10005_reconstruye_publicacion_tras_crash_sin_reemitir(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
) -> None:
    """Un crash tras cerrar 10005 publica desde lote fallido sin otra FECAE."""

    class FakeWSFEClient:
        """Rechaza exactamente una solicitud con el 10005 global canónico."""

        consultas_numeracion = 0
        solicitudes_fecae = 0

        def __init__(self, *args, **kwargs) -> None:
            """Acepta la firma productiva sin abrir red."""

        async def fe_comp_ultimo_autorizado(self, punto_venta_numero, tipo):
            """Mantiene estable la numeración sintética."""
            FakeWSFEClient.consultas_numeracion += 1
            return 0

        async def fe_cae_solicitar(self, arca_request):
            """Devuelve un rechazo global exacto sin detalle ni CAE."""
            FakeWSFEClient.solicitudes_fecae += 1
            raise ArcaErrorGlobalEstructurado(
                cabecera=CabeceraRespuestaFecae(
                    cuit=int(test_empresa.cuit),
                    punto_venta=arca_request.punto_venta,
                    tipo_comprobante=arca_request.tipo_cbte,
                    cantidad=1,
                    resultado="R",
                ),
                errores=(MensajeArcaEstructurado(10005, "mensaje privado ARCA"),),
                eventos=(),
                detalles_presentes=False,
                senales_cae_presentes=False,
                request_cuit=int(test_empresa.cuit),
                request_punto_venta=arca_request.punto_venta,
                request_tipo_comprobante=arca_request.tipo_cbte,
                request_cantidad=1,
                request_rangos=((arca_request.cbte_desde, arca_request.cbte_hasta),),
            )

    lote_id, [grupo] = await _preparar_reintento_manual_pf02b2(
        client,
        auth_headers,
        monkeypatch,
        db_session,
        test_empresa,
        test_punto_venta,
        FakeWSFEClient,
        nombre_archivo="lote-reintento-pf19c-10005-crash.xlsx",
    )
    idempotency_key = "idem-reintento-pf19c-10005-crash"
    headers_reintento = await _confirmacion_fecha_fiscal_header_lote(
        db_session,
        lote_id=lote_id,
        estados={"fallido"},
        grupo_ids=[int(grupo.id)],
        idempotency_key=idempotency_key,
    )
    headers = {**auth_headers, **headers_reintento}
    body = {"grupo_ids": [int(grupo.id)]}
    guardar_original = IdempotenciaFiscalService.guardar_resultado_operacion_sync
    publicaciones_fallidas = 0

    async def fallar_primera_publicacion(
        self,
        operacion,
        *,
        response_json,
        estado,
    ):
        """Simula el crash posterior al commit del grafo y previo al response CAS."""
        nonlocal publicaciones_fallidas
        if publicaciones_fallidas == 0 and estado == "rechazado_arca":
            publicaciones_fallidas += 1
            raise SQLAlchemyTimeoutError("crash sintético post-cierre")
        return await guardar_original(
            self,
            operacion,
            response_json=response_json,
            estado=estado,
        )

    monkeypatch.setattr(
        IdempotenciaFiscalService,
        "guardar_resultado_operacion_sync",
        fallar_primera_publicacion,
    )

    cierre_sin_publicar = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/reintentar-fallidos",
        headers=headers,
        json=body,
    )
    assert cierre_sin_publicar.status_code == 409, cierre_sin_publicar.text
    db_session.expire_all()
    lote_cerrado = await db_session.get(LoteComprobante, lote_id)
    operacion = await db_session.scalar(
        select(OperacionIdempotente).where(
            OperacionIdempotente.idempotency_key == idempotency_key
        )
    )
    assert lote_cerrado is not None
    assert lote_cerrado.estado == "fallido"
    assert operacion is not None
    assert operacion.estado == "en_proceso"
    assert operacion.response_json is None

    reconstruida = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/reintentar-fallidos",
        headers=headers,
        json=body,
    )
    replay = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/reintentar-fallidos",
        headers=headers,
        json=body,
    )

    assert reconstruida.status_code == 200, reconstruida.text
    assert replay.status_code == 200, replay.text
    assert replay.json() == reconstruida.json()
    assert reconstruida.json()["lote"]["estado"] == "fallido"
    assert reconstruida.json()["errores_arca"] == [
        {
            "codigo": 10005,
            "alcance": "global",
            "mensaje": "El punto de venta no está dado de alta como RECE en ARCA.",
        }
    ]
    assert FakeWSFEClient.solicitudes_fecae == 1
    await db_session.refresh(operacion)
    assert operacion.estado == "rechazado_arca"
    assert operacion.response_json == reconstruida.json()


@pytest.mark.asyncio
@pytest.mark.parametrize(
    ("body", "caso"),
    [
        pytest.param({}, "omitido", id="grupo_ids-omitido"),
        pytest.param({"grupo_ids": []}, "vacio", id="grupo_ids-vacio"),
    ],
)
async def test_reintentar_fallidos_replay_terminal_sin_seleccion_no_reemite(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
    body: dict[str, list[int]],
    caso: str,
) -> None:
    """Omitir o vaciar la selección reejecuta la misma respuesta terminal sin I/O."""

    class FakeWSFEClient:
        """Autoriza una vez y contabiliza cualquier I/O fiscal posterior."""

        consultas_fecomp = 0
        solicitudes_fecae = 0

        def __init__(self, *args, **kwargs) -> None:
            """Acepta la firma del cliente real sin usar red."""

        async def fe_comp_ultimo_autorizado(self, punto_venta_numero, tipo):
            """Devuelve una historia vacía y contabiliza FEComp."""
            FakeWSFEClient.consultas_fecomp += 1
            return 0

        async def fe_cae_solicitar(self, arca_request):
            """Autoriza una única solicitud fiscal sintética."""
            FakeWSFEClient.solicitudes_fecae += 1
            return CAEResponse(
                cae=CAE_TEST_NO_REAL_36,
                cae_vencimiento="20260831",
                numero_comprobante=arca_request.cbte_desde,
                tipo_cbte=arca_request.tipo_cbte,
                punto_venta=arca_request.punto_venta,
                resultado="A",
            )

    lote_id, [grupo] = await _preparar_reintento_manual_pf02b2(
        client,
        auth_headers,
        monkeypatch,
        db_session,
        test_empresa,
        test_punto_venta,
        FakeWSFEClient,
        nombre_archivo=f"lote-replay-sin-seleccion-{caso}.xlsx",
    )
    grupo_id = int(grupo.id)
    idempotency_key = f"idem-replay-sin-seleccion-{caso}"
    headers_reintento = await _confirmacion_fecha_fiscal_header_lote(
        db_session,
        lote_id=lote_id,
        estados={"fallido"},
        idempotency_key=idempotency_key,
    )
    llamadas_ticket = 0

    async def fake_ticket(self, empresa, certificado):
        nonlocal llamadas_ticket
        llamadas_ticket += 1
        return SimpleNamespace(token="token-test", sign="sign-test")

    monkeypatch.setattr(FacturacionService, "_obtener_ticket_acceso", fake_ticket)

    primera = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/reintentar-fallidos",
        headers={**auth_headers, **headers_reintento},
        json=body,
    )
    assert primera.status_code == 200, primera.text
    operacion = await db_session.scalar(
        select(OperacionIdempotente).where(
            OperacionIdempotente.idempotency_key == idempotency_key
        )
    )
    assert operacion is not None
    operacion_id = int(operacion.id)
    conteos_antes = {
        "operaciones": int(
            await db_session.scalar(
                select(func.count(OperacionIdempotente.id)).where(
                    OperacionIdempotente.idempotency_key == idempotency_key
                )
            )
            or 0
        ),
        "guardas": int(
            await db_session.scalar(
                select(func.count(PuntoVentaGuardaEmisionRece.id)).where(
                    PuntoVentaGuardaEmisionRece.operacion_id == operacion_id
                )
            )
            or 0
        ),
        "intentos": int(
            await db_session.scalar(
                select(func.count(IntentoEmisionFiscal.id)).where(
                    IntentoEmisionFiscal.operacion_id == operacion_id
                )
            )
            or 0
        ),
    }
    assert conteos_antes == {"operaciones": 1, "guardas": 1, "intentos": 1}
    llamadas_ticket = 0
    FakeWSFEClient.consultas_fecomp = 0
    FakeWSFEClient.solicitudes_fecae = 0

    replay = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/reintentar-fallidos",
        headers={**auth_headers, **headers_reintento},
        json=body,
    )

    assert replay.status_code == 200, replay.text
    assert replay.json() == primera.json()
    assert llamadas_ticket == 0
    assert FakeWSFEClient.consultas_fecomp == 0
    assert FakeWSFEClient.solicitudes_fecae == 0
    db_session.expire_all()
    assert (await db_session.get(LoteComprobanteGrupo, grupo_id)).estado == (
        "autorizado"
    )
    conteos_despues = {
        "operaciones": int(
            await db_session.scalar(
                select(func.count(OperacionIdempotente.id)).where(
                    OperacionIdempotente.idempotency_key == idempotency_key
                )
            )
            or 0
        ),
        "guardas": int(
            await db_session.scalar(
                select(func.count(PuntoVentaGuardaEmisionRece.id)).where(
                    PuntoVentaGuardaEmisionRece.operacion_id == operacion_id
                )
            )
            or 0
        ),
        "intentos": int(
            await db_session.scalar(
                select(func.count(IntentoEmisionFiscal.id)).where(
                    IntentoEmisionFiscal.operacion_id == operacion_id
                )
            )
            or 0
        ),
    }
    assert conteos_despues == conteos_antes


@pytest.mark.asyncio
@pytest.mark.parametrize(
    ("segundo_preflight", "categoria_error"),
    [
        ("avanza", "numeracion_arca_cambio_pre_arca"),
        ("falla", "preflight_arca_no_disponible"),
    ],
)
async def test_reintentar_fallidos_aborta_seleccion_si_falla_segundo_preflight(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
    segundo_preflight: str,
    categoria_error: str,
) -> None:
    """La numeración inestable cierra el intento y no reclama otro grupo."""

    class FakeWSFEClient:
        """Desestabiliza el segundo preflight y prohíbe continuar la selección."""

        consultas_numeracion = 0
        llamadas_cae = 0

        def __init__(self, *args, **kwargs) -> None:
            """Acepta la firma del cliente real sin usar red."""

        async def fe_comp_ultimo_autorizado(self, punto_venta_numero, tipo):
            """Cambia o falla después de reservar el primer grupo."""
            FakeWSFEClient.consultas_numeracion += 1
            if FakeWSFEClient.consultas_numeracion == 1:
                return 5
            if FakeWSFEClient.consultas_numeracion == 2:
                if segundo_preflight == "avanza":
                    return 6
                raise ArcaServiceError("preflight sintético no disponible")
            raise AssertionError("No debe consultar la numeración de otro grupo")

        async def fe_cae_solicitar(self, arca_request):
            """No debe invocarse con una reserva no reconfirmada."""
            FakeWSFEClient.llamadas_cae += 1
            raise AssertionError("No debe solicitar CAE")

    lote_id, grupos = await _preparar_reintento_manual_pf02b2(
        client,
        auth_headers,
        monkeypatch,
        db_session,
        test_empresa,
        test_punto_venta,
        FakeWSFEClient,
        nombre_archivo=f"lote-reintento-{segundo_preflight}.xlsx",
        total_grupos=2,
    )
    grupo_ids = [grupo.id for grupo in grupos]
    mensajes_segundo = list(grupos[1].mensajes_json or [])
    headers_reintento = await _confirmacion_fecha_fiscal_header_lote(
        db_session,
        lote_id=lote_id,
        estados={"fallido"},
        grupo_ids=grupo_ids,
        idempotency_key=f"idem-reintento-{segundo_preflight}",
    )

    response = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/reintentar-fallidos",
        headers={**auth_headers, **headers_reintento},
        json={"grupo_ids": grupo_ids},
    )

    assert response.status_code == 200, response.text
    assert FakeWSFEClient.consultas_numeracion == 2
    assert FakeWSFEClient.llamadas_cae == 0
    db_session.expire_all()
    primero = await db_session.get(LoteComprobanteGrupo, grupo_ids[0])
    segundo = await db_session.get(LoteComprobanteGrupo, grupo_ids[1])
    assert primero is not None
    assert segundo is not None
    assert primero.estado == "fallido"
    assert primero.numero_asignado is None
    assert primero.cae is None
    assert primero.comprobante_id is None
    assert segundo.estado == "fallido"
    assert segundo.mensajes_json == mensajes_segundo
    intentos = (
        (
            await db_session.execute(
                select(IntentoEmisionFiscal).order_by(IntentoEmisionFiscal.id)
            )
        )
        .scalars()
        .all()
    )
    assert len(intentos) == 1
    assert intentos[0].grupo_id == primero.id
    assert intentos[0].estado == "fallido_verificado"
    assert intentos[0].categoria_error == categoria_error


@pytest.mark.asyncio
async def test_reintentar_fallidos_detiene_seleccion_ante_respuesta_incierta(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
) -> None:
    """Una respuesta ambigua bloquea el lote y deja intactos los demás grupos."""

    class FakeWSFEClient:
        """Simula una excepción incierta después de iniciar FECAE."""

        consultas_numeracion = 0
        llamadas_cae = 0

        def __init__(self, *args, **kwargs) -> None:
            """Acepta la firma del cliente real sin usar red."""

        async def fe_comp_ultimo_autorizado(self, punto_venta_numero, tipo):
            """Mantiene estable la numeración antes del primer FECAE."""
            FakeWSFEClient.consultas_numeracion += 1
            return 0

        async def fe_cae_solicitar(self, arca_request):
            """Falla sin confirmar si ARCA autorizó el comprobante."""
            FakeWSFEClient.llamadas_cae += 1
            raise ArcaServiceError("respuesta sintética incierta")

    lote_id, grupos = await _preparar_reintento_manual_pf02b2(
        client,
        auth_headers,
        monkeypatch,
        db_session,
        test_empresa,
        test_punto_venta,
        FakeWSFEClient,
        nombre_archivo="lote-reintento-incierto.xlsx",
        total_grupos=2,
    )
    grupo_ids = [grupo.id for grupo in grupos]
    headers_reintento = await _confirmacion_fecha_fiscal_header_lote(
        db_session,
        lote_id=lote_id,
        estados={"fallido"},
        grupo_ids=grupo_ids,
        idempotency_key="idem-reintento-incierto",
    )

    response = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/reintentar-fallidos",
        headers={**auth_headers, **headers_reintento},
        json={"grupo_ids": grupo_ids},
    )

    assert response.status_code == 200, response.text
    assert response.json()["lote"]["estado"] == "requiere_reconciliacion"
    assert FakeWSFEClient.consultas_numeracion == 2
    assert FakeWSFEClient.llamadas_cae == 1
    db_session.expire_all()
    primero = await db_session.get(LoteComprobanteGrupo, grupo_ids[0])
    segundo = await db_session.get(LoteComprobanteGrupo, grupo_ids[1])
    assert primero is not None
    assert segundo is not None
    assert primero.estado == "requiere_reconciliacion"
    assert primero.numero_asignado == 1
    assert segundo.estado == "requiere_reconciliacion"
    assert segundo.numero_asignado is None
    assert segundo.cae is None
    assert segundo.comprobante_id is None
    assert "No se enviaron grupos posteriores" in segundo.mensajes_json[0]
    intentos = (
        (
            await db_session.execute(
                select(IntentoEmisionFiscal).order_by(IntentoEmisionFiscal.id)
            )
        )
        .scalars()
        .all()
    )
    assert len(intentos) == 1
    assert intentos[0].estado == "requiere_reconciliacion"
    operacion = await db_session.scalar(
        select(OperacionIdempotente).where(
            OperacionIdempotente.idempotency_key == "idem-reintento-incierto"
        )
    )
    assert operacion is not None
    assert operacion.estado == "requiere_reconciliacion"


@pytest.mark.asyncio
async def test_reintentar_fallidos_no_degrada_autorizacion_si_falla_capa_lote(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
) -> None:
    """Un fallo local posterior al CAE debe bloquear, nunca volver a fallido."""

    class FakeWSFEClient:
        """Autoriza el primer comprobante antes del fallo local inyectado."""

        llamadas_cae = 0

        def __init__(self, *args, **kwargs) -> None:
            """Acepta la firma del cliente real sin usar red."""

        async def fe_comp_ultimo_autorizado(self, punto_venta_numero, tipo):
            """Mantiene estable la numeración sintética."""
            return 0

        async def fe_cae_solicitar(self, arca_request):
            """Devuelve un CAE sintético válido."""
            FakeWSFEClient.llamadas_cae += 1
            return CAEResponse(
                cae=CAE_TEST_NO_REAL_ALT,
                cae_vencimiento="20260831",
                numero_comprobante=arca_request.cbte_desde,
                tipo_cbte=arca_request.tipo_cbte,
                punto_venta=arca_request.punto_venta,
                resultado="A",
            )

    lote_id, grupos = await _preparar_reintento_manual_pf02b2(
        client,
        auth_headers,
        monkeypatch,
        db_session,
        test_empresa,
        test_punto_venta,
        FakeWSFEClient,
        nombre_archivo="lote-reintento-fallo-capa-lote.xlsx",
        total_grupos=2,
    )
    grupo_ids = [grupo.id for grupo in grupos]
    mensajes_segundo = list(grupos[1].mensajes_json or [])

    async def fail_aplicar_resultado(self, grupo, resultado):
        raise RuntimeError("detalle interno sintético")

    monkeypatch.setattr(
        LoteComprobantesService,
        "_aplicar_resultado_emision_grupo",
        fail_aplicar_resultado,
    )
    headers_reintento = await _confirmacion_fecha_fiscal_header_lote(
        db_session,
        lote_id=lote_id,
        estados={"fallido"},
        grupo_ids=grupo_ids,
        idempotency_key="idem-reintento-fallo-capa-lote",
    )

    response = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/reintentar-fallidos",
        headers={**auth_headers, **headers_reintento},
        json={"grupo_ids": grupo_ids},
    )

    assert response.status_code == 200, response.text
    assert response.json()["lote"]["estado"] == "requiere_reconciliacion"
    assert "detalle interno sintético" not in response.text
    assert FakeWSFEClient.llamadas_cae == 1
    db_session.expire_all()
    primero = await db_session.get(LoteComprobanteGrupo, grupo_ids[0])
    segundo = await db_session.get(LoteComprobanteGrupo, grupo_ids[1])
    assert primero is not None
    assert segundo is not None
    assert primero.estado == "requiere_reconciliacion"
    assert primero.numero_asignado == 1
    assert primero.cae == CAE_TEST_NO_REAL_ALT
    assert primero.comprobante_id is None
    assert segundo.estado == "fallido"
    assert segundo.mensajes_json == mensajes_segundo
    intentos = (
        (
            await db_session.execute(
                select(IntentoEmisionFiscal).order_by(IntentoEmisionFiscal.id)
            )
        )
        .scalars()
        .all()
    )
    assert len(intentos) == 1
    assert intentos[0].estado == "requiere_reconciliacion"
    assert intentos[0].cae == CAE_TEST_NO_REAL_ALT
    comprobantes = (await db_session.execute(select(Comprobante))).scalars().all()
    assert comprobantes == []


@pytest.mark.asyncio
async def test_reintentar_fallidos_continua_solo_tras_rechazo_arca_explicito(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
) -> None:
    """Un rechazo explícito libera el número y permite el siguiente grupo."""

    class FakeWSFEClient:
        """Rechaza el primer request y autoriza el segundo de forma explícita."""

        consultas_numeracion = 0
        llamadas_cae = 0

        def __init__(self, *args, **kwargs) -> None:
            """Acepta la firma del cliente real sin usar red."""

        async def fe_comp_ultimo_autorizado(self, punto_venta_numero, tipo):
            """Mantiene disponible el mismo número tras el rechazo."""
            FakeWSFEClient.consultas_numeracion += 1
            return 0

        async def fe_cae_solicitar(self, arca_request):
            """Devuelve primero R y luego A con respuestas completas."""
            FakeWSFEClient.llamadas_cae += 1
            if FakeWSFEClient.llamadas_cae == 1:
                return CAEResponse(
                    cae=None,
                    cae_vencimiento=None,
                    numero_comprobante=arca_request.cbte_desde,
                    tipo_cbte=arca_request.tipo_cbte,
                    punto_venta=arca_request.punto_venta,
                    resultado="R",
                    errores=[
                        {
                            "code": 10016,
                            "msg": "Rechazo sintético explícito",
                        }
                    ],
                )
            return CAEResponse(
                cae=CAE_TEST_NO_REAL_ALT,
                cae_vencimiento="20260831",
                numero_comprobante=arca_request.cbte_desde,
                tipo_cbte=arca_request.tipo_cbte,
                punto_venta=arca_request.punto_venta,
                resultado="A",
            )

    lote_id, grupos = await _preparar_reintento_manual_pf02b2(
        client,
        auth_headers,
        monkeypatch,
        db_session,
        test_empresa,
        test_punto_venta,
        FakeWSFEClient,
        nombre_archivo="lote-reintento-rechazo-explicito.xlsx",
        total_grupos=2,
    )
    grupo_ids = [grupo.id for grupo in grupos]
    headers_reintento = await _confirmacion_fecha_fiscal_header_lote(
        db_session,
        lote_id=lote_id,
        estados={"fallido"},
        grupo_ids=grupo_ids,
        idempotency_key="idem-reintento-rechazo-explicito",
    )

    response = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/reintentar-fallidos",
        headers={**auth_headers, **headers_reintento},
        json={"grupo_ids": grupo_ids},
    )

    assert response.status_code == 200, response.text
    assert FakeWSFEClient.consultas_numeracion == 4
    assert FakeWSFEClient.llamadas_cae == 2
    db_session.expire_all()
    primero = await db_session.get(LoteComprobanteGrupo, grupo_ids[0])
    segundo = await db_session.get(LoteComprobanteGrupo, grupo_ids[1])
    assert primero is not None
    assert segundo is not None
    assert primero.estado == "fallido"
    assert primero.numero_asignado is None
    assert segundo.estado == "autorizado"
    assert segundo.numero_asignado == 1
    intentos = list(
        (
            await db_session.scalars(
                select(IntentoEmisionFiscal).order_by(IntentoEmisionFiscal.id)
            )
        ).all()
    )
    assert [intento.estado for intento in intentos] == [
        "rechazado_arca",
        "autorizado",
    ]


@pytest.mark.asyncio
@pytest.mark.parametrize(
    "bloqueo",
    ["local_adelantada", "en_proceso", "requiere_reconciliacion"],
)
async def test_reintentar_fallidos_detiene_seleccion_ante_bloqueo_propio(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
    bloqueo: str,
) -> None:
    """La historia local o un intento propio incierto bloquean toda selección."""

    class FakeWSFEClient:
        """Expone historia ARCA solo para el caso local adelantado."""

        consultas_numeracion = 0
        llamadas_cae = 0

        def __init__(self, *args, **kwargs) -> None:
            """Acepta la firma del cliente real sin usar red."""

        async def fe_comp_ultimo_autorizado(self, punto_venta_numero, tipo):
            """Deja a la historia local exactamente un número adelantada."""
            FakeWSFEClient.consultas_numeracion += 1
            if bloqueo != "local_adelantada":
                raise AssertionError("Un intento propio debe bloquear antes de ARCA")
            return 4

        async def fe_cae_solicitar(self, arca_request):
            """No debe solicitar CAE bajo ningún bloqueo."""
            FakeWSFEClient.llamadas_cae += 1
            raise AssertionError("No debe solicitar CAE")

    lote_id, grupos = await _preparar_reintento_manual_pf02b2(
        client,
        auth_headers,
        monkeypatch,
        db_session,
        test_empresa,
        test_punto_venta,
        FakeWSFEClient,
        nombre_archivo=f"lote-reintento-bloqueo-{bloqueo}.xlsx",
        total_grupos=2,
        ultimo_local=5 if bloqueo == "local_adelantada" else None,
    )
    grupo_ids = [grupo.id for grupo in grupos]
    mensajes_segundo = list(grupos[1].mensajes_json or [])
    if bloqueo != "local_adelantada":
        db_session.add(
            IntentoEmisionFiscal(
                tipo_comprobante=6,
                punto_venta_numero=test_punto_venta.numero,
                numero_planificado=1,
                fecha_emision=FECHA_FISCAL_PF02B2,
                total=Decimal("1210.00"),
                receptor_tipo_documento=80,
                receptor_numero_documento=CUIT_RECEPTOR_TEST_NO_REAL,
                receptor_razon_social="Cliente Lote SA",
                payload_hash=f"payload-bloqueante-{bloqueo}",
                huella_logica=f"huella-bloqueante-{bloqueo}",
                estado=bloqueo,
                empresa_id=test_empresa.id,
                punto_venta_id=test_punto_venta.id,
            )
        )
        await db_session.commit()
    headers_reintento = await _confirmacion_fecha_fiscal_header_lote(
        db_session,
        lote_id=lote_id,
        estados={"fallido"},
        grupo_ids=grupo_ids,
        idempotency_key=f"idem-reintento-bloqueo-{bloqueo}",
    )

    response = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/reintentar-fallidos",
        headers={**auth_headers, **headers_reintento},
        json={"grupo_ids": grupo_ids},
    )

    assert response.status_code == 200, response.text
    assert FakeWSFEClient.consultas_numeracion == (
        1 if bloqueo == "local_adelantada" else 0
    )
    assert FakeWSFEClient.llamadas_cae == 0
    db_session.expire_all()
    primero = await db_session.get(LoteComprobanteGrupo, grupo_ids[0])
    segundo = await db_session.get(LoteComprobanteGrupo, grupo_ids[1])
    assert primero is not None
    assert segundo is not None
    assert primero.estado == "fallido"
    assert primero.numero_asignado is None
    assert primero.cae is None
    assert segundo.estado == "fallido"
    assert segundo.mensajes_json == mensajes_segundo
    intentos_grupo = list(
        (
            await db_session.scalars(
                select(IntentoEmisionFiscal).where(
                    IntentoEmisionFiscal.grupo_id.in_(grupo_ids)
                )
            )
        ).all()
    )
    assert intentos_grupo == []


@pytest.mark.asyncio
@pytest.mark.parametrize(
    "error_type",
    [SQLAlchemyTimeoutError, OperationalError],
    ids=["timeout", "operational"],
)
async def test_reintentar_lote_db_temporal_pre_arca_restaura_grupo_exacto(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
    error_type: type[Exception],
) -> None:
    """La caída DB pre-ARCA restaura solo el grupo reclamado y abre replay."""
    lote_id = await _crear_lote_validado_por_api(
        client,
        auth_headers,
        test_empresa.cuit,
        nombre_archivo="lote-reintento-db-temporal.xlsx",
        total_grupos=2,
    )
    grupos = await _marcar_grupos_lote(db_session, lote_id, ["fallido", "fallido"])
    grupo = grupos[0]
    otro_grupo = grupos[1]
    grupo_id = grupo.id
    otro_grupo_id = otro_grupo.id
    mensajes_previos = list(grupo.mensajes_json or [])

    async def fail_emitir_locked(self, request, **kwargs):
        assert kwargs["fase_solicitud_arca"].iniciada is False
        raise _crear_error_db_temporal(error_type)

    monkeypatch.setattr(
        FacturacionService,
        "_emitir_comprobante_locked",
        fail_emitir_locked,
    )
    confirmacion = await _confirmacion_fecha_fiscal_header_lote(
        db_session,
        lote_id=lote_id,
        estados={"fallido"},
        grupo_ids=[grupo_id],
    )

    response = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/reintentar-fallidos",
        headers={**auth_headers, **confirmacion},
        json={"grupo_ids": [grupo_id]},
    )

    assert response.status_code == 503, response.text
    assert response.headers["Retry-After"] == "2"
    assert "UPDATE lotes_comprobantes" not in response.text
    db_session.expire_all()
    grupo_actual = await db_session.get(LoteComprobanteGrupo, grupo_id)
    assert grupo_actual is not None
    assert grupo_actual.estado == "fallido"
    assert grupo_actual.mensajes_json == mensajes_previos
    otro_actual = await db_session.get(LoteComprobanteGrupo, otro_grupo_id)
    assert otro_actual is not None
    assert otro_actual.estado == "fallido"
    operacion = await db_session.scalar(
        select(OperacionIdempotente).where(
            OperacionIdempotente.idempotency_key == "idem-lote-test"
        )
    )
    assert operacion is not None
    assert operacion.estado == "interrumpida_pre_arca"
    intentos_actuales = (
        (
            await db_session.execute(
                select(IntentoEmisionFiscal).where(
                    IntentoEmisionFiscal.operacion_id == operacion.id
                )
            )
        )
        .scalars()
        .all()
    )
    assert intentos_actuales == []
    assert not any(
        "UPDATE lotes_comprobantes" in mensaje
        for mensaje in (grupo_actual.mensajes_json or [])
    )


@pytest.mark.asyncio
@pytest.mark.parametrize(
    "metodo_fallido",
    [
        "obtener_resumen_operativo_lote",
        "obtener_confirmacion_duplicado_logico_grupos",
    ],
)
async def test_procesar_lote_recupera_consultas_db_post_operacion(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
    metodo_fallido: str,
) -> None:
    """Resumen y duplicados quedan dentro de la recuperación atómica pre-ARCA."""
    lote_id = await _crear_lote_validado_por_api(
        client,
        auth_headers,
        test_empresa.cuit,
        nombre_archivo=f"lote-fallo-{metodo_fallido}.xlsx",
    )

    async def fail_db(self, *args, **kwargs):
        raise SQLAlchemyTimeoutError()

    headers_procesar = await _confirmacion_fecha_fiscal_header_lote(
        db_session,
        lote_id=lote_id,
        estados={"validado"},
    )
    monkeypatch.setattr(LoteComprobantesService, metodo_fallido, fail_db)
    response = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/procesar",
        headers={**auth_headers, **headers_procesar},
    )

    assert response.status_code == 503, response.text
    lote = await db_session.get(LoteComprobante, lote_id)
    assert lote is not None
    assert lote.estado == "validado"
    operacion = await db_session.scalar(
        select(OperacionIdempotente).where(
            OperacionIdempotente.idempotency_key == "idem-lote-test"
        )
    )
    assert operacion is not None
    assert operacion.estado == "interrumpida_pre_arca"
    intentos = await db_session.scalars(
        select(IntentoEmisionFiscal).where(
            IntentoEmisionFiscal.operacion_id == operacion.id
        )
    )
    assert intentos.all() == []


@pytest.mark.asyncio
async def test_reintentar_fallo_db_antes_de_operacion_devuelve_503_sin_unboundlocal(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
) -> None:
    """Una caída antes del resolver se propaga sin leer variables inexistentes."""
    lote_id = await _crear_lote_validado_por_api(
        client,
        auth_headers,
        test_empresa.cuit,
        nombre_archivo="lote-reintento-fallo-temprano.xlsx",
    )
    headers_reintento = await _confirmacion_fecha_fiscal_header_lote(
        db_session,
        lote_id=lote_id,
        estados={"validado"},
    )

    async def fail_lookup(self, *args, **kwargs):
        raise OperationalError("SELECT lote", {}, RuntimeError("db caída"))

    monkeypatch.setattr(
        LoteComprobantesService,
        "obtener_lote_resumen",
        fail_lookup,
    )
    response = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/reintentar-fallidos",
        headers={**auth_headers, **headers_reintento},
        json={"grupo_ids": []},
    )

    assert response.status_code == 503
    assert "UnboundLocalError" not in response.text
    assert response.headers["Retry-After"] == "2"


@pytest.mark.asyncio
async def test_reintentar_commit_ambiguo_confirmado_habilita_replay(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
) -> None:
    """El reintento recupera un commit ambiguo y la misma clave puede reclamarlo."""
    lote_id = await _crear_lote_validado_por_api(
        client,
        auth_headers,
        test_empresa.cuit,
        nombre_archivo="lote-reintento-commit-ambiguo.xlsx",
    )
    grupos = await _marcar_grupos_lote(db_session, lote_id, ["fallido"])
    original_commit = db_session.commit
    fallo_inyectado = False

    async def commit_ambiguo():
        nonlocal fallo_inyectado
        await original_commit()
        if not fallo_inyectado:
            fallo_inyectado = True
            raise SQLAlchemyTimeoutError()

    monkeypatch.setattr(db_session, "commit", commit_ambiguo)

    async def fake_reintentar(self, lote_id, empresa_id, **kwargs):
        return await self.obtener_lote(lote_id, empresa_id)

    monkeypatch.setattr(
        LoteComprobantesService,
        "reintentar_grupos_fallidos",
        fake_reintentar,
    )
    grupo_ids = [grupos[0].id]
    headers_reintento = await _confirmacion_fecha_fiscal_header_lote(
        db_session,
        lote_id=lote_id,
        estados={"fallido"},
        grupo_ids=grupo_ids,
        idempotency_key="idem-reintento-create-ambiguo",
    )
    headers = {**auth_headers, **headers_reintento}
    body = {"grupo_ids": grupo_ids}

    primera = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/reintentar-fallidos",
        headers=headers,
        json=body,
    )
    segunda = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/reintentar-fallidos",
        headers=headers,
        json=body,
    )

    assert primera.status_code == 503, primera.text
    assert segunda.status_code == 200, segunda.text
    operacion = await db_session.scalar(
        select(OperacionIdempotente).where(
            OperacionIdempotente.idempotency_key == "idem-reintento-create-ambiguo"
        )
    )
    assert operacion is not None
    assert operacion.estado == "finalizado"
    assert fallo_inyectado is True


@pytest.mark.asyncio
@pytest.mark.parametrize("caso", ["inexistente", "cruzado"])
async def test_reconciliar_externo_valida_ownership_antes_de_arca(
    caso: str,
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
) -> None:
    """Un lote inexistente o ajeno corta antes de WSAA/FEComp/FECAE."""
    lote_id = 999_999
    if caso == "cruzado":
        otra_empresa = Empresa(
            razon_social="Empresa lote cruzado sintética",
            cuit="20304050607",
            condicion_iva="RI",
            domicilio="Domicilio sintético 456",
            localidad="Ciudad de prueba",
            provincia="Buenos Aires",
            codigo_postal="1000",
            inicio_actividades=date(2020, 1, 1),
        )
        db_session.add(otra_empresa)
        await db_session.flush()
        lote_ajeno = LoteComprobante(
            empresa_id=otra_empresa.id,
            nombre_archivo="lote-reconciliacion-cruzado.xlsx",
            archivo_hash="d" * 64,
            estado="con_errores",
        )
        db_session.add(lote_ajeno)
        await db_session.commit()
        lote_id = int(lote_ajeno.id)

    llamadas = {"wsaa": 0, "fecomp": 0, "fecae": 0}

    class FakeWSFEClient:
        """Registra cualquier cruce indebido de la frontera ARCA."""

        async def fe_comp_consultar(self, **kwargs):
            """Registra una consulta FEComp indebida."""
            llamadas["fecomp"] += 1
            raise AssertionError("No debe consultar ARCA para un lote ajeno")

        async def fe_cae_solicitar(self, request):
            """Registra una solicitud FECAE indebida."""
            llamadas["fecae"] += 1
            raise AssertionError("La reconciliación nunca solicita CAE")

    async def fake_get_wsfe_client(*args, **kwargs):
        """Representa la autenticación WSAA que debe quedar en cero."""
        llamadas["wsaa"] += 1
        return FakeWSFEClient()

    monkeypatch.setattr(
        "app.api.lotes_comprobantes.get_wsfe_client",
        fake_get_wsfe_client,
    )

    response = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/reconciliar-externos",
        headers=auth_headers,
        json={
            "comprobantes": [
                {
                    "grupo_id": 1,
                    "tipo_comprobante": 6,
                    "punto_venta_numero": 1,
                    "numero": 1,
                    "fecha_emision": "08/08/2026",
                    "total": 121.0,
                    "cae": CAE_TEST_NO_REAL,
                    "motivo": "Reconciliación sintética",
                }
            ]
        },
    )

    assert response.status_code == 400, response.text
    assert "No se encontró el lote solicitado" in response.json()["detail"]
    assert llamadas == {"wsaa": 0, "fecomp": 0, "fecae": 0}


@pytest.mark.asyncio
async def test_reconciliar_externo_verifica_arca_y_crea_comprobante(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """Un comprobante manual se reconcilia solo si ARCA confirma los datos."""
    lote_id = await _crear_lote_validado_por_api(
        client,
        auth_headers,
        test_empresa.cuit,
        nombre_archivo="lote-reconciliar-externo.xlsx",
    )
    grupos = await _marcar_grupos_lote(db_session, lote_id, ["fallido"])
    grupo = grupos[0]

    class FakeWsfeClient:
        async def fe_comp_consultar(
            self,
            punto_venta: int,
            tipo_cbte: int,
            numero: int,
        ) -> ArcaComprobanteResponse:
            return ArcaComprobanteResponse(
                punto_venta=punto_venta,
                tipo_cbte=tipo_cbte,
                numero=numero,
                cuit_emisor=test_empresa.cuit,
                cae=CAE_TEST_NO_REAL_ALT,
                cae_vencimiento="20260630",
                fecha_cbte=str(grupo.fecha_emision).replace("-", ""),
                fecha_proceso="20260601",
                imp_total=1210.0,
                imp_neto=1000.0,
                imp_iva=210.0,
                imp_op_ex=0.0,
                imp_tot_conc=0.0,
                imp_trib=0.0,
                moneda_id="PES",
                moneda_cotiz=1.0,
                tipo_doc=80,
                nro_doc=CUIT_RECEPTOR_TEST_NO_REAL_INT,
                resultado="A",
            )

    async def fake_get_wsfe_client(*args, **kwargs):
        return FakeWsfeClient()

    monkeypatch.setattr(
        "app.api.lotes_comprobantes.get_wsfe_client",
        fake_get_wsfe_client,
    )

    response = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/reconciliar-externos",
        headers=auth_headers,
        json={
            "comprobantes": [
                {
                    "grupo_id": grupo.id,
                    "tipo_comprobante": grupo.tipo_comprobante,
                    "punto_venta_numero": grupo.punto_venta_numero,
                    "numero": 456,
                    "fecha_emision": _fecha_argentina(grupo.fecha_emision),
                    "total": 1210.0,
                    "cae": CAE_TEST_NO_REAL_ALT,
                    "motivo": "Emitido manualmente por ARCA Web",
                }
            ]
        },
    )

    assert response.status_code == 200, response.text
    data = response.json()["lote"]
    assert data["estado"] == "cerrado_reconciliado"
    assert data["grupos_reconciliados_externos"] == 1

    await db_session.refresh(grupo)
    assert grupo.estado == "autorizado_externo"
    assert grupo.numero_asignado == 456
    assert grupo.comprobante_id is not None
    comprobante = await db_session.get(Comprobante, grupo.comprobante_id)
    assert comprobante.origen_emision == "arca_web"
    assert comprobante.estado == "autorizado"


@pytest.mark.asyncio
async def test_reconciliar_externo_rechaza_receptor_distinto_en_arca(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """No se debe vincular un comprobante ARCA emitido a otro receptor."""
    lote_id = await _crear_lote_validado_por_api(
        client,
        auth_headers,
        test_empresa.cuit,
        nombre_archivo="lote-reconciliar-receptor-distinto.xlsx",
    )
    grupos = await _marcar_grupos_lote(db_session, lote_id, ["fallido"])
    grupo = grupos[0]

    class FakeWsfeClient:
        async def fe_comp_consultar(
            self,
            punto_venta: int,
            tipo_cbte: int,
            numero: int,
        ) -> ArcaComprobanteResponse:
            return ArcaComprobanteResponse(
                punto_venta=punto_venta,
                tipo_cbte=tipo_cbte,
                numero=numero,
                cuit_emisor=test_empresa.cuit,
                cae=CAE_TEST_NO_REAL_38,
                cae_vencimiento="20260630",
                fecha_cbte=str(grupo.fecha_emision).replace("-", ""),
                fecha_proceso="20260601",
                imp_total=1210.0,
                imp_neto=1000.0,
                imp_iva=210.0,
                imp_op_ex=0.0,
                imp_tot_conc=0.0,
                imp_trib=0.0,
                moneda_id="PES",
                moneda_cotiz=1.0,
                tipo_doc=80,
                nro_doc=30712345678,
                resultado="A",
            )

    async def fake_get_wsfe_client(*args, **kwargs):
        return FakeWsfeClient()

    monkeypatch.setattr(
        "app.api.lotes_comprobantes.get_wsfe_client",
        fake_get_wsfe_client,
    )

    response = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/reconciliar-externos",
        headers=auth_headers,
        json={
            "comprobantes": [
                {
                    "grupo_id": grupo.id,
                    "tipo_comprobante": grupo.tipo_comprobante,
                    "punto_venta_numero": grupo.punto_venta_numero,
                    "numero": 456,
                    "fecha_emision": _fecha_argentina(grupo.fecha_emision),
                    "total": 1210.0,
                    "cae": CAE_TEST_NO_REAL_38,
                    "motivo": "Emitido manualmente por ARCA Web",
                }
            ]
        },
    )

    assert response.status_code == 400
    assert "receptor informado por ARCA no coincide" in response.json()["detail"]
    await db_session.refresh(grupo)
    assert grupo.estado == "fallido"
    assert grupo.comprobante_id is None


@pytest.mark.asyncio
async def test_reconciliar_externo_rechaza_comprobante_ya_vinculado(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """Un comprobante externo no puede cerrar dos grupos distintos."""
    lote_id = await _crear_lote_validado_por_api(
        client,
        auth_headers,
        test_empresa.cuit,
        nombre_archivo="lote-reconciliar-externo-duplicado.xlsx",
        total_grupos=2,
    )
    grupos = await _marcar_grupos_lote(db_session, lote_id, ["fallido", "fallido"])

    class FakeWsfeClient:
        async def fe_comp_consultar(
            self,
            punto_venta: int,
            tipo_cbte: int,
            numero: int,
        ) -> ArcaComprobanteResponse:
            return ArcaComprobanteResponse(
                punto_venta=punto_venta,
                tipo_cbte=tipo_cbte,
                numero=numero,
                cuit_emisor=test_empresa.cuit,
                cae=CAE_TEST_NO_REAL_39,
                cae_vencimiento="20260630",
                fecha_cbte=str(grupos[0].fecha_emision).replace("-", ""),
                fecha_proceso="20260601",
                imp_total=1210.0,
                imp_neto=1000.0,
                imp_iva=210.0,
                imp_op_ex=0.0,
                imp_tot_conc=0.0,
                imp_trib=0.0,
                moneda_id="PES",
                moneda_cotiz=1.0,
                tipo_doc=80,
                nro_doc=CUIT_RECEPTOR_TEST_NO_REAL_INT,
                resultado="A",
            )

    async def fake_get_wsfe_client(*args, **kwargs):
        return FakeWsfeClient()

    monkeypatch.setattr(
        "app.api.lotes_comprobantes.get_wsfe_client",
        fake_get_wsfe_client,
    )

    payload_base = {
        "tipo_comprobante": grupos[0].tipo_comprobante,
        "punto_venta_numero": grupos[0].punto_venta_numero,
        "numero": 456,
        "fecha_emision": str(grupos[0].fecha_emision),
        "total": 1210.0,
        "cae": CAE_TEST_NO_REAL_39,
        "motivo": "Emitido manualmente por ARCA Web",
    }
    primera = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/reconciliar-externos",
        headers=auth_headers,
        json={"comprobantes": [{**payload_base, "grupo_id": grupos[0].id}]},
    )
    assert primera.status_code == 200, primera.text

    segunda = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/reconciliar-externos",
        headers=auth_headers,
        json={"comprobantes": [{**payload_base, "grupo_id": grupos[1].id}]},
    )

    assert segunda.status_code == 400
    assert "ya está vinculado a otro grupo" in segunda.json()["detail"]
    await db_session.refresh(grupos[1])
    assert grupos[1].estado == "fallido"
    assert grupos[1].comprobante_id is None


@pytest.mark.asyncio
async def test_reconciliar_externo_resuelve_lote_con_reconciliacion_tecnica(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """Los grupos inciertos deben poder cerrarse con verificación de ARCA."""
    lote_id = await _crear_lote_validado_por_api(
        client,
        auth_headers,
        test_empresa.cuit,
        nombre_archivo="lote-requiere-reconciliacion.xlsx",
    )
    grupos = await _marcar_grupos_lote(
        db_session,
        lote_id,
        ["requiere_reconciliacion"],
    )
    grupo = grupos[0]

    class FakeWsfeClient:
        async def fe_comp_consultar(
            self,
            punto_venta: int,
            tipo_cbte: int,
            numero: int,
        ) -> ArcaComprobanteResponse:
            return ArcaComprobanteResponse(
                punto_venta=punto_venta,
                tipo_cbte=tipo_cbte,
                numero=numero,
                cuit_emisor=test_empresa.cuit,
                cae=CAE_TEST_NO_REAL_36,
                cae_vencimiento="20260630",
                fecha_cbte=str(grupo.fecha_emision).replace("-", ""),
                fecha_proceso="20260601",
                imp_total=1210.0,
                imp_neto=1000.0,
                imp_iva=210.0,
                imp_op_ex=0.0,
                imp_tot_conc=0.0,
                imp_trib=0.0,
                moneda_id="PES",
                moneda_cotiz=1.0,
                tipo_doc=80,
                nro_doc=CUIT_RECEPTOR_TEST_NO_REAL_INT,
                resultado="A",
            )

    async def fake_get_wsfe_client(*args, **kwargs):
        return FakeWsfeClient()

    monkeypatch.setattr(
        "app.api.lotes_comprobantes.get_wsfe_client",
        fake_get_wsfe_client,
    )

    response = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/reconciliar-externos",
        headers=auth_headers,
        json={
            "comprobantes": [
                {
                    "grupo_id": grupo.id,
                    "tipo_comprobante": grupo.tipo_comprobante,
                    "punto_venta_numero": grupo.punto_venta_numero,
                    "numero": 789,
                    "fecha_emision": _fecha_argentina(grupo.fecha_emision),
                    "total": 1210.0,
                    "motivo": "Recuperación luego de corte post-ARCA",
                }
            ]
        },
    )

    assert response.status_code == 200, response.text
    assert response.json()["lote"]["estado"] == "cerrado_reconciliado"


@pytest.mark.asyncio
async def test_reconciliar_externo_resuelve_reintento_interrumpido(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """Un grupo reclamado para reintento debe cerrarse verificando ARCA."""
    lote_id = await _crear_lote_validado_por_api(
        client,
        auth_headers,
        test_empresa.cuit,
        nombre_archivo="lote-reintento-interrumpido.xlsx",
    )
    grupos = await _marcar_grupos_lote(db_session, lote_id, ["reintentando"])
    grupo = grupos[0]

    class FakeWsfeClient:
        async def fe_comp_consultar(
            self,
            punto_venta: int,
            tipo_cbte: int,
            numero: int,
        ) -> ArcaComprobanteResponse:
            return ArcaComprobanteResponse(
                punto_venta=punto_venta,
                tipo_cbte=tipo_cbte,
                numero=numero,
                cuit_emisor=test_empresa.cuit,
                cae=CAE_TEST_NO_REAL_40,
                cae_vencimiento="20260630",
                fecha_cbte=str(grupo.fecha_emision).replace("-", ""),
                fecha_proceso="20260601",
                imp_total=1210.0,
                imp_neto=1000.0,
                imp_iva=210.0,
                imp_op_ex=0.0,
                imp_tot_conc=0.0,
                imp_trib=0.0,
                moneda_id="PES",
                moneda_cotiz=1.0,
                tipo_doc=80,
                nro_doc=CUIT_RECEPTOR_TEST_NO_REAL_INT,
                resultado="A",
            )

    async def fake_get_wsfe_client(*args, **kwargs):
        return FakeWsfeClient()

    monkeypatch.setattr(
        "app.api.lotes_comprobantes.get_wsfe_client",
        fake_get_wsfe_client,
    )

    response = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/reconciliar-externos",
        headers=auth_headers,
        json={
            "comprobantes": [
                {
                    "grupo_id": grupo.id,
                    "tipo_comprobante": grupo.tipo_comprobante,
                    "punto_venta_numero": grupo.punto_venta_numero,
                    "numero": 790,
                    "fecha_emision": _fecha_argentina(grupo.fecha_emision),
                    "total": 1210.0,
                    "motivo": "Recuperación de reintento interrumpido",
                }
            ]
        },
    )

    assert response.status_code == 200, response.text
    assert response.json()["lote"]["estado"] == "cerrado_reconciliado"


@pytest.mark.asyncio
async def test_reconciliar_externo_error_arca_responde_400_controlado(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """Un error esperado de consulta ARCA no debe escapar como 500."""
    lote_id = await _crear_lote_validado_por_api(
        client,
        auth_headers,
        test_empresa.cuit,
        nombre_archivo="lote-reconciliar-error-arca.xlsx",
    )
    grupos = await _marcar_grupos_lote(db_session, lote_id, ["fallido"])
    grupo = grupos[0]

    class FakeWsfeClient:
        async def fe_comp_consultar(self, *args, **kwargs):
            raise ArcaServiceError("Comprobante inexistente")

    async def fake_get_wsfe_client(*args, **kwargs):
        return FakeWsfeClient()

    monkeypatch.setattr(
        "app.api.lotes_comprobantes.get_wsfe_client",
        fake_get_wsfe_client,
    )

    response = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/reconciliar-externos",
        headers=auth_headers,
        json={
            "comprobantes": [
                {
                    "grupo_id": grupo.id,
                    "tipo_comprobante": grupo.tipo_comprobante,
                    "punto_venta_numero": grupo.punto_venta_numero,
                    "numero": 456,
                    "fecha_emision": _fecha_argentina(grupo.fecha_emision),
                    "total": 1210.0,
                    "motivo": "Emitido manualmente por ARCA Web",
                }
            ]
        },
    )

    assert response.status_code == 400
    assert (
        "No se pudo verificar el comprobante externo contra ARCA"
        in response.json()["detail"]
    )


@pytest.mark.asyncio
async def test_reconciliar_externo_rechaza_arca_sin_cae(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """Una consulta ARCA autorizada pero sin CAE no puede cerrar el grupo."""
    lote_id = await _crear_lote_validado_por_api(
        client,
        auth_headers,
        test_empresa.cuit,
        nombre_archivo="lote-reconciliar-sin-cae.xlsx",
    )
    grupos = await _marcar_grupos_lote(db_session, lote_id, ["fallido"])
    grupo = grupos[0]

    class FakeWsfeClient:
        async def fe_comp_consultar(
            self,
            punto_venta: int,
            tipo_cbte: int,
            numero: int,
        ) -> ArcaComprobanteResponse:
            return ArcaComprobanteResponse(
                punto_venta=punto_venta,
                tipo_cbte=tipo_cbte,
                numero=numero,
                cuit_emisor=test_empresa.cuit,
                cae="",
                cae_vencimiento="20260630",
                fecha_cbte=str(grupo.fecha_emision).replace("-", ""),
                fecha_proceso="20260601",
                imp_total=1210.0,
                imp_neto=1000.0,
                imp_iva=210.0,
                imp_op_ex=0.0,
                imp_tot_conc=0.0,
                imp_trib=0.0,
                moneda_id="PES",
                moneda_cotiz=1.0,
                tipo_doc=80,
                nro_doc=CUIT_RECEPTOR_TEST_NO_REAL_INT,
                resultado="A",
            )

    async def fake_get_wsfe_client(*args, **kwargs):
        return FakeWsfeClient()

    monkeypatch.setattr(
        "app.api.lotes_comprobantes.get_wsfe_client",
        fake_get_wsfe_client,
    )

    response = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/reconciliar-externos",
        headers=auth_headers,
        json={
            "comprobantes": [
                {
                    "grupo_id": grupo.id,
                    "tipo_comprobante": grupo.tipo_comprobante,
                    "punto_venta_numero": grupo.punto_venta_numero,
                    "numero": 456,
                    "fecha_emision": _fecha_argentina(grupo.fecha_emision),
                    "total": 1210.0,
                    "motivo": "Emitido manualmente por ARCA Web",
                }
            ]
        },
    )

    assert response.status_code == 400
    assert "ARCA no devolvió un CAE válido" in response.json()["detail"]
    await db_session.refresh(grupo)
    assert grupo.estado == "fallido"
    assert grupo.comprobante_id is None


@pytest.mark.asyncio
async def test_reconciliar_externo_con_error_incompleto_responde_400(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """Un grupo observado sin payload fiscal completo no debe escapar como 500."""
    lote_id = await _crear_lote_validado_por_api(
        client,
        auth_headers,
        test_empresa.cuit,
        nombre_archivo="lote-reconciliar-con-error-incompleto.xlsx",
    )
    grupos = await _marcar_grupos_lote(db_session, lote_id, ["con_error"])
    grupo = grupos[0]
    fecha_emision = str(grupo.fecha_emision)
    grupo.payload_json = {}
    await db_session.commit()

    class FakeWsfeClient:
        async def fe_comp_consultar(self, *args, **kwargs):  # pragma: no cover
            raise AssertionError("No debe consultar ARCA con payload incompleto")

    async def fake_get_wsfe_client(*args, **kwargs):
        return FakeWsfeClient()

    monkeypatch.setattr(
        "app.api.lotes_comprobantes.get_wsfe_client",
        fake_get_wsfe_client,
    )

    response = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/reconciliar-externos",
        headers=auth_headers,
        json={
            "comprobantes": [
                {
                    "grupo_id": grupo.id,
                    "tipo_comprobante": grupo.tipo_comprobante,
                    "punto_venta_numero": grupo.punto_venta_numero,
                    "numero": 456,
                    "fecha_emision": fecha_emision,
                    "total": 1210.0,
                    "motivo": "Emitido manualmente por ARCA Web",
                }
            ]
        },
    )

    assert response.status_code == 400
    assert "datos fiscales completos" in response.json()["detail"]


@pytest.mark.asyncio
async def test_reconciliar_externos_multi_item_es_atomico_si_un_item_falla(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """Una reconciliación parcial fallida no debe dejar comprobantes huérfanos."""
    lote_id = await _crear_lote_validado_por_api(
        client,
        auth_headers,
        test_empresa.cuit,
        nombre_archivo="lote-reconciliar-atomico.xlsx",
        total_grupos=2,
    )
    grupos = await _marcar_grupos_lote(db_session, lote_id, ["fallido", "fallido"])
    consultas = 0

    class FakeWsfeClient:
        async def fe_comp_consultar(
            self,
            punto_venta: int,
            tipo_cbte: int,
            numero: int,
        ) -> ArcaComprobanteResponse:
            nonlocal consultas
            consultas += 1
            if consultas == 2:
                raise ArcaServiceError("Comprobante inexistente")
            return ArcaComprobanteResponse(
                punto_venta=punto_venta,
                tipo_cbte=tipo_cbte,
                numero=numero,
                cuit_emisor=test_empresa.cuit,
                cae=CAE_TEST_NO_REAL_37,
                cae_vencimiento="20260630",
                fecha_cbte=str(grupos[0].fecha_emision).replace("-", ""),
                fecha_proceso="20260601",
                imp_total=1210.0,
                imp_neto=1000.0,
                imp_iva=210.0,
                imp_op_ex=0.0,
                imp_tot_conc=0.0,
                imp_trib=0.0,
                moneda_id="PES",
                moneda_cotiz=1.0,
                tipo_doc=80,
                nro_doc=CUIT_RECEPTOR_TEST_NO_REAL_INT,
                resultado="A",
            )

    async def fake_get_wsfe_client(*args, **kwargs):
        return FakeWsfeClient()

    monkeypatch.setattr(
        "app.api.lotes_comprobantes.get_wsfe_client",
        fake_get_wsfe_client,
    )

    response = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/reconciliar-externos",
        headers=auth_headers,
        json={
            "comprobantes": [
                {
                    "grupo_id": grupos[0].id,
                    "tipo_comprobante": grupos[0].tipo_comprobante,
                    "punto_venta_numero": grupos[0].punto_venta_numero,
                    "numero": 456,
                    "fecha_emision": str(grupos[0].fecha_emision),
                    "total": 1210.0,
                    "motivo": "Emitido manualmente por ARCA Web",
                },
                {
                    "grupo_id": grupos[1].id,
                    "tipo_comprobante": grupos[1].tipo_comprobante,
                    "punto_venta_numero": grupos[1].punto_venta_numero,
                    "numero": 457,
                    "fecha_emision": str(grupos[1].fecha_emision),
                    "total": 1210.0,
                    "motivo": "Emitido manualmente por ARCA Web",
                },
            ]
        },
    )

    assert response.status_code == 400
    assert consultas == 2

    comprobantes = await db_session.scalar(
        select(func.count())
        .select_from(Comprobante)
        .where(Comprobante.origen_emision == "arca_web")
    )
    assert comprobantes == 0

    for grupo in grupos:
        await db_session.refresh(grupo)
        assert grupo.estado == "fallido"
        assert grupo.comprobante_id is None


@pytest.mark.asyncio
async def test_compactar_lote_cerrado_elimina_filas_y_bloquea_observado(
    client: AsyncClient,
    auth_headers: dict,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """Compactar debe eliminar filas pesadas sin borrar grupos ni resumen."""
    lote_id = await _crear_lote_validado_por_api(
        client,
        auth_headers,
        test_empresa.cuit,
        nombre_archivo="lote-compactar.xlsx",
    )
    await _marcar_grupos_lote(db_session, lote_id, ["autorizado"])

    response = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/compactar",
        headers=auth_headers,
    )

    assert response.status_code == 200, response.text
    lote_data = response.json()["lote"]
    assert lote_data["compactado_at"] is not None

    filas = await db_session.scalar(
        select(func.count())
        .select_from(LoteComprobanteFila)
        .where(LoteComprobanteFila.lote_id == lote_id)
    )
    grupos = await db_session.scalar(
        select(func.count())
        .select_from(LoteComprobanteGrupo)
        .where(LoteComprobanteGrupo.lote_id == lote_id)
    )
    assert filas == 0
    assert grupos == 1
    evento = (
        await db_session.execute(
            select(LoteComprobanteEvento).where(
                LoteComprobanteEvento.accion == "compactar_lote"
            )
        )
    ).scalar_one()
    assert evento.motivo == "Compactación para ahorro de almacenamiento"

    observado = await client.get(
        f"/api/lotes-comprobantes/{lote_id}/archivo-observado",
        headers=auth_headers,
    )
    assert observado.status_code == 400
    assert "compactado" in observado.json()["detail"]


@pytest.mark.asyncio
async def test_eliminar_lote_sin_emision_permite_y_conserva_evento(
    client: AsyncClient,
    auth_headers: dict,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """Solo se elimina físicamente un lote sin emisión ni incertidumbre."""
    lote_id = await _crear_lote_validado_por_api(
        client,
        auth_headers,
        test_empresa.cuit,
        nombre_archivo="lote-eliminar-sin-emision.xlsx",
    )
    await _marcar_grupos_lote(db_session, lote_id, ["con_error"])

    response = await client.request(
        "DELETE",
        f"/api/lotes-comprobantes/{lote_id}",
        headers=auth_headers,
        json={"motivo": "Carga con archivo equivocado"},
    )

    assert response.status_code == 204, response.text
    assert await db_session.get(LoteComprobante, lote_id) is None

    evento = (
        await db_session.execute(
            select(LoteComprobanteEvento).where(
                LoteComprobanteEvento.accion == "eliminar_lote"
            )
        )
    ).scalar_one()
    assert evento.lote_id is None
    assert evento.metadata_json["lote_id_original"] == lote_id


@pytest.mark.asyncio
async def test_eliminar_lote_rechaza_emitidos_o_inciertos(
    client: AsyncClient,
    auth_headers: dict,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """Un lote con emisión o incertidumbre fiscal no puede borrarse."""
    lote_id = await _crear_lote_validado_por_api(
        client,
        auth_headers,
        test_empresa.cuit,
        nombre_archivo="lote-no-eliminar-emitido.xlsx",
    )
    await _marcar_grupos_lote(db_session, lote_id, ["autorizado"])

    response = await client.request(
        "DELETE",
        f"/api/lotes-comprobantes/{lote_id}",
        headers=auth_headers,
        json={"motivo": "No quiero conservarlo"},
    )

    assert response.status_code == 400
    assert "comprobantes emitidos" in response.json()["detail"]
    assert await db_session.get(LoteComprobante, lote_id) is not None


@pytest.mark.asyncio
async def test_eliminar_lote_rechaza_cualquier_intento_con_conflicto(
    client: AsyncClient,
    auth_headers: dict,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """Todo intento fiscal preserva el lote aunque haya fallado de forma segura."""
    lote_id = await _crear_lote_validado_por_api(
        client,
        auth_headers,
        test_empresa.cuit,
        nombre_archivo="lote-no-eliminar-intento.xlsx",
    )
    await _marcar_grupos_lote(db_session, lote_id, ["requiere_reconciliacion"])
    grupo = (
        await db_session.execute(
            select(LoteComprobanteGrupo).where(LoteComprobanteGrupo.lote_id == lote_id)
        )
    ).scalar_one()
    db_session.add(
        IntentoEmisionFiscal(
            tipo_comprobante=6,
            punto_venta_numero=test_punto_venta.numero,
            fecha_emision=date(2026, 8, 8),
            total=Decimal("121.00"),
            payload_hash="1" * 64,
            huella_logica="2" * 64,
            estado="fallido_verificado",
            empresa_id=test_empresa.id,
            punto_venta_id=test_punto_venta.id,
            lote_id=lote_id,
            grupo_id=grupo.id,
        )
    )
    await db_session.commit()

    response = await client.request(
        "DELETE",
        f"/api/lotes-comprobantes/{lote_id}",
        headers=auth_headers,
        json={"motivo": "No quiero conservarlo"},
    )

    assert response.status_code == 409
    assert "intentos fiscales" in response.json()["detail"]
    assert await db_session.get(LoteComprobante, lote_id) is not None


@pytest.mark.asyncio
async def test_eliminar_lote_reconciliacion_sin_intentos_devuelve_conflicto(
    client: AsyncClient,
    auth_headers: dict,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
) -> None:
    """Un lote incierto devuelve 409 aun sin una fila de intento asociada."""
    lote_id = await _crear_lote_validado_por_api(
        client,
        auth_headers,
        test_empresa.cuit,
        nombre_archivo="lote-no-eliminar-reconciliacion.xlsx",
    )
    lote = await db_session.get(LoteComprobante, lote_id)
    assert lote is not None
    lote.estado = "requiere_reconciliacion"
    grupos = list(
        (
            await db_session.scalars(
                select(LoteComprobanteGrupo).where(
                    LoteComprobanteGrupo.lote_id == lote_id
                )
            )
        ).all()
    )
    filas = list(
        (
            await db_session.scalars(
                select(LoteComprobanteFila).where(
                    LoteComprobanteFila.lote_id == lote_id
                )
            )
        ).all()
    )
    for grupo in grupos:
        grupo.estado = "requiere_reconciliacion"
    for fila in filas:
        fila.estado = "requiere_reconciliacion"
    await db_session.commit()

    response = await client.request(
        "DELETE",
        f"/api/lotes-comprobantes/{lote_id}",
        headers=auth_headers,
        json={"motivo": "No quiero conservarlo"},
    )

    assert response.status_code == 409
    assert "requiere reconciliación" in response.json()["detail"]
    assert await db_session.get(LoteComprobante, lote_id) is not None


@pytest.mark.asyncio
async def test_reanudar_lote_vincula_comprobante_ya_guardado_sin_reemitir(
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
):
    """Si el comprobante ya fue guardado, reanudar no debe volver a emitirlo."""
    fecha_fiscal = date(2026, 3, 20)
    emitir_request = {
        "empresa_id": test_empresa.id,
        "punto_venta_id": test_punto_venta.id,
        "tipo_comprobante": 6,
        "concepto": 1,
        "fecha_emision": fecha_fiscal.isoformat(),
        "confirmacion_fecha_fiscal": True,
        "tipo_documento": 80,
        "numero_documento": CUIT_RECEPTOR_TEST_NO_REAL,
        "razon_social": "Cliente Lote SA",
        "condicion_iva": "RI",
        "domicilio": "Av. Siempre Viva 123",
        "moneda": "PES",
        "cotizacion": "1",
        "guardar_cliente": False,
        "items": [
            {
                "descripcion": "Servicio mensual",
                "cantidad": "1",
                "unidad": "unidad",
                "precio_unitario": "1000",
                "iva_porcentaje": "21",
            }
        ],
    }
    lote = LoteComprobante(
        nombre_archivo="lote-reanudar-idempotente.xlsx",
        archivo_hash="hash-reanudar-idempotente",
        estado="procesando",
        total_filas=1,
        total_grupos=1,
        grupos_validos=1,
        empresa_id=test_empresa.id,
        metadata_json={
            "opciones_concepto": {"concepto_modo": "archivo"},
            "opciones_descripcion_item": {"descripcion_item_modo": "archivo"},
        },
        updated_at=datetime.utcnow()
        - timedelta(minutes=settings.batch_processing_stale_minutes + 1),
    )
    grupo = LoteComprobanteGrupo(
        lote=lote,
        comprobante_ref="LOTE-001",
        orden=1,
        estado="validado",
        tipo_comprobante=6,
        punto_venta_numero=test_punto_venta.numero,
        cliente_documento=CUIT_RECEPTOR_TEST_NO_REAL,
        cliente_razon_social="Cliente Lote SA",
        total_estimado=Decimal("1210.00"),
        payload_json=emitir_request,
        mensajes_json=["Validado correctamente. Listo para emitir."],
    )
    fila = LoteComprobanteFila(
        lote=lote,
        grupo=grupo,
        fila_excel=2,
        comprobante_ref="LOTE-001",
        estado="validado",
        datos_json={},
        mensajes_json=["Validado correctamente. Listo para emitir."],
    )
    comprobante = Comprobante(
        tipo_comprobante=6,
        concepto=1,
        numero=77,
        fecha_emision=fecha_fiscal,
        subtotal=Decimal("1000.00"),
        descuento=Decimal("0.00"),
        iva_21=Decimal("210.00"),
        iva_10_5=Decimal("0.00"),
        iva_27=Decimal("0.00"),
        otros_impuestos=Decimal("0.00"),
        total=Decimal("1210.00"),
        cae=CAE_TEST_NO_REAL,
        cae_vencimiento=date(2026, 5, 26),
        estado="autorizado",
        moneda="PES",
        cotizacion=Decimal("1"),
        empresa_id=test_empresa.id,
        punto_venta_id=test_punto_venta.id,
        receptor_tipo_documento=80,
        receptor_numero_documento=CUIT_RECEPTOR_TEST_NO_REAL,
        receptor_razon_social="Cliente Lote SA",
        receptor_condicion_iva="RI",
        receptor_domicilio="Av. Siempre Viva 123",
    )
    comprobante.items = [
        ComprobanteItem(
            descripcion="Servicio mensual",
            cantidad=Decimal("1"),
            unidad="unidad",
            precio_unitario=Decimal("1000"),
            descuento_porcentaje=Decimal("0"),
            iva_porcentaje=Decimal("21"),
            subtotal=Decimal("1000.00"),
            orden=0,
        )
    ]
    request_model = EmitirComprobanteRequest.model_validate(emitir_request)
    payload_hash, huella = _hashes_fiscales_request(
        request_model,
        test_punto_venta.numero,
        Decimal("1210.00"),
    )
    db_session.add_all([lote, grupo, fila, comprobante])
    await db_session.flush()
    intento = IntentoEmisionFiscal(
        tipo_comprobante=6,
        punto_venta_numero=test_punto_venta.numero,
        numero_planificado=77,
        fecha_emision=fecha_fiscal,
        total=Decimal("1210.00"),
        receptor_tipo_documento=80,
        receptor_numero_documento=CUIT_RECEPTOR_TEST_NO_REAL,
        receptor_razon_social="Cliente Lote SA",
        payload_hash=payload_hash,
        huella_logica=huella,
        cae=comprobante.cae,
        cae_vencimiento=comprobante.cae_vencimiento,
        estado="autorizado",
        empresa_id=test_empresa.id,
        punto_venta_id=test_punto_venta.id,
        comprobante_id=comprobante.id,
        lote_id=lote.id,
        grupo_id=grupo.id,
    )
    db_session.add(intento)
    await db_session.commit()
    await db_session.refresh(lote)
    lote_id = lote.id
    empresa_id = test_empresa.id
    comprobante_id = comprobante.id
    comprobante_numero = comprobante.numero
    db_session.expire_all()

    service = LoteComprobantesService(db_session)

    async def fail_emitir(_request, **kwargs):
        raise AssertionError("No debe reemitir un grupo ya guardado")

    service.facturacion_service.emitir_comprobante = fail_emitir

    resultado = await service.procesar_lote(lote_id, empresa_id, reanudar=True)

    assert resultado.estado == "completado"
    assert resultado.finished_at is not None
    detalle = await service.obtener_lote(lote_id, empresa_id)
    assert detalle.grupos[0].estado == "autorizado"
    assert detalle.grupos[0].comprobante_id == comprobante_id
    assert detalle.grupos[0].numero_asignado == comprobante_numero


@pytest.mark.asyncio
async def test_reanudar_lote_stale_con_grupos_autorizados_cierra_sin_reconciliacion(
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
) -> None:
    """Un lote stale localmente emitido se cierra sin pedir nuevos CAE."""
    fecha_fiscal = date(2026, 3, 20)
    emitir_request = {
        "empresa_id": test_empresa.id,
        "punto_venta_id": test_punto_venta.id,
        "tipo_comprobante": 6,
        "concepto": 1,
        "fecha_emision": fecha_fiscal.isoformat(),
        "confirmacion_fecha_fiscal": True,
        "tipo_documento": 80,
        "numero_documento": CUIT_RECEPTOR_TEST_NO_REAL,
        "razon_social": "Cliente Lote SA",
        "condicion_iva": "RI",
        "domicilio": "Av. Siempre Viva 123",
        "moneda": "PES",
        "cotizacion": "1",
        "guardar_cliente": False,
        "items": [
            {
                "descripcion": "Servicio mensual",
                "cantidad": "1",
                "unidad": "unidad",
                "precio_unitario": "1000",
                "iva_porcentaje": "21",
            }
        ],
    }
    lote = LoteComprobante(
        nombre_archivo="lote-stale-ya-autorizado.xlsx",
        archivo_hash="hash-stale-ya-autorizado",
        estado="procesando",
        total_filas=1,
        total_grupos=1,
        empresa_id=test_empresa.id,
        updated_at=datetime.utcnow()
        - timedelta(minutes=settings.batch_processing_stale_minutes + 1),
    )
    comprobante = Comprobante(
        tipo_comprobante=6,
        concepto=1,
        numero=77,
        fecha_emision=fecha_fiscal,
        subtotal=Decimal("1000.00"),
        descuento=Decimal("0.00"),
        iva_21=Decimal("210.00"),
        iva_10_5=Decimal("0.00"),
        iva_27=Decimal("0.00"),
        otros_impuestos=Decimal("0.00"),
        total=Decimal("1210.00"),
        cae=CAE_TEST_NO_REAL,
        cae_vencimiento=date(2026, 5, 26),
        estado="autorizado",
        moneda="PES",
        cotizacion=Decimal("1"),
        empresa_id=test_empresa.id,
        punto_venta_id=test_punto_venta.id,
        receptor_tipo_documento=80,
        receptor_numero_documento=CUIT_RECEPTOR_TEST_NO_REAL,
        receptor_razon_social="Cliente Lote SA",
        receptor_condicion_iva="RI",
        receptor_domicilio="Av. Siempre Viva 123",
    )
    comprobante.items = [
        ComprobanteItem(
            descripcion="Servicio mensual",
            cantidad=Decimal("1"),
            unidad="unidad",
            precio_unitario=Decimal("1000"),
            descuento_porcentaje=Decimal("0"),
            iva_porcentaje=Decimal("21"),
            subtotal=Decimal("1000.00"),
            orden=0,
        )
    ]
    request_model = EmitirComprobanteRequest.model_validate(emitir_request)
    payload_hash, huella = _hashes_fiscales_request(
        request_model,
        test_punto_venta.numero,
        Decimal("1210.00"),
    )
    db_session.add_all([lote, comprobante])
    await db_session.flush()
    grupo = LoteComprobanteGrupo(
        lote=lote,
        comprobante_ref="LOTE-001",
        orden=1,
        estado="autorizado",
        tipo_comprobante=6,
        punto_venta_numero=test_punto_venta.numero,
        cliente_documento=CUIT_RECEPTOR_TEST_NO_REAL,
        cliente_razon_social="Cliente Lote SA",
        total_estimado=Decimal("1210.00"),
        payload_json=emitir_request,
        cae=comprobante.cae,
        numero_asignado=comprobante.numero,
        comprobante_id=comprobante.id,
        mensajes_json=["Comprobante autorizado."],
    )
    fila = LoteComprobanteFila(
        lote=lote,
        grupo=grupo,
        fila_excel=2,
        comprobante_ref="LOTE-001",
        estado="autorizado",
        datos_json={},
        mensajes_json=["Comprobante autorizado."],
    )
    db_session.add_all([grupo, fila])
    await db_session.flush()
    intento = IntentoEmisionFiscal(
        tipo_comprobante=6,
        punto_venta_numero=test_punto_venta.numero,
        numero_planificado=comprobante.numero,
        fecha_emision=fecha_fiscal,
        total=Decimal("1210.00"),
        receptor_tipo_documento=80,
        receptor_numero_documento=CUIT_RECEPTOR_TEST_NO_REAL,
        receptor_razon_social="Cliente Lote SA",
        payload_hash=payload_hash,
        huella_logica=huella,
        cae=comprobante.cae,
        cae_vencimiento=comprobante.cae_vencimiento,
        estado="autorizado",
        empresa_id=test_empresa.id,
        punto_venta_id=test_punto_venta.id,
        comprobante_id=comprobante.id,
        lote_id=lote.id,
        grupo_id=grupo.id,
    )
    db_session.add(intento)
    await db_session.commit()
    await db_session.refresh(lote)
    lote_id = lote.id
    empresa_id = test_empresa.id
    db_session.expire_all()

    service = LoteComprobantesService(db_session)
    llamadas_arca = _instalar_oraculos_stale_sin_arca(service)

    resultado = await service.procesar_lote(lote_id, empresa_id, reanudar=True)

    assert resultado.estado == "completado"
    assert resultado.finished_at is not None
    assert resultado.grupos_emitidos == 1
    assert "Todos los comprobantes" in resultado.mensaje_resumen

    eventos = (
        (
            await db_session.execute(
                select(LoteComprobanteEvento).where(
                    LoteComprobanteEvento.lote_id == lote_id,
                    LoteComprobanteEvento.accion == "reconciliacion_local_stale",
                )
            )
        )
        .scalars()
        .all()
    )
    assert len(eventos) == 1
    assert eventos[0].metadata_json["grupos_reconciliados"] == 0
    assert llamadas_arca == {"wsaa": 0, "fecomp": 0, "fecae": 0}


@pytest.mark.asyncio
async def test_reanudar_lote_stale_legacy_no_consulta_arca_y_exige_reconciliacion(
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
) -> None:
    """Un lote stale legacy nunca consulta ARCA ni vuelve automáticamente a cola."""
    fecha_fiscal = date(2026, 3, 20)
    payload_autorizado = _payload_lote_basico(
        test_empresa.id,
        test_punto_venta.id,
        fecha_fiscal,
        razon_social="Cliente Autorizado SA",
    )
    payload_pendiente = _payload_lote_basico(
        test_empresa.id,
        test_punto_venta.id,
        fecha_fiscal,
        razon_social="Cliente Pendiente SA",
    )
    lote = LoteComprobante(
        nombre_archivo="lote-stale-parcial-intacto.xlsx",
        archivo_hash="hash-stale-parcial-intacto",
        estado="procesando",
        total_filas=2,
        total_grupos=2,
        empresa_id=test_empresa.id,
        metadata_json={
            "opciones_concepto": {"concepto_modo": "archivo"},
            "opciones_descripcion_item": {"descripcion_item_modo": "archivo"},
        },
        updated_at=datetime.utcnow()
        - timedelta(minutes=settings.batch_processing_stale_minutes + 1),
    )
    comprobante = Comprobante(
        tipo_comprobante=6,
        concepto=1,
        numero=77,
        fecha_emision=fecha_fiscal,
        subtotal=Decimal("1000.00"),
        descuento=Decimal("0.00"),
        iva_21=Decimal("210.00"),
        iva_10_5=Decimal("0.00"),
        iva_27=Decimal("0.00"),
        otros_impuestos=Decimal("0.00"),
        total=Decimal("1210.00"),
        cae=CAE_TEST_NO_REAL,
        cae_vencimiento=date(2026, 5, 26),
        estado="autorizado",
        moneda="PES",
        cotizacion=Decimal("1"),
        empresa_id=test_empresa.id,
        punto_venta_id=test_punto_venta.id,
        receptor_tipo_documento=80,
        receptor_numero_documento=CUIT_RECEPTOR_TEST_NO_REAL,
        receptor_razon_social="Cliente Autorizado SA",
        receptor_condicion_iva="RI",
        receptor_domicilio="Av. Siempre Viva 123",
    )
    comprobante.items = [
        ComprobanteItem(
            descripcion="Servicio mensual",
            cantidad=Decimal("1"),
            unidad="unidad",
            precio_unitario=Decimal("1000"),
            descuento_porcentaje=Decimal("0"),
            iva_porcentaje=Decimal("21"),
            subtotal=Decimal("1000.00"),
            orden=0,
        )
    ]
    db_session.add_all([lote, comprobante])
    await db_session.flush()

    grupo_autorizado = LoteComprobanteGrupo(
        lote=lote,
        comprobante_ref="LOTE-001",
        orden=1,
        estado="autorizado",
        tipo_comprobante=6,
        punto_venta_numero=test_punto_venta.numero,
        cliente_documento=CUIT_RECEPTOR_TEST_NO_REAL,
        cliente_razon_social="Cliente Autorizado SA",
        total_estimado=Decimal("1210.00"),
        payload_json=payload_autorizado,
        cae=comprobante.cae,
        numero_asignado=comprobante.numero,
        comprobante_id=comprobante.id,
        mensajes_json=["Comprobante autorizado."],
    )
    grupo_pendiente = LoteComprobanteGrupo(
        lote=lote,
        comprobante_ref="LOTE-002",
        orden=2,
        estado="validado",
        tipo_comprobante=6,
        punto_venta_numero=test_punto_venta.numero,
        cliente_documento=CUIT_RECEPTOR_TEST_NO_REAL,
        cliente_razon_social="Cliente Pendiente SA",
        total_estimado=Decimal("1210.00"),
        payload_json=payload_pendiente,
        mensajes_json=["Validado correctamente. Listo para emitir."],
    )
    filas = [
        LoteComprobanteFila(
            lote=lote,
            grupo=grupo_autorizado,
            fila_excel=2,
            comprobante_ref="LOTE-001",
            estado="autorizado",
            datos_json={},
            mensajes_json=["Comprobante autorizado."],
        ),
        LoteComprobanteFila(
            lote=lote,
            grupo=grupo_pendiente,
            fila_excel=3,
            comprobante_ref="LOTE-002",
            estado="validado",
            datos_json={},
            mensajes_json=["Validado correctamente. Listo para emitir."],
        ),
    ]
    db_session.add_all([grupo_autorizado, grupo_pendiente, *filas])
    await db_session.flush()

    request_autorizado = EmitirComprobanteRequest.model_validate(payload_autorizado)
    payload_hash, huella = _hashes_fiscales_request(
        request_autorizado,
        test_punto_venta.numero,
        Decimal("1210.00"),
    )
    db_session.add(
        IntentoEmisionFiscal(
            tipo_comprobante=6,
            punto_venta_numero=test_punto_venta.numero,
            numero_planificado=comprobante.numero,
            fecha_emision=fecha_fiscal,
            total=Decimal("1210.00"),
            receptor_tipo_documento=80,
            receptor_numero_documento=CUIT_RECEPTOR_TEST_NO_REAL,
            receptor_razon_social="Cliente Autorizado SA",
            payload_hash=payload_hash,
            huella_logica=huella,
            cae=comprobante.cae,
            cae_vencimiento=comprobante.cae_vencimiento,
            estado="autorizado",
            empresa_id=test_empresa.id,
            punto_venta_id=test_punto_venta.id,
            comprobante_id=comprobante.id,
            lote_id=lote.id,
            grupo_id=grupo_autorizado.id,
        )
    )
    lote_id = lote.id
    empresa_id = test_empresa.id
    await db_session.commit()
    db_session.expire_all()

    service = LoteComprobantesService(db_session)
    llamadas_arca = _instalar_oraculos_stale_sin_arca(service)

    resultado = await service.bloquear_lote_procesando_stale(
        lote_id,
        empresa_id,
    )

    assert resultado.estado == "requiere_reconciliacion"
    assert resultado.finished_at is not None
    assert resultado.grupos_emitidos == 1
    assert resultado.grupos_validos == 0

    segundo_resultado = await service.bloquear_lote_procesando_stale(
        lote_id,
        empresa_id,
    )
    assert segundo_resultado.estado == "requiere_reconciliacion"

    detalle = await service.obtener_lote(lote_id, empresa_id)
    pendiente = next(
        grupo for grupo in detalle.grupos if grupo.comprobante_ref == "LOTE-002"
    )
    assert pendiente.estado == "requiere_reconciliacion"
    assert pendiente.cae is None
    assert pendiente.numero_asignado is None
    assert pendiente.comprobante_id is None
    intentos_pendiente = await db_session.scalar(
        select(func.count(IntentoEmisionFiscal.id)).where(
            IntentoEmisionFiscal.grupo_id == pendiente.id
        )
    )
    assert intentos_pendiente == 0

    eventos = (
        (
            await db_session.execute(
                select(LoteComprobanteEvento).where(
                    LoteComprobanteEvento.lote_id == lote_id,
                    LoteComprobanteEvento.accion == "bloqueo_operativo_no_reemitir",
                )
            )
        )
        .scalars()
        .all()
    )
    assert len(eventos) == 1
    assert eventos[0].metadata_json["estado_nuevo"] == "requiere_reconciliacion"
    assert eventos[0].metadata_json["grupos_marcados_reconciliacion"] == 1
    assert eventos[0].metadata_json["preflight_arca"] == []
    assert eventos[0].metadata_json["preflight_error"] == (
        "operacion_o_snapshot_rece_legacy"
    )
    assert llamadas_arca == {"wsaa": 0, "fecomp": 0, "fecae": 0}


@pytest.mark.asyncio
@pytest.mark.parametrize("segunda_combinacion_segura", [True, False])
async def test_preflight_stale_exige_todas_las_combinaciones_seguras(
    db_session: AsyncSession,
    test_empresa,
    test_user,
    test_punto_venta,
    segunda_combinacion_segura: bool,
) -> None:
    """Un lote mixto solo supera el preflight si todas sus combinaciones son seguras."""
    fecha_fiscal = date(2026, 8, 4)
    empresa_id = int(test_empresa.id)
    primer_punto_venta_id = int(test_punto_venta.id)
    primer_punto_venta_numero = int(test_punto_venta.numero)
    segundo_punto_venta = await _crear_punto_venta_rece_verificado(
        db_session,
        test_empresa,
        usuario_id=int(test_user.id),
        numero=2,
        nombre="Punto stale 2",
        documento_emitido_en=FECHA_DOCUMENTO_RECE_TEST,
        vigente_hasta=FECHA_VIGENCIA_RECE_TEST,
        observado_en=INSTANTE_RECE_TEST,
    )
    segundo_punto_venta_id = int(segundo_punto_venta.id)
    lote, grupos = await _crear_lote_stale_moderno_intacto(
        db_session,
        test_empresa,
        grupos_payload=[
            (
                "MIXTO-001",
                _payload_lote_basico(
                    empresa_id,
                    primer_punto_venta_id,
                    fecha_fiscal,
                    razon_social="Cliente Alineado SA",
                ),
            ),
            (
                "MIXTO-002",
                _payload_lote_basico(
                    empresa_id,
                    segundo_punto_venta_id,
                    fecha_fiscal,
                    razon_social="Cliente Historia Externa SA",
                ),
            ),
        ],
        idempotency_key=(
            f"idem-stale-combinaciones-{str(segunda_combinacion_segura).lower()}"
        ),
    )
    service = LoteComprobantesService(db_session)
    llamadas: list[dict[str, int]] = []

    async def fake_preflight(**kwargs):
        llamadas.append(dict(kwargs))
        if kwargs["punto_venta_id"] == segundo_punto_venta_id:
            if not segunda_combinacion_segura:
                raise RuntimeError("segunda combinación no verificable")
            return {
                **kwargs,
                "punto_venta_numero": 2,
                "ultimo_local": 70,
                "ultimo_arca": 75,
                "proximo_local": 71,
                "proximo_arca": 76,
                "proximo_numero": 76,
                "estado": "arca_adelantada",
            }
        return {
            **kwargs,
            "punto_venta_numero": primer_punto_venta_numero,
            "ultimo_local": 70,
            "ultimo_arca": 70,
            "proximo_local": 71,
            "proximo_arca": 71,
            "proximo_numero": 71,
            "estado": "alineada",
        }

    service.facturacion_service.verificar_numeracion_segura_para_emision = (
        fake_preflight
    )

    (
        preflight_ok,
        checks,
        error,
    ) = await service._preflight_reanudar_grupos_intactos_stale(lote, grupos)

    assert llamadas == [
        {
            "empresa_id": empresa_id,
            "punto_venta_id": primer_punto_venta_id,
            "tipo_comprobante": 6,
        },
        {
            "empresa_id": empresa_id,
            "punto_venta_id": segundo_punto_venta_id,
            "tipo_comprobante": 6,
        },
    ]
    if segunda_combinacion_segura:
        assert preflight_ok is True
        assert error is None
        assert [check["estado"] for check in checks] == [
            "alineada",
            "arca_adelantada",
        ]
    else:
        assert preflight_ok is False
        assert error == "numeracion_no_verificable"
        assert [check["estado"] for check in checks] == ["alineada"]


@pytest.mark.asyncio
async def test_preflight_stale_bloquea_payload_con_clave_superior_desconocida(
    db_session: AsyncSession,
    test_empresa,
    test_user,
    test_punto_venta,
) -> None:
    """Un payload no canónico bloquea todo el conjunto antes del preflight."""
    fecha_fiscal = date(2026, 8, 5)
    empresa_id = int(test_empresa.id)
    primer_punto_venta_id = int(test_punto_venta.id)
    segundo_punto_venta = await _crear_punto_venta_rece_verificado(
        db_session,
        test_empresa,
        usuario_id=int(test_user.id),
        numero=2,
        nombre="Punto stale payload 2",
        documento_emitido_en=FECHA_DOCUMENTO_RECE_TEST,
        vigente_hasta=FECHA_VIGENCIA_RECE_TEST,
        observado_en=INSTANTE_RECE_TEST,
    )
    segundo_punto_venta_id = int(segundo_punto_venta.id)
    payload_invalido = _payload_lote_basico(
        empresa_id,
        segundo_punto_venta_id,
        fecha_fiscal,
        razon_social="Cliente con payload no canónico SA",
    )
    payload_invalido["monedaa"] = "USD"
    lote, grupos = await _crear_lote_stale_moderno_intacto(
        db_session,
        test_empresa,
        grupos_payload=[
            (
                "MIXTO-VALIDO",
                _payload_lote_basico(
                    empresa_id,
                    primer_punto_venta_id,
                    fecha_fiscal,
                ),
            ),
            ("MIXTO-INVALIDO", payload_invalido),
        ],
        idempotency_key="idem-stale-payload-invalido",
    )
    service = LoteComprobantesService(db_session)
    llamadas_preflight = 0

    async def fail_preflight(**kwargs):
        nonlocal llamadas_preflight
        llamadas_preflight += 1
        raise AssertionError("No debe consultar numeración con payload inválido")

    service.facturacion_service.verificar_numeracion_segura_para_emision = (
        fail_preflight
    )

    (
        preflight_ok,
        checks,
        error,
    ) = await service._preflight_reanudar_grupos_intactos_stale(lote, grupos)

    assert preflight_ok is False
    assert checks == []
    assert error == "payload_fiscal_invalido"
    assert llamadas_preflight == 0


@pytest.mark.asyncio
async def test_reanudar_lote_stale_autorizado_sin_evidencia_requiere_reconciliacion(
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
) -> None:
    """Un estado autorizado sin evidencia fiscal no alcanza para cerrar stale."""
    fecha_fiscal = date(2026, 3, 20)
    lote = LoteComprobante(
        nombre_archivo="lote-stale-autorizado-sin-evidencia.xlsx",
        archivo_hash="hash-stale-autorizado-sin-evidencia",
        estado="procesando",
        total_filas=1,
        total_grupos=1,
        empresa_id=test_empresa.id,
        updated_at=datetime.utcnow()
        - timedelta(minutes=settings.batch_processing_stale_minutes + 1),
    )
    comprobante = Comprobante(
        tipo_comprobante=6,
        concepto=1,
        numero=77,
        fecha_emision=fecha_fiscal,
        subtotal=Decimal("1000.00"),
        descuento=Decimal("0.00"),
        iva_21=Decimal("210.00"),
        iva_10_5=Decimal("0.00"),
        iva_27=Decimal("0.00"),
        otros_impuestos=Decimal("0.00"),
        total=Decimal("1210.00"),
        cae=CAE_TEST_NO_REAL,
        cae_vencimiento=date(2026, 5, 26),
        estado="autorizado",
        moneda="PES",
        cotizacion=Decimal("1"),
        empresa_id=test_empresa.id,
        punto_venta_id=test_punto_venta.id,
        receptor_tipo_documento=80,
        receptor_numero_documento=CUIT_RECEPTOR_TEST_NO_REAL,
        receptor_razon_social="Cliente Lote SA",
        receptor_condicion_iva="RI",
        receptor_domicilio="Av. Siempre Viva 123",
    )
    db_session.add_all([lote, comprobante])
    await db_session.flush()
    grupo = LoteComprobanteGrupo(
        lote=lote,
        comprobante_ref="LOTE-001",
        orden=1,
        estado="autorizado",
        tipo_comprobante=6,
        punto_venta_numero=test_punto_venta.numero,
        cliente_documento=CUIT_RECEPTOR_TEST_NO_REAL,
        cliente_razon_social="Cliente Lote SA",
        total_estimado=Decimal("1210.00"),
        cae=comprobante.cae,
        numero_asignado=comprobante.numero,
        comprobante_id=comprobante.id,
        mensajes_json=["Comprobante autorizado."],
    )
    fila = LoteComprobanteFila(
        lote=lote,
        grupo=grupo,
        fila_excel=2,
        comprobante_ref="LOTE-001",
        estado="autorizado",
        datos_json={},
        mensajes_json=["Comprobante autorizado."],
    )
    db_session.add_all([grupo, fila])
    await db_session.commit()
    await db_session.refresh(lote)

    service = LoteComprobantesService(db_session)
    llamadas_arca = _instalar_oraculos_stale_sin_arca(service)

    resultado = await service.procesar_lote(lote.id, test_empresa.id, reanudar=True)

    assert resultado.estado == "requiere_reconciliacion"
    assert resultado.grupos_emitidos == 1
    assert "reconciliar contra ARCA" in resultado.mensaje_resumen
    assert llamadas_arca == {"wsaa": 0, "fecomp": 0, "fecae": 0}


@pytest.mark.asyncio
@pytest.mark.parametrize(
    "estado_intento",
    ["en_proceso", "requiere_reconciliacion"],
)
async def test_reanudar_lote_stale_autorizado_con_intento_incierto_requiere_reconciliacion(
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    estado_intento: str,
) -> None:
    """Un lote localmente autorizado no se cierra si conserva intentos inciertos."""
    fecha_fiscal = date(2026, 3, 20)
    empresa_id = int(test_empresa.id)
    punto_venta_id = int(test_punto_venta.id)
    punto_venta_numero = int(test_punto_venta.numero)
    lote = LoteComprobante(
        nombre_archivo="lote-stale-autorizado-incierto.xlsx",
        archivo_hash="hash-stale-autorizado-incierto",
        estado="procesando",
        total_filas=1,
        total_grupos=1,
        empresa_id=empresa_id,
        updated_at=datetime.utcnow()
        - timedelta(minutes=settings.batch_processing_stale_minutes + 1),
    )
    comprobante = Comprobante(
        tipo_comprobante=6,
        concepto=1,
        numero=77,
        fecha_emision=fecha_fiscal,
        subtotal=Decimal("1000.00"),
        descuento=Decimal("0.00"),
        iva_21=Decimal("210.00"),
        iva_10_5=Decimal("0.00"),
        iva_27=Decimal("0.00"),
        otros_impuestos=Decimal("0.00"),
        total=Decimal("1210.00"),
        cae=CAE_TEST_NO_REAL,
        cae_vencimiento=date(2026, 5, 26),
        estado="autorizado",
        moneda="PES",
        cotizacion=Decimal("1"),
        empresa_id=empresa_id,
        punto_venta_id=punto_venta_id,
        receptor_tipo_documento=80,
        receptor_numero_documento=CUIT_RECEPTOR_TEST_NO_REAL,
        receptor_razon_social="Cliente Lote SA",
        receptor_condicion_iva="RI",
        receptor_domicilio="Av. Siempre Viva 123",
    )
    db_session.add_all([lote, comprobante])
    await db_session.flush()
    lote_id = int(lote.id)
    grupo = LoteComprobanteGrupo(
        lote=lote,
        empresa_id=empresa_id,
        comprobante_ref="LOTE-001",
        orden=1,
        estado="autorizado",
        tipo_comprobante=6,
        punto_venta_numero=punto_venta_numero,
        cliente_documento=CUIT_RECEPTOR_TEST_NO_REAL,
        cliente_razon_social="Cliente Lote SA",
        total_estimado=Decimal("1210.00"),
        cae=comprobante.cae,
        numero_asignado=comprobante.numero,
        comprobante_id=comprobante.id,
        mensajes_json=["Comprobante autorizado."],
    )
    fila = LoteComprobanteFila(
        lote=lote,
        grupo=grupo,
        fila_excel=2,
        comprobante_ref="LOTE-001",
        estado="autorizado",
        datos_json={},
        mensajes_json=["Comprobante autorizado."],
    )
    db_session.add_all([grupo, fila])
    await db_session.flush()
    grupo_id = int(grupo.id)
    intento = IntentoEmisionFiscal(
        tipo_comprobante=6,
        punto_venta_numero=punto_venta_numero,
        numero_planificado=78,
        fecha_emision=fecha_fiscal,
        total=Decimal("1210.00"),
        receptor_tipo_documento=80,
        receptor_numero_documento=CUIT_RECEPTOR_TEST_NO_REAL,
        receptor_razon_social="Cliente Lote SA",
        payload_hash="hash-payload-incierto",
        huella_logica="hash-huella-incierta",
        estado=estado_intento,
        operacion_id=None,
        empresa_id=empresa_id,
        punto_venta_id=punto_venta_id,
        lote_id=lote_id,
        grupo_id=grupo_id,
        ambiente=None,
        punto_venta_elegibilidad_revision_id=None,
        punto_venta_revision_fiscal=None,
        guarda_rece_id=None,
    )
    db_session.add(intento)
    assert intento.operacion_id is None
    assert intento.ambiente is None
    assert intento.punto_venta_elegibilidad_revision_id is None
    assert intento.punto_venta_revision_fiscal is None
    assert intento.guarda_rece_id is None
    await db_session.commit()

    service = LoteComprobantesService(db_session)
    llamadas_arca = _instalar_oraculos_stale_sin_arca(service)

    await service.procesar_lote(lote_id, empresa_id, reanudar=True)

    async with AsyncSession(bind=db_session.bind, expire_on_commit=False) as observador:
        lote_actual = await observador.get(LoteComprobante, lote_id)
    assert lote_actual is not None
    assert lote_actual.estado == "requiere_reconciliacion"
    assert lote_actual.grupos_emitidos == 1
    assert "reconciliar contra ARCA" in lote_actual.mensaje_resumen
    assert llamadas_arca == {"wsaa": 0, "fecomp": 0, "fecae": 0}


@pytest.mark.asyncio
async def test_reanudar_lote_no_vincula_comprobante_sin_intento_del_grupo(
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
):
    """Un comprobante parecido pero sin intento del grupo no cierra el lote."""
    fecha_fiscal = date(2026, 3, 20)
    empresa_id = int(test_empresa.id)
    punto_venta_id = int(test_punto_venta.id)
    punto_venta_numero = int(test_punto_venta.numero)
    emitir_request = {
        "empresa_id": empresa_id,
        "punto_venta_id": punto_venta_id,
        "tipo_comprobante": 6,
        "concepto": 1,
        "fecha_emision": fecha_fiscal.isoformat(),
        "confirmacion_fecha_fiscal": True,
        "tipo_documento": 80,
        "numero_documento": CUIT_RECEPTOR_TEST_NO_REAL,
        "razon_social": "Cliente Lote SA",
        "condicion_iva": "RI",
        "domicilio": "Av. Siempre Viva 123",
        "moneda": "PES",
        "cotizacion": "1",
        "guardar_cliente": False,
        "items": [
            {
                "descripcion": "Servicio mensual",
                "cantidad": "1",
                "unidad": "unidad",
                "precio_unitario": "1000",
                "iva_porcentaje": "21",
            }
        ],
    }
    lote = LoteComprobante(
        nombre_archivo="lote-reanudar-sin-intento.xlsx",
        archivo_hash="hash-reanudar-sin-intento",
        estado="procesando",
        total_filas=1,
        total_grupos=1,
        grupos_validos=1,
        empresa_id=empresa_id,
        metadata_json={
            "opciones_concepto": {"concepto_modo": "archivo"},
            "opciones_descripcion_item": {"descripcion_item_modo": "archivo"},
        },
        updated_at=datetime.utcnow()
        - timedelta(minutes=settings.batch_processing_stale_minutes + 1),
    )
    grupo = LoteComprobanteGrupo(
        lote=lote,
        empresa_id=empresa_id,
        comprobante_ref="LOTE-001",
        orden=1,
        estado="validado",
        tipo_comprobante=6,
        punto_venta_numero=punto_venta_numero,
        cliente_documento=CUIT_RECEPTOR_TEST_NO_REAL,
        cliente_razon_social="Cliente Lote SA",
        total_estimado=Decimal("1210.00"),
        payload_json=emitir_request,
        mensajes_json=["Validado correctamente. Listo para emitir."],
    )
    fila = LoteComprobanteFila(
        lote=lote,
        grupo=grupo,
        fila_excel=2,
        comprobante_ref="LOTE-001",
        estado="validado",
        datos_json={},
        mensajes_json=["Validado correctamente. Listo para emitir."],
    )
    comprobante = Comprobante(
        tipo_comprobante=6,
        concepto=1,
        numero=77,
        fecha_emision=fecha_fiscal,
        subtotal=Decimal("1000.00"),
        descuento=Decimal("0.00"),
        iva_21=Decimal("210.00"),
        iva_10_5=Decimal("0.00"),
        iva_27=Decimal("0.00"),
        otros_impuestos=Decimal("0.00"),
        total=Decimal("1210.00"),
        cae=CAE_TEST_NO_REAL,
        cae_vencimiento=date(2026, 5, 26),
        estado="autorizado",
        moneda="PES",
        cotizacion=Decimal("1"),
        empresa_id=empresa_id,
        punto_venta_id=punto_venta_id,
        receptor_tipo_documento=80,
        receptor_numero_documento=CUIT_RECEPTOR_TEST_NO_REAL,
        receptor_razon_social="Cliente Lote SA",
        receptor_condicion_iva="RI",
        receptor_domicilio="Av. Siempre Viva 123",
    )
    db_session.add_all([lote, grupo, fila, comprobante])
    await db_session.flush()
    lote_id = int(lote.id)
    assert grupo.ambiente is None
    assert grupo.punto_venta_elegibilidad_revision_id is None
    assert grupo.punto_venta_revision_fiscal is None
    await db_session.commit()

    service = LoteComprobantesService(db_session)
    llamadas_arca = _instalar_oraculos_stale_sin_arca(service)

    await service.procesar_lote(lote_id, empresa_id, reanudar=True)

    async with AsyncSession(bind=db_session.bind, expire_on_commit=False) as observador:
        detalle = await LoteComprobantesService(observador).obtener_lote(
            lote_id,
            empresa_id,
        )
    assert detalle.estado == "requiere_reconciliacion"
    assert detalle.grupos[0].estado == "requiere_reconciliacion"
    assert detalle.grupos[0].comprobante_id is None
    assert detalle.grupos[0].numero_asignado is None
    assert "reconciliar" in detalle.mensaje_resumen
    assert llamadas_arca == {"wsaa": 0, "fecomp": 0, "fecae": 0}


@pytest.mark.asyncio
async def test_reanudar_lote_no_reconcilia_intentos_autorizados_duplicados(
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
) -> None:
    """Múltiples intentos autorizados del mismo grupo requieren auditoría."""
    fecha_fiscal = date(2026, 3, 20)
    empresa_id = int(test_empresa.id)
    punto_venta_id = int(test_punto_venta.id)
    punto_venta_numero = int(test_punto_venta.numero)
    emitir_request = {
        "empresa_id": empresa_id,
        "punto_venta_id": punto_venta_id,
        "tipo_comprobante": 6,
        "concepto": 1,
        "fecha_emision": fecha_fiscal.isoformat(),
        "confirmacion_fecha_fiscal": True,
        "tipo_documento": 80,
        "numero_documento": CUIT_RECEPTOR_TEST_NO_REAL,
        "razon_social": "Cliente Lote SA",
        "condicion_iva": "RI",
        "domicilio": "Av. Siempre Viva 123",
        "moneda": "PES",
        "cotizacion": "1",
        "guardar_cliente": False,
        "items": [
            {
                "descripcion": "Servicio mensual",
                "cantidad": "1",
                "unidad": "unidad",
                "precio_unitario": "1000",
                "iva_porcentaje": "21",
            }
        ],
    }
    lote = LoteComprobante(
        nombre_archivo="lote-reanudar-intentos-duplicados.xlsx",
        archivo_hash="hash-reanudar-intentos-duplicados",
        estado="procesando",
        total_filas=1,
        total_grupos=1,
        grupos_validos=1,
        empresa_id=empresa_id,
        metadata_json={
            "opciones_concepto": {"concepto_modo": "archivo"},
            "opciones_descripcion_item": {"descripcion_item_modo": "archivo"},
        },
        updated_at=datetime.utcnow()
        - timedelta(minutes=settings.batch_processing_stale_minutes + 1),
    )
    grupo = LoteComprobanteGrupo(
        lote=lote,
        empresa_id=empresa_id,
        comprobante_ref="LOTE-001",
        orden=1,
        estado="validado",
        tipo_comprobante=6,
        punto_venta_numero=punto_venta_numero,
        cliente_documento=CUIT_RECEPTOR_TEST_NO_REAL,
        cliente_razon_social="Cliente Lote SA",
        total_estimado=Decimal("1210.00"),
        payload_json=emitir_request,
        mensajes_json=["Validado correctamente. Listo para emitir."],
    )
    fila = LoteComprobanteFila(
        lote=lote,
        grupo=grupo,
        fila_excel=2,
        comprobante_ref="LOTE-001",
        estado="validado",
        datos_json={},
        mensajes_json=["Validado correctamente. Listo para emitir."],
    )
    comprobante_1 = Comprobante(
        tipo_comprobante=6,
        concepto=1,
        numero=77,
        fecha_emision=fecha_fiscal,
        subtotal=Decimal("1000.00"),
        descuento=Decimal("0.00"),
        iva_21=Decimal("210.00"),
        iva_10_5=Decimal("0.00"),
        iva_27=Decimal("0.00"),
        otros_impuestos=Decimal("0.00"),
        total=Decimal("1210.00"),
        cae=CAE_TEST_NO_REAL,
        cae_vencimiento=date(2026, 5, 26),
        estado="autorizado",
        moneda="PES",
        cotizacion=Decimal("1"),
        empresa_id=empresa_id,
        punto_venta_id=punto_venta_id,
        receptor_tipo_documento=80,
        receptor_numero_documento=CUIT_RECEPTOR_TEST_NO_REAL,
        receptor_razon_social="Cliente Lote SA",
        receptor_condicion_iva="RI",
        receptor_domicilio="Av. Siempre Viva 123",
    )
    comprobante_2 = Comprobante(
        tipo_comprobante=6,
        concepto=1,
        numero=78,
        fecha_emision=fecha_fiscal,
        subtotal=Decimal("1000.00"),
        descuento=Decimal("0.00"),
        iva_21=Decimal("210.00"),
        iva_10_5=Decimal("0.00"),
        iva_27=Decimal("0.00"),
        otros_impuestos=Decimal("0.00"),
        total=Decimal("1210.00"),
        cae=CAE_TEST_NO_REAL_ALT,
        cae_vencimiento=date(2026, 5, 26),
        estado="autorizado",
        moneda="PES",
        cotizacion=Decimal("1"),
        empresa_id=empresa_id,
        punto_venta_id=punto_venta_id,
        receptor_tipo_documento=80,
        receptor_numero_documento=CUIT_RECEPTOR_TEST_NO_REAL,
        receptor_razon_social="Cliente Lote SA",
        receptor_condicion_iva="RI",
        receptor_domicilio="Av. Siempre Viva 123",
    )
    db_session.add_all([lote, grupo, fila, comprobante_1, comprobante_2])
    await db_session.flush()
    lote_id = int(lote.id)
    grupo_id = int(grupo.id)
    db_session.add_all(
        [
            IntentoEmisionFiscal(
                tipo_comprobante=6,
                punto_venta_numero=punto_venta_numero,
                numero_planificado=comprobante_1.numero,
                fecha_emision=fecha_fiscal,
                total=Decimal("1210.00"),
                receptor_tipo_documento=80,
                receptor_numero_documento=CUIT_RECEPTOR_TEST_NO_REAL,
                receptor_razon_social="Cliente Lote SA",
                payload_hash="hash-payload-grupo-001-a",
                huella_logica="hash-logica-grupo-001",
                cae=comprobante_1.cae,
                cae_vencimiento=comprobante_1.cae_vencimiento,
                estado="autorizado",
                operacion_id=None,
                empresa_id=empresa_id,
                punto_venta_id=punto_venta_id,
                comprobante_id=comprobante_1.id,
                lote_id=lote_id,
                grupo_id=grupo_id,
                ambiente=None,
                punto_venta_elegibilidad_revision_id=None,
                punto_venta_revision_fiscal=None,
                guarda_rece_id=None,
            ),
            IntentoEmisionFiscal(
                tipo_comprobante=6,
                punto_venta_numero=punto_venta_numero,
                numero_planificado=comprobante_2.numero,
                fecha_emision=fecha_fiscal,
                total=Decimal("1210.00"),
                receptor_tipo_documento=80,
                receptor_numero_documento=CUIT_RECEPTOR_TEST_NO_REAL,
                receptor_razon_social="Cliente Lote SA",
                payload_hash="hash-payload-grupo-001-b",
                huella_logica="hash-logica-grupo-001-b",
                cae=comprobante_2.cae,
                cae_vencimiento=comprobante_2.cae_vencimiento,
                estado="autorizado",
                operacion_id=None,
                empresa_id=empresa_id,
                punto_venta_id=punto_venta_id,
                comprobante_id=comprobante_2.id,
                lote_id=lote_id,
                grupo_id=grupo_id,
                ambiente=None,
                punto_venta_elegibilidad_revision_id=None,
                punto_venta_revision_fiscal=None,
                guarda_rece_id=None,
            ),
        ]
    )
    await db_session.commit()

    service = LoteComprobantesService(db_session)
    llamadas_arca = _instalar_oraculos_stale_sin_arca(service)

    await service.procesar_lote(lote_id, empresa_id, reanudar=True)

    async with AsyncSession(bind=db_session.bind, expire_on_commit=False) as observador:
        detalle = await LoteComprobantesService(observador).obtener_lote(
            lote_id,
            empresa_id,
        )
    assert detalle.estado == "requiere_reconciliacion"
    assert detalle.grupos[0].estado == "requiere_reconciliacion"
    assert detalle.grupos[0].comprobante_id is None
    assert detalle.grupos[0].numero_asignado is None
    assert "reconciliar" in detalle.mensaje_resumen
    assert llamadas_arca == {"wsaa": 0, "fecomp": 0, "fecae": 0}


@pytest.mark.asyncio
async def test_reconciliacion_local_rechaza_intento_incierto_del_grupo(
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
) -> None:
    """Un intento incierto del mismo grupo bloquea el cierre local automático."""
    fecha_fiscal = date(2026, 3, 20)
    emitir_request = {
        "empresa_id": test_empresa.id,
        "punto_venta_id": test_punto_venta.id,
        "tipo_comprobante": 6,
        "concepto": 1,
        "fecha_emision": fecha_fiscal.isoformat(),
        "confirmacion_fecha_fiscal": True,
        "tipo_documento": 80,
        "numero_documento": CUIT_RECEPTOR_TEST_NO_REAL,
        "razon_social": "Cliente Lote SA",
        "condicion_iva": "RI",
        "domicilio": "Av. Siempre Viva 123",
        "moneda": "PES",
        "cotizacion": "1",
        "guardar_cliente": False,
        "items": [
            {
                "descripcion": "Servicio mensual",
                "cantidad": "1",
                "unidad": "unidad",
                "precio_unitario": "1000",
                "iva_porcentaje": "21",
            }
        ],
    }
    request = EmitirComprobanteRequest.model_validate(emitir_request)
    payload_hash, huella = _hashes_fiscales_request(
        request,
        test_punto_venta.numero,
        Decimal("1210.00"),
    )
    lote = LoteComprobante(
        nombre_archivo="lote-intento-incierto.xlsx",
        archivo_hash="hash-intento-incierto",
        estado="procesando",
        total_filas=1,
        total_grupos=1,
        grupos_validos=1,
        empresa_id=test_empresa.id,
    )
    grupo = LoteComprobanteGrupo(
        lote=lote,
        comprobante_ref="LOTE-001",
        orden=1,
        estado="validado",
        tipo_comprobante=6,
        punto_venta_numero=test_punto_venta.numero,
        cliente_documento=CUIT_RECEPTOR_TEST_NO_REAL,
        cliente_razon_social="Cliente Lote SA",
        total_estimado=Decimal("1210.00"),
        payload_json=emitir_request,
        mensajes_json=["Validado correctamente. Listo para emitir."],
    )
    comprobante = Comprobante(
        tipo_comprobante=6,
        concepto=1,
        numero=77,
        fecha_emision=fecha_fiscal,
        subtotal=Decimal("1000.00"),
        descuento=Decimal("0.00"),
        iva_21=Decimal("210.00"),
        iva_10_5=Decimal("0.00"),
        iva_27=Decimal("0.00"),
        otros_impuestos=Decimal("0.00"),
        total=Decimal("1210.00"),
        cae=CAE_TEST_NO_REAL,
        cae_vencimiento=date(2026, 5, 26),
        estado="autorizado",
        moneda="PES",
        cotizacion=Decimal("1"),
        empresa_id=test_empresa.id,
        punto_venta_id=test_punto_venta.id,
        receptor_tipo_documento=80,
        receptor_numero_documento=CUIT_RECEPTOR_TEST_NO_REAL,
        receptor_razon_social="Cliente Lote SA",
        receptor_condicion_iva="RI",
        receptor_domicilio="Av. Siempre Viva 123",
    )
    comprobante.items = [
        ComprobanteItem(
            descripcion="Servicio mensual",
            cantidad=Decimal("1"),
            unidad="unidad",
            precio_unitario=Decimal("1000"),
            descuento_porcentaje=Decimal("0"),
            iva_porcentaje=Decimal("21"),
            subtotal=Decimal("1000.00"),
            orden=0,
        )
    ]
    db_session.add_all([lote, grupo, comprobante])
    await db_session.flush()
    db_session.add_all(
        [
            IntentoEmisionFiscal(
                tipo_comprobante=6,
                punto_venta_numero=test_punto_venta.numero,
                numero_planificado=77,
                fecha_emision=fecha_fiscal,
                total=Decimal("1210.00"),
                receptor_tipo_documento=80,
                receptor_numero_documento=CUIT_RECEPTOR_TEST_NO_REAL,
                receptor_razon_social="Cliente Lote SA",
                payload_hash=payload_hash,
                huella_logica=huella,
                cae=comprobante.cae,
                cae_vencimiento=comprobante.cae_vencimiento,
                estado="autorizado",
                empresa_id=test_empresa.id,
                punto_venta_id=test_punto_venta.id,
                comprobante_id=comprobante.id,
                lote_id=lote.id,
                grupo_id=grupo.id,
            ),
            IntentoEmisionFiscal(
                tipo_comprobante=6,
                punto_venta_numero=test_punto_venta.numero,
                numero_planificado=78,
                fecha_emision=fecha_fiscal,
                total=Decimal("1210.00"),
                receptor_tipo_documento=80,
                receptor_numero_documento=CUIT_RECEPTOR_TEST_NO_REAL,
                receptor_razon_social="Cliente Lote SA",
                payload_hash=payload_hash,
                huella_logica=huella,
                estado="en_proceso",
                empresa_id=test_empresa.id,
                punto_venta_id=test_punto_venta.id,
                lote_id=lote.id,
                grupo_id=grupo.id,
            ),
        ]
    )
    await db_session.commit()

    assert (
        await LoteComprobantesService(
            db_session
        )._reconciliar_grupo_autorizado_existente(grupo, request)
        is False
    )


@pytest.mark.asyncio
async def test_reconciliacion_local_rechaza_payload_drift(
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
) -> None:
    """La reconciliación local exige que la huella fiscal completa coincida."""
    request_original = EmitirComprobanteRequest.model_validate(
        {
            "empresa_id": test_empresa.id,
            "punto_venta_id": test_punto_venta.id,
            "tipo_comprobante": 6,
            "concepto": 1,
            "fecha_emision": "2026-03-20",
            "confirmacion_fecha_fiscal": True,
            "tipo_documento": 80,
            "numero_documento": CUIT_RECEPTOR_TEST_NO_REAL,
            "razon_social": "Cliente Lote SA",
            "condicion_iva": "RI",
            "domicilio": "Av. Siempre Viva 123",
            "moneda": "PES",
            "cotizacion": "1",
            "guardar_cliente": False,
            "items": [
                {
                    "descripcion": "Servicio mensual original",
                    "cantidad": "1",
                    "unidad": "unidad",
                    "precio_unitario": "1000",
                    "iva_porcentaje": "21",
                }
            ],
        }
    )
    request_drift = request_original.model_copy(deep=True)
    request_drift.items[0].descripcion = "Servicio mensual cambiado"
    payload_hash, huella = _hashes_fiscales_request(
        request_original,
        test_punto_venta.numero,
        Decimal("1210.00"),
    )
    intento = IntentoEmisionFiscal(
        tipo_comprobante=6,
        punto_venta_numero=test_punto_venta.numero,
        numero_planificado=77,
        fecha_emision=request_original.fecha_emision,
        total=Decimal("1210.00"),
        receptor_tipo_documento=request_original.tipo_documento,
        receptor_numero_documento=request_original.numero_documento,
        receptor_razon_social=request_original.razon_social,
        payload_hash=payload_hash,
        huella_logica=huella,
        cae=CAE_TEST_NO_REAL,
        cae_vencimiento=date(2026, 5, 26),
        estado="autorizado",
        empresa_id=test_empresa.id,
        punto_venta_id=test_punto_venta.id,
        comprobante_id=123,
        lote_id=456,
        grupo_id=789,
    )
    comprobante = Comprobante(
        id=123,
        tipo_comprobante=6,
        concepto=1,
        numero=77,
        fecha_emision=request_original.fecha_emision,
        subtotal=Decimal("1000.00"),
        descuento=Decimal("0.00"),
        iva_21=Decimal("210.00"),
        iva_10_5=Decimal("0.00"),
        iva_27=Decimal("0.00"),
        otros_impuestos=Decimal("0.00"),
        total=Decimal("1210.00"),
        cae=intento.cae,
        cae_vencimiento=intento.cae_vencimiento,
        estado="autorizado",
        moneda="PES",
        cotizacion=Decimal("1"),
        empresa_id=test_empresa.id,
        punto_venta_id=test_punto_venta.id,
        receptor_tipo_documento=request_original.tipo_documento,
        receptor_numero_documento=request_original.numero_documento,
        receptor_razon_social=request_original.razon_social,
        receptor_condicion_iva="RI",
        receptor_domicilio="Av. Siempre Viva 123",
    )
    comprobante.punto_venta = test_punto_venta
    comprobante.items = [
        ComprobanteItem(
            descripcion="Servicio mensual original",
            cantidad=Decimal("1"),
            unidad="unidad",
            precio_unitario=Decimal("1000"),
            descuento_porcentaje=Decimal("0"),
            iva_porcentaje=Decimal("21"),
            subtotal=Decimal("1000.00"),
            orden=0,
        )
    ]

    assert (
        LoteComprobantesService(db_session)._intento_local_coincide_con_grupo(
            intento=intento,
            comprobante=comprobante,
            request=request_drift,
            total=Decimal("1210.00"),
        )
        is False
    )


@pytest.mark.asyncio
async def test_reconciliacion_local_nota_usa_huella_del_intento_con_asociado(
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
) -> None:
    """La huella autorizada del intento conserva el asociado fiscal de la nota."""
    request = EmitirComprobanteRequest.model_validate(
        {
            "empresa_id": test_empresa.id,
            "punto_venta_id": test_punto_venta.id,
            "tipo_comprobante": 13,
            "concepto": 2,
            "fecha_emision": "2026-06-01",
            "fecha_servicio_desde": "2026-06-01",
            "fecha_servicio_hasta": "2026-06-01",
            "fecha_vto_pago": "2026-06-10",
            "confirmacion_fecha_fiscal": True,
            "tipo_documento": 99,
            "numero_documento": "0",
            "razon_social": "A CONSUMIDOR FINAL",
            "condicion_iva": "CF",
            "moneda": "PES",
            "cotizacion": "1",
            "guardar_cliente": False,
            "comprobantes_asociados": [
                {
                    "tipo_comprobante": 11,
                    "punto_venta": test_punto_venta.numero,
                    "numero": 1645,
                    "fecha": "2026-04-30",
                    "cuit": test_empresa.cuit,
                }
            ],
            "items": [
                {
                    "descripcion": "Anulación por duplicado",
                    "cantidad": "1",
                    "unidad": "unidad",
                    "precio_unitario": "59500",
                    "iva_porcentaje": "0",
                }
            ],
        }
    )
    total = Decimal("59500.00")
    payload_hash, huella = _hashes_fiscales_request(
        request,
        test_punto_venta.numero,
        total,
    )
    intento = IntentoEmisionFiscal(
        tipo_comprobante=request.tipo_comprobante,
        punto_venta_numero=test_punto_venta.numero,
        numero_planificado=27,
        fecha_emision=request.fecha_emision,
        total=total,
        receptor_tipo_documento=request.tipo_documento,
        receptor_numero_documento=request.numero_documento,
        receptor_razon_social=request.razon_social,
        payload_hash=payload_hash,
        huella_logica=huella,
        cae=CAE_TEST_NO_REAL,
        cae_vencimiento=date(2026, 6, 11),
        estado="autorizado",
        empresa_id=test_empresa.id,
        punto_venta_id=test_punto_venta.id,
        comprobante_id=123,
        lote_id=456,
        grupo_id=789,
    )
    comprobante = Comprobante(
        id=123,
        tipo_comprobante=request.tipo_comprobante,
        concepto=request.concepto,
        numero=27,
        fecha_emision=request.fecha_emision,
        fecha_servicio_desde=request.fecha_servicio_desde,
        fecha_servicio_hasta=request.fecha_servicio_hasta,
        fecha_vto_pago=request.fecha_vto_pago,
        fecha_vencimiento=request.fecha_vto_pago,
        subtotal=total,
        descuento=Decimal("0.00"),
        iva_21=Decimal("0.00"),
        iva_10_5=Decimal("0.00"),
        iva_27=Decimal("0.00"),
        otros_impuestos=Decimal("0.00"),
        total=total,
        cae=intento.cae,
        cae_vencimiento=intento.cae_vencimiento,
        estado="autorizado",
        moneda="PES",
        cotizacion=Decimal("1"),
        empresa_id=test_empresa.id,
        punto_venta_id=test_punto_venta.id,
        receptor_tipo_documento=request.tipo_documento,
        receptor_numero_documento=request.numero_documento,
        receptor_razon_social=request.razon_social,
        receptor_condicion_iva="CF",
    )
    comprobante.punto_venta = test_punto_venta
    comprobante.items = [
        ComprobanteItem(
            descripcion="Anulación por duplicado",
            cantidad=Decimal("1"),
            unidad="unidad",
            precio_unitario=Decimal("59500"),
            descuento_porcentaje=Decimal("0"),
            iva_porcentaje=Decimal("0"),
            subtotal=total,
            orden=0,
        )
    ]

    assert LoteComprobantesService(db_session)._intento_local_coincide_con_grupo(
        intento=intento,
        comprobante=comprobante,
        request=request,
        total=total,
    )


@pytest.mark.asyncio
async def test_reconciliacion_local_rechaza_snapshot_comprobante_distinto(
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
) -> None:
    """El comprobante local debe conservar el snapshot completo del request."""
    request = EmitirComprobanteRequest.model_validate(
        {
            "empresa_id": test_empresa.id,
            "punto_venta_id": test_punto_venta.id,
            "tipo_comprobante": 6,
            "concepto": 1,
            "fecha_emision": "2026-03-20",
            "confirmacion_fecha_fiscal": True,
            "tipo_documento": 80,
            "numero_documento": CUIT_RECEPTOR_TEST_NO_REAL,
            "razon_social": "Cliente Lote SA",
            "condicion_iva": "RI",
            "domicilio": "Av. Siempre Viva 123",
            "moneda": "PES",
            "cotizacion": "1",
            "guardar_cliente": False,
            "items": [
                {
                    "descripcion": "Servicio mensual",
                    "cantidad": "1",
                    "unidad": "unidad",
                    "precio_unitario": "1000",
                    "iva_porcentaje": "21",
                }
            ],
        }
    )
    payload_hash, huella = _hashes_fiscales_request(
        request,
        test_punto_venta.numero,
        Decimal("1210.00"),
    )
    intento = IntentoEmisionFiscal(
        tipo_comprobante=6,
        punto_venta_numero=test_punto_venta.numero,
        numero_planificado=77,
        fecha_emision=request.fecha_emision,
        total=Decimal("1210.00"),
        receptor_tipo_documento=request.tipo_documento,
        receptor_numero_documento=request.numero_documento,
        receptor_razon_social=request.razon_social,
        payload_hash=payload_hash,
        huella_logica=huella,
        cae=CAE_TEST_NO_REAL,
        cae_vencimiento=date(2026, 5, 26),
        estado="autorizado",
        empresa_id=test_empresa.id,
        punto_venta_id=test_punto_venta.id,
        comprobante_id=123,
        lote_id=456,
        grupo_id=789,
    )
    comprobante = Comprobante(
        id=123,
        tipo_comprobante=6,
        concepto=1,
        numero=77,
        fecha_emision=request.fecha_emision,
        subtotal=Decimal("1000.00"),
        descuento=Decimal("0.00"),
        iva_21=Decimal("210.00"),
        iva_10_5=Decimal("0.00"),
        iva_27=Decimal("0.00"),
        otros_impuestos=Decimal("0.00"),
        total=Decimal("1210.00"),
        cae=intento.cae,
        cae_vencimiento=intento.cae_vencimiento,
        estado="autorizado",
        moneda="PES",
        cotizacion=Decimal("1"),
        empresa_id=test_empresa.id,
        punto_venta_id=test_punto_venta.id,
        receptor_tipo_documento=request.tipo_documento,
        receptor_numero_documento=request.numero_documento,
        receptor_razon_social=request.razon_social,
        receptor_condicion_iva=request.condicion_iva,
        receptor_domicilio=request.domicilio,
    )
    comprobante.punto_venta = test_punto_venta
    comprobante.items = [
        ComprobanteItem(
            descripcion="Servicio mensual",
            cantidad=Decimal("1"),
            unidad="hora",
            precio_unitario=Decimal("1000"),
            descuento_porcentaje=Decimal("0"),
            iva_porcentaje=Decimal("21"),
            subtotal=Decimal("1000.00"),
            orden=0,
        )
    ]

    assert (
        LoteComprobantesService(db_session)._intento_local_coincide_con_grupo(
            intento=intento,
            comprobante=comprobante,
            request=request,
            total=Decimal("1210.00"),
        )
        is False
    )


@pytest.mark.asyncio
async def test_reconciliacion_local_rechaza_cae_vencimiento_distinto(
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
) -> None:
    """El vencimiento de CAE local debe coincidir con el intento autorizado."""
    request = EmitirComprobanteRequest.model_validate(
        {
            "empresa_id": test_empresa.id,
            "punto_venta_id": test_punto_venta.id,
            "tipo_comprobante": 6,
            "concepto": 1,
            "fecha_emision": "2026-03-20",
            "confirmacion_fecha_fiscal": True,
            "tipo_documento": 80,
            "numero_documento": CUIT_RECEPTOR_TEST_NO_REAL,
            "razon_social": "Cliente Lote SA",
            "condicion_iva": "RI",
            "domicilio": "Av. Siempre Viva 123",
            "moneda": "PES",
            "cotizacion": "1",
            "guardar_cliente": False,
            "items": [
                {
                    "descripcion": "Servicio mensual",
                    "cantidad": "1",
                    "unidad": "unidad",
                    "precio_unitario": "1000",
                    "iva_porcentaje": "21",
                }
            ],
        }
    )
    payload_hash, huella = _hashes_fiscales_request(
        request,
        test_punto_venta.numero,
        Decimal("1210.00"),
    )
    intento = IntentoEmisionFiscal(
        tipo_comprobante=6,
        punto_venta_numero=test_punto_venta.numero,
        numero_planificado=77,
        fecha_emision=request.fecha_emision,
        total=Decimal("1210.00"),
        receptor_tipo_documento=request.tipo_documento,
        receptor_numero_documento=request.numero_documento,
        receptor_razon_social=request.razon_social,
        payload_hash=payload_hash,
        huella_logica=huella,
        cae=CAE_TEST_NO_REAL,
        cae_vencimiento=date(2026, 5, 26),
        estado="autorizado",
        empresa_id=test_empresa.id,
        punto_venta_id=test_punto_venta.id,
        comprobante_id=123,
        lote_id=456,
        grupo_id=789,
    )
    comprobante = Comprobante(
        id=123,
        tipo_comprobante=6,
        concepto=1,
        numero=77,
        fecha_emision=request.fecha_emision,
        subtotal=Decimal("1000.00"),
        descuento=Decimal("0.00"),
        iva_21=Decimal("210.00"),
        iva_10_5=Decimal("0.00"),
        iva_27=Decimal("0.00"),
        otros_impuestos=Decimal("0.00"),
        total=Decimal("1210.00"),
        cae=intento.cae,
        cae_vencimiento=date(2026, 5, 27),
        estado="autorizado",
        moneda="PES",
        cotizacion=Decimal("1"),
        empresa_id=test_empresa.id,
        punto_venta_id=test_punto_venta.id,
        receptor_tipo_documento=request.tipo_documento,
        receptor_numero_documento=request.numero_documento,
        receptor_razon_social=request.razon_social,
        receptor_condicion_iva=request.condicion_iva,
        receptor_domicilio=request.domicilio,
    )
    comprobante.punto_venta = test_punto_venta
    comprobante.items = [
        ComprobanteItem(
            descripcion="Servicio mensual",
            cantidad=Decimal("1"),
            unidad="unidad",
            precio_unitario=Decimal("1000"),
            descuento_porcentaje=Decimal("0"),
            iva_porcentaje=Decimal("21"),
            subtotal=Decimal("1000.00"),
            orden=0,
        )
    ]

    assert (
        LoteComprobantesService(db_session)._intento_local_coincide_con_grupo(
            intento=intento,
            comprobante=comprobante,
            request=request,
            total=Decimal("1210.00"),
        )
        is False
    )


@pytest.mark.asyncio
async def test_reconciliacion_local_rechaza_fechas_servicio_distintas(
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
) -> None:
    """Las fechas fiscales de servicio deben coincidir con el request."""
    request = EmitirComprobanteRequest.model_validate(
        {
            "empresa_id": test_empresa.id,
            "punto_venta_id": test_punto_venta.id,
            "tipo_comprobante": 6,
            "concepto": 2,
            "fecha_emision": "2026-03-20",
            "fecha_servicio_desde": "2026-03-01",
            "fecha_servicio_hasta": "2026-03-31",
            "fecha_vto_pago": "2026-04-10",
            "confirmacion_fecha_fiscal": True,
            "tipo_documento": 80,
            "numero_documento": CUIT_RECEPTOR_TEST_NO_REAL,
            "razon_social": "Cliente Lote SA",
            "condicion_iva": "RI",
            "domicilio": "Av. Siempre Viva 123",
            "moneda": "PES",
            "cotizacion": "1",
            "guardar_cliente": False,
            "items": [
                {
                    "descripcion": "Servicio mensual",
                    "cantidad": "1",
                    "unidad": "unidad",
                    "precio_unitario": "1000",
                    "iva_porcentaje": "21",
                }
            ],
        }
    )
    payload_hash, huella = _hashes_fiscales_request(
        request,
        test_punto_venta.numero,
        Decimal("1210.00"),
    )
    intento = IntentoEmisionFiscal(
        tipo_comprobante=6,
        punto_venta_numero=test_punto_venta.numero,
        numero_planificado=77,
        fecha_emision=request.fecha_emision,
        total=Decimal("1210.00"),
        receptor_tipo_documento=request.tipo_documento,
        receptor_numero_documento=request.numero_documento,
        receptor_razon_social=request.razon_social,
        payload_hash=payload_hash,
        huella_logica=huella,
        cae=CAE_TEST_NO_REAL,
        cae_vencimiento=date(2026, 5, 26),
        estado="autorizado",
        empresa_id=test_empresa.id,
        punto_venta_id=test_punto_venta.id,
        comprobante_id=123,
        lote_id=456,
        grupo_id=789,
    )
    comprobante = Comprobante(
        id=123,
        tipo_comprobante=6,
        concepto=2,
        numero=77,
        fecha_emision=request.fecha_emision,
        fecha_servicio_desde=date(2026, 3, 2),
        fecha_servicio_hasta=request.fecha_servicio_hasta,
        fecha_vto_pago=request.fecha_vto_pago,
        fecha_vencimiento=request.fecha_vto_pago,
        subtotal=Decimal("1000.00"),
        descuento=Decimal("0.00"),
        iva_21=Decimal("210.00"),
        iva_10_5=Decimal("0.00"),
        iva_27=Decimal("0.00"),
        otros_impuestos=Decimal("0.00"),
        total=Decimal("1210.00"),
        cae=intento.cae,
        cae_vencimiento=intento.cae_vencimiento,
        estado="autorizado",
        moneda="PES",
        cotizacion=Decimal("1"),
        empresa_id=test_empresa.id,
        punto_venta_id=test_punto_venta.id,
        receptor_tipo_documento=request.tipo_documento,
        receptor_numero_documento=request.numero_documento,
        receptor_razon_social=request.razon_social,
        receptor_condicion_iva=request.condicion_iva,
        receptor_domicilio=request.domicilio,
    )
    comprobante.punto_venta = test_punto_venta
    comprobante.items = [
        ComprobanteItem(
            descripcion="Servicio mensual",
            cantidad=Decimal("1"),
            unidad="unidad",
            precio_unitario=Decimal("1000"),
            descuento_porcentaje=Decimal("0"),
            iva_porcentaje=Decimal("21"),
            subtotal=Decimal("1000.00"),
            orden=0,
        )
    ]

    assert (
        LoteComprobantesService(db_session)._intento_local_coincide_con_grupo(
            intento=intento,
            comprobante=comprobante,
            request=request,
            total=Decimal("1210.00"),
        )
        is False
    )


@pytest.mark.asyncio
async def test_reconciliacion_local_rechaza_comprobante_con_tipo_doc_distinto(
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
) -> None:
    """La reconciliación local exige identidad fiscal completa del receptor."""
    request = EmitirComprobanteRequest.model_validate(
        {
            "empresa_id": test_empresa.id,
            "punto_venta_id": test_punto_venta.id,
            "tipo_comprobante": 6,
            "concepto": 1,
            "fecha_emision": "2026-03-20",
            "confirmacion_fecha_fiscal": True,
            "tipo_documento": 80,
            "numero_documento": CUIT_RECEPTOR_TEST_NO_REAL,
            "razon_social": "Cliente Lote SA",
            "condicion_iva": "RI",
            "domicilio": "Av. Siempre Viva 123",
            "moneda": "PES",
            "cotizacion": "1",
            "guardar_cliente": False,
            "items": [
                {
                    "descripcion": "Servicio mensual",
                    "cantidad": "1",
                    "unidad": "unidad",
                    "precio_unitario": "1000",
                    "iva_porcentaje": "21",
                }
            ],
        }
    )
    payload_hash, huella = _hashes_fiscales_request(
        request,
        test_punto_venta.numero,
        Decimal("1210.00"),
    )
    intento = IntentoEmisionFiscal(
        tipo_comprobante=6,
        punto_venta_numero=test_punto_venta.numero,
        numero_planificado=77,
        fecha_emision=request.fecha_emision,
        total=Decimal("1210.00"),
        receptor_tipo_documento=request.tipo_documento,
        receptor_numero_documento=request.numero_documento,
        receptor_razon_social=request.razon_social,
        payload_hash=payload_hash,
        huella_logica=huella,
        cae=CAE_TEST_NO_REAL,
        cae_vencimiento=date(2026, 5, 26),
        estado="autorizado",
        empresa_id=test_empresa.id,
        punto_venta_id=test_punto_venta.id,
        comprobante_id=123,
        lote_id=456,
        grupo_id=789,
    )
    comprobante = Comprobante(
        id=123,
        tipo_comprobante=6,
        concepto=1,
        numero=77,
        fecha_emision=request.fecha_emision,
        subtotal=Decimal("1000.00"),
        descuento=Decimal("0.00"),
        iva_21=Decimal("210.00"),
        iva_10_5=Decimal("0.00"),
        iva_27=Decimal("0.00"),
        otros_impuestos=Decimal("0.00"),
        total=Decimal("1210.00"),
        cae=intento.cae,
        cae_vencimiento=intento.cae_vencimiento,
        estado="autorizado",
        moneda="PES",
        cotizacion=Decimal("1"),
        empresa_id=test_empresa.id,
        punto_venta_id=test_punto_venta.id,
        receptor_tipo_documento=96,
        receptor_numero_documento=request.numero_documento,
        receptor_razon_social=request.razon_social,
        receptor_condicion_iva="RI",
        receptor_domicilio="Av. Siempre Viva 123",
    )
    comprobante.punto_venta = test_punto_venta
    comprobante.items = [
        ComprobanteItem(
            descripcion="Servicio mensual",
            cantidad=Decimal("1"),
            unidad="unidad",
            precio_unitario=Decimal("1000"),
            descuento_porcentaje=Decimal("0"),
            iva_porcentaje=Decimal("21"),
            subtotal=Decimal("1000.00"),
            orden=0,
        )
    ]

    assert (
        LoteComprobantesService(db_session)._intento_local_coincide_con_grupo(
            intento=intento,
            comprobante=comprobante,
            request=request,
            total=Decimal("1210.00"),
        )
        is False
    )


@pytest.mark.asyncio
async def test_procesar_lote_exige_confirmacion_fecha_fiscal(
    client: AsyncClient,
    auth_headers: dict,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """No debe procesar lotes por API sin confirmacion fiscal final."""
    test_certificado.ambiente = settings.arca_env
    validar = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data=_opciones_fechas(),
        files={
            "archivo": (
                "lote-sin-confirmacion.xlsx",
                _build_lote_excel(test_empresa.cuit),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert validar.status_code == 200, validar.text
    lote_id = validar.json()["lote"]["id"]

    response = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/procesar",
        headers={**auth_headers, "X-Idempotency-Key": "idem-lote-sin-confirmacion"},
    )

    assert response.status_code == 400
    detalle = response.json()["detail"]
    assert "confirmar la fecha fiscal exacta" in detalle["mensaje"]
    assert FECHA_FISCAL_CONTROLADA_PF19B.strftime("%d/%m/%y") in detalle["mensaje"]
    assert "0001" in detalle["mensaje"]
    assert "XX/XX/XX" not in detalle["mensaje"]


@pytest.mark.asyncio
async def test_procesar_lote_exige_idempotency_key(
    client: AsyncClient,
    auth_headers: dict,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """No debe procesar un lote confirmado sin clave de idempotencia."""
    test_certificado.ambiente = settings.arca_env
    validar = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data=_opciones_fechas(),
        files={
            "archivo": (
                "lote-sin-idem.xlsx",
                _build_lote_excel(test_empresa.cuit),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert validar.status_code == 200, validar.text
    lote_id = validar.json()["lote"]["id"]

    response = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/procesar",
        headers={
            **auth_headers,
            "X-Confirmacion-Fecha-Fiscal": (
                f"fechas={FECHA_FISCAL_CONTROLADA_PF19B.isoformat()};puntos_venta=1"
            ),
        },
    )

    assert response.status_code == 400
    assert "X-Idempotency-Key" in response.json()["detail"]["mensaje"]


@pytest.mark.asyncio
@pytest.mark.parametrize(
    "estado_lote",
    ["en_cola", "procesando", "requiere_reconciliacion"],
)
async def test_lote_activo_o_incierto_no_admite_reintento_manual(
    db_session: AsyncSession,
    estado_lote: str,
) -> None:
    """El worker y el reintento manual no pueden resolver el mismo lote."""
    lote = SimpleNamespace(estado=estado_lote)
    service = LoteComprobantesService(db_session)

    with pytest.raises(LoteComprobanteError):
        service._validar_lote_resoluble(lote)


@pytest.mark.asyncio
async def test_tomar_lote_para_procesamiento_es_atomico(
    client: AsyncClient,
    auth_headers: dict,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """Un lote ya tomado no puede volver a tomarse para emisión."""
    test_certificado.ambiente = settings.arca_env
    validar = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data=_opciones_fechas(),
        files={
            "archivo": (
                "lote-lock.xlsx",
                _build_lote_excel(test_empresa.cuit),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert validar.status_code == 200, validar.text
    lote_id = validar.json()["lote"]["id"]
    service = LoteComprobantesService(db_session)

    await service._tomar_lote_para_procesamiento(
        lote_id=lote_id,
        empresa_id=test_empresa.id,
        procesamiento_async=False,
        modo_procesamiento="sincronico",
    )
    await db_session.commit()

    with pytest.raises(LoteComprobanteError, match="ya está siendo procesado"):
        await service._tomar_lote_para_procesamiento(
            lote_id=lote_id,
            empresa_id=test_empresa.id,
            procesamiento_async=False,
            modo_procesamiento="sincronico",
        )


@pytest.mark.asyncio
async def test_procesar_background_no_reencola_lote_en_proceso(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """Un lote en procesamiento no debe volver a estado en_cola."""
    test_certificado.ambiente = settings.arca_env
    empresa_id = int(test_empresa.id)
    worker_iniciado = False

    def fake_ensure_worker(_app):
        nonlocal worker_iniciado
        worker_iniciado = True
        return True

    monkeypatch.setattr(
        "app.api.lotes_comprobantes.ensure_lote_worker_running",
        fake_ensure_worker,
    )

    validar = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data=_opciones_fechas(),
        files={
            "archivo": (
                "lote-procesando-background.xlsx",
                _build_lote_excel(test_empresa.cuit),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert validar.status_code == 200, validar.text
    lote_id = validar.json()["lote"]["id"]
    idempotency_key = "idem-lote-procesando-background"
    headers_procesar = await _confirmacion_fecha_fiscal_header_lote(
        db_session,
        lote_id=lote_id,
        estados={"validado"},
        idempotency_key=idempotency_key,
    )
    encolado = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/procesar?background=true",
        headers={**auth_headers, **headers_procesar},
    )
    assert encolado.status_code == 200, encolado.text

    db_session.expire_all()
    lote = await db_session.get(LoteComprobante, lote_id)
    operacion = await db_session.scalar(
        select(OperacionIdempotente).where(
            OperacionIdempotente.idempotency_key == idempotency_key
        )
    )
    assert lote is not None
    assert operacion is not None
    operacion_id = int(operacion.id)
    material_rece = deepcopy(lote.metadata_json["pf19b_rece_material"])
    assert lote.estado == "en_cola"
    assert lote.metadata_json["operacion_idempotente_id"] == operacion_id
    assert operacion.estado == "en_proceso"
    assert operacion.response_json["en_progreso"] is True
    assert (
        operacion.response_json["lote"]["metadata_json"]["pf19b_rece_material"]
        == material_rece
    )

    service = LoteComprobantesService(db_session)
    await service._tomar_lote_para_procesamiento(
        lote_id=lote_id,
        empresa_id=empresa_id,
        procesamiento_async=True,
        modo_procesamiento="background",
    )
    await db_session.commit()
    db_session.expire_all()
    lote = await db_session.get(LoteComprobante, lote_id)
    assert lote is not None
    await service._guardar_respuesta_operacion_background(lote, operacion_id)
    await db_session.commit()

    db_session.expire_all()
    operacion = await db_session.get(OperacionIdempotente, operacion_id)
    assert operacion is not None
    assert operacion.estado == "en_proceso"
    assert operacion.response_json["en_progreso"] is True
    assert operacion.response_json["lote"]["estado"] == "procesando"
    assert (
        operacion.response_json["lote"]["metadata_json"]["pf19b_rece_material"]
        == material_rece
    )
    worker_iniciado = False

    procesar = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/procesar?background=true",
        headers={**auth_headers, **headers_procesar},
    )

    assert procesar.status_code == 200, procesar.text
    data = procesar.json()
    assert data["en_progreso"] is True
    assert data["lote"]["estado"] == "procesando"
    assert data["mensaje"] == "Procesando comprobantes..."
    assert worker_iniciado is False


@pytest.mark.asyncio
async def test_tomar_lote_no_reanuda_procesando_stale(
    client: AsyncClient,
    auth_headers: dict,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """Un lote procesando vencido no debe volver a tomarse para emitir."""
    test_certificado.ambiente = settings.arca_env
    validar = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data=_opciones_fechas(),
        files={
            "archivo": (
                "lote-procesando-stale.xlsx",
                _build_lote_excel(test_empresa.cuit),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert validar.status_code == 200, validar.text
    lote_id = validar.json()["lote"]["id"]
    service = LoteComprobantesService(db_session)

    await service._tomar_lote_para_procesamiento(
        lote_id=lote_id,
        empresa_id=test_empresa.id,
        procesamiento_async=True,
        modo_procesamiento="background",
    )
    await db_session.commit()

    with pytest.raises(LoteComprobanteError, match="ya está siendo procesado"):
        await service._tomar_lote_para_procesamiento(
            lote_id=lote_id,
            empresa_id=test_empresa.id,
            procesamiento_async=True,
            modo_procesamiento="background",
            reanudar=True,
        )

    lote = await db_session.get(LoteComprobante, lote_id)
    lote.updated_at = datetime.utcnow() - timedelta(
        minutes=settings.batch_processing_stale_minutes + 1
    )
    await db_session.commit()

    with pytest.raises(LoteComprobanteError, match="ya está siendo procesado"):
        await service._tomar_lote_para_procesamiento(
            lote_id=lote_id,
            empresa_id=test_empresa.id,
            procesamiento_async=True,
            modo_procesamiento="background",
            reanudar=True,
        )


@pytest.mark.asyncio
async def test_procesar_lote_procesando_stale_bloquea_y_preserva_intactos_sin_emitir(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """El worker bloquea si no puede comprobar una reanudación segura."""
    test_certificado.ambiente = settings.arca_env
    monkeypatch.setattr(settings, "batch_sync_limit", 0)
    llamadas_emitir = 0

    async def fake_emitir(self, request, **kwargs):
        nonlocal llamadas_emitir
        llamadas_emitir += 1
        return EmitirComprobanteResponse(
            exito=True,
            comprobante_id=432,
            tipo_comprobante=request.tipo_comprobante,
            punto_venta=1,
            numero=987,
            fecha=request.fecha_emision,
            cae=CAE_TEST_NO_REAL_36,
            cae_vencimiento=date(2026, 3, 31),
            total=Decimal("1210.00"),
            mensaje="Comprobante autorizado",
            errores=[],
        )

    monkeypatch.setattr(
        "app.services.facturacion_service.FacturacionService.emitir_comprobante",
        fake_emitir,
    )

    validar = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data=_opciones_fechas(),
        files={
            "archivo": (
                "lote-procesando-stale-worker.xlsx",
                _build_lote_excel(test_empresa.cuit),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert validar.status_code == 200, validar.text
    lote_id = validar.json()["lote"]["id"]
    service = LoteComprobantesService(db_session)

    async def fail_preflight(**_kwargs):
        raise RuntimeError(
            "preflight ARCA no disponible en C:\\privado\\certificado.key"
        )

    service.facturacion_service.verificar_numeracion_segura_para_emision = (
        fail_preflight
    )

    await service._tomar_lote_para_procesamiento(
        lote_id=lote_id,
        empresa_id=test_empresa.id,
        procesamiento_async=True,
        modo_procesamiento="background",
    )
    lote = await db_session.get(LoteComprobante, lote_id)
    lote.updated_at = datetime.utcnow() - timedelta(
        minutes=settings.batch_processing_stale_minutes + 1
    )
    await db_session.commit()

    lote = await service.procesar_lote(lote_id, test_empresa.id, reanudar=True)

    assert lote.estado == "requiere_reconciliacion"
    assert lote.grupos_validos == 0
    assert lote.grupos_emitidos == 0
    assert llamadas_emitir == 0
    assert "reconciliar contra ARCA" in lote.mensaje_resumen
    grupo = (
        (
            await db_session.execute(
                select(LoteComprobanteGrupo).where(
                    LoteComprobanteGrupo.lote_id == lote_id
                )
            )
        )
        .scalars()
        .one()
    )
    assert grupo.estado == "requiere_reconciliacion"
    assert grupo.cae is None
    assert grupo.numero_asignado is None
    assert grupo.comprobante_id is None
    eventos = (
        (
            await db_session.execute(
                select(LoteComprobanteEvento).where(
                    LoteComprobanteEvento.lote_id == lote_id,
                    LoteComprobanteEvento.accion == "bloqueo_operativo_no_reemitir",
                )
            )
        )
        .scalars()
        .all()
    )
    assert len(eventos) == 1
    assert eventos[0].metadata_json["estado_nuevo"] == "requiere_reconciliacion"
    assert eventos[0].metadata_json["grupos_marcados_reconciliacion"] == 1
    assert eventos[0].metadata_json["grupos_intactos_preservados"] == 1
    assert eventos[0].metadata_json["preflight_error"] == (
        "operacion_o_snapshot_rece_legacy"
    )
    metadata_serializada = str(
        {
            "lote": lote.metadata_json,
            "evento": eventos[0].metadata_json,
        }
    )
    assert "preflight ARCA no disponible" not in metadata_serializada
    assert "certificado.key" not in metadata_serializada


@pytest.mark.asyncio
async def test_worker_usa_factory_dedicada_e_instrumenta_sin_mutar_lote(
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
) -> None:
    """El ciclo usa conexiones worker y guarda métricas solo en memoria."""
    monkeypatch.setattr(settings, "batch_worker_enabled", True)
    lote = LoteComprobante(
        nombre_archivo="lote-worker-instrumentado.xlsx",
        archivo_hash="hash-lote-worker-instrumentado",
        estado="en_cola",
        modo_procesamiento="background",
        procesamiento_async=True,
        total_filas=1,
        total_grupos=1,
        grupos_validos=1,
        empresa_id=test_empresa.id,
    )
    db_session.add(lote)
    await db_session.commit()
    await db_session.refresh(lote)
    updated_at_antes = lote.updated_at

    class SessionFactory:
        aperturas = 0

        async def __aenter__(self):
            SessionFactory.aperturas += 1
            return db_session

        async def __aexit__(self, exc_type, exc, tb):
            return False

    roles: list[str] = []
    reanudar_recibido: list[bool] = []

    async def fake_acquire(session: AsyncSession, role: str) -> None:
        assert session is db_session
        roles.append(role)

    async def fake_procesar(
        self: LoteComprobantesService,
        lote_id: int,
        empresa_id: int,
        **kwargs,
    ) -> LoteComprobante:
        reanudar_recibido.append(bool(kwargs.get("reanudar")))
        return await self.obtener_lote_resumen(lote_id, empresa_id)

    monkeypatch.setattr(
        "app.services.lote_worker.WorkerSessionLocal",
        SessionFactory,
    )
    monkeypatch.setattr(
        "app.services.lote_worker.acquire_database_connection",
        fake_acquire,
    )
    monkeypatch.setattr(
        LoteComprobantesService,
        "procesar_lote",
        fake_procesar,
    )

    worker = LoteWorker()
    resultado = await worker.procesar_pendientes()
    app = SimpleNamespace(
        state=SimpleNamespace(
            lote_worker=worker,
            lote_worker_task=SimpleNamespace(done=lambda: False),
        )
    )
    runtime = get_lote_worker_status(app)

    assert SessionFactory.aperturas == 3
    assert roles == ["worker", "worker", "worker"]
    assert reanudar_recibido == [True]
    assert resultado.stale_detectados == 0
    assert resultado.lotes_en_cola_detectados == 1
    assert resultado.lotes_procesados == 1
    assert resultado.tuvo_error is False
    assert runtime["estado"] == "esperando"
    assert runtime["ocupado"] is False
    assert runtime["ultimo_resultado"] == "exitoso"
    assert runtime["ciclo_iniciado_at"] is not None
    assert runtime["ciclo_finalizado_at"] is not None
    assert runtime["ultima_duracion_ms"] is not None
    assert runtime["stale_detectados_ultimo_ciclo"] == 0
    assert runtime["lotes_en_cola_ultimo_ciclo"] == 1
    assert runtime["lotes_procesados_ultimo_ciclo"] == 1
    assert not any("mensaje" in key for key in runtime)
    await db_session.refresh(lote)
    assert lote.updated_at == updated_at_antes


@pytest.mark.asyncio
@pytest.mark.parametrize(
    "error_type",
    [SQLAlchemyTimeoutError, OperationalError],
    ids=["timeout", "operational"],
)
async def test_worker_corta_ciclo_tras_db_temporal_del_primer_lote(
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    error_type: type[Exception],
) -> None:
    """El worker no avanza al segundo lote si la base falla en el primero."""

    class SessionFactory:
        """Reutiliza la sesión aislada de pruebas como sesión worker."""

        async def __aenter__(self):
            return db_session

        async def __aexit__(self, exc_type, exc, tb):
            return False

    lotes = [
        LoteComprobante(
            nombre_archivo=f"lote-worker-db-{indice}.xlsx",
            archivo_hash=f"hash-lote-worker-db-{indice}",
            estado="en_cola",
            modo_procesamiento="background",
            procesamiento_async=True,
            total_filas=1,
            total_grupos=1,
            grupos_validos=1,
            empresa_id=test_empresa.id,
        )
        for indice in (1, 2)
    ]
    db_session.add_all(lotes)
    await db_session.flush()
    operacion_worker = OperacionIdempotente(
        empresa_id=test_empresa.id,
        usuario_id=None,
        lote_id=lotes[0].id,
        idempotency_key="idem-worker-pre-arca",
        tipo_operacion="procesar_lote",
        payload_hash="payload-worker-pre-arca",
        estado="en_proceso",
    )
    db_session.add(operacion_worker)
    await db_session.commit()
    lote_ids = [lote.id for lote in lotes]
    operacion_worker_id = operacion_worker.id

    procesados: list[int] = []

    async def fake_acquire(session: AsyncSession, role: str) -> None:
        assert session is db_session
        assert role == "worker"

    async def fail_primer_lote(self, lote_id, empresa_id, **kwargs):
        procesados.append(lote_id)
        raise _crear_error_db_temporal(error_type)

    monkeypatch.setattr("app.services.lote_worker.WorkerSessionLocal", SessionFactory)
    monkeypatch.setattr(
        "app.services.lote_worker.acquire_database_connection",
        fake_acquire,
    )
    monkeypatch.setattr(
        LoteComprobantesService,
        "procesar_lote",
        fail_primer_lote,
    )

    resultado = await LoteWorker().procesar_pendientes()

    assert procesados == [lote_ids[0]]
    assert resultado.lotes_en_cola_detectados == 1
    assert resultado.lotes_procesados == 0
    assert resultado.tuvo_error is True
    db_session.expire_all()
    primer_lote = await db_session.get(LoteComprobante, lote_ids[0])
    segundo_lote = await db_session.get(LoteComprobante, lote_ids[1])
    operacion_actual = await db_session.get(
        OperacionIdempotente,
        operacion_worker_id,
    )
    assert primer_lote is not None
    assert segundo_lote is not None
    assert operacion_actual is not None
    assert primer_lote.estado == "en_cola"
    assert segundo_lote.estado == "en_cola"
    assert operacion_actual.estado == "en_proceso"


@pytest.mark.asyncio
@pytest.mark.parametrize(
    "mutacion",
    ["json_null", "progreso_adulterado", "terminal"],
)
async def test_worker_rechaza_ownership_invalido_antes_de_consultar_arca(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
    mutacion: str,
) -> None:
    """Ownership inválido deja el lote en cola y corta toda capacidad/FECAE."""
    llamadas_wsaa = 0
    llamadas_fecomp = 0
    llamadas_fecae = 0

    async def fake_ticket(empresa, certificado):
        nonlocal llamadas_wsaa
        llamadas_wsaa += 1
        raise AssertionError("No debe solicitar WSAA sin ownership worker")

    class FakeWSFEClient:
        """Hace observables FECompTotXRequest y FECAE si se cruza la compuerta."""

        def __init__(self, *args, **kwargs) -> None:
            """Acepta la firma productiva sin abrir red."""

        async def fe_comp_tot_x_request(self):
            """Registra una consulta de capacidad indebida."""
            nonlocal llamadas_fecomp
            llamadas_fecomp += 1
            return 1

        async def fe_cae_solicitar(self, arca_request):
            """Registra una solicitud FECAE indebida."""
            nonlocal llamadas_fecae
            llamadas_fecae += 1
            raise AssertionError("No debe solicitar FECAE sin ownership worker")

    async def fail_emitir(request, **kwargs):
        nonlocal llamadas_fecae
        llamadas_fecae += 1
        raise AssertionError("No debe solicitar FECAE sin ownership worker")

    monkeypatch.setattr(
        "app.api.lotes_comprobantes.ensure_lote_worker_running",
        lambda app: True,
    )
    monkeypatch.setattr(settings, "arca_fecaesolicitar_batch_enabled", True)
    monkeypatch.setattr(
        "app.services.facturacion_service.WSFEv1Client",
        FakeWSFEClient,
    )
    test_certificado.ambiente = settings.arca_env
    await db_session.commit()
    lote_id = await _crear_lote_validado_por_api(
        client,
        auth_headers,
        test_empresa.cuit,
        nombre_archivo=f"lote-worker-ownership-{mutacion}.xlsx",
    )
    idempotency_key = f"idem-worker-ownership-{mutacion}"
    headers = await _confirmacion_fecha_fiscal_header_lote(
        db_session,
        lote_id=lote_id,
        estados={"validado"},
        idempotency_key=idempotency_key,
    )
    encolado = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/procesar?background=true",
        headers={**auth_headers, **headers},
    )
    assert encolado.status_code == 200, encolado.text
    operacion = await db_session.scalar(
        select(OperacionIdempotente).where(
            OperacionIdempotente.idempotency_key == idempotency_key
        )
    )
    assert operacion is not None
    if mutacion == "json_null":
        operacion.response_json = JSON.NULL
    elif mutacion == "progreso_adulterado":
        respuesta = deepcopy(operacion.response_json)
        respuesta["lote"]["metadata_json"]["pf19b_rece_material"]["grupos_hash"] = (
            "0" * 64
        )
        operacion.response_json = respuesta
    else:
        operacion.estado = "finalizado"
    await db_session.commit()

    service = LoteComprobantesService(db_session)
    monkeypatch.setattr(
        service.facturacion_service, "_obtener_ticket_acceso", fake_ticket
    )
    monkeypatch.setattr(service.facturacion_service, "emitir_comprobante", fail_emitir)

    with pytest.raises(LoteComprobanteError, match="perdió el ownership"):
        await service.procesar_lote(lote_id, test_empresa.id, reanudar=True)

    async with AsyncSession(bind=db_session.bind, expire_on_commit=False) as observador:
        lote = await observador.get(LoteComprobante, lote_id)
        assert lote is not None
        assert lote.estado == "en_cola"
        assert lote.started_at is None
    assert llamadas_wsaa == 0
    assert llamadas_fecomp == 0
    assert llamadas_fecae == 0


@pytest.mark.asyncio
@pytest.mark.parametrize("mutacion", ["progreso_adulterado", "terminal"])
async def test_worker_revalida_ownership_post_claim_antes_de_capacidad(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
    mutacion: str,
) -> None:
    """Una segunda sesión que cambia la operación post-claim corta toda ARCA."""
    llamadas_wsaa = 0
    llamadas_fecomp = 0
    llamadas_fecae = 0

    class FakeWSFEClient:
        """Hace observables capacidad y FECAE después del segundo gate."""

        def __init__(self, *args, **kwargs) -> None:
            """Acepta la firma productiva sin abrir red."""

        async def fe_comp_tot_x_request(self):
            """Registra una consulta de capacidad indebida."""
            nonlocal llamadas_fecomp
            llamadas_fecomp += 1
            return 1

        async def fe_cae_solicitar(self, arca_request):
            """Registra una solicitud FECAE indebida."""
            nonlocal llamadas_fecae
            llamadas_fecae += 1
            raise AssertionError("No debe solicitar FECAE tras perder ownership")

    async def fake_ticket(empresa, certificado):
        nonlocal llamadas_wsaa
        llamadas_wsaa += 1
        raise AssertionError("No debe solicitar WSAA tras perder ownership")

    async def fail_emitir(request, **kwargs):
        nonlocal llamadas_fecae
        llamadas_fecae += 1
        raise AssertionError("No debe emitir tras perder ownership")

    monkeypatch.setattr(
        "app.api.lotes_comprobantes.ensure_lote_worker_running",
        lambda app: True,
    )
    monkeypatch.setattr(settings, "arca_fecaesolicitar_batch_enabled", True)
    monkeypatch.setattr(
        "app.services.facturacion_service.WSFEv1Client",
        FakeWSFEClient,
    )
    test_certificado.ambiente = settings.arca_env
    await db_session.commit()
    lote_id = await _crear_lote_validado_por_api(
        client,
        auth_headers,
        test_empresa.cuit,
        nombre_archivo=f"lote-worker-post-claim-{mutacion}.xlsx",
    )
    key = f"idem-worker-post-claim-{mutacion}"
    headers = await _confirmacion_fecha_fiscal_header_lote(
        db_session,
        lote_id=lote_id,
        estados={"validado"},
        idempotency_key=key,
    )
    encolado = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/procesar?background=true",
        headers={**auth_headers, **headers},
    )
    assert encolado.status_code == 200, encolado.text
    operacion = await db_session.scalar(
        select(OperacionIdempotente).where(OperacionIdempotente.idempotency_key == key)
    )
    assert operacion is not None
    operacion_id = int(operacion.id)
    commit_original = db_session.commit
    mutada = False

    async def commit_claim_con_carrera() -> None:
        """Publica el claim y luego muta ownership desde otra sesión."""
        nonlocal mutada
        await commit_original()
        if mutada:
            return
        mutada = True
        async with AsyncSession(
            bind=db_session.bind,
            expire_on_commit=False,
        ) as competidora:
            operacion_competidora = await competidora.get(
                OperacionIdempotente,
                operacion_id,
            )
            assert operacion_competidora is not None
            respuesta = deepcopy(operacion_competidora.response_json)
            if mutacion == "terminal":
                operacion_competidora.estado = "finalizado"
                respuesta["en_progreso"] = False
                respuesta["mensaje"] = "Resultado terminal sintético"
            else:
                respuesta["lote"]["metadata_json"]["pf19b_rece_material"][
                    "grupos_hash"
                ] = ("f" * 64)
            operacion_competidora.response_json = respuesta
            await competidora.commit()

    service = LoteComprobantesService(db_session)
    monkeypatch.setattr(db_session, "commit", commit_claim_con_carrera)
    monkeypatch.setattr(
        service.facturacion_service, "_obtener_ticket_acceso", fake_ticket
    )
    monkeypatch.setattr(service.facturacion_service, "emitir_comprobante", fail_emitir)

    with pytest.raises(LoteComprobanteError, match="perdió el ownership"):
        await service.procesar_lote(lote_id, test_empresa.id, reanudar=True)

    assert mutada is True
    assert llamadas_wsaa == 0
    assert llamadas_fecomp == 0
    assert llamadas_fecae == 0
    async with AsyncSession(bind=db_session.bind, expire_on_commit=False) as observador:
        lote = await observador.get(LoteComprobante, lote_id)
        grupos = list(
            (
                await observador.scalars(
                    select(LoteComprobanteGrupo).where(
                        LoteComprobanteGrupo.lote_id == lote_id
                    )
                )
            ).all()
        )
        filas = list(
            (
                await observador.scalars(
                    select(LoteComprobanteFila).where(
                        LoteComprobanteFila.lote_id == lote_id
                    )
                )
            ).all()
        )
        operacion_visible = await observador.get(
            OperacionIdempotente,
            operacion_id,
        )
        assert lote is not None
        assert operacion_visible is not None
        assert lote.estado == "requiere_reconciliacion"
        assert {grupo.estado for grupo in grupos} == {"requiere_reconciliacion"}
        assert {fila.estado for fila in filas} == {"requiere_reconciliacion"}
        if mutacion == "terminal":
            assert operacion_visible.estado == "finalizado"
        else:
            assert operacion_visible.estado == "en_proceso"


@pytest.mark.asyncio
async def test_worker_background_recupera_guarda_pre_arca_con_ownership_durable(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
) -> None:
    """Endpoint y worker cierran una guarda pre-ARCA sin publicar ownership parcial."""

    class SessionFactory:
        """Reutiliza la sesión del test en todos los ciclos del worker."""

        async def __aenter__(self):
            return db_session

        async def __aexit__(self, exc_type, exc, tb):
            return False

    class FakeWSFEClient:
        """Permite lecturas de numeración y hace observable cualquier FECAE."""

        llamadas_fecae = 0

        def __init__(self, *args, **kwargs) -> None:
            """Acepta la firma productiva sin abrir red."""

        async def fe_comp_ultimo_autorizado(self, punto_venta_numero, tipo):
            """Mantiene estable el doble preflight seguro de numeración."""
            return 0

        async def fe_cae_solicitar(self, arca_request):
            """Registra una violación: este caso nunca debe llegar a FECAE."""
            FakeWSFEClient.llamadas_fecae += 1
            raise AssertionError("No debe solicitar FECAE después del fallo del CAS")

    async def fake_acquire(session: AsyncSession, role: str) -> None:
        assert session is db_session
        assert role == "worker"

    async def fake_ticket(self, empresa, certificado):
        return SimpleNamespace(token="token", sign="sign")

    async def fake_validar_punto(self, wsfe_client, punto_venta_numero):
        return None

    async def fallar_cas_pre_fecae(self, **kwargs):
        raise SQLAlchemyTimeoutError()

    monkeypatch.setattr(
        "app.api.lotes_comprobantes.ensure_lote_worker_running",
        lambda app: True,
    )
    monkeypatch.setattr("app.services.lote_worker.WorkerSessionLocal", SessionFactory)
    monkeypatch.setattr(
        "app.services.lote_worker.acquire_database_connection",
        fake_acquire,
    )
    monkeypatch.setattr(
        "app.services.facturacion_service.WSFEv1Client",
        FakeWSFEClient,
    )
    monkeypatch.setattr(FacturacionService, "_obtener_ticket_acceso", fake_ticket)
    monkeypatch.setattr(
        FacturacionService,
        "_validar_punto_venta_habilitado",
        fake_validar_punto,
    )
    monkeypatch.setattr(
        ElegibilidadReceService,
        "marcar_arca_iniciada",
        fallar_cas_pre_fecae,
    )

    lote_id = await _crear_lote_validado_por_api(
        client,
        auth_headers,
        test_empresa.cuit,
        nombre_archivo="lote-worker-ownership-durable.xlsx",
    )
    confirmacion = await _confirmacion_fecha_fiscal_header_lote(
        db_session,
        lote_id=lote_id,
        estados={"validado"},
        idempotency_key="idem-worker-ownership-durable",
    )
    encolado = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/procesar?background=true",
        headers={**auth_headers, **confirmacion},
    )
    assert encolado.status_code == 200, encolado.text
    assert encolado.json()["en_progreso"] is True

    resultado = await LoteWorker().procesar_pendientes()

    assert resultado.tuvo_error is True
    assert resultado.lotes_procesados == 0
    assert FakeWSFEClient.llamadas_fecae == 0
    async with AsyncSession(bind=db_session.bind, expire_on_commit=False) as observador:
        lote = await observador.get(LoteComprobante, lote_id)
        operacion = await observador.scalar(
            select(OperacionIdempotente).where(
                OperacionIdempotente.idempotency_key == "idem-worker-ownership-durable"
            )
        )
        assert lote is not None
        assert operacion is not None
        guardas = list(
            (
                await observador.scalars(
                    select(PuntoVentaGuardaEmisionRece).where(
                        PuntoVentaGuardaEmisionRece.operacion_id == operacion.id
                    )
                )
            ).all()
        )
        intentos = list(
            (
                await observador.scalars(
                    select(IntentoEmisionFiscal).where(
                        IntentoEmisionFiscal.operacion_id == operacion.id
                    )
                )
            ).all()
        )
        assert len(guardas) == 1
        assert guardas[0].fase == "cerrada_pre_arca"
        assert guardas[0].arca_iniciada_en is None
        assert len(intentos) == 1
        assert intentos[0].estado == "fallido_verificado"
        assert lote.estado == "en_cola"
        assert operacion.estado == "en_proceso"
        assert operacion.response_json["en_progreso"] is True
        assert operacion.response_json["lote"]["id"] == lote_id
        assert lote.metadata_json["operacion_idempotente_id"] == operacion.id
        comprobantes = list((await observador.scalars(select(Comprobante))).all())
        assert comprobantes == []


@pytest.mark.asyncio
async def test_procesar_lote_recupera_fallo_durable_al_cerrar_pre_arca_batch(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
) -> None:
    """El cierre batch fallido se recupera completo sin publicar ni emitir CAE."""
    consultas_numeracion = 0
    consultas_capacidad = 0
    llamadas_fecae = 0
    cierres = 0

    class FakeWSFEClient:
        """Fuerza cambio de rango luego de reservar el sublote completo."""

        def __init__(self, *args, **kwargs) -> None:
            """Acepta la firma productiva sin abrir red."""

        async def fe_comp_tot_x_request(self):
            """Permite agrupar los dos comprobantes en una única guarda."""
            nonlocal consultas_capacidad
            consultas_capacidad += 1
            return 2

        async def fe_comp_ultimo_autorizado(self, punto_venta_numero, tipo):
            """Cambia el próximo número en el segundo preflight seguro."""
            nonlocal consultas_numeracion
            consultas_numeracion += 1
            return 0 if consultas_numeracion == 1 else 1

        async def fe_cae_solicitar_lote(self, arca_requests):
            """Hace observable cualquier FECAE batch indebido."""
            nonlocal llamadas_fecae
            llamadas_fecae += 1
            raise AssertionError("No debe solicitar FECAE batch en este escenario")

        async def fe_cae_solicitar(self, arca_request):
            """Hace observable un fallback unitario indebido."""
            nonlocal llamadas_fecae
            llamadas_fecae += 1
            raise AssertionError("No debe solicitar FECAE unitario en este escenario")

    async def fake_ticket(self, empresa, certificado):
        return SimpleNamespace(token="token", sign="sign")

    async def fake_validar_punto(self, wsfe_client, punto_venta_numero):
        return None

    cierre_original = ElegibilidadReceService.cerrar_pre_arca

    async def fallar_primer_cierre(self, guarda, **kwargs):
        nonlocal cierres
        cierres += 1
        if cierres == 1:
            raise RuntimeError("fallo sintético al confirmar cierre batch")
        return await cierre_original(self, guarda, **kwargs)

    monkeypatch.setattr(settings, "arca_fecaesolicitar_batch_enabled", True)
    monkeypatch.setattr(settings, "arca_fecaesolicitar_batch_max_registros", 2)
    monkeypatch.setattr(
        "app.services.facturacion_service.WSFEv1Client",
        FakeWSFEClient,
    )
    monkeypatch.setattr(FacturacionService, "_obtener_ticket_acceso", fake_ticket)
    monkeypatch.setattr(
        FacturacionService,
        "_validar_punto_venta_habilitado",
        fake_validar_punto,
    )
    monkeypatch.setattr(
        ElegibilidadReceService,
        "cerrar_pre_arca",
        fallar_primer_cierre,
    )
    lote_id = await _crear_lote_validado_por_api(
        client,
        auth_headers,
        test_empresa.cuit,
        nombre_archivo="lote-fallo-cierre-pre-arca-batch.xlsx",
        total_grupos=2,
    )
    key = "idem-fallo-cierre-pre-arca-batch"
    headers = await _confirmacion_fecha_fiscal_header_lote(
        db_session,
        lote_id=lote_id,
        estados={"validado"},
        idempotency_key=key,
    )
    resumen = await client.get(
        f"/api/lotes-comprobantes/{lote_id}/resumen",
        headers=auth_headers,
    )
    assert resumen.status_code == 200, resumen.text
    duplicado = resumen.json()["confirmacion_duplicado_logico"]
    if duplicado:
        headers["X-Confirmacion-Duplicado-Logico"] = duplicado

    response = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/procesar",
        headers={**auth_headers, **headers},
    )

    assert response.status_code == 503, response.text
    assert consultas_capacidad == 1
    assert consultas_numeracion == 2
    assert llamadas_fecae == 0
    assert cierres == 1
    async with AsyncSession(bind=db_session.bind, expire_on_commit=False) as observador:
        operacion = await observador.scalar(
            select(OperacionIdempotente).where(
                OperacionIdempotente.idempotency_key == key
            )
        )
        assert operacion is not None
        lote = await observador.get(LoteComprobante, lote_id)
        guardas = list(
            (
                await observador.scalars(
                    select(PuntoVentaGuardaEmisionRece).where(
                        PuntoVentaGuardaEmisionRece.operacion_id == operacion.id
                    )
                )
            ).all()
        )
        intentos = list(
            (
                await observador.scalars(
                    select(IntentoEmisionFiscal)
                    .where(IntentoEmisionFiscal.operacion_id == operacion.id)
                    .order_by(IntentoEmisionFiscal.id)
                )
            ).all()
        )
        assert lote is not None
        assert lote.estado == "validado"
        assert len(guardas) == 1
        assert guardas[0].fase == "cerrada_pre_arca"
        assert guardas[0].arca_iniciada_en is None
        assert len(intentos) == 2
        assert {intento.estado for intento in intentos} == {"fallido_verificado"}
        assert operacion.estado == "interrumpida_pre_arca"
        assert operacion.response_json is None
        comprobantes = list((await observador.scalars(select(Comprobante))).all())
        assert comprobantes == []


@pytest.mark.asyncio
async def test_procesar_lote_recupera_segundo_chunk_pre_arca_automaticamente(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
) -> None:
    """El caller conserva el primer CAE y cierra la segunda guarda sin reemitir."""

    class FakeWSFEClient:
        """Autoriza el primer grupo y no recibe la segunda solicitud FECAE."""

        consultas_numeracion = 0
        llamadas_fecae = 0

        def __init__(self, *args, **kwargs) -> None:
            """Acepta la firma real sin abrir red."""

        async def fe_comp_ultimo_autorizado(self, punto_venta_numero, tipo):
            """Refleja el primer CAE antes de evaluar el segundo grupo."""
            FakeWSFEClient.consultas_numeracion += 1
            return 0 if FakeWSFEClient.consultas_numeracion <= 2 else 1

        async def fe_cae_solicitar(self, arca_request):
            """Autoriza solamente el primer chunk unitario."""
            FakeWSFEClient.llamadas_fecae += 1
            return CAEResponse(
                cae=CAE_TEST_NO_REAL,
                cae_vencimiento="20260819",
                numero_comprobante=arca_request.cbte_desde,
                tipo_cbte=arca_request.tipo_cbte,
                punto_venta=arca_request.punto_venta,
                resultado="A",
            )

    async def fake_ticket(self, empresa, certificado):
        return SimpleNamespace(token="token", sign="sign")

    async def fake_validar_punto(self, wsfe_client, punto_venta_numero):
        return None

    marcar_original = ElegibilidadReceService.marcar_arca_iniciada
    guardas_evaluadas = 0

    async def fallar_segundo_cas(self, **kwargs):
        nonlocal guardas_evaluadas
        guardas_evaluadas += 1
        if guardas_evaluadas == 2:
            raise SQLAlchemyTimeoutError()
        return await marcar_original(self, **kwargs)

    monkeypatch.setattr(
        "app.services.facturacion_service.WSFEv1Client",
        FakeWSFEClient,
    )
    monkeypatch.setattr(FacturacionService, "_obtener_ticket_acceso", fake_ticket)
    monkeypatch.setattr(
        FacturacionService,
        "_validar_punto_venta_habilitado",
        fake_validar_punto,
    )
    monkeypatch.setattr(
        ElegibilidadReceService,
        "marcar_arca_iniciada",
        fallar_segundo_cas,
    )

    lote_id = await _crear_lote_validado_por_api(
        client,
        auth_headers,
        test_empresa.cuit,
        nombre_archivo="lote-dos-chunks-recovery-automatico.xlsx",
        total_grupos=2,
    )
    headers = await _confirmacion_fecha_fiscal_header_lote(
        db_session,
        lote_id=lote_id,
        estados={"validado"},
        idempotency_key="idem-dos-chunks-recovery-automatico",
    )
    resumen = await client.get(
        f"/api/lotes-comprobantes/{lote_id}/resumen",
        headers=auth_headers,
    )
    assert resumen.status_code == 200, resumen.text
    duplicado = resumen.json()["confirmacion_duplicado_logico"]
    if duplicado:
        headers["X-Confirmacion-Duplicado-Logico"] = duplicado

    response = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/procesar",
        headers={**auth_headers, **headers},
    )

    assert response.status_code == 409, response.text
    assert response.json()["detail"]["categoria_error"] == "post_arca_persistencia"
    assert guardas_evaluadas == 2
    assert FakeWSFEClient.llamadas_fecae == 1
    async with AsyncSession(bind=db_session.bind, expire_on_commit=False) as observador:
        operacion = await observador.scalar(
            select(OperacionIdempotente).where(
                OperacionIdempotente.idempotency_key
                == "idem-dos-chunks-recovery-automatico"
            )
        )
        assert operacion is not None
        lote = await observador.get(LoteComprobante, lote_id)
        grupos = list(
            (
                await observador.scalars(
                    select(LoteComprobanteGrupo)
                    .where(LoteComprobanteGrupo.lote_id == lote_id)
                    .order_by(LoteComprobanteGrupo.orden)
                )
            ).all()
        )
        guardas = list(
            (
                await observador.scalars(
                    select(PuntoVentaGuardaEmisionRece)
                    .where(PuntoVentaGuardaEmisionRece.operacion_id == operacion.id)
                    .order_by(PuntoVentaGuardaEmisionRece.id)
                )
            ).all()
        )
        intentos = list(
            (
                await observador.scalars(
                    select(IntentoEmisionFiscal)
                    .where(IntentoEmisionFiscal.operacion_id == operacion.id)
                    .order_by(IntentoEmisionFiscal.id)
                )
            ).all()
        )
        assert lote.estado == "requiere_reconciliacion"
        assert [grupo.estado for grupo in grupos] == [
            "autorizado",
            "requiere_reconciliacion",
        ]
        assert [guarda.fase for guarda in guardas] == [
            "cerrada_terminal",
            "cerrada_pre_arca",
        ]
        assert [intento.estado for intento in intentos] == [
            "autorizado",
            "fallido_verificado",
        ]
        assert operacion.estado == "requiere_reconciliacion"
        comprobantes = list((await observador.scalars(select(Comprobante))).all())
        assert len(comprobantes) == 1


@pytest.mark.asyncio
async def test_worker_propaga_guarda_actual_a_recovery_aunque_hubo_fecae_previo(
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
) -> None:
    """El worker usa id/token de la guarda actual, no el bit global acumulado."""

    class SessionFactory:
        """Reutiliza la sesión aislada como conexión del worker."""

        async def __aenter__(self):
            return db_session

        async def __aexit__(self, exc_type, exc, tb):
            return False

    lote = LoteComprobante(
        nombre_archivo="lote-worker-guarda-actual.xlsx",
        archivo_hash="hash-lote-worker-guarda-actual",
        estado="en_cola",
        modo_procesamiento="background",
        procesamiento_async=True,
        total_filas=1,
        total_grupos=1,
        grupos_validos=1,
        empresa_id=test_empresa.id,
    )
    db_session.add(lote)
    await db_session.commit()
    lote_id = lote.id
    empresa_id = test_empresa.id
    guarda_id = 9876
    guarda_token = "f" * 64
    recovery_recibido: list[tuple[int | None, str | None]] = []

    async def fake_acquire(session: AsyncSession, role: str) -> None:
        assert session is db_session
        assert role == "worker"

    async def fail_segundo_chunk(self, lote_id, empresa_id, **kwargs):
        fase = kwargs["fase_solicitud_arca"]
        fase.marcar_iniciada()
        fase.registrar_guarda_pre_arca(
            SimpleNamespace(id=guarda_id, token=guarda_token)
        )
        assert fase.iniciada is True
        assert fase.guarda_actual_iniciada is False
        raise SQLAlchemyTimeoutError()

    async def fake_recovery(
        self,
        *,
        lote_id,
        empresa_id,
        guarda_rece_id,
        guarda_rece_token,
    ):
        recovery_recibido.append((guarda_rece_id, guarda_rece_token))
        return "requiere_reconciliacion"

    monkeypatch.setattr("app.services.lote_worker.WorkerSessionLocal", SessionFactory)
    monkeypatch.setattr(
        "app.services.lote_worker.acquire_database_connection",
        fake_acquire,
    )
    monkeypatch.setattr(
        LoteComprobantesService,
        "procesar_lote",
        fail_segundo_chunk,
    )
    monkeypatch.setattr(
        LoteComprobantesService,
        "recuperar_lote_worker_interrumpido_pre_arca",
        fake_recovery,
    )

    resultado = await LoteWorker().procesar_pendientes()

    assert resultado.lotes_en_cola_detectados == 1
    assert resultado.lotes_procesados == 0
    assert resultado.tuvo_error is True
    assert recovery_recibido == [(guarda_id, guarda_token)]
    assert (await db_session.get(LoteComprobante, lote_id)).empresa_id == empresa_id


@pytest.mark.asyncio
async def test_worker_no_procesa_en_cola_si_falla_bloqueo_stale(
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
):
    """Si no puede bloquear un stale, el worker no sigue con nuevos CAE."""

    class SessionFactory:
        async def __aenter__(self):
            return db_session

        async def __aexit__(self, exc_type, exc, tb):
            return False

    stale = LoteComprobante(
        nombre_archivo="lote-stale-worker.xlsx",
        archivo_hash="hash-stale-worker",
        estado="procesando",
        total_filas=1,
        total_grupos=1,
        grupos_validos=1,
        empresa_id=test_empresa.id,
        updated_at=datetime.utcnow()
        - timedelta(minutes=settings.batch_processing_stale_minutes + 1),
    )
    en_cola = LoteComprobante(
        nombre_archivo="lote-en-cola-worker.xlsx",
        archivo_hash="hash-en-cola-worker",
        estado="en_cola",
        total_filas=1,
        total_grupos=1,
        grupos_validos=1,
        empresa_id=test_empresa.id,
    )
    db_session.add_all([stale, en_cola])
    await db_session.commit()

    bloqueados: list[int] = []
    procesados: list[int] = []

    async def fail_bloquear(self, lote_id, empresa_id, **kwargs):
        bloqueados.append(lote_id)
        raise RuntimeError("fallo controlado de bloqueo stale")

    async def record_procesar(self, lote_id, empresa_id, **kwargs):
        procesados.append(lote_id)
        return await self.obtener_lote_resumen(lote_id, empresa_id)

    monkeypatch.setattr("app.services.lote_worker.WorkerSessionLocal", SessionFactory)
    monkeypatch.setattr(
        LoteComprobantesService,
        "bloquear_lote_procesando_stale",
        fail_bloquear,
    )
    monkeypatch.setattr(LoteComprobantesService, "procesar_lote", record_procesar)

    await LoteWorker().procesar_pendientes()

    assert bloqueados == [stale.id]
    assert procesados == []


@pytest.mark.asyncio
async def test_worker_no_procesa_en_cola_si_quedan_stale_fuera_del_batch(
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
) -> None:
    """Si quedan stale fuera del batch, el worker posterga nuevos CAE."""

    class SessionFactory:
        async def __aenter__(self):
            return db_session

        async def __aexit__(self, exc_type, exc, tb):
            return False

    monkeypatch.setattr(settings, "batch_worker_batch_size", 1)
    vencido = datetime.utcnow() - timedelta(
        minutes=settings.batch_processing_stale_minutes + 3
    )
    stale_1 = LoteComprobante(
        nombre_archivo="lote-stale-worker-1.xlsx",
        archivo_hash="hash-stale-worker-1",
        estado="procesando",
        total_filas=1,
        total_grupos=1,
        grupos_validos=1,
        empresa_id=test_empresa.id,
        updated_at=vencido,
    )
    stale_2 = LoteComprobante(
        nombre_archivo="lote-stale-worker-2.xlsx",
        archivo_hash="hash-stale-worker-2",
        estado="procesando",
        total_filas=1,
        total_grupos=1,
        grupos_validos=1,
        empresa_id=test_empresa.id,
        updated_at=vencido + timedelta(minutes=1),
    )
    en_cola = LoteComprobante(
        nombre_archivo="lote-en-cola-worker-overflow.xlsx",
        archivo_hash="hash-en-cola-worker-overflow",
        estado="en_cola",
        total_filas=1,
        total_grupos=1,
        grupos_validos=1,
        empresa_id=test_empresa.id,
    )
    db_session.add_all([stale_1, stale_2, en_cola])
    await db_session.commit()

    bloqueados: list[int] = []
    procesados: list[int] = []

    async def fake_bloquear(self, lote_id, empresa_id, **kwargs):
        bloqueados.append(lote_id)
        lote = await self.db.get(LoteComprobante, lote_id)
        lote.estado = "requiere_reconciliacion"
        await self.db.commit()
        return lote

    async def record_procesar(self, lote_id, empresa_id, **kwargs):
        procesados.append(lote_id)
        return await self.obtener_lote_resumen(lote_id, empresa_id)

    monkeypatch.setattr("app.services.lote_worker.WorkerSessionLocal", SessionFactory)
    monkeypatch.setattr(
        LoteComprobantesService,
        "bloquear_lote_procesando_stale",
        fake_bloquear,
    )
    monkeypatch.setattr(LoteComprobantesService, "procesar_lote", record_procesar)

    await LoteWorker().procesar_pendientes()

    assert bloqueados == [stale_1.id]
    assert procesados == []


@pytest.mark.asyncio
async def test_procesar_lote_legacy_sin_descripcion_item_bloquea_emision(
    client: AsyncClient,
    auth_headers: dict,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    """No debe emitir lotes validados antes de confirmar descripción facturada."""
    validar = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data=_opciones_fechas(),
        files={
            "archivo": (
                "lote-legacy-descripcion.xlsx",
                _build_lote_excel(test_empresa.cuit),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert validar.status_code == 200, validar.text
    lote_id = validar.json()["lote"]["id"]

    lote = await db_session.get(LoteComprobante, lote_id)
    assert lote is not None
    metadata = dict(lote.metadata_json or {})
    metadata.pop("opciones_descripcion_item", None)
    lote.metadata_json = metadata
    await db_session.commit()
    headers_procesar = await _confirmacion_fecha_fiscal_header_lote(
        db_session,
        lote_id=lote_id,
        estados={"validado"},
    )

    procesar = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/procesar",
        headers={**auth_headers, **headers_procesar},
    )

    assert procesar.status_code == 400
    detail = procesar.json()["detail"]
    assert "descripción facturada" in detail["mensaje"]


@pytest.mark.asyncio
async def test_procesar_lote_grande_encola_y_se_reanuda(
    client: AsyncClient,
    auth_headers: dict,
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_punto_venta,
    test_certificado,
):
    test_certificado.ambiente = settings.arca_env
    monkeypatch.setattr(settings, "batch_sync_limit", 0)

    async def fake_emitir(self, request, **kwargs):
        comprobante_id = await _persistir_comprobante_autorizado(
            db_session,
            test_empresa,
            test_punto_venta,
            tipo_comprobante=request.tipo_comprobante,
            numero=654,
            fecha_emision=request.fecha_emision,
            cae=CAE_TEST_NO_REAL_ALT,
            cae_vencimiento=date(2026, 3, 31),
            total=Decimal("1210.00"),
        )
        return EmitirComprobanteResponse(
            exito=True,
            comprobante_id=comprobante_id,
            tipo_comprobante=request.tipo_comprobante,
            punto_venta=1,
            numero=654,
            fecha=request.fecha_emision,
            cae=CAE_TEST_NO_REAL_ALT,
            cae_vencimiento=date(2026, 3, 31),
            total=Decimal("1210.00"),
            mensaje="Comprobante autorizado",
            errores=[],
        )

    monkeypatch.setattr(
        "app.services.facturacion_service.FacturacionService.emitir_comprobante",
        fake_emitir,
    )

    validar = await client.post(
        "/api/lotes-comprobantes/validar",
        headers=auth_headers,
        data=_opciones_fechas(),
        files={
            "archivo": (
                "lote-background.xlsx",
                _build_lote_excel(test_empresa.cuit),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert validar.status_code == 200, validar.text
    lote_id = validar.json()["lote"]["id"]
    headers_procesar = await _confirmacion_fecha_fiscal_header_lote(
        db_session,
        lote_id=lote_id,
        estados={"validado"},
    )

    procesar = await client.post(
        f"/api/lotes-comprobantes/{lote_id}/procesar",
        headers={**auth_headers, **headers_procesar},
    )

    assert procesar.status_code == 200, procesar.text
    data = procesar.json()
    assert data["en_progreso"] is True
    assert data["lote"]["estado"] == "en_cola"

    service = LoteComprobantesService(db_session)
    lote = await service.procesar_lote(lote_id, test_empresa.id, reanudar=True)

    assert lote.estado == "completado"
    assert lote.grupos_emitidos == 1
