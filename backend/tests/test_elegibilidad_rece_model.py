"""Pruebas de integridad persistente del modelo RECE PF-19B."""

from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy import func, select, update
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.models.elegibilidad_rece import (
    OperacionIdempotenteElegibilidadRece,
    PuntoVentaElegibilidadReceActual,
    PuntoVentaElegibilidadReceRevision,
    PuntoVentaGuardaEmisionRece,
)
from app.models.idempotencia_fiscal import IntentoEmisionFiscal, OperacionIdempotente
from app.models.lote_comprobante import (
    LoteComprobante,
    LoteComprobanteFila,
    LoteComprobanteGrupo,
)
from app.models.punto_venta import PuntoVenta
from app.services.elegibilidad_rece_service import (
    ContextoElegibilidadRece,
    ElegibilidadReceError,
    ElegibilidadReceService,
)
from app.services.idempotencia_fiscal_service import IdempotenciaFiscalService
from app.services.lote_comprobantes_service import LoteComprobantesService


async def crear_contexto_rece_verificado_sintetico(
    db: AsyncSession,
    *,
    empresa_id: int,
    empresa_cuit: str,
    usuario_id: int,
    numero: int = 41,
    idempotency_key: str = "rece-sintetico",
    guarda_fase: str = "arca_iniciada",
) -> tuple[
    PuntoVenta,
    PuntoVentaElegibilidadReceRevision,
    OperacionIdempotente,
    PuntoVentaGuardaEmisionRece,
]:
    """Crea evidencia positiva sintética solo para pruebas automatizadas."""
    punto = PuntoVenta(
        empresa_id=empresa_id,
        numero=numero,
        nombre="Punto RECE sintético",
        es_webservice=True,
        activo=True,
        bloqueado=False,
        revision_fiscal=1,
    )
    db.add(punto)
    await db.flush()

    ahora = datetime(2026, 8, 8, 12, 0, 0)
    revision = PuntoVentaElegibilidadReceRevision(
        empresa_id=empresa_id,
        punto_venta_id=punto.id,
        ambiente="produccion",
        revision=1,
        estado="verificado_rece",
        fuente="constancia_arca_atestada",
        evidencia_tipo="rece_aplicativo_web_services_v1",
        evidencia_sha256="a" * 64,
        clasificador_version="rece-v1-test",
        empresa_cuit_snapshot=empresa_cuit,
        punto_venta_numero_snapshot=numero,
        punto_revision_fiscal=1,
        documento_emitido_en=date(2026, 8, 8),
        vigente_hasta=date(2026, 8, 15),
        observado_en=ahora,
        verificado_en=ahora,
        creado_por_usuario_id=usuario_id,
        actor_usuario_id_snapshot=usuario_id,
        created_at=ahora,
    )
    db.add(revision)
    await db.flush()
    db.add(
        PuntoVentaElegibilidadReceActual(
            empresa_id=empresa_id,
            punto_venta_id=punto.id,
            ambiente="produccion",
            revision_actual_id=revision.id,
        )
    )

    operacion = OperacionIdempotente(
        empresa_id=empresa_id,
        usuario_id=usuario_id,
        idempotency_key=idempotency_key,
        tipo_operacion="emitir_comprobante",
        payload_hash="b" * 64,
        estado="en_proceso",
        rece_snapshot_hash=ElegibilidadReceService.calcular_digest_contextos(
            [
                ContextoElegibilidadRece(
                    empresa_id=empresa_id,
                    punto_venta_id=punto.id,
                    punto_venta_numero=numero,
                    ambiente="produccion",
                    elegibilidad_revision_id=revision.id,
                    punto_venta_revision_fiscal=1,
                )
            ]
        ),
    )
    db.add(operacion)
    await db.flush()
    db.add(
        OperacionIdempotenteElegibilidadRece(
            operacion_id=operacion.id,
            empresa_id=empresa_id,
            punto_venta_id=punto.id,
            ambiente="produccion",
            elegibilidad_revision_id=revision.id,
            punto_venta_revision_fiscal=1,
        )
    )
    await db.flush()

    guarda = PuntoVentaGuardaEmisionRece(
        token="d" * 64,
        fase=guarda_fase,
        operacion_id=operacion.id,
        empresa_id=empresa_id,
        punto_venta_id=punto.id,
        ambiente="produccion",
        elegibilidad_revision_id=revision.id,
        punto_venta_revision_fiscal=1,
        arca_iniciada_en=(ahora if guarda_fase == "arca_iniciada" else None),
    )
    db.add(guarda)
    await db.flush()
    return punto, revision, operacion, guarda


def crear_intento_sintetico_rece(
    *,
    empresa_id: int,
    usuario_id: int,
    punto: PuntoVenta,
    revision: PuntoVentaElegibilidadReceRevision,
    operacion: OperacionIdempotente,
    guarda: PuntoVentaGuardaEmisionRece,
    numero_planificado: int = 1,
    lote: LoteComprobante | None = None,
    grupo: LoteComprobanteGrupo | None = None,
) -> IntentoEmisionFiscal:
    """Construye un intento individual sintético con snapshot RECE exacto."""
    return IntentoEmisionFiscal(
        operacion_id=operacion.id,
        empresa_id=empresa_id,
        usuario_id=usuario_id,
        punto_venta_id=punto.id,
        punto_venta_numero=punto.numero,
        tipo_comprobante=6,
        numero_planificado=numero_planificado,
        fecha_emision=date(2026, 8, 8),
        total=Decimal("121.00"),
        payload_hash=f"{numero_planificado + 30:064d}",
        huella_logica=f"{numero_planificado + 40:064d}",
        estado="en_proceso",
        ambiente="produccion",
        punto_venta_elegibilidad_revision_id=revision.id,
        punto_venta_revision_fiscal=1,
        guarda_rece_id=guarda.id,
        lote_id=lote.id if lote is not None else None,
        grupo_id=grupo.id if grupo is not None else None,
    )


async def crear_lote_recovery_sintetico(
    db: AsyncSession,
    *,
    empresa_id: int,
    usuario_id: int,
    punto: PuntoVenta,
    revision: PuntoVentaElegibilidadReceRevision,
    operacion: OperacionIdempotente,
    guarda: PuntoVentaGuardaEmisionRece,
    lote_estado: str = "procesando",
    grupo_estado: str = "validado",
    numero_planificado: int = 1,
    con_fila: bool = True,
    respuesta_worker_en_progreso: bool = False,
) -> tuple[LoteComprobante, LoteComprobanteGrupo, IntentoEmisionFiscal]:
    """Crea lote, grupo e intento batch con identidad RECE exacta."""
    lote = LoteComprobante(
        empresa_id=empresa_id,
        usuario_id=usuario_id,
        nombre_archivo=f"recovery-{operacion.id}-{numero_planificado}.xlsx",
        archivo_hash=f"{operacion.id + numero_planificado + 10:064d}",
        estado=lote_estado,
        total_filas=1 if con_fila else 0,
        total_grupos=1,
        metadata_json={"operacion_idempotente_id": operacion.id},
        compactado_at=(datetime.utcnow() if not con_fila else None),
        procesamiento_async=respuesta_worker_en_progreso,
        modo_procesamiento=("background" if respuesta_worker_en_progreso else "sync"),
    )
    db.add(lote)
    await db.flush()
    operacion.lote_id = lote.id
    grupo = LoteComprobanteGrupo(
        lote_id=lote.id,
        empresa_id=empresa_id,
        comprobante_ref=f"REC-{operacion.id}-{numero_planificado}",
        estado=grupo_estado,
        tipo_comprobante=6,
        punto_venta_numero=punto.numero,
        total_estimado=Decimal("121.00"),
        payload_json={"fecha_emision": "2026-08-08"},
        mensajes_json=["Estado previo sintético."],
        punto_venta_id=punto.id,
        ambiente="produccion",
        punto_venta_elegibilidad_revision_id=revision.id,
        punto_venta_revision_fiscal=1,
    )
    db.add(grupo)
    await db.flush()
    if con_fila:
        db.add(
            LoteComprobanteFila(
                lote_id=lote.id,
                grupo_id=grupo.id,
                fila_excel=2,
                comprobante_ref=grupo.comprobante_ref,
                estado=grupo_estado,
                datos_json={},
                mensajes_json=list(grupo.mensajes_json),
            )
        )
    intento = crear_intento_sintetico_rece(
        empresa_id=empresa_id,
        usuario_id=usuario_id,
        punto=punto,
        revision=revision,
        operacion=operacion,
        guarda=guarda,
        numero_planificado=numero_planificado,
        lote=lote,
        grupo=grupo,
    )
    db.add(intento)
    material_rece = await LoteComprobantesService(
        db
    ).calcular_material_idempotente_grupos(
        lote_id=lote.id,
        empresa_id=empresa_id,
        grupo_ids=[grupo.id],
    )
    lote.metadata_json = {
        "operacion_idempotente_id": operacion.id,
        "pf19b_rece_material": material_rece,
    }
    if respuesta_worker_en_progreso:
        operacion.tipo_operacion = "procesar_lote"
        operacion.response_json = {
            "lote": {
                "id": lote.id,
                "empresa_id": empresa_id,
                "estado": "en_cola",
                "modo_procesamiento": "background",
                "procesamiento_async": True,
                "metadata_json": dict(lote.metadata_json or {}),
            },
            "mensaje": "El lote quedó en cola para procesamiento en segundo plano.",
            "en_progreso": True,
        }
    await db.commit()
    return lote, grupo, intento


@pytest.mark.asyncio
async def test_sqlite_fks_rechazan_intento_con_guarda_de_otra_operacion(
    db_session: AsyncSession,
    test_empresa,
    test_user,
) -> None:
    """La guarda y el intento deben compartir operación y snapshot exactos."""
    punto, revision, operacion, guarda = await crear_contexto_rece_verificado_sintetico(
        db_session,
        empresa_id=test_empresa.id,
        empresa_cuit=test_empresa.cuit,
        usuario_id=test_user.id,
    )
    intento = IntentoEmisionFiscal(
        operacion_id=operacion.id,
        empresa_id=test_empresa.id,
        usuario_id=test_user.id,
        punto_venta_id=punto.id,
        punto_venta_numero=punto.numero,
        tipo_comprobante=6,
        numero_planificado=1,
        fecha_emision=date(2026, 8, 8),
        total=Decimal("121.00"),
        payload_hash="e" * 64,
        huella_logica="f" * 64,
        estado="en_proceso",
        ambiente="produccion",
        punto_venta_elegibilidad_revision_id=revision.id,
        punto_venta_revision_fiscal=1,
        guarda_rece_id=guarda.id,
    )
    db_session.add(intento)
    await db_session.commit()

    otra_operacion = OperacionIdempotente(
        empresa_id=test_empresa.id,
        usuario_id=test_user.id,
        idempotency_key="rece-sintetico-otra-op",
        tipo_operacion="emitir_comprobante",
        payload_hash="1" * 64,
        estado="en_proceso",
        rece_snapshot_hash="2" * 64,
    )
    db_session.add(otra_operacion)
    await db_session.commit()

    db_session.add(
        IntentoEmisionFiscal(
            operacion_id=otra_operacion.id,
            empresa_id=test_empresa.id,
            usuario_id=test_user.id,
            punto_venta_id=punto.id,
            punto_venta_numero=punto.numero,
            tipo_comprobante=6,
            numero_planificado=2,
            fecha_emision=date(2026, 8, 8),
            total=Decimal("121.00"),
            payload_hash="3" * 64,
            huella_logica="4" * 64,
            estado="en_proceso",
            ambiente="produccion",
            punto_venta_elegibilidad_revision_id=revision.id,
            punto_venta_revision_fiscal=1,
            guarda_rece_id=guarda.id,
        )
    )
    with pytest.raises(IntegrityError):
        await db_session.commit()
    await db_session.rollback()


@pytest.mark.asyncio
async def test_recuperacion_cierra_guarda_pre_arca_y_habilita_reclamo_seguro(
    db_session: AsyncSession,
    test_empresa,
    test_user,
) -> None:
    """Una guarda pre-ARCA propia se cierra con intento y operación atómicos."""
    punto, revision, operacion, guarda = await crear_contexto_rece_verificado_sintetico(
        db_session,
        empresa_id=test_empresa.id,
        empresa_cuit=test_empresa.cuit,
        usuario_id=test_user.id,
        idempotency_key="rece-recuperacion-pre-arca",
        guarda_fase="pre_arca",
    )
    intento = crear_intento_sintetico_rece(
        empresa_id=test_empresa.id,
        usuario_id=test_user.id,
        punto=punto,
        revision=revision,
        operacion=operacion,
        guarda=guarda,
    )
    db_session.add(intento)
    await db_session.commit()

    resultado = await ElegibilidadReceService(
        db_session
    ).recuperar_guarda_interrumpida_pre_arca(
        operacion_id=operacion.id,
        guarda_id=guarda.id,
        token=guarda.token,
    )

    assert resultado == "recuperada_pre_arca"
    await db_session.refresh(guarda)
    await db_session.refresh(intento)
    await db_session.refresh(operacion)
    assert guarda.fase == "cerrada_pre_arca"
    assert guarda.arca_iniciada_en is None
    assert guarda.cerrada_en is not None
    assert intento.estado == "fallido_verificado"
    assert intento.categoria_error == "interrumpida_pre_arca_recuperada"
    assert operacion.estado == "interrumpida_pre_arca"


@pytest.mark.asyncio
async def test_recuperacion_nunca_libera_guarda_con_inicio_arca_commiteado(
    db_session: AsyncSession,
    test_empresa,
    test_user,
) -> None:
    """Un commit ambiguo de arca_iniciada conserva guarda e impone reconciliación."""
    punto, revision, operacion, guarda = await crear_contexto_rece_verificado_sintetico(
        db_session,
        empresa_id=test_empresa.id,
        empresa_cuit=test_empresa.cuit,
        usuario_id=test_user.id,
        idempotency_key="rece-recuperacion-arca-iniciada",
    )
    intento = crear_intento_sintetico_rece(
        empresa_id=test_empresa.id,
        usuario_id=test_user.id,
        punto=punto,
        revision=revision,
        operacion=operacion,
        guarda=guarda,
    )
    db_session.add(intento)
    await db_session.commit()

    resultado = await ElegibilidadReceService(
        db_session
    ).recuperar_guarda_interrumpida_pre_arca(
        operacion_id=operacion.id,
        guarda_id=guarda.id,
        token=guarda.token,
    )

    assert resultado == "requiere_reconciliacion"
    await db_session.refresh(guarda)
    await db_session.refresh(intento)
    await db_session.refresh(operacion)
    assert guarda.fase == "requiere_reconciliacion"
    assert guarda.arca_iniciada_en is not None
    assert guarda.cerrada_en is None
    assert intento.estado == "requiere_reconciliacion"
    assert operacion.estado == "requiere_reconciliacion"


@pytest.mark.asyncio
async def test_recuperacion_rechaza_operacion_con_intento_fuera_de_su_guarda(
    db_session: AsyncSession,
    test_empresa,
    test_user,
) -> None:
    """No abre replay si la operación conserva otra evidencia fiscal."""
    punto, revision, operacion, guarda = await crear_contexto_rece_verificado_sintetico(
        db_session,
        empresa_id=test_empresa.id,
        empresa_cuit=test_empresa.cuit,
        usuario_id=test_user.id,
        idempotency_key="rece-recuperacion-evidencia-ajena",
        guarda_fase="pre_arca",
    )
    intento = crear_intento_sintetico_rece(
        empresa_id=test_empresa.id,
        usuario_id=test_user.id,
        punto=punto,
        revision=revision,
        operacion=operacion,
        guarda=guarda,
    )
    intento_legacy = IntentoEmisionFiscal(
        operacion_id=operacion.id,
        empresa_id=test_empresa.id,
        usuario_id=test_user.id,
        punto_venta_id=punto.id,
        punto_venta_numero=punto.numero,
        tipo_comprobante=6,
        numero_planificado=2,
        fecha_emision=date(2026, 8, 8),
        total=Decimal("121.00"),
        payload_hash="7" * 64,
        huella_logica="8" * 64,
        estado="fallido_verificado",
    )
    db_session.add_all([intento, intento_legacy])
    await db_session.commit()

    resultado = await ElegibilidadReceService(
        db_session
    ).recuperar_guarda_interrumpida_pre_arca(
        operacion_id=operacion.id,
        guarda_id=guarda.id,
        token=guarda.token,
    )

    assert resultado == "no_recuperable"
    await db_session.refresh(guarda)
    await db_session.refresh(intento)
    await db_session.refresh(operacion)
    assert guarda.fase == "pre_arca"
    assert intento.estado == "en_proceso"
    assert operacion.estado == "en_proceso"


@pytest.mark.asyncio
async def test_recuperacion_rechaza_guarda_huerfana_de_la_misma_operacion(
    db_session: AsyncSession,
    test_empresa,
    test_user,
) -> None:
    """Recovery no ignora una guarda durable sin intento que la referencie."""
    punto, revision, operacion, guarda = await crear_contexto_rece_verificado_sintetico(
        db_session,
        empresa_id=test_empresa.id,
        empresa_cuit=test_empresa.cuit,
        usuario_id=test_user.id,
        idempotency_key="rece-recovery-guarda-huerfana",
        guarda_fase="pre_arca",
    )
    intento = crear_intento_sintetico_rece(
        empresa_id=test_empresa.id,
        usuario_id=test_user.id,
        punto=punto,
        revision=revision,
        operacion=operacion,
        guarda=guarda,
    )
    guarda_huerfana = PuntoVentaGuardaEmisionRece(
        token="e" * 64,
        fase="cerrada_pre_arca",
        operacion_id=operacion.id,
        empresa_id=test_empresa.id,
        punto_venta_id=punto.id,
        ambiente="produccion",
        elegibilidad_revision_id=revision.id,
        punto_venta_revision_fiscal=1,
        cerrada_en=datetime(2026, 8, 8, 13, 0, 0),
    )
    db_session.add_all([intento, guarda_huerfana])
    await db_session.commit()
    operacion_id = int(operacion.id)
    guarda_id = int(guarda.id)
    intento_id = int(intento.id)

    resultado = await ElegibilidadReceService(
        db_session
    ).recuperar_guarda_interrumpida_pre_arca(
        operacion_id=operacion_id,
        guarda_id=guarda_id,
        token="d" * 64,
    )

    assert resultado == "no_recuperable"
    assert (await db_session.get(PuntoVentaGuardaEmisionRece, guarda_id)).fase == (
        "pre_arca"
    )
    assert (
        await db_session.get(IntentoEmisionFiscal, intento_id)
    ).estado == "en_proceso"
    assert (await db_session.get(OperacionIdempotente, operacion_id)).estado == (
        "en_proceso"
    )


@pytest.mark.asyncio
async def test_cas_arca_rechaza_guarda_huerfana_de_la_misma_operacion(
    db_session: AsyncSession,
    test_empresa,
    test_user,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """El CAS final exige correspondencia exacta entre guardas e intentos."""
    monkeypatch.setattr(settings, "arca_env", "produccion")
    punto, revision, operacion, guarda = await crear_contexto_rece_verificado_sintetico(
        db_session,
        empresa_id=test_empresa.id,
        empresa_cuit=test_empresa.cuit,
        usuario_id=test_user.id,
        idempotency_key="rece-cas-guarda-huerfana",
        guarda_fase="pre_arca",
    )
    intento = crear_intento_sintetico_rece(
        empresa_id=test_empresa.id,
        usuario_id=test_user.id,
        punto=punto,
        revision=revision,
        operacion=operacion,
        guarda=guarda,
    )
    db_session.add_all(
        [
            intento,
            PuntoVentaGuardaEmisionRece(
                token="e" * 64,
                fase="cerrada_pre_arca",
                operacion_id=operacion.id,
                empresa_id=test_empresa.id,
                punto_venta_id=punto.id,
                ambiente="produccion",
                elegibilidad_revision_id=revision.id,
                punto_venta_revision_fiscal=1,
                cerrada_en=datetime(2026, 8, 8, 13, 0, 0),
            ),
        ]
    )
    await db_session.commit()
    guarda_id = int(guarda.id)
    intento_id = int(intento.id)
    contexto = ContextoElegibilidadRece(
        empresa_id=test_empresa.id,
        punto_venta_id=punto.id,
        punto_venta_numero=punto.numero,
        ambiente="produccion",
        elegibilidad_revision_id=revision.id,
        punto_venta_revision_fiscal=1,
    )

    with pytest.raises(ElegibilidadReceError, match="guardas RECE sin intentos"):
        await ElegibilidadReceService(db_session).marcar_arca_iniciada(
            guarda=guarda,
            contexto=contexto,
            tipo_comprobante=6,
        )

    assert (await db_session.get(PuntoVentaGuardaEmisionRece, guarda_id)).fase == (
        "pre_arca"
    )
    assert (
        await db_session.get(IntentoEmisionFiscal, intento_id)
    ).estado == "en_proceso"


@pytest.mark.asyncio
async def test_cas_arca_rechaza_owner_lote_adulterado_antes_de_fecae(
    db_session: AsyncSession,
    test_empresa,
    test_user,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """El CAS final no inicia ARCA si el lote dejó de pertenecer a la operación."""
    monkeypatch.setattr(settings, "arca_env", "produccion")
    punto, revision, operacion, guarda = await crear_contexto_rece_verificado_sintetico(
        db_session,
        empresa_id=test_empresa.id,
        empresa_cuit=test_empresa.cuit,
        usuario_id=test_user.id,
        idempotency_key="rece-cas-owner-lote-adulterado",
        guarda_fase="pre_arca",
    )
    operacion.tipo_operacion = "reintentar_fallidos_lote"
    lote, _, intento = await crear_lote_recovery_sintetico(
        db_session,
        empresa_id=test_empresa.id,
        usuario_id=test_user.id,
        punto=punto,
        revision=revision,
        operacion=operacion,
        guarda=guarda,
        lote_estado="con_errores",
        grupo_estado="reintentando",
    )
    lote_id = int(lote.id)
    operacion_id = int(operacion.id)
    guarda_id = int(guarda.id)
    intento_id = int(intento.id)
    empresa_id = int(test_empresa.id)
    punto_id = int(punto.id)
    punto_numero = int(punto.numero)
    revision_id = int(revision.id)
    metadata = dict(lote.metadata_json or {})
    metadata["operacion_idempotente_id"] = operacion_id + 999
    lote.metadata_json = metadata
    await db_session.commit()
    await db_session.refresh(guarda)
    contexto = ContextoElegibilidadRece(
        empresa_id=empresa_id,
        punto_venta_id=punto_id,
        punto_venta_numero=punto_numero,
        ambiente="produccion",
        elegibilidad_revision_id=revision_id,
        punto_venta_revision_fiscal=1,
    )

    with pytest.raises(ElegibilidadReceError, match="operación propietaria"):
        await ElegibilidadReceService(db_session).marcar_arca_iniciada(
            guarda=guarda,
            contexto=contexto,
            tipo_comprobante=6,
        )

    db_session.expire_all()
    assert (await db_session.get(PuntoVentaGuardaEmisionRece, guarda_id)).fase == (
        "pre_arca"
    )
    assert (
        await db_session.get(IntentoEmisionFiscal, intento_id)
    ).estado == "en_proceso"
    assert (await db_session.get(LoteComprobante, lote_id)).metadata_json[
        "operacion_idempotente_id"
    ] != operacion_id


@pytest.mark.asyncio
async def test_stale_batch_no_reclama_operacion_con_guarda_huerfana_cerrada(
    db_session: AsyncSession,
    test_empresa,
    test_user,
) -> None:
    """El stale batch falla cerrado ante una guarda ajena al grafo de intentos."""
    punto, revision, operacion, guarda = await crear_contexto_rece_verificado_sintetico(
        db_session,
        empresa_id=test_empresa.id,
        empresa_cuit=test_empresa.cuit,
        usuario_id=test_user.id,
        idempotency_key="rece-stale-batch-guarda-huerfana",
        guarda_fase="pre_arca",
    )
    lote, _, intento = await crear_lote_recovery_sintetico(
        db_session,
        empresa_id=test_empresa.id,
        usuario_id=test_user.id,
        punto=punto,
        revision=revision,
        operacion=operacion,
        guarda=guarda,
        respuesta_worker_en_progreso=True,
    )
    guarda.fase = "cerrada_pre_arca"
    guarda.cerrada_en = datetime(2026, 8, 8, 13, 0, 0)
    intento.estado = "fallido_verificado"
    await db_session.commit()
    db_session.add(
        PuntoVentaGuardaEmisionRece(
            token="e" * 64,
            fase="cerrada_pre_arca",
            operacion_id=operacion.id,
            empresa_id=test_empresa.id,
            punto_venta_id=punto.id,
            ambiente="produccion",
            elegibilidad_revision_id=revision.id,
            punto_venta_revision_fiscal=1,
            cerrada_en=datetime(2026, 8, 8, 13, 5, 0),
        )
    )
    await db_session.commit()
    await db_session.execute(
        update(LoteComprobante)
        .where(LoteComprobante.id == lote.id)
        .values(
            updated_at=datetime.utcnow()
            - timedelta(minutes=settings.batch_processing_stale_minutes + 1)
        )
    )
    await db_session.commit()

    resultado = await LoteComprobantesService(
        db_session
    )._lote_tiene_guarda_rece_activa(lote)

    assert resultado == "legacy_invalida"
    lote_actual = await db_session.get(LoteComprobante, lote.id)
    assert lote_actual.estado == "procesando"


@pytest.mark.asyncio
async def test_recovery_lote_revierte_todo_si_grupo_cambio_de_estado(
    db_session: AsyncSession,
    test_empresa,
    test_user,
) -> None:
    """Un grupo fuera de `validado` impide cerrar guarda, intento u operación."""
    punto, revision, operacion, guarda = await crear_contexto_rece_verificado_sintetico(
        db_session,
        empresa_id=test_empresa.id,
        empresa_cuit=test_empresa.cuit,
        usuario_id=test_user.id,
        idempotency_key="rece-recovery-grupo-alterado",
        guarda_fase="pre_arca",
    )
    lote, grupo, intento = await crear_lote_recovery_sintetico(
        db_session,
        empresa_id=test_empresa.id,
        usuario_id=test_user.id,
        punto=punto,
        revision=revision,
        operacion=operacion,
        guarda=guarda,
        grupo_estado="fallido",
    )
    guarda_id = guarda.id
    guarda_token = guarda.token
    intento_id = intento.id
    operacion_id = operacion.id
    lote_id = lote.id
    grupo_id = grupo.id
    empresa_id = test_empresa.id

    resultado = await LoteComprobantesService(
        db_session
    ).recuperar_lote_interrumpido_pre_arca(
        lote_id=lote_id,
        empresa_id=empresa_id,
        operacion_id=operacion_id,
        estado_reanudable="validado",
        estados_claim={"procesando"},
        mensaje_seguro="Recovery sintético.",
        guarda_rece_id=guarda_id,
        guarda_rece_token=guarda_token,
    )

    assert resultado == "no_recuperable"
    async with AsyncSession(bind=db_session.bind, expire_on_commit=False) as observador:
        assert (
            await observador.get(PuntoVentaGuardaEmisionRece, guarda_id)
        ).fase == "pre_arca"
        assert (
            await observador.get(IntentoEmisionFiscal, intento_id)
        ).estado == "en_proceso"
        assert (
            await observador.get(OperacionIdempotente, operacion_id)
        ).estado == "en_proceso"
        assert (await observador.get(LoteComprobante, lote_id)).estado == "procesando"
        assert (
            await observador.get(LoteComprobanteGrupo, grupo_id)
        ).estado == "fallido"


@pytest.mark.asyncio
async def test_recovery_worker_retiene_operacion_antes_de_reencolar(
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    test_empresa,
    test_user,
) -> None:
    """El worker conserva ownership idempotente al reencolar tras recovery seguro."""
    punto, revision, operacion, guarda = await crear_contexto_rece_verificado_sintetico(
        db_session,
        empresa_id=test_empresa.id,
        empresa_cuit=test_empresa.cuit,
        usuario_id=test_user.id,
        idempotency_key="rece-recovery-worker",
        guarda_fase="pre_arca",
    )
    lote, grupo, intento = await crear_lote_recovery_sintetico(
        db_session,
        empresa_id=test_empresa.id,
        usuario_id=test_user.id,
        punto=punto,
        revision=revision,
        operacion=operacion,
        guarda=guarda,
        respuesta_worker_en_progreso=True,
    )
    guarda_id = guarda.id
    guarda_token = guarda.token
    intento_id = intento.id
    operacion_id = operacion.id
    lote_id = lote.id
    grupo_id = grupo.id
    empresa_id = test_empresa.id
    resultados_guarda: list[str] = []
    recuperar_guarda_original = (
        ElegibilidadReceService.recuperar_guarda_interrumpida_pre_arca
    )

    async def registrar_resultado_guarda(self, **kwargs):
        resultado_guarda = await recuperar_guarda_original(self, **kwargs)
        resultados_guarda.append(resultado_guarda)
        return resultado_guarda

    monkeypatch.setattr(
        ElegibilidadReceService,
        "recuperar_guarda_interrumpida_pre_arca",
        registrar_resultado_guarda,
    )

    service = LoteComprobantesService(db_session)
    assert service._respuesta_worker_en_progreso_valida(
        operacion.response_json,
        lote_id=lote_id,
        empresa_id=empresa_id,
        operacion_id=operacion_id,
        material_rece=lote.metadata_json["pf19b_rece_material"],
    )
    resultado = await service.recuperar_lote_worker_interrumpido_pre_arca(
        lote_id=lote_id,
        empresa_id=empresa_id,
        guarda_rece_id=guarda_id,
        guarda_rece_token=guarda_token,
    )

    assert resultados_guarda == ["recuperada_pre_arca"]
    assert resultado == "recuperada_pre_arca"
    async with AsyncSession(bind=db_session.bind, expire_on_commit=False) as observador:
        guarda_actual = await observador.get(PuntoVentaGuardaEmisionRece, guarda_id)
        intento_actual = await observador.get(IntentoEmisionFiscal, intento_id)
        operacion_actual = await observador.get(OperacionIdempotente, operacion_id)
        lote_actual = await observador.get(LoteComprobante, lote_id)
        grupo_actual = await observador.get(LoteComprobanteGrupo, grupo_id)
        assert guarda_actual.fase == "cerrada_pre_arca"
        assert intento_actual.estado == "fallido_verificado"
        assert operacion_actual.estado == "en_proceso"
        assert operacion_actual.response_json is not None
        assert operacion_actual.response_json["en_progreso"] is True
        assert operacion_actual.response_json["lote"]["id"] == lote_id
        assert lote_actual.estado == "en_cola"
        assert grupo_actual.estado == "validado"

        _, reclamada_por_http = await IdempotenciaFiscalService(
            observador
        ).reclamar_operacion_interrumpida_pre_arca(operacion_actual)
        assert reclamada_por_http is False


@pytest.mark.asyncio
async def test_recovery_ambiguo_actualiza_detalle_y_archivo_observado(
    db_session: AsyncSession,
    test_empresa,
    test_user,
) -> None:
    """Lote, grupo y filas reflejan la misma reconciliación durable."""
    punto, revision, operacion, guarda = await crear_contexto_rece_verificado_sintetico(
        db_session,
        empresa_id=test_empresa.id,
        empresa_cuit=test_empresa.cuit,
        usuario_id=test_user.id,
        idempotency_key="rece-recovery-observado",
        guarda_fase="arca_iniciada",
    )
    lote, grupo, intento = await crear_lote_recovery_sintetico(
        db_session,
        empresa_id=test_empresa.id,
        usuario_id=test_user.id,
        punto=punto,
        revision=revision,
        operacion=operacion,
        guarda=guarda,
    )
    service = LoteComprobantesService(db_session)
    guarda_id = guarda.id
    guarda_token = guarda.token
    intento_id = intento.id
    operacion_id = operacion.id
    lote_id = lote.id
    empresa_id = test_empresa.id

    resultado = await service.recuperar_lote_interrumpido_pre_arca(
        lote_id=lote_id,
        empresa_id=empresa_id,
        operacion_id=operacion_id,
        estado_reanudable="validado",
        estados_claim={"procesando"},
        mensaje_seguro="No debe usarse.",
        guarda_rece_id=guarda_id,
        guarda_rece_token=guarda_token,
    )

    assert resultado == "requiere_reconciliacion"
    async with AsyncSession(bind=db_session.bind, expire_on_commit=False) as observador:
        servicio_observador = LoteComprobantesService(observador)
        detalle = await servicio_observador.obtener_lote(lote_id, empresa_id)
        assert detalle.estado == "requiere_reconciliacion"
        assert detalle.grupos[0].estado == "requiere_reconciliacion"
        assert detalle.filas[0].estado == "requiere_reconciliacion"
        mensaje = detalle.grupos[0].mensajes_json[0]
        assert "No reintentes" in mensaje
        assert detalle.filas[0].mensajes_json == [mensaje]
        assert (
            await observador.get(IntentoEmisionFiscal, intento_id)
        ).estado == "requiere_reconciliacion"

        observado = load_workbook(
            BytesIO(
                await servicio_observador.generar_archivo_observado(lote_id, empresa_id)
            )
        )
        hoja = observado["Resultados"]
        assert hoja.cell(row=2, column=hoja.max_column - 1).value == (
            "requiere_reconciliacion"
        )
        assert "No reintentes" in hoja.cell(row=2, column=hoja.max_column).value


@pytest.mark.asyncio
async def test_recovery_segundo_chunk_pre_arca_cierra_local_y_bloquea_replay_global(
    db_session: AsyncSession,
    test_empresa,
    test_user,
) -> None:
    """Una guarda previa cerrada no deja activa la segunda tras una caída DB."""
    (
        punto,
        revision,
        operacion,
        primera_guarda,
    ) = await crear_contexto_rece_verificado_sintetico(
        db_session,
        empresa_id=test_empresa.id,
        empresa_cuit=test_empresa.cuit,
        usuario_id=test_user.id,
        idempotency_key="rece-recovery-dos-chunks",
        guarda_fase="pre_arca",
    )
    lote, primer_grupo, primer_intento = await crear_lote_recovery_sintetico(
        db_session,
        empresa_id=test_empresa.id,
        usuario_id=test_user.id,
        punto=punto,
        revision=revision,
        operacion=operacion,
        guarda=primera_guarda,
        grupo_estado="fallido",
        numero_planificado=1,
    )
    ahora = datetime.utcnow()
    primera_guarda.fase = "cerrada_pre_arca"
    primera_guarda.cerrada_en = ahora
    primer_intento.estado = "fallido_verificado"
    segunda_guarda = PuntoVentaGuardaEmisionRece(
        token="e" * 64,
        fase="pre_arca",
        operacion_id=operacion.id,
        empresa_id=test_empresa.id,
        punto_venta_id=punto.id,
        ambiente="produccion",
        elegibilidad_revision_id=revision.id,
        punto_venta_revision_fiscal=1,
    )
    db_session.add(segunda_guarda)
    await db_session.flush()
    segundo_grupo = LoteComprobanteGrupo(
        lote_id=lote.id,
        empresa_id=test_empresa.id,
        comprobante_ref="REC-SEGUNDO-CHUNK",
        estado="validado",
        tipo_comprobante=6,
        punto_venta_numero=punto.numero,
        total_estimado=Decimal("121.00"),
        payload_json={"fecha_emision": "2026-08-08"},
        mensajes_json=["Estado previo sintético."],
        punto_venta_id=punto.id,
        ambiente="produccion",
        punto_venta_elegibilidad_revision_id=revision.id,
        punto_venta_revision_fiscal=1,
    )
    db_session.add(segundo_grupo)
    await db_session.flush()
    db_session.add(
        LoteComprobanteFila(
            lote_id=lote.id,
            grupo_id=segundo_grupo.id,
            fila_excel=3,
            comprobante_ref=segundo_grupo.comprobante_ref,
            estado="validado",
            datos_json={},
            mensajes_json=["Estado previo sintético."],
        )
    )
    segundo_intento = crear_intento_sintetico_rece(
        empresa_id=test_empresa.id,
        usuario_id=test_user.id,
        punto=punto,
        revision=revision,
        operacion=operacion,
        guarda=segunda_guarda,
        numero_planificado=2,
        lote=lote,
        grupo=segundo_grupo,
    )
    db_session.add(segundo_intento)
    await db_session.commit()
    primera_guarda_id = primera_guarda.id
    segunda_guarda_id = segunda_guarda.id
    segunda_guarda_token = segunda_guarda.token
    primer_intento_id = primer_intento.id
    segundo_intento_id = segundo_intento.id
    operacion_id = operacion.id
    lote_id = lote.id
    primer_grupo_id = primer_grupo.id
    segundo_grupo_id = segundo_grupo.id
    empresa_id = test_empresa.id

    resultado = await LoteComprobantesService(
        db_session
    ).recuperar_lote_interrumpido_pre_arca(
        lote_id=lote_id,
        empresa_id=empresa_id,
        operacion_id=operacion_id,
        estado_reanudable="validado",
        estados_claim={"procesando"},
        mensaje_seguro="No debe reencolar.",
        guarda_rece_id=segunda_guarda_id,
        guarda_rece_token=segunda_guarda_token,
    )

    assert resultado == "requiere_reconciliacion"
    async with AsyncSession(bind=db_session.bind, expire_on_commit=False) as observador:
        assert (
            await observador.get(PuntoVentaGuardaEmisionRece, primera_guarda_id)
        ).fase == "cerrada_pre_arca"
        assert (
            await observador.get(PuntoVentaGuardaEmisionRece, segunda_guarda_id)
        ).fase == "cerrada_pre_arca"
        assert (
            await observador.get(IntentoEmisionFiscal, primer_intento_id)
        ).estado == "fallido_verificado"
        assert (
            await observador.get(IntentoEmisionFiscal, segundo_intento_id)
        ).estado == "fallido_verificado"
        assert (
            await observador.get(OperacionIdempotente, operacion_id)
        ).estado == "requiere_reconciliacion"
        assert (
            await observador.get(LoteComprobante, lote_id)
        ).estado == "requiere_reconciliacion"
        assert (
            await observador.get(LoteComprobanteGrupo, primer_grupo_id)
        ).estado == "fallido"
        assert (
            await observador.get(LoteComprobanteGrupo, segundo_grupo_id)
        ).estado == "requiere_reconciliacion"


@pytest.mark.asyncio
async def test_sqlite_fk_rechaza_intento_batch_con_grupo_o_snapshot_cruzado(
    db_session: AsyncSession,
    test_empresa,
    test_user,
) -> None:
    """El intento batch debe coincidir con lote e identidad fiscal del grupo."""
    punto, revision, operacion, guarda = await crear_contexto_rece_verificado_sintetico(
        db_session,
        empresa_id=test_empresa.id,
        empresa_cuit=test_empresa.cuit,
        usuario_id=test_user.id,
        idempotency_key="rece-sintetico-batch",
    )
    lote = LoteComprobante(
        empresa_id=test_empresa.id,
        usuario_id=test_user.id,
        nombre_archivo="rece-batch.xlsx",
        archivo_hash="5" * 64,
    )
    otro_lote = LoteComprobante(
        empresa_id=test_empresa.id,
        usuario_id=test_user.id,
        nombre_archivo="rece-batch-otro.xlsx",
        archivo_hash="6" * 64,
    )
    db_session.add_all([lote, otro_lote])
    await db_session.flush()
    grupo = LoteComprobanteGrupo(
        lote_id=lote.id,
        empresa_id=test_empresa.id,
        comprobante_ref="RECE-BATCH-1",
        tipo_comprobante=6,
        punto_venta_numero=punto.numero,
        total_estimado=Decimal("121.00"),
        punto_venta_id=punto.id,
        ambiente="produccion",
        punto_venta_elegibilidad_revision_id=revision.id,
        punto_venta_revision_fiscal=1,
    )
    db_session.add(grupo)
    await db_session.commit()
    operacion_id = operacion.id
    empresa_id = test_empresa.id
    usuario_id = test_user.id
    punto_venta_id = punto.id
    punto_venta_numero = punto.numero
    revision_id = revision.id
    guarda_id = guarda.id
    lote_id_correcto = lote.id
    lote_id_alternativo = otro_lote.id
    grupo_id = grupo.id

    def nuevo_intento(
        *,
        operacion_snapshot_id: int | None = operacion_id,
        lote_id: int | None = lote_id_correcto,
        grupo_snapshot_id: int | None = grupo_id,
        tipo_comprobante: int = 6,
        punto_numero: int = punto_venta_numero,
        ambiente: str | None = "produccion",
        revision_snapshot_id: int | None = revision_id,
        revision_fiscal: int | None = 1,
        guarda_snapshot_id: int | None = guarda_id,
        numero_planificado: int = 6,
        estado: str = "en_proceso",
    ) -> IntentoEmisionFiscal:
        """Construye un intento batch sintético ligado al contexto RECE."""
        return IntentoEmisionFiscal(
            operacion_id=operacion_snapshot_id,
            empresa_id=empresa_id,
            usuario_id=usuario_id,
            punto_venta_id=punto_venta_id,
            punto_venta_numero=punto_numero,
            tipo_comprobante=tipo_comprobante,
            numero_planificado=numero_planificado,
            fecha_emision=date(2026, 8, 8),
            total=Decimal("121.00"),
            payload_hash=f"{numero_planificado + 10:064d}",
            huella_logica=f"{numero_planificado + 20:064d}",
            estado=estado,
            ambiente=ambiente,
            punto_venta_elegibilidad_revision_id=revision_snapshot_id,
            punto_venta_revision_fiscal=revision_fiscal,
            guarda_rece_id=guarda_snapshot_id,
            lote_id=lote_id,
            grupo_id=grupo_snapshot_id,
        )

    async def rechazar_cruce(**cambios) -> None:
        """Afirma que SQLite aplica CHECK/FK mediante DML real."""
        db_session.add(nuevo_intento(**cambios))
        with pytest.raises(IntegrityError):
            await db_session.commit()
        await db_session.rollback()

    db_session.add(nuevo_intento())
    await db_session.commit()

    await rechazar_cruce(lote_id=lote_id_alternativo, numero_planificado=7)
    await rechazar_cruce(tipo_comprobante=1, numero_planificado=8)
    await rechazar_cruce(punto_numero=punto_venta_numero + 1, numero_planificado=9)
    await rechazar_cruce(ambiente="homologacion", numero_planificado=10)
    await rechazar_cruce(
        revision_snapshot_id=revision_id + 1000,
        numero_planificado=11,
    )
    await rechazar_cruce(revision_fiscal=2, numero_planificado=12)
    await rechazar_cruce(ambiente=None, numero_planificado=13)
    await rechazar_cruce(grupo_snapshot_id=None, numero_planificado=14)
    await rechazar_cruce(revision_fiscal=None, numero_planificado=17)
    await rechazar_cruce(operacion_snapshot_id=None, numero_planificado=18)
    await rechazar_cruce(revision_snapshot_id=None, numero_planificado=19)
    await rechazar_cruce(guarda_snapshot_id=None, numero_planificado=20)

    db_session.add(
        nuevo_intento(
            lote_id=None,
            grupo_snapshot_id=None,
            numero_planificado=15,
            estado="fallido_verificado",
        )
    )
    db_session.add(
        nuevo_intento(
            operacion_snapshot_id=None,
            ambiente=None,
            revision_snapshot_id=None,
            revision_fiscal=None,
            guarda_snapshot_id=None,
            numero_planificado=16,
            estado="fallido_verificado",
        )
    )
    await db_session.commit()

    intentos_validos = list(
        (
            await db_session.scalars(
                select(IntentoEmisionFiscal).where(
                    IntentoEmisionFiscal.numero_planificado.in_({6, 15, 16})
                )
            )
        ).all()
    )
    assert {intento.numero_planificado for intento in intentos_validos} == {6, 15, 16}
    legacy = next(
        intento for intento in intentos_validos if intento.numero_planificado == 16
    )
    assert legacy.operacion_id is None
    assert legacy.ambiente is None
    assert legacy.guarda_rece_id is None


@pytest.mark.asyncio
async def test_sqlite_check_rechaza_snapshot_parcial_en_grupo(
    db_session: AsyncSession,
    test_empresa,
    test_user,
) -> None:
    """El snapshot RECE del grupo debe persistirse completo o totalmente nulo."""
    (
        punto,
        revision,
        _operacion,
        _guarda,
    ) = await crear_contexto_rece_verificado_sintetico(
        db_session,
        empresa_id=test_empresa.id,
        empresa_cuit=test_empresa.cuit,
        usuario_id=test_user.id,
        idempotency_key="rece-grupo-snapshot-completo",
    )
    lote = LoteComprobante(
        empresa_id=test_empresa.id,
        usuario_id=test_user.id,
        nombre_archivo="rece-grupo-snapshot.xlsx",
        archivo_hash="7" * 64,
    )
    db_session.add(lote)
    await db_session.flush()

    grupo_valido = LoteComprobanteGrupo(
        lote_id=lote.id,
        empresa_id=test_empresa.id,
        comprobante_ref="RECE-GRUPO-VALIDO",
        tipo_comprobante=6,
        punto_venta_numero=punto.numero,
        total_estimado=Decimal("121.00"),
        punto_venta_id=punto.id,
        ambiente="produccion",
        punto_venta_elegibilidad_revision_id=revision.id,
        punto_venta_revision_fiscal=1,
    )
    db_session.add(grupo_valido)
    await db_session.commit()
    grupo_valido_id = int(grupo_valido.id)
    punto_id = int(punto.id)
    punto_numero = int(punto.numero)
    revision_id = int(revision.id)

    db_session.add(
        LoteComprobanteGrupo(
            lote_id=lote.id,
            empresa_id=test_empresa.id,
            comprobante_ref="RECE-GRUPO-PARCIAL",
            tipo_comprobante=6,
            punto_venta_numero=punto.numero,
            total_estimado=Decimal("121.00"),
            punto_venta_id=punto.id,
            ambiente=None,
            punto_venta_elegibilidad_revision_id=revision.id,
            punto_venta_revision_fiscal=1,
        )
    )
    with pytest.raises(IntegrityError):
        await db_session.commit()
    await db_session.rollback()

    for atributo in (
        "punto_venta_revision_fiscal",
        "punto_venta_numero",
        "tipo_comprobante",
    ):
        grupo_actual = await db_session.get(LoteComprobanteGrupo, grupo_valido_id)
        assert grupo_actual is not None
        setattr(grupo_actual, atributo, None)
        with pytest.raises(IntegrityError):
            await db_session.commit()
        await db_session.rollback()

    grupo_persistido = await db_session.get(LoteComprobanteGrupo, grupo_valido_id)
    assert grupo_persistido is not None
    assert grupo_persistido.ambiente == "produccion"
    assert grupo_persistido.punto_venta_id == punto_id
    assert grupo_persistido.punto_venta_numero == punto_numero
    assert grupo_persistido.tipo_comprobante == 6
    assert grupo_persistido.punto_venta_elegibilidad_revision_id == revision_id
    assert grupo_persistido.punto_venta_revision_fiscal == 1


@pytest.mark.asyncio
async def test_relacion_lote_completa_empresa_en_grupo(
    db_session: AsyncSession,
    test_empresa,
    test_user,
) -> None:
    """Agregar un grupo mediante la relación debe conservar ownership durable."""
    lote = LoteComprobante(
        empresa_id=test_empresa.id,
        usuario_id=test_user.id,
        nombre_archivo="rece-sintetico.xlsx",
        archivo_hash="5" * 64,
    )
    grupo = LoteComprobanteGrupo(
        comprobante_ref="RECE-1",
        total_estimado=Decimal("0"),
    )
    lote.grupos.append(grupo)
    db_session.add(lote)
    await db_session.commit()

    assert grupo.lote_id == lote.id
    assert grupo.empresa_id == test_empresa.id


@pytest.mark.asyncio
async def test_mutacion_con_guarda_activa_aborta_sin_cambios_parciales(
    db_session: AsyncSession,
    test_empresa,
    test_user,
) -> None:
    """Una guarda activa inmoviliza punto, head y ledger antes de toda escritura."""
    (
        punto,
        revision,
        _operacion,
        _guarda,
    ) = await crear_contexto_rece_verificado_sintetico(
        db_session,
        empresa_id=test_empresa.id,
        empresa_cuit=test_empresa.cuit,
        usuario_id=test_user.id,
        idempotency_key="rece-mutar-bloqueado",
    )
    await db_session.commit()
    revision_id = revision.id

    with pytest.raises(ElegibilidadReceError) as error:
        await ElegibilidadReceService(db_session).aplicar_cambios_punto(
            punto,
            {"sistema": "Sistema modificado"},
            fuente="edicion",
            actor_usuario_id=test_user.id,
        )

    assert error.value.categoria == "conflicto_guarda_rece_activa"
    await db_session.refresh(punto)
    assert punto.sistema is None
    assert punto.revision_fiscal == 1
    revisiones = await db_session.scalar(
        select(func.count())
        .select_from(PuntoVentaElegibilidadReceRevision)
        .where(PuntoVentaElegibilidadReceRevision.punto_venta_id == punto.id)
    )
    assert revisiones == 1
    head = (
        await db_session.execute(
            select(PuntoVentaElegibilidadReceActual).where(
                PuntoVentaElegibilidadReceActual.punto_venta_id == punto.id,
                PuntoVentaElegibilidadReceActual.ambiente == "produccion",
            )
        )
    ).scalar_one()
    assert head.revision_actual_id == revision_id
