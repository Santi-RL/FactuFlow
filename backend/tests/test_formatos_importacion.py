"""Pruebas de administración visual de plantillas de carga masiva."""

from copy import deepcopy
from datetime import date
from io import BytesIO

import pytest
from httpx import AsyncClient
from openpyxl import Workbook, load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.empresa import Empresa
from app.models.formato_importacion import FormatoImportacionVersion
from app.services.formatos_importacion_service import (
    FORMATO_BANCARIO_CONFIG,
    FormatoImportacionError,
    FormatosImportacionService,
)


def _config_plantilla_basica(tipo_comprobante: int = 1) -> dict:
    """Devuelve una configuración visual mínima de plantilla."""
    return {
        "tipo": "plantilla_visual",
        "header_row": 1,
        "modo_agrupacion": "fila",
        "plantilla": {
            "nombre_publico": "Plantilla simple",
            "columnas": [
                {
                    "campo_destino": "tipo_comprobante",
                    "etiqueta": "Tipo",
                    "origen": "constante",
                    "valor": tipo_comprobante,
                    "requerido": True,
                },
                {
                    "campo_destino": "punto_venta_numero",
                    "etiqueta": "Punto de venta",
                    "origen": "constante",
                    "valor": 1,
                    "requerido": True,
                },
                {
                    "campo_destino": "fecha_emision",
                    "etiqueta": "Fecha",
                    "origen": "header",
                    "transformacion": "fecha",
                    "requerido": True,
                    "ejemplo": "2026-05-31",
                },
                {
                    "campo_destino": "item_descripcion",
                    "etiqueta": "Descripción",
                    "origen": "header",
                    "transformacion": "texto",
                    "requerido": True,
                    "ejemplo": "Servicio mensual",
                },
                {
                    "campo_destino": "item_precio_unitario",
                    "etiqueta": "Importe",
                    "origen": "header",
                    "transformacion": "decimal",
                    "requerido": True,
                    "ejemplo": "10000.00",
                },
                {
                    "campo_destino": "item_cantidad",
                    "etiqueta": "Cantidad",
                    "origen": "constante",
                    "valor": 1,
                    "requerido": True,
                },
                {
                    "campo_destino": "item_iva_porcentaje",
                    "etiqueta": "IVA",
                    "origen": "constante",
                    "valor": 0 if tipo_comprobante in {11, 12, 13} else 21,
                    "requerido": True,
                },
            ],
        },
    }


def _xlsx_headers(headers: list[str]) -> bytes:
    """Genera un Excel simple con encabezados."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Comprobantes"
    sheet.append(headers)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _xlsx_con_filas(headers: list[str], rows: list[list[object]]) -> bytes:
    """Genera un Excel simple con encabezados y filas de datos."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Comprobantes"
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


@pytest.mark.asyncio
async def test_plantilla_global_visible_y_plantilla_emisor_aislada(
    client: AsyncClient,
    auth_headers: dict,
    admin_auth_headers: dict,
    db_session: AsyncSession,
    test_empresa: Empresa,
):
    """Una plantilla global debe verse desde otro emisor y una local no."""
    otra_empresa = Empresa(
        razon_social="Otra Empresa S.A.",
        cuit="20333444556",
        condicion_iva="RI",
        domicilio="Av. Otra 123",
        localidad="Buenos Aires",
        provincia="Buenos Aires",
        codigo_postal="1000",
        inicio_actividades=date(2021, 1, 1),
    )
    db_session.add(otra_empresa)
    await db_session.commit()
    await db_session.refresh(otra_empresa)

    global_response = await client.post(
        "/api/formatos-importacion",
        headers={**admin_auth_headers, "X-Empresa-Id": str(test_empresa.id)},
        json={
            "nombre": "Plantilla global simple",
            "descripcion": "Disponible para todos",
            "alcance": "global",
            "configuracion_json": _config_plantilla_basica(),
        },
    )
    assert global_response.status_code == 201, global_response.text

    local_response = await client.post(
        "/api/formatos-importacion",
        headers=auth_headers,
        json={
            "nombre": "Plantilla local simple",
            "descripcion": "Solo emisor original",
            "alcance": "emisor",
            "configuracion_json": _config_plantilla_basica(),
        },
    )
    assert local_response.status_code == 201, local_response.text

    usuario_otra_empresa = await client.get(
        "/api/formatos-importacion",
        headers={**auth_headers, "X-Empresa-Id": str(otra_empresa.id)},
    )
    assert usuario_otra_empresa.status_code == 403

    otra_lista = await client.get(
        "/api/formatos-importacion",
        headers={**admin_auth_headers, "X-Empresa-Id": str(otra_empresa.id)},
    )
    assert otra_lista.status_code == 200
    nombres = {item["nombre"] for item in otra_lista.json()}
    assert "Plantilla global simple" in nombres
    assert "Plantilla local simple" not in nombres


@pytest.mark.asyncio
async def test_usuario_comun_no_muta_plantillas_globales(
    client: AsyncClient,
    auth_headers: dict,
    admin_auth_headers: dict,
):
    """Las plantillas globales quedan reservadas a administradores."""
    crear_global = await client.post(
        "/api/formatos-importacion",
        headers=auth_headers,
        json={
            "nombre": "Global no autorizada",
            "descripcion": "No debe crearse",
            "alcance": "global",
            "configuracion_json": _config_plantilla_basica(),
        },
    )
    assert crear_global.status_code == 403

    crear_local = await client.post(
        "/api/formatos-importacion",
        headers=auth_headers,
        json={
            "nombre": "Local editable",
            "descripcion": "Sí debe crearse",
            "alcance": "emisor",
            "configuracion_json": _config_plantilla_basica(),
        },
    )
    assert crear_local.status_code == 201, crear_local.text

    promover_local = await client.put(
        f"/api/formatos-importacion/{crear_local.json()['id']}",
        headers=auth_headers,
        json={
            "nombre": "Local promovida",
            "descripcion": "No debe promoverse",
            "alcance": "global",
            "configuracion_json": _config_plantilla_basica(),
        },
    )
    assert promover_local.status_code == 403

    global_admin = await client.post(
        "/api/formatos-importacion",
        headers=admin_auth_headers,
        json={
            "nombre": "Global administrada",
            "descripcion": "Creada por admin",
            "alcance": "global",
            "configuracion_json": _config_plantilla_basica(),
        },
    )
    assert global_admin.status_code == 201, global_admin.text

    editar_global = await client.put(
        f"/api/formatos-importacion/{global_admin.json()['id']}",
        headers=auth_headers,
        json={
            "nombre": "Global editada por común",
            "descripcion": "No debe editarse",
            "alcance": "global",
            "configuracion_json": _config_plantilla_basica(),
        },
    )
    assert editar_global.status_code == 403

    clonar_a_global = await client.post(
        f"/api/formatos-importacion/{global_admin.json()['id']}/clonar",
        headers=auth_headers,
        json={"nombre": "Clon global no autorizado", "alcance": "global"},
    )
    assert clonar_a_global.status_code == 403

    clonar_a_emisor = await client.post(
        f"/api/formatos-importacion/{global_admin.json()['id']}/clonar",
        headers=auth_headers,
        json={"nombre": "Clon local permitido", "alcance": "emisor"},
    )
    assert clonar_a_emisor.status_code == 201, clonar_a_emisor.text
    assert clonar_a_emisor.json()["alcance"] == "emisor"

    desactivar_global = await client.delete(
        f"/api/formatos-importacion/{global_admin.json()['id']}",
        headers=auth_headers,
    )
    assert desactivar_global.status_code == 403


@pytest.mark.asyncio
async def test_no_cambia_alcance_de_plantilla_existente(
    client: AsyncClient,
    admin_auth_headers: dict,
    test_empresa: Empresa,
):
    """Una plantilla global no debe degradarse en la misma fila."""
    headers = {**admin_auth_headers, "X-Empresa-Id": str(test_empresa.id)}
    crear_global = await client.post(
        "/api/formatos-importacion",
        headers=headers,
        json={
            "nombre": "Global estable",
            "descripcion": "No debe cambiar de alcance",
            "alcance": "global",
            "configuracion_json": _config_plantilla_basica(),
        },
    )
    assert crear_global.status_code == 201, crear_global.text

    degradar_global = await client.put(
        f"/api/formatos-importacion/{crear_global.json()['id']}",
        headers=headers,
        json={
            "nombre": "Global degradada",
            "descripcion": "No debe permitirse",
            "alcance": "emisor",
            "configuracion_json": _config_plantilla_basica(),
        },
    )
    assert degradar_global.status_code == 400
    assert "No se puede cambiar el alcance" in degradar_global.json()["detail"]

    detalle = await client.get(
        f"/api/formatos-importacion/{crear_global.json()['id']}",
        headers=headers,
    )
    assert detalle.status_code == 200
    assert detalle.json()["alcance"] == "global"


@pytest.mark.asyncio
async def test_crear_formato_no_visual_exige_tipo_e_iva_explicitos(
    client: AsyncClient,
    auth_headers: dict,
):
    """Un formato legacy nuevo no debe depender de defaults fiscales ocultos."""
    response = await client.post(
        "/api/formatos-importacion",
        headers=auth_headers,
        json={
            "nombre": "Legacy incompleto",
            "descripcion": "No debe persistirse",
            "alcance": "emisor",
            "configuracion_json": {
                "tipo": "legacy_externo",
                "header_row": 1,
                "modo_agrupacion": "fila",
                "campos": {
                    "item_precio_unitario": {
                        "origen": "header",
                        "encabezados": ["Importe"],
                        "transformacion": "decimal",
                        "requerido": True,
                    }
                },
            },
        },
    )

    assert response.status_code == 400
    assert "Tipo de comprobante" in response.json()["detail"]


@pytest.mark.asyncio
async def test_actualizar_plantilla_crea_nueva_version(
    client: AsyncClient,
    auth_headers: dict,
):
    """Editar una plantilla debe reemplazar la versión vigente."""
    crear = await client.post(
        "/api/formatos-importacion",
        headers=auth_headers,
        json={
            "nombre": "Plantilla versionada",
            "descripcion": "v1",
            "alcance": "emisor",
            "configuracion_json": _config_plantilla_basica(),
        },
    )
    assert crear.status_code == 201, crear.text
    formato_id = crear.json()["id"]
    version_inicial = crear.json()["version_vigente"]["id"]

    config_v2 = _config_plantilla_basica()
    config_v2["plantilla"]["columnas"].append(
        {
            "campo_destino": "observaciones",
            "etiqueta": "Observaciones",
            "origen": "header",
            "transformacion": "texto",
            "requerido": False,
        }
    )
    actualizar = await client.put(
        f"/api/formatos-importacion/{formato_id}",
        headers=auth_headers,
        json={
            "nombre": "Plantilla versionada",
            "descripcion": "v2",
            "alcance": "emisor",
            "configuracion_json": config_v2,
        },
    )
    assert actualizar.status_code == 200, actualizar.text
    assert actualizar.json()["version_vigente"]["version"] == 2
    assert actualizar.json()["version_vigente"]["id"] != version_inicial

    limpiar_descripcion = await client.put(
        f"/api/formatos-importacion/{formato_id}",
        headers=auth_headers,
        json={
            "nombre": "Plantilla versionada",
            "descripcion": None,
            "alcance": "emisor",
        },
    )
    assert limpiar_descripcion.status_code == 200, limpiar_descripcion.text
    assert limpiar_descripcion.json()["descripcion"] is None


@pytest.mark.asyncio
async def test_actualizar_metadatos_no_reemplaza_version_si_mapeo_no_cambia(
    client: AsyncClient,
    auth_headers: dict,
):
    """Guardar el mismo mapeo no debe volver obsoletos perfiles existentes."""
    configuracion = _config_plantilla_basica()
    crear = await client.post(
        "/api/formatos-importacion",
        headers=auth_headers,
        json={
            "nombre": "Plantilla estable",
            "descripcion": "v1",
            "alcance": "emisor",
            "configuracion_json": configuracion,
        },
    )
    assert crear.status_code == 201, crear.text
    formato_id = crear.json()["id"]
    version_id = crear.json()["version_vigente"]["id"]

    actualizar = await client.put(
        f"/api/formatos-importacion/{formato_id}",
        headers=auth_headers,
        json={
            "nombre": "Plantilla estable renombrada",
            "descripcion": "Solo cambian metadatos",
            "alcance": "emisor",
            "configuracion_json": configuracion,
        },
    )
    assert actualizar.status_code == 200, actualizar.text
    assert actualizar.json()["nombre"] == "Plantilla estable renombrada"
    assert actualizar.json()["version_vigente"]["id"] == version_id
    assert actualizar.json()["version_vigente"]["version"] == 1


@pytest.mark.asyncio
async def test_version_reemplazada_no_se_usa_en_importaciones_nuevas(
    client: AsyncClient,
    auth_headers: dict,
    db_session: AsyncSession,
    test_empresa: Empresa,
):
    """Una versión reemplazada queda como historial, no como plantilla usable."""
    empresa_id = test_empresa.id
    crear = await client.post(
        "/api/formatos-importacion",
        headers=auth_headers,
        json={
            "nombre": "Plantilla con versión vieja",
            "descripcion": None,
            "alcance": "emisor",
            "configuracion_json": _config_plantilla_basica(),
        },
    )
    assert crear.status_code == 201, crear.text
    formato_id = crear.json()["id"]
    version_reemplazada_id = crear.json()["version_vigente"]["id"]

    config_v2 = _config_plantilla_basica()
    config_v2["plantilla"]["columnas"][2]["etiqueta"] = "Fecha nueva"
    actualizar = await client.put(
        f"/api/formatos-importacion/{formato_id}",
        headers=auth_headers,
        json={
            "nombre": "Plantilla con versión vieja",
            "descripcion": None,
            "alcance": "emisor",
            "configuracion_json": config_v2,
        },
    )
    assert actualizar.status_code == 200, actualizar.text

    db_session.expire_all()
    service = FormatosImportacionService(db_session)
    with pytest.raises(FormatoImportacionError, match="ya no está vigente"):
        await service.obtener_version(version_reemplazada_id, empresa_id)

    version_historial = await service.obtener_version(
        version_reemplazada_id,
        empresa_id,
        exigir_vigente=False,
    )
    assert version_historial.estado == "reemplazada"


@pytest.mark.asyncio
async def test_actualizar_metadatos_formato_base_no_reemplaza_version_existente(
    client: AsyncClient,
    auth_headers: dict,
    db_session: AsyncSession,
    test_empresa: Empresa,
):
    """Agregar metadatos visuales al formato base no debe romper perfiles."""
    empresa_id = test_empresa.id
    listar = await client.get("/api/formatos-importacion", headers=auth_headers)
    assert listar.status_code == 200, listar.text
    base = next(
        item
        for item in listar.json()
        if item["nombre"] == FormatosImportacionService.FORMATO_BANCARIO_NOMBRE
    )
    version_id = base["version_vigente"]["id"]
    version = await db_session.get(FormatoImportacionVersion, version_id)
    assert version is not None

    config_legacy = deepcopy(FORMATO_BANCARIO_CONFIG)
    config_legacy.pop("plantilla_sistema_protegida", None)
    config_legacy.pop("plantilla", None)
    version.configuracion_json = config_legacy
    await db_session.commit()

    service = FormatosImportacionService(db_session)
    await service.asegurar_formatos_base()
    db_session.expire_all()

    version_actualizada = await db_session.get(FormatoImportacionVersion, version_id)
    assert version_actualizada is not None
    assert version_actualizada.estado == "vigente"
    assert version_actualizada.configuracion_json["plantilla_sistema_protegida"] is True
    assert version_actualizada.configuracion_json["plantilla"]

    usable = await service.obtener_version(version_id, empresa_id)
    assert usable.id == version_id


@pytest.mark.asyncio
async def test_plantilla_protegida_sistema_no_editable_y_clonable(
    client: AsyncClient,
    auth_headers: dict,
    admin_auth_headers: dict,
    db_session: AsyncSession,
    test_empresa: Empresa,
):
    """Las plantillas internas se protegen contra edición directa."""
    listar = await client.get("/api/formatos-importacion", headers=auth_headers)
    assert listar.status_code == 200, listar.text
    protegida = next(
        item
        for item in listar.json()
        if item["version_vigente"]["configuracion_json"].get(
            "plantilla_sistema_protegida"
        )
    )

    editar = await client.put(
        f"/api/formatos-importacion/{protegida['id']}",
        headers=admin_auth_headers,
        json={
            "nombre": protegida["nombre"],
            "descripcion": protegida["descripcion"],
            "alcance": protegida["alcance"],
            "configuracion_json": protegida["version_vigente"]["configuracion_json"],
        },
    )
    assert editar.status_code == 400
    assert "clonarla" in editar.json()["detail"]

    clonar = await client.post(
        f"/api/formatos-importacion/{protegida['id']}/clonar",
        headers=auth_headers,
        json={"nombre": "Copia editable", "alcance": "emisor"},
    )
    assert clonar.status_code == 201, clonar.text
    assert (
        clonar.json()["version_vigente"]["configuracion_json"].get(
            "plantilla_sistema_protegida"
        )
        is None
    )
    aliases_importe = clonar.json()["version_vigente"]["configuracion_json"]["campos"][
        "importe_total"
    ]["encabezados"]
    assert "Creditos" in aliases_importe
    assert "Créditos" in aliases_importe

    test_empresa.condicion_iva = "Exento"
    await db_session.commit()
    compatibilidad = await client.post(
        "/api/formatos-importacion/compatibilidad",
        headers=auth_headers,
        json={
            "configuracion_json": clonar.json()["version_vigente"]["configuracion_json"]
        },
    )
    assert compatibilidad.status_code == 200, compatibilidad.text
    data_compatibilidad = compatibilidad.json()
    assert data_compatibilidad["estado"] != "incompatible"
    assert all(
        mensaje["campo"] != "item_precio_unitario"
        for mensaje in data_compatibilidad["faltantes"]
    )


@pytest.mark.asyncio
async def test_analizar_y_descargar_excel_de_plantilla(
    client: AsyncClient,
    auth_headers: dict,
):
    """La API debe leer encabezados y descargar un XLSX generado."""
    analizar = await client.post(
        "/api/formatos-importacion/analizar-excel",
        headers=auth_headers,
        files={
            "archivo": (
                "ejemplo.xlsx",
                _xlsx_headers(["Fecha", "Descripción", "Importe"]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert analizar.status_code == 200, analizar.text
    assert [item["encabezado"] for item in analizar.json()["columnas"]] == [
        "Fecha",
        "Descripción",
        "Importe",
    ]

    crear = await client.post(
        "/api/formatos-importacion",
        headers=auth_headers,
        json={
            "nombre": "Plantilla descargable",
            "descripcion": None,
            "alcance": "emisor",
            "configuracion_json": _config_plantilla_basica(),
        },
    )
    assert crear.status_code == 201, crear.text

    descargar = await client.get(
        f"/api/formatos-importacion/{crear.json()['id']}/descargar",
        headers=auth_headers,
    )
    assert descargar.status_code == 200, descargar.text
    workbook = load_workbook(BytesIO(descargar.content), data_only=True)
    assert "Comprobantes" in workbook.sheetnames
    assert "Instrucciones" in workbook.sheetnames
    assert workbook["_factuflow"].sheet_state == "hidden"
    headers = [
        workbook["Comprobantes"].cell(row=1, column=index).value
        for index in range(1, 4)
    ]
    assert headers == ["Fecha", "Descripción", "Importe"]
    assert [
        workbook["Comprobantes"].cell(row=2, column=index).value
        for index in range(1, 4)
    ] == [None, None, None]
    assert workbook["Instrucciones"].cell(row=7, column=4).value == "Ejemplo"


@pytest.mark.asyncio
async def test_descarga_excel_neutraliza_textos_con_prefijo_formula(
    client: AsyncClient,
    auth_headers: dict,
):
    """Los textos configurables del XLSX no deben quedar como fórmulas."""
    config = _config_plantilla_basica()
    config["plantilla"]["columnas"][4]["etiqueta"] = '=HYPERLINK("http://x")'
    config["plantilla"]["columnas"][4]["ejemplo"] = "+SUM(1,1)"

    crear = await client.post(
        "/api/formatos-importacion",
        headers=auth_headers,
        json={
            "nombre": "=Plantilla peligrosa",
            "descripcion": None,
            "alcance": "emisor",
            "configuracion_json": config,
        },
    )
    assert crear.status_code == 201, crear.text

    descargar = await client.get(
        f"/api/formatos-importacion/{crear.json()['id']}/descargar",
        headers=auth_headers,
    )
    assert descargar.status_code == 200, descargar.text
    workbook = load_workbook(BytesIO(descargar.content), data_only=False)

    assert workbook["Comprobantes"].cell(row=1, column=3).value.startswith("'=")
    assert workbook["Instrucciones"].cell(row=10, column=1).value.startswith("'=")
    assert workbook["Instrucciones"].cell(row=10, column=4).value.startswith("'+")
    assert workbook["_factuflow"]["B5"].value.startswith("'=")


@pytest.mark.asyncio
async def test_descarga_excel_respeta_columnas_fijas(
    client: AsyncClient,
    auth_headers: dict,
):
    """El XLSX generado debe conservar letras de columna configuradas."""
    config = _config_plantilla_basica()
    for columna in config["plantilla"]["columnas"]:
        if columna["campo_destino"] == "item_precio_unitario":
            columna["origen"] = "columna"
            columna["letra_columna"] = "D"

    crear = await client.post(
        "/api/formatos-importacion",
        headers=auth_headers,
        json={
            "nombre": "Plantilla con columna fija",
            "descripcion": None,
            "alcance": "emisor",
            "configuracion_json": config,
        },
    )
    assert crear.status_code == 201, crear.text

    descargar = await client.get(
        f"/api/formatos-importacion/{crear.json()['id']}/descargar",
        headers=auth_headers,
    )
    assert descargar.status_code == 200, descargar.text
    workbook = load_workbook(BytesIO(descargar.content), data_only=False)
    sheet = workbook["Comprobantes"]

    assert sheet["A1"].value == "Fecha"
    assert sheet["B1"].value == "Descripción"
    assert sheet["C1"].value is None
    assert sheet["D1"].value == "Importe"


@pytest.mark.asyncio
async def test_descarga_excel_normaliza_orden_visual_invalido(
    client: AsyncClient,
    auth_headers: dict,
):
    """Un orden visual inválido no debe persistirse como futuro 500."""
    config = _config_plantilla_basica()
    for columna in config["plantilla"]["columnas"]:
        if columna.get("origen", "header") == "header":
            columna["orden"] = "abc"

    crear = await client.post(
        "/api/formatos-importacion",
        headers=auth_headers,
        json={
            "nombre": "Plantilla con orden inválido",
            "descripcion": None,
            "alcance": "emisor",
            "configuracion_json": config,
        },
    )
    assert crear.status_code == 201, crear.text
    columnas_guardadas = crear.json()["version_vigente"]["configuracion_json"][
        "plantilla"
    ]["columnas"]
    assert all(isinstance(columna["orden"], int) for columna in columnas_guardadas)

    descargar = await client.get(
        f"/api/formatos-importacion/{crear.json()['id']}/descargar",
        headers=auth_headers,
    )
    assert descargar.status_code == 200, descargar.text
    workbook = load_workbook(BytesIO(descargar.content), data_only=False)
    sheet = workbook["Comprobantes"]

    assert sheet["A1"].value == "Fecha"
    assert sheet["B1"].value == "Descripción"
    assert sheet["C1"].value == "Importe"


@pytest.mark.asyncio
async def test_detectar_no_sugiere_metadata_si_hay_ambiguedad_fiscal(
    client: AsyncClient,
    auth_headers: dict,
):
    """La metadata oculta no debe decidir entre constantes fiscales distintas."""
    config_factura_c = _config_plantilla_basica(tipo_comprobante=11)
    for columna in config_factura_c["plantilla"]["columnas"]:
        if columna["campo_destino"] == "item_iva_porcentaje":
            columna["valor"] = 0
    crear_factura_c = await client.post(
        "/api/formatos-importacion",
        headers=auth_headers,
        json={
            "nombre": "A - Misma estructura Factura C",
            "descripcion": None,
            "alcance": "emisor",
            "configuracion_json": config_factura_c,
        },
    )
    assert crear_factura_c.status_code == 201, crear_factura_c.text

    config_factura_b = _config_plantilla_basica(tipo_comprobante=6)
    crear_factura_b = await client.post(
        "/api/formatos-importacion",
        headers=auth_headers,
        json={
            "nombre": "Z - Misma estructura Factura B",
            "descripcion": None,
            "alcance": "emisor",
            "configuracion_json": config_factura_b,
        },
    )
    assert crear_factura_b.status_code == 201, crear_factura_b.text

    descargar_b = await client.get(
        f"/api/formatos-importacion/{crear_factura_b.json()['id']}/descargar",
        headers=auth_headers,
    )
    assert descargar_b.status_code == 200, descargar_b.text

    detectar = await client.post(
        "/api/formatos-importacion/detectar",
        headers=auth_headers,
        files={
            "archivo": (
                "plantilla-factura-b.xlsx",
                descargar_b.content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert detectar.status_code == 200, detectar.text
    assert detectar.json()["formato_sugerido_version_id"] is None
    candidatos = {
        item["formato_version_id"]
        for item in detectar.json()["candidatos"]
        if item["confianza"] == "alta"
    }
    assert crear_factura_c.json()["version_vigente"]["id"] in candidatos
    assert crear_factura_b.json()["version_vigente"]["id"] in candidatos


@pytest.mark.asyncio
async def test_importacion_conserva_cuit_emisor_mapeado(
    client: AsyncClient,
    auth_headers: dict,
    db_session: AsyncSession,
    test_empresa: Empresa,
):
    """El CUIT del archivo debe llegar al validador sin ser sobrescrito."""
    config = _config_plantilla_basica(tipo_comprobante=6)
    config["plantilla"]["columnas"].insert(
        0,
        {
            "campo_destino": "empresa_cuit",
            "etiqueta": "CUIT Emisor",
            "origen": "header",
            "transformacion": "documento",
            "requerido": True,
            "ejemplo": test_empresa.cuit,
        },
    )
    crear = await client.post(
        "/api/formatos-importacion",
        headers=auth_headers,
        json={
            "nombre": "Plantilla con CUIT emisor",
            "descripcion": None,
            "alcance": "emisor",
            "configuracion_json": config,
        },
    )
    assert crear.status_code == 201, crear.text

    version_id = crear.json()["version_vigente"]["id"]
    service = FormatosImportacionService(db_session)
    version = await service.obtener_version(version_id, test_empresa.id)
    cuit_de_otro_emisor = "20333333330"
    contenido = _xlsx_con_filas(
        ["CUIT Emisor", "Fecha", "Descripción", "Importe"],
        [[cuit_de_otro_emisor, "2026-05-31", "Servicio mensual", 100]],
    )

    importacion = await service.importar_con_version(contenido, test_empresa, version)

    assert importacion.filas[0]["empresa_cuit"] == cuit_de_otro_emisor
    assert importacion.filas[0]["empresa_cuit"] != test_empresa.cuit


@pytest.mark.asyncio
async def test_compatibilidad_exige_iva_explicito(
    client: AsyncClient,
    auth_headers: dict,
):
    """La plantilla no debe depender del fallback oculto de IVA en cero."""
    config_sin_iva = _config_plantilla_basica()
    config_sin_iva["plantilla"]["columnas"] = [
        columna
        for columna in config_sin_iva["plantilla"]["columnas"]
        if columna["campo_destino"] != "item_iva_porcentaje"
    ]

    response = await client.post(
        "/api/formatos-importacion/compatibilidad",
        headers=auth_headers,
        json={"configuracion_json": config_sin_iva},
    )
    assert response.status_code == 200, response.text
    data = response.json()
    assert data["estado"] == "incompatible"
    assert any(
        mensaje["codigo"] == "falta_item_iva_porcentaje"
        for mensaje in data["faltantes"]
    )

    crear = await client.post(
        "/api/formatos-importacion",
        headers=auth_headers,
        json={
            "nombre": "Plantilla sin IVA explícito",
            "descripcion": "No debe persistirse",
            "alcance": "emisor",
            "configuracion_json": config_sin_iva,
        },
    )
    assert crear.status_code == 400
    assert "IVA" in crear.json()["detail"]


@pytest.mark.asyncio
async def test_constantes_requeridas_vacias_no_resuelven_compatibilidad(
    client: AsyncClient,
    auth_headers: dict,
):
    """Una constante vacía no debe contar como campo fiscal resuelto."""
    config = _config_plantilla_basica()
    for columna in config["plantilla"]["columnas"]:
        if columna["campo_destino"] in {"tipo_comprobante", "item_iva_porcentaje"}:
            columna["valor"] = ""

    compatibilidad = await client.post(
        "/api/formatos-importacion/compatibilidad",
        headers=auth_headers,
        json={"configuracion_json": config},
    )
    assert compatibilidad.status_code == 200, compatibilidad.text
    data = compatibilidad.json()
    assert data["estado"] == "incompatible"
    codigos_faltantes = {mensaje["codigo"] for mensaje in data["faltantes"]}
    assert "falta_tipo_comprobante" in codigos_faltantes
    assert "falta_item_iva_porcentaje" in codigos_faltantes

    crear = await client.post(
        "/api/formatos-importacion",
        headers=auth_headers,
        json={
            "nombre": "Plantilla con constantes vacías",
            "descripcion": "No debe persistirse",
            "alcance": "emisor",
            "configuracion_json": config,
        },
    )
    assert crear.status_code == 400
    assert "debe informar un valor constante" in crear.json()["detail"]


@pytest.mark.asyncio
async def test_constantes_fiscales_invalidas_no_son_compatibles(
    client: AsyncClient,
    auth_headers: dict,
):
    """Los valores fijos fiscales deben pertenecer a dominios soportados."""
    config_tipo_invalido = _config_plantilla_basica()
    for columna in config_tipo_invalido["plantilla"]["columnas"]:
        if columna["campo_destino"] == "tipo_comprobante":
            columna["valor"] = 99

    compatibilidad = await client.post(
        "/api/formatos-importacion/compatibilidad",
        headers=auth_headers,
        json={"configuracion_json": config_tipo_invalido},
    )
    assert compatibilidad.status_code == 400
    assert "tipo de comprobante fijo" in compatibilidad.json()["detail"]

    config_tipo_fraccionario = _config_plantilla_basica()
    for columna in config_tipo_fraccionario["plantilla"]["columnas"]:
        if columna["campo_destino"] == "tipo_comprobante":
            columna["valor"] = "1.9"

    compatibilidad_fraccionaria = await client.post(
        "/api/formatos-importacion/compatibilidad",
        headers=auth_headers,
        json={"configuracion_json": config_tipo_fraccionario},
    )
    assert compatibilidad_fraccionaria.status_code == 400
    assert "tipo de comprobante fijo" in compatibilidad_fraccionaria.json()["detail"]

    config_concepto_fraccionario = _config_plantilla_basica()
    config_concepto_fraccionario["plantilla"]["columnas"].append(
        {
            "campo_destino": "concepto",
            "etiqueta": "Concepto",
            "origen": "constante",
            "valor": "2.5",
            "requerido": True,
        }
    )

    crear_concepto = await client.post(
        "/api/formatos-importacion",
        headers=auth_headers,
        json={
            "nombre": "Plantilla con concepto inválido",
            "descripcion": "No debe persistirse",
            "alcance": "emisor",
            "configuracion_json": config_concepto_fraccionario,
        },
    )
    assert crear_concepto.status_code == 400
    assert "concepto fijo" in crear_concepto.json()["detail"]

    config_iva_invalido = _config_plantilla_basica()
    for columna in config_iva_invalido["plantilla"]["columnas"]:
        if columna["campo_destino"] == "item_iva_porcentaje":
            columna["valor"] = 5

    crear = await client.post(
        "/api/formatos-importacion",
        headers=auth_headers,
        json={
            "nombre": "Plantilla con IVA inválido",
            "descripcion": "No debe persistirse",
            "alcance": "emisor",
            "configuracion_json": config_iva_invalido,
        },
    )
    assert crear.status_code == 400
    assert "alícuota de IVA fija" in crear.json()["detail"]

    config_factura_c_iva = _config_plantilla_basica(tipo_comprobante=11)
    for columna in config_factura_c_iva["plantilla"]["columnas"]:
        if columna["campo_destino"] == "item_iva_porcentaje":
            columna["valor"] = 21

    crear_factura_c = await client.post(
        "/api/formatos-importacion",
        headers=auth_headers,
        json={
            "nombre": "Factura C con IVA",
            "descripcion": "No debe persistirse",
            "alcance": "emisor",
            "configuracion_json": config_factura_c_iva,
        },
    )
    assert crear_factura_c.status_code == 400
    assert "tipo C deben usar IVA 0" in crear_factura_c.json()["detail"]


@pytest.mark.asyncio
async def test_compatibilidad_detecta_conflictos_de_perfil_y_condicion_iva(
    client: AsyncClient,
    auth_headers: dict,
    test_empresa: Empresa,
    db_session: AsyncSession,
):
    """La compatibilidad debe cruzar plantilla, perfil y condición fiscal."""
    config_sin_descripcion = _config_plantilla_basica(tipo_comprobante=11)
    config_sin_descripcion["plantilla"]["columnas"] = [
        columna
        for columna in config_sin_descripcion["plantilla"]["columnas"]
        if columna["campo_destino"] != "item_descripcion"
    ]

    response = await client.post(
        "/api/formatos-importacion/compatibilidad",
        headers=auth_headers,
        json={
            "configuracion_json": config_sin_descripcion,
            "perfil_configuracion_json": {
                "descripcion_item_modo": "archivo",
                "punto_venta": {"modo": "fijo", "numero": 1},
                "fecha_emision": {"modo": "archivo"},
            },
        },
    )
    assert response.status_code == 200, response.text
    data = response.json()
    assert data["estado"] == "incompatible"
    assert any(
        mensaje["codigo"] == "perfil_descripcion_archivo_sin_columna"
        for mensaje in data["conflictos"]
    )
    assert any(mensaje["codigo"] == "ri_no_emite_c" for mensaje in data["conflictos"])

    test_empresa.condicion_iva = "Monotributo"
    await db_session.commit()
    config_factura_a = _config_plantilla_basica(tipo_comprobante=1)
    response_monotributo = await client.post(
        "/api/formatos-importacion/compatibilidad",
        headers=auth_headers,
        json={"configuracion_json": config_factura_a},
    )
    assert response_monotributo.status_code == 200, response_monotributo.text
    assert any(
        mensaje["codigo"] == "monotributo_exento_no_emite_a"
        for mensaje in response_monotributo.json()["conflictos"]
    )


@pytest.mark.asyncio
async def test_compatibilidad_sin_perfil_no_exige_punto_desde_archivo(
    client: AsyncClient,
    auth_headers: dict,
):
    """Sin perfil explícito, el punto de venta puede completarse manualmente."""
    config_sin_punto = _config_plantilla_basica()
    config_sin_punto["plantilla"]["columnas"] = [
        columna
        for columna in config_sin_punto["plantilla"]["columnas"]
        if columna["campo_destino"] != "punto_venta_numero"
    ]

    response = await client.post(
        "/api/formatos-importacion/compatibilidad",
        headers=auth_headers,
        json={"configuracion_json": config_sin_punto},
    )
    assert response.status_code == 200, response.text
    data = response.json()
    assert data["estado"] == "advertencia"
    assert not any(
        mensaje["codigo"] == "perfil_punto_venta_archivo_sin_columna"
        for mensaje in data["conflictos"]
    )
    assert any(
        mensaje["codigo"] == "punto_venta_manual" for mensaje in data["advertencias"]
    )


@pytest.mark.asyncio
async def test_origen_empresa_solo_se_permite_en_campos_resueltos(
    client: AsyncClient,
    auth_headers: dict,
):
    """El origen emisor no debe prometer campos que el importador no resuelve."""
    config = _config_plantilla_basica()
    for columna in config["plantilla"]["columnas"]:
        if columna["campo_destino"] == "punto_venta_numero":
            columna["origen"] = "empresa"
            columna.pop("valor", None)

    compatibilidad = await client.post(
        "/api/formatos-importacion/compatibilidad",
        headers=auth_headers,
        json={"configuracion_json": config},
    )
    assert compatibilidad.status_code == 400
    assert "no puede usar datos del emisor" in compatibilidad.json()["detail"]

    crear = await client.post(
        "/api/formatos-importacion",
        headers=auth_headers,
        json={
            "nombre": "Plantilla inválida por origen emisor",
            "descripcion": "No debe persistirse",
            "alcance": "emisor",
            "configuracion_json": config,
        },
    )
    assert crear.status_code == 400
    assert "no puede usar datos del emisor" in crear.json()["detail"]


@pytest.mark.asyncio
async def test_plantilla_no_permite_campos_destino_duplicados(
    client: AsyncClient,
    auth_headers: dict,
):
    """Dos columnas visuales no deben mapear el mismo campo fiscal."""
    config = _config_plantilla_basica()
    config["plantilla"]["columnas"].append(
        {
            "campo_destino": "item_precio_unitario",
            "etiqueta": "Importe duplicado",
            "origen": "header",
            "transformacion": "decimal",
            "requerido": False,
        }
    )

    compatibilidad = await client.post(
        "/api/formatos-importacion/compatibilidad",
        headers=auth_headers,
        json={"configuracion_json": config},
    )
    assert compatibilidad.status_code == 400
    assert "repite el campo item_precio_unitario" in compatibilidad.json()["detail"]

    crear = await client.post(
        "/api/formatos-importacion",
        headers=auth_headers,
        json={
            "nombre": "Plantilla con campo duplicado",
            "descripcion": "No debe persistirse",
            "alcance": "emisor",
            "configuracion_json": config,
        },
    )
    assert crear.status_code == 400
    assert "repite el campo item_precio_unitario" in crear.json()["detail"]


@pytest.mark.asyncio
async def test_fecha_emision_no_admite_origen_constante(
    client: AsyncClient,
    auth_headers: dict,
):
    """La fecha fiscal debe venir del archivo o quedar definida fuera del formato."""
    config = _config_plantilla_basica()
    for columna in config["plantilla"]["columnas"]:
        if columna["campo_destino"] == "fecha_emision":
            columna["origen"] = "constante"
            columna["valor"] = "2026-05-31"
            columna.pop("transformacion", None)

    compatibilidad = await client.post(
        "/api/formatos-importacion/compatibilidad",
        headers=auth_headers,
        json={"configuracion_json": config},
    )
    assert compatibilidad.status_code == 400
    assert "fecha_emision no admite origen constante" in compatibilidad.json()["detail"]

    crear = await client.post(
        "/api/formatos-importacion",
        headers=auth_headers,
        json={
            "nombre": "Plantilla con fecha fija oculta",
            "descripcion": "No debe persistirse",
            "alcance": "emisor",
            "configuracion_json": config,
        },
    )
    assert crear.status_code == 400
    assert "fecha_emision no admite origen constante" in crear.json()["detail"]


@pytest.mark.asyncio
async def test_campo_no_admite_transformacion_incompatible(
    client: AsyncClient,
    auth_headers: dict,
):
    """Cada campo debe usar solo transformaciones admitidas por el catálogo."""
    config = _config_plantilla_basica()
    for columna in config["plantilla"]["columnas"]:
        if columna["campo_destino"] == "item_iva_porcentaje":
            columna["transformacion"] = "fecha"

    compatibilidad = await client.post(
        "/api/formatos-importacion/compatibilidad",
        headers=auth_headers,
        json={"configuracion_json": config},
    )
    assert compatibilidad.status_code == 400
    assert (
        "item_iva_porcentaje no admite transformación fecha"
        in compatibilidad.json()["detail"]
    )

    crear = await client.post(
        "/api/formatos-importacion",
        headers=auth_headers,
        json={
            "nombre": "Plantilla con transformación inválida",
            "descripcion": "No debe persistirse",
            "alcance": "emisor",
            "configuracion_json": config,
        },
    )
    assert crear.status_code == 400
    assert (
        "item_iva_porcentaje no admite transformación fecha" in crear.json()["detail"]
    )


@pytest.mark.asyncio
async def test_origen_columna_exige_posicion_explicita(
    client: AsyncClient,
    auth_headers: dict,
):
    """Una columna fija sin letra o índice no debe generar un XLSX inválido."""
    config = _config_plantilla_basica()
    for columna in config["plantilla"]["columnas"]:
        if columna["campo_destino"] == "item_precio_unitario":
            columna["origen"] = "columna"
            columna.pop("letra_columna", None)
            columna.pop("indice_columna", None)

    compatibilidad = await client.post(
        "/api/formatos-importacion/compatibilidad",
        headers=auth_headers,
        json={"configuracion_json": config},
    )
    assert compatibilidad.status_code == 400
    assert "debe declarar letra o índice de columna" in compatibilidad.json()["detail"]

    crear = await client.post(
        "/api/formatos-importacion",
        headers=auth_headers,
        json={
            "nombre": "Plantilla con columna fija incompleta",
            "descripcion": "No debe persistirse",
            "alcance": "emisor",
            "configuracion_json": config,
        },
    )
    assert crear.status_code == 400
    assert "debe declarar letra o índice de columna" in crear.json()["detail"]


@pytest.mark.asyncio
async def test_origen_columna_rechaza_indice_fuera_de_rango_excel(
    client: AsyncClient,
    auth_headers: dict,
):
    """Una columna fija no debe superar el rango real de columnas de Excel."""
    config = _config_plantilla_basica()
    for columna in config["plantilla"]["columnas"]:
        if columna["campo_destino"] == "item_precio_unitario":
            columna["origen"] = "columna"
            columna["indice_columna"] = 20000

    compatibilidad = await client.post(
        "/api/formatos-importacion/compatibilidad",
        headers=auth_headers,
        json={"configuracion_json": config},
    )
    assert compatibilidad.status_code == 400
    assert "configuración de columna" in compatibilidad.json()["detail"]

    crear = await client.post(
        "/api/formatos-importacion",
        headers=auth_headers,
        json={
            "nombre": "Plantilla con columna fuera de rango",
            "descripcion": "No debe persistirse",
            "alcance": "emisor",
            "configuracion_json": config,
        },
    )
    assert crear.status_code == 400
    assert "configuración de columna" in crear.json()["detail"]
