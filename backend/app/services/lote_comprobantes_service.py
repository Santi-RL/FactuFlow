"""Servicios para validación y procesamiento de comprobantes masivos."""

from __future__ import annotations

import hashlib
import json
import logging
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from io import BytesIO
from typing import Any
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.utils.datetime import from_excel
from openpyxl.styles import Font, PatternFill
from sqlalchemy import delete, exists, func, select, update
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.arca.config import ArcaAmbiente
from app.arca.exceptions import ArcaServiceError, ArcaValidationError
from app.arca.utils import clean_cuit, validate_cuit
from app.core.config import settings
from app.core.database import DATABASE_TEMPORARILY_UNAVAILABLE_ERRORS
from app.models.certificado import Certificado
from app.models.comprobante import Comprobante
from app.models.empresa import Empresa
from app.models.lote_comprobante import (
    LoteComprobante,
    LoteComprobanteEvento,
    LoteComprobanteFila,
    LoteComprobanteGrupo,
)
from app.models.idempotencia_fiscal import IntentoEmisionFiscal, OperacionIdempotente
from app.models.punto_venta import PuntoVenta
from app.models.usuario import Usuario
from app.schemas.comprobante import (
    ComprobanteAsociadoCreate,
    EmitirComprobanteRequest,
    EmitirComprobanteResponse,
    ItemComprobanteCreate,
)
from app.schemas.lote_comprobante import (
    LoteComprobanteResponse,
    LoteComprobanteSeguimientoResponse,
    LoteProcesamientoResponse,
)
from app.services.facturacion_service import (
    FacturacionService,
    FaseSolicitudArca,
    ValidationError,
)
from app.services.formatos_importacion_service import (
    FormatoImportacionError,
    FormatosImportacionService,
    ImportacionNormalizada,
)
from app.services.idempotencia_fiscal_service import IdempotenciaFiscalService

logger = logging.getLogger(__name__)


class LoteComprobanteError(Exception):
    """Error funcional durante la validación o emisión del lote."""


@dataclass(frozen=True)
class OpcionesFechasLote:
    """Define cómo tomar fechas fiscales al validar un lote."""

    fecha_emision_modo: str
    fecha_servicio_desde_modo: str
    fecha_servicio_hasta_modo: str
    fecha_vto_pago_modo: str
    fecha_emision_fija: date | None = None
    fecha_servicio_desde_fija: date | None = None
    fecha_servicio_hasta_fija: date | None = None
    fecha_vto_pago_fija: date | None = None


@dataclass(frozen=True)
class OpcionesConceptoLote:
    """Define cómo tomar el concepto fiscal al validar un lote."""

    concepto_modo: str


@dataclass(frozen=True)
class OpcionesDescripcionItemLote:
    """Define cómo tomar la descripción facturada de los ítems del lote."""

    descripcion_item_modo: str
    descripcion_item_fija: str | None = None


@dataclass(frozen=True)
class OpcionesPuntoVentaLote:
    """Define cómo tomar el punto de venta al validar un lote."""

    punto_venta_modo: str = "archivo"
    punto_venta_numero: int | None = None


@dataclass(frozen=True)
class GrupoPendienteEmision:
    """Agrupa el registro del lote con el request listo para emitir."""

    grupo: LoteComprobanteGrupo
    request: EmitirComprobanteRequest


@dataclass(frozen=True)
class ConfiguracionBatchArca:
    """Configuración efectiva de emisión por sublotes ARCA."""

    habilitado: bool
    reg_x_req: int | None
    chunk_size: int
    fallback_unitario: bool = False
    fallback_motivo: str | None = None


@dataclass(frozen=True)
class ClasificacionGruposStale:
    """Clasifica grupos válidos de un lote interrumpido según evidencia fiscal."""

    intactos: tuple[LoteComprobanteGrupo, ...]
    con_evidencia_fiscal: tuple[LoteComprobanteGrupo, ...]
    total_validos: int


@dataclass(frozen=True)
class ResultadoArcaExterno:
    """Datos mínimos de ARCA para guardar un comprobante reconciliado."""

    cae: str
    cae_vencimiento: str


class LoteComprobantesService:
    """Servicio de carga y emisión masiva de comprobantes."""

    ESTADOS_TERMINALES = {
        "completado",
        "cerrado_reconciliado",
        "cerrado_con_descartes",
        "fallido",
        "requiere_reconciliacion",
    }
    ESTADOS_PROCESABLES = {"validado", "en_cola"}
    ESTADOS_RESOLUBLES = {"validado", "con_error", "fallido"}
    ESTADOS_RECONCILIABLES = ESTADOS_RESOLUBLES | {
        "requiere_reconciliacion",
        "reintentando",
    }
    ALICUOTAS_IVA_PERMITIDAS = {
        Decimal("0"),
        Decimal("10.5"),
        Decimal("21"),
        Decimal("27"),
    }

    TEMPLATE_COLUMNS = [
        "comprobante_ref",
        "empresa_cuit",
        "punto_venta_numero",
        "tipo_comprobante",
        "concepto",
        "fecha_emision",
        "cliente_tipo_documento",
        "cliente_numero_documento",
        "cliente_razon_social",
        "cliente_condicion_iva",
        "cliente_domicilio",
        "fecha_servicio_desde",
        "fecha_servicio_hasta",
        "fecha_vto_pago",
        "item_codigo",
        "item_descripcion",
        "item_cantidad",
        "item_unidad",
        "item_precio_unitario",
        "item_descuento_porcentaje",
        "item_iva_porcentaje",
        "observaciones",
        "asociado_tipo_comprobante",
        "asociado_punto_venta",
        "asociado_numero",
        "asociado_fecha",
        "asociado_cuit",
    ]
    HEADER_FIELDS = [
        "empresa_cuit",
        "punto_venta_numero",
        "tipo_comprobante",
        "concepto",
        "fecha_emision",
        "cliente_tipo_documento",
        "cliente_numero_documento",
        "cliente_razon_social",
        "cliente_condicion_iva",
        "cliente_domicilio",
        "fecha_servicio_desde",
        "fecha_servicio_hasta",
        "fecha_vto_pago",
        "observaciones",
        "asociado_tipo_comprobante",
        "asociado_punto_venta",
        "asociado_numero",
        "asociado_fecha",
        "asociado_cuit",
    ]
    TIPO_DOCUMENTO_MAP = {
        "CUIT": 80,
        "80": 80,
        "CUIL": 86,
        "86": 86,
        "DNI": 96,
        "96": 96,
        "LE": 89,
        "89": 89,
        "LC": 90,
        "90": 90,
        "PASAPORTE": 94,
        "94": 94,
        "CI": 99,
        "99": 99,
        "CONSUMIDOR FINAL": 99,
    }
    CONDICION_IVA_MAP = {
        "RESPONSABLE INSCRIPTO": "RI",
        "RI": "RI",
        "MONOTRIBUTO": "Monotributo",
        "EXENTO": "Exento",
        "CONSUMIDOR FINAL": "CF",
        "CF": "CF",
    }

    def __init__(self, db: AsyncSession):
        self.db = db
        self.facturacion_service = FacturacionService(db)

    async def generar_plantilla(self, empresa: Empresa) -> bytes:
        """Genera la plantilla fija de Excel para emisión masiva."""
        workbook = Workbook()
        instrucciones = workbook.active
        instrucciones.title = "Instrucciones"
        instrucciones["A1"] = "Plantilla de emisión masiva FactuFlow"
        instrucciones["A1"].font = Font(bold=True, size=14)
        instrucciones["A3"] = "1. Completá una fila por ítem."
        instrucciones[
            "A4"
        ] = "2. Repetí los datos del comprobante en todas las filas que compartan el mismo comprobante_ref."
        instrucciones[
            "A5"
        ] = "3. Definí Productos/Servicios en pantalla o completá concepto con Producto/Servicio."
        instrucciones[
            "A6"
        ] = "4. La descripción facturada del ítem es independiente: completá item_descripcion o elegí una descripción fija en pantalla."
        instrucciones[
            "A7"
        ] = "5. El lote debe corresponder a una única empresa. Usá el CUIT activo en la columna empresa_cuit."
        instrucciones[
            "A8"
        ] = "6. Completá fecha_emision o elegí una fecha fija en la pantalla antes de validar."
        instrucciones[
            "A9"
        ] = "7. Para consumidor final de bajo importe podés dejar documento, nombre y domicilio vacíos."
        instrucciones[
            "A10"
        ] = "8. Si informás documento, tipos sugeridos: CUIT, DNI, CUIL, Pasaporte."
        instrucciones[
            "A11"
        ] = "9. Condición IVA sugerida: Responsable Inscripto, Monotributo, Exento, Consumidor Final."

        hoja = workbook.create_sheet("Comprobantes")
        hoja.append(self.TEMPLATE_COLUMNS)
        for cell in hoja[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")

        hoja.append(
            [
                "LOTE-001",
                empresa.cuit,
                "",
                6,
                "",
                "",
                "",
                "",
                "",
                "Consumidor Final",
                "",
                "",
                "",
                "",
                "SERV-001",
                "Servicio mensual",
                1,
                "unidad",
                1000,
                0,
                21,
                "Factura generada por lote",
            ]
        )
        hoja.append(
            [
                "LOTE-001",
                empresa.cuit,
                1,
                6,
                1,
                "",
                "",
                "",
                "",
                "Consumidor Final",
                "",
                "",
                "",
                "",
                "SERV-002",
                "Soporte adicional",
                2,
                "hora",
                500,
                0,
                21,
                "Factura generada por lote",
            ]
        )

        stream = BytesIO()
        workbook.save(stream)
        return stream.getvalue()

    async def validar_y_registrar_lote(
        self,
        file_bytes: bytes,
        filename: str,
        empresa: Empresa,
        usuario: Usuario,
        opciones_fechas: OpcionesFechasLote,
        opciones_concepto: OpcionesConceptoLote,
        opciones_descripcion_item: OpcionesDescripcionItemLote,
        opciones_punto_venta: OpcionesPuntoVentaLote | None = None,
        formato_version_id: int | None = None,
        perfil_carga_masiva_snapshot: dict[str, Any] | None = None,
    ) -> LoteComprobante:
        """Valida un archivo Excel y persiste el lote con su detalle."""
        if not file_bytes:
            raise LoteComprobanteError("El archivo está vacío")
        opciones_punto_venta = opciones_punto_venta or OpcionesPuntoVentaLote()
        self._validar_opciones_concepto(opciones_concepto)
        self._validar_opciones_descripcion_item(opciones_descripcion_item)
        self._validar_opciones_punto_venta(opciones_punto_venta)
        self._validar_opciones_fechas(
            opciones_fechas,
            requiere_fechas_servicio=opciones_concepto.concepto_modo != "productos",
        )

        file_hash = self._calcular_hash_lote(
            file_bytes,
            formato_version_id,
            opciones_fechas,
            opciones_concepto,
            opciones_descripcion_item,
            opciones_punto_venta,
        )
        await self._preparar_idempotencia(empresa.id, file_hash)

        importacion = await self._leer_excel_para_lote(
            file_bytes=file_bytes,
            empresa=empresa,
            formato_version_id=formato_version_id,
        )
        filas_excel = importacion["filas"]
        if len(filas_excel) > settings.batch_max_rows:
            raise LoteComprobanteError(
                f"El archivo supera el máximo permitido de {settings.batch_max_rows} filas"
            )
        self._validar_empresa_del_lote(filas_excel, empresa)
        self._validar_concepto_archivo_si_corresponde(importacion, opciones_concepto)
        self._validar_descripcion_item_archivo_si_corresponde(
            importacion, opciones_descripcion_item
        )
        filas_excel = [
            self._aplicar_opciones_concepto(row, opciones_concepto)
            for row in filas_excel
        ]
        filas_excel = [
            self._aplicar_opciones_descripcion_item(row, opciones_descripcion_item)
            for row in filas_excel
        ]
        filas_excel = [
            self._aplicar_opciones_fechas(row, opciones_fechas) for row in filas_excel
        ]

        puntos_venta = await self._obtener_puntos_venta(empresa.id)
        self._validar_punto_venta_fijo(opciones_punto_venta, puntos_venta)
        filas_excel = [
            self._aplicar_opciones_punto_venta(row, opciones_punto_venta)
            for row in filas_excel
        ]
        certificado = await self._obtener_certificado_activo(empresa.id)
        if certificado is None:
            raise LoteComprobanteError(
                "No hay un certificado activo para la empresa en el ambiente configurado"
            )
        logger.info(
            "Validando lote '%s' para empresa %s (%s) con %s filas",
            filename,
            empresa.id,
            empresa.cuit,
            len(filas_excel),
        )

        lote = LoteComprobante(
            nombre_archivo=filename,
            archivo_hash=file_hash,
            estado="cargado",
            modo_procesamiento="sincronico",
            procesamiento_async=False,
            total_filas=len(filas_excel),
            empresa_id=empresa.id,
            usuario_id=usuario.id,
            formato_importacion_id=importacion["formato_importacion_id"],
            formato_importacion_version_id=importacion[
                "formato_importacion_version_id"
            ],
            headers_detectados_json=importacion["headers_detectados"],
            mapeo_usado_json=importacion["mapeo_usado"],
            metadata_json={
                "empresa_cuit": empresa.cuit,
                "formato_importacion": importacion["formato_nombre"],
                "perfil_carga_masiva": perfil_carga_masiva_snapshot,
                "opciones_fechas": self._serializar_opciones_fechas(opciones_fechas),
                "opciones_concepto": self._serializar_opciones_concepto(
                    opciones_concepto
                ),
                "opciones_descripcion_item": (
                    self._serializar_opciones_descripcion_item(
                        opciones_descripcion_item
                    )
                ),
                "opciones_punto_venta": self._serializar_opciones_punto_venta(
                    opciones_punto_venta
                ),
            },
        )
        self.db.add(lote)
        await self.db.flush()

        grupos_por_ref: dict[str, list[tuple[int, dict[str, Any]]]] = defaultdict(list)
        for fila_excel, row in enumerate(filas_excel, start=2):
            comprobante_ref = str(row["comprobante_ref"]).strip()
            if not comprobante_ref:
                raise LoteComprobanteError(
                    "Todas las filas deben incluir comprobante_ref"
                )
            grupos_por_ref[comprobante_ref].append((fila_excel, row))

        if len(grupos_por_ref) > settings.batch_max_groups:
            raise LoteComprobanteError(
                f"El archivo supera el máximo permitido de {settings.batch_max_groups} comprobantes"
            )

        orden = 0
        for comprobante_ref, row_group in grupos_por_ref.items():
            group_result = self._validar_grupo(
                comprobante_ref=comprobante_ref,
                rows=row_group,
                empresa=empresa,
                puntos_venta=puntos_venta,
            )
            orden += 1
            grupo = LoteComprobanteGrupo(
                lote_id=lote.id,
                comprobante_ref=comprobante_ref,
                orden=orden,
                estado=group_result["estado"],
                tipo_comprobante=group_result.get("tipo_comprobante"),
                punto_venta_numero=group_result.get("punto_venta_numero"),
                cliente_documento=group_result.get("cliente_documento"),
                cliente_razon_social=group_result.get("cliente_razon_social"),
                total_estimado=group_result.get("total_estimado", Decimal("0")),
                payload_json=group_result.get("payload"),
                mensajes_json=group_result["mensajes"],
            )
            self.db.add(grupo)
            await self.db.flush()

            for fila_excel, row in row_group:
                fila = LoteComprobanteFila(
                    lote_id=lote.id,
                    grupo_id=grupo.id,
                    fila_excel=fila_excel,
                    comprobante_ref=comprobante_ref,
                    estado=grupo.estado,
                    datos_json=row,
                    mensajes_json=group_result["mensajes"],
                )
                self.db.add(fila)

        await self.db.flush()
        duplicados_logicos = await self.obtener_confirmacion_duplicado_logico_grupos(
            lote_id=lote.id,
            empresa_id=empresa.id,
            estados={"validado"},
        )
        if duplicados_logicos["cantidad_duplicados_logicos"]:
            metadata = dict(lote.metadata_json or {})
            metadata["duplicados_logicos"] = {
                "cantidad": duplicados_logicos["cantidad_duplicados_logicos"],
                "confirmacion": duplicados_logicos["confirmacion_duplicado_logico"],
            }
            lote.metadata_json = metadata
            await self._agregar_advertencia_duplicado_logico(
                grupo_ids=duplicados_logicos["ids_grupos"],
                mensaje=duplicados_logicos["mensaje_confirmacion_duplicado_logico"],
            )
            await self.db.flush()
        await self._actualizar_estado_lote(lote)
        try:
            await self.db.commit()
        except IntegrityError as exc:
            await self.db.rollback()
            raise LoteComprobanteError(
                "Ese archivo ya fue cargado previamente. Revisá el lote existente antes de volver a subirlo."
            ) from exc
        await self.db.refresh(lote)
        logger.info(
            "Lote %s validado: %s grupos, %s validos, %s con error",
            lote.id,
            lote.total_grupos,
            lote.grupos_validos,
            lote.grupos_con_error,
        )
        return lote

    def _calcular_hash_lote(
        self,
        file_bytes: bytes,
        formato_version_id: int | None,
        opciones_fechas: OpcionesFechasLote,
        opciones_concepto: OpcionesConceptoLote,
        opciones_descripcion_item: OpcionesDescripcionItemLote,
        opciones_punto_venta: OpcionesPuntoVentaLote,
    ) -> str:
        """Calcula idempotencia por archivo, formato y políticas operativas."""
        formato_key = str(formato_version_id or "plantilla_oficial").encode("utf-8")
        fechas_key = repr(
            sorted(self._serializar_opciones_fechas(opciones_fechas).items())
        ).encode("utf-8")
        concepto_key = repr(
            sorted(self._serializar_opciones_concepto(opciones_concepto).items())
        ).encode("utf-8")
        descripcion_key = repr(
            sorted(
                self._serializar_opciones_descripcion_item(
                    opciones_descripcion_item
                ).items()
            )
        ).encode("utf-8")
        punto_venta_key = repr(
            sorted(self._serializar_opciones_punto_venta(opciones_punto_venta).items())
        ).encode("utf-8")
        return hashlib.sha256(
            file_bytes
            + b":factuflow-format:"
            + formato_key
            + b":factuflow-dates:"
            + fechas_key
            + b":factuflow-concept:"
            + concepto_key
            + b":factuflow-item-description:"
            + descripcion_key
            + b":factuflow-point-of-sale:"
            + punto_venta_key
        ).hexdigest()

    def _validar_opciones_concepto(self, opciones: OpcionesConceptoLote) -> None:
        """Valida que el concepto fiscal haya sido elegido explícitamente."""
        if opciones.concepto_modo not in {"productos", "servicios", "archivo"}:
            raise LoteComprobanteError(
                "Debes indicar si el lote corresponde a productos, servicios o si el concepto sale del archivo."
            )

    def _validar_opciones_descripcion_item(
        self, opciones: OpcionesDescripcionItemLote
    ) -> None:
        """Valida que la descripción facturada haya sido elegida explícitamente."""
        if opciones.descripcion_item_modo not in {"archivo", "fija"}:
            raise LoteComprobanteError(
                "Debes indicar si la descripción facturada sale del archivo o se fija para todo el lote."
            )
        if (
            opciones.descripcion_item_modo == "fija"
            and not (opciones.descripcion_item_fija or "").strip()
        ):
            raise LoteComprobanteError(
                "Debes indicar la descripción facturada fija antes de validar el lote."
            )

    def _validar_opciones_punto_venta(self, opciones: OpcionesPuntoVentaLote) -> None:
        """Valida que la política de punto de venta sea explícita y consistente."""
        if opciones.punto_venta_modo not in {"archivo", "fijo"}:
            raise LoteComprobanteError(
                "Debes indicar si el punto de venta sale del archivo o se fija para todo el lote."
            )
        if opciones.punto_venta_modo == "fijo" and not opciones.punto_venta_numero:
            raise LoteComprobanteError(
                "Debes indicar el punto de venta fijo antes de validar el lote."
            )

    def _validar_punto_venta_fijo(
        self,
        opciones: OpcionesPuntoVentaLote,
        puntos_venta: dict[int, PuntoVenta],
    ) -> None:
        """Verifica que el punto de venta fijo esté habilitado para FactuFlow."""
        if opciones.punto_venta_modo != "fijo":
            return
        if opciones.punto_venta_numero not in puntos_venta:
            raise LoteComprobanteError(
                "El punto de venta elegido no está habilitado para usar en FactuFlow. "
                "Primero completá Puntos de venta para este emisor."
            )

    def _validar_opciones_fechas(
        self, opciones: OpcionesFechasLote, requiere_fechas_servicio: bool
    ) -> None:
        """Valida que la política de fechas haya sido elegida explícitamente."""
        modos_validos = {"archivo", "fija"}
        if opciones.fecha_emision_modo not in modos_validos:
            raise LoteComprobanteError(
                "Debes indicar si la fecha de emisión sale del archivo o se fija manualmente."
            )
        if opciones.fecha_emision_modo == "fija" and not opciones.fecha_emision_fija:
            raise LoteComprobanteError(
                "Debes indicar la fecha de emisión fija antes de validar el lote."
            )

        if not requiere_fechas_servicio:
            return

        for nombre, modo, valor in [
            (
                "fecha desde del servicio",
                opciones.fecha_servicio_desde_modo,
                opciones.fecha_servicio_desde_fija,
            ),
            (
                "fecha hasta del servicio",
                opciones.fecha_servicio_hasta_modo,
                opciones.fecha_servicio_hasta_fija,
            ),
            (
                "vencimiento de pago",
                opciones.fecha_vto_pago_modo,
                opciones.fecha_vto_pago_fija,
            ),
        ]:
            if modo not in modos_validos:
                raise LoteComprobanteError(
                    f"Debes indicar si la {nombre} sale del archivo o se fija manualmente."
                )
            if modo == "fija" and not valor:
                raise LoteComprobanteError(
                    f"Debes indicar la {nombre} fija antes de validar el lote."
                )

    def _serializar_opciones_concepto(
        self, opciones: OpcionesConceptoLote
    ) -> dict[str, str]:
        """Convierte la política de concepto a un JSON persistible."""
        return {"concepto_modo": opciones.concepto_modo}

    def _serializar_opciones_descripcion_item(
        self, opciones: OpcionesDescripcionItemLote
    ) -> dict[str, str | None]:
        """Convierte la política de descripción de ítem a JSON persistible."""
        return {
            "descripcion_item_modo": opciones.descripcion_item_modo,
            "descripcion_item_fija": (
                opciones.descripcion_item_fija.strip()
                if opciones.descripcion_item_fija
                else None
            ),
        }

    def _serializar_opciones_punto_venta(
        self, opciones: OpcionesPuntoVentaLote
    ) -> dict[str, str | int | None]:
        """Convierte la política de punto de venta a JSON persistible."""
        return {
            "punto_venta_modo": opciones.punto_venta_modo,
            "punto_venta_numero": opciones.punto_venta_numero,
        }

    def _serializar_opciones_fechas(
        self, opciones: OpcionesFechasLote
    ) -> dict[str, str | None]:
        """Convierte la política de fechas a un JSON persistible."""
        return {
            "fecha_emision_modo": opciones.fecha_emision_modo,
            "fecha_emision_fija": (
                opciones.fecha_emision_fija.isoformat()
                if opciones.fecha_emision_fija
                else None
            ),
            "fecha_servicio_desde_modo": opciones.fecha_servicio_desde_modo,
            "fecha_servicio_desde_fija": (
                opciones.fecha_servicio_desde_fija.isoformat()
                if opciones.fecha_servicio_desde_fija
                else None
            ),
            "fecha_servicio_hasta_modo": opciones.fecha_servicio_hasta_modo,
            "fecha_servicio_hasta_fija": (
                opciones.fecha_servicio_hasta_fija.isoformat()
                if opciones.fecha_servicio_hasta_fija
                else None
            ),
            "fecha_vto_pago_modo": opciones.fecha_vto_pago_modo,
            "fecha_vto_pago_fija": (
                opciones.fecha_vto_pago_fija.isoformat()
                if opciones.fecha_vto_pago_fija
                else None
            ),
        }

    def _validar_concepto_archivo_si_corresponde(
        self,
        importacion: dict[str, Any],
        opciones: OpcionesConceptoLote,
    ) -> None:
        """Verifica que el Excel tenga columna de concepto cuando se la elige."""
        if opciones.concepto_modo != "archivo":
            return

        concepto = importacion["mapeo_usado"].get("campos", {}).get("concepto")
        if not concepto or concepto.get("origen") not in {"header", "columna"}:
            raise LoteComprobanteError(
                "Elegiste 'Definido por el archivo', pero el Excel no tiene una columna de concepto fiscal. Agrega una columna con Producto o Servicio en todas las filas, o elegí Productos/Servicios para todo el lote."
            )
        if concepto.get("encontrado") is False:
            raise LoteComprobanteError(
                "La columna de concepto fiscal no fue encontrada en el archivo. Debe indicar Producto o Servicio en todas las filas."
            )

    def _aplicar_opciones_concepto(
        self, row: dict[str, Any], opciones: OpcionesConceptoLote
    ) -> dict[str, Any]:
        """Completa el concepto fiscal según la elección del usuario."""
        updated = dict(row)
        if opciones.concepto_modo == "productos":
            updated["concepto"] = 1
        elif opciones.concepto_modo == "servicios":
            updated["concepto"] = 2
        else:
            updated["concepto"] = self._resolver_concepto_archivo(
                updated.get("concepto")
            )
        return updated

    def _validar_descripcion_item_archivo_si_corresponde(
        self,
        importacion: dict[str, Any],
        opciones: OpcionesDescripcionItemLote,
    ) -> None:
        """Verifica que el Excel tenga descripción de ítem si se eligió archivo."""
        if opciones.descripcion_item_modo != "archivo":
            return

        descripcion = (
            importacion["mapeo_usado"].get("campos", {}).get("item_descripcion")
        )
        if not descripcion or descripcion.get("origen") not in {"header", "columna"}:
            raise LoteComprobanteError(
                "Elegiste que la descripción facturada salga del archivo, pero el Excel no tiene una columna de descripción del ítem. Agregá una columna de descripción o elegí una descripción fija para todo el lote."
            )
        if descripcion.get("encontrado") is False:
            raise LoteComprobanteError(
                "La columna de descripción facturada no fue encontrada en el archivo. Debe tener valor en todas las filas o debes elegir una descripción fija."
            )

    def _aplicar_opciones_descripcion_item(
        self,
        row: dict[str, Any],
        opciones: OpcionesDescripcionItemLote,
    ) -> dict[str, Any]:
        """Completa la descripción facturada según la elección del usuario."""
        updated = dict(row)
        if opciones.descripcion_item_modo == "fija":
            updated["item_descripcion"] = (opciones.descripcion_item_fija or "").strip()
        return updated

    def _aplicar_opciones_punto_venta(
        self,
        row: dict[str, Any],
        opciones: OpcionesPuntoVentaLote,
    ) -> dict[str, Any]:
        """Completa el punto de venta según la elección del usuario."""
        updated = dict(row)
        if opciones.punto_venta_modo == "fijo":
            updated["punto_venta_numero"] = opciones.punto_venta_numero
        return updated

    def _resolver_concepto_archivo(self, value: Any) -> int | str:
        """Convierte Producto/Servicio del archivo al código ARCA."""
        normalized = self._normalizar_texto(value)
        if normalized in {"producto", "productos"}:
            return 1
        if normalized in {"servicio", "servicios"}:
            return 2
        return value or ""

    def _normalizar_texto(self, value: Any) -> str:
        """Normaliza texto de Excel para comparar etiquetas fiscales."""
        text = str(value or "").strip().lower()
        decomposed = unicodedata.normalize("NFKD", text)
        return decomposed.encode("ascii", "ignore").decode("ascii")

    def _aplicar_opciones_fechas(
        self, row: dict[str, Any], opciones: OpcionesFechasLote
    ) -> dict[str, Any]:
        """Completa las fechas fiscales del comprobante según la elección del usuario."""
        updated = dict(row)
        fuente_archivo = updated.get("fecha_emision") or updated.get("fecha_origen")
        updated["fecha_emision"] = self._resolver_fecha_lote(
            modo=opciones.fecha_emision_modo,
            fecha_fija=opciones.fecha_emision_fija,
            valor_archivo=fuente_archivo,
        )
        updated["fecha_servicio_desde"] = self._resolver_fecha_lote(
            modo=opciones.fecha_servicio_desde_modo,
            fecha_fija=opciones.fecha_servicio_desde_fija,
            valor_archivo=updated.get("fecha_servicio_desde")
            or updated.get("fecha_origen"),
        )
        updated["fecha_servicio_hasta"] = self._resolver_fecha_lote(
            modo=opciones.fecha_servicio_hasta_modo,
            fecha_fija=opciones.fecha_servicio_hasta_fija,
            valor_archivo=updated.get("fecha_servicio_hasta")
            or updated.get("fecha_origen"),
        )
        updated["fecha_vto_pago"] = self._resolver_fecha_lote(
            modo=opciones.fecha_vto_pago_modo,
            fecha_fija=opciones.fecha_vto_pago_fija,
            valor_archivo=updated.get("fecha_vto_pago") or updated.get("fecha_origen"),
        )
        return updated

    def _resolver_fecha_lote(
        self, modo: str, fecha_fija: date | None, valor_archivo: Any
    ) -> str:
        """Resuelve una fecha de lote y la deja en formato ISO o vacía."""
        if not modo:
            return ""
        if modo == "fija":
            return fecha_fija.isoformat() if fecha_fija else ""
        fecha = self._parse_date(valor_archivo)
        return fecha.isoformat() if fecha else ""

    async def encolar_lote(
        self,
        lote_id: int,
        empresa_id: int,
        operacion_id: int | None = None,
        confirmacion_duplicado_logico: bool = False,
    ) -> LoteComprobante:
        """Deja un lote validado en cola persistente para el worker."""
        lote = await self.obtener_lote_resumen(lote_id, empresa_id)
        if lote.estado == "procesando":
            return lote
        if lote.estado in self.ESTADOS_TERMINALES:
            return lote
        if lote.grupos_validos == 0:
            raise LoteComprobanteError(
                "El lote no tiene comprobantes válidos para emitir"
            )

        lote.estado = "en_cola"
        lote.procesamiento_async = True
        lote.modo_procesamiento = "background"
        lote.mensaje_resumen = (
            "El lote quedó en cola para procesamiento en segundo plano."
        )
        if operacion_id is not None:
            metadata = dict(lote.metadata_json or {})
            metadata["operacion_idempotente_id"] = operacion_id
            metadata["confirmacion_duplicado_logico"] = confirmacion_duplicado_logico
            lote.metadata_json = metadata
        await self.db.commit()
        await self.db.refresh(lote)
        return lote

    async def bloquear_lote_procesando_stale(
        self,
        lote_id: int,
        empresa_id: int,
        *,
        motivo: str | None = None,
        commit: bool = True,
    ) -> LoteComprobante:
        """Bloquea un lote vencido para evitar reemisión fiscal automática."""
        stale_before = datetime.utcnow() - timedelta(
            minutes=settings.batch_processing_stale_minutes
        )
        lote = await self.obtener_lote_resumen(lote_id, empresa_id)
        if lote.estado != "procesando" or lote.updated_at >= stale_before:
            return lote

        grupos_reconciliados = await self._reconciliar_grupos_autorizados_existentes(
            lote_id
        )
        lote.estado = "cargado"
        await self.db.flush()
        await self._actualizar_estado_lote(lote)
        if (
            lote.estado
            in {
                "completado",
                "cerrado_reconciliado",
                "cerrado_con_descartes",
            }
            and not await self._lote_tiene_intentos_fiscales_inciertos(lote.id)
            and await self._lote_tiene_evidencia_fiscal_local_coherente(lote.id)
        ):
            if lote.finished_at is None:
                lote.finished_at = datetime.utcnow()
            motivo_cierre = (
                (
                    "El worker vinculó comprobantes locales ya autorizados "
                    "sin solicitar nuevos CAE."
                )
                if grupos_reconciliados
                else (
                    "El worker cerró un lote stale con comprobantes locales "
                    "ya autorizados sin solicitar nuevos CAE."
                )
            )
            self._registrar_evento_lote(
                lote_id=lote.id,
                accion="reconciliacion_local_stale",
                usuario_id=None,
                motivo=motivo_cierre,
                metadata_json={
                    "grupos_reconciliados": grupos_reconciliados,
                    "estado_final": lote.estado,
                },
            )
            operacion_id = (lote.metadata_json or {}).get("operacion_idempotente_id")
            if operacion_id is not None:
                await self._guardar_respuesta_operacion_background(lote, operacion_id)
            if commit:
                await self.db.commit()
                await self.db.refresh(lote)
            else:
                await self.db.flush()
            return lote

        clasificacion_stale = await self._clasificar_grupos_validos_stale(lote)
        preflight_checks: list[dict[str, Any]] = []
        preflight_error: str | None = None
        if (
            clasificacion_stale.total_validos
            and not clasificacion_stale.con_evidencia_fiscal
            and not await self._lote_tiene_intentos_fiscales_inciertos(lote.id)
            and await self._lote_tiene_evidencia_fiscal_local_coherente(lote.id)
        ):
            (
                preflight_ok,
                preflight_checks,
                preflight_error,
            ) = await self._preflight_reanudar_grupos_intactos_stale(
                lote,
                clasificacion_stale.intactos,
            )
            if preflight_ok:
                await self._actualizar_contadores_lote(lote)
                lote.estado = "en_cola"
                lote.finished_at = None
                lote.procesamiento_async = True
                lote.modo_procesamiento = "background"
                lote.mensaje_resumen = (
                    "El lote se recuperó de una interrupción sin evidencia fiscal "
                    "en los comprobantes pendientes y quedó en cola para continuar."
                )
                metadata = dict(lote.metadata_json or {})
                metadata["recuperacion_stale_intacta"] = {
                    "motivo": "pendientes_sin_intento_fiscal_y_numeracion_alineada",
                    "stale_minutes": settings.batch_processing_stale_minutes,
                    "requeued_at": datetime.utcnow().isoformat(),
                    "grupos_reconciliados_localmente": grupos_reconciliados,
                    "grupos_intactos": len(clasificacion_stale.intactos),
                    "preflight_arca": preflight_checks,
                }
                lote.metadata_json = metadata
                self._registrar_evento_lote(
                    lote_id=lote.id,
                    accion="reanudacion_segura_stale",
                    usuario_id=None,
                    motivo=(
                        "El worker verificó que los comprobantes pendientes no "
                        "tenían intento fiscal ni numeración asignada y reencoló "
                        "el lote tras validar la numeración contra ARCA."
                    ),
                    metadata_json={
                        "estado_anterior": "procesando",
                        "estado_nuevo": "en_cola",
                        "stale_minutes": settings.batch_processing_stale_minutes,
                        "grupos_reconciliados_localmente": grupos_reconciliados,
                        "grupos_intactos": len(clasificacion_stale.intactos),
                        "preflight_arca": preflight_checks,
                    },
                )
                operacion_id = metadata.get("operacion_idempotente_id")
                if operacion_id is not None:
                    await self._guardar_respuesta_operacion_background(
                        lote, operacion_id
                    )
                if commit:
                    await self.db.commit()
                    await self.db.refresh(lote)
                else:
                    await self.db.flush()
                return lote

        grupos_marcados_reconciliacion = (
            await self._marcar_grupos_validos_stale_para_reconciliacion(
                lote,
                grupos=clasificacion_stale.con_evidencia_fiscal,
            )
        )
        await self._actualizar_contadores_lote(lote)
        lote.estado = "requiere_reconciliacion"
        lote.finished_at = datetime.utcnow()
        lote.mensaje_resumen = (
            "El procesamiento quedó detenido y no se pudo demostrar una "
            "reanudación automática segura. No reintentes este lote: primero "
            "hay que reconciliar contra ARCA para confirmar qué comprobantes "
            "fueron autorizados."
        )
        metadata = dict(lote.metadata_json or {})
        metadata["bloqueo_operativo"] = {
            "motivo": "procesamiento_stale_requiere_reconciliacion",
            "stale_minutes": settings.batch_processing_stale_minutes,
            "blocked_at": lote.finished_at.isoformat(),
            "grupos_intactos_preservados": len(clasificacion_stale.intactos),
            "grupos_con_evidencia_fiscal": len(
                clasificacion_stale.con_evidencia_fiscal
            ),
            "preflight_error": preflight_error,
        }
        lote.metadata_json = metadata
        self._registrar_evento_lote(
            lote_id=lote.id,
            accion="bloqueo_operativo_no_reemitir",
            usuario_id=None,
            motivo=motivo
            or (
                "El worker detectó un lote procesando vencido y lo bloqueó para "
                "reconciliación antes de cualquier nuevo CAE."
            ),
            metadata_json={
                "estado_anterior": "procesando",
                "estado_nuevo": "requiere_reconciliacion",
                "stale_minutes": settings.batch_processing_stale_minutes,
                "grupos_reconciliados_localmente": grupos_reconciliados,
                "grupos_marcados_reconciliacion": grupos_marcados_reconciliacion,
                "grupos_intactos_preservados": len(clasificacion_stale.intactos),
                "grupos_con_evidencia_fiscal": len(
                    clasificacion_stale.con_evidencia_fiscal
                ),
                "preflight_arca": preflight_checks,
                "preflight_error": preflight_error,
            },
        )
        operacion_id = metadata.get("operacion_idempotente_id")
        if operacion_id is not None:
            await self._guardar_respuesta_operacion_background(lote, operacion_id)
        if commit:
            await self.db.commit()
            await self.db.refresh(lote)
        else:
            await self.db.flush()
        return lote

    async def listar_lotes(
        self, empresa_id: int, limit: int = 20
    ) -> list[LoteComprobante]:
        """Lista lotes recientes de una empresa."""
        result = await self.db.execute(
            select(LoteComprobante)
            .where(LoteComprobante.empresa_id == empresa_id)
            .order_by(LoteComprobante.created_at.desc())
            .limit(limit)
        )
        return list(result.scalars().all())

    async def calcular_material_idempotente_grupos(
        self,
        *,
        lote_id: int,
        empresa_id: int,
        estados: set[str] | None = None,
        grupo_ids: list[int] | None = None,
    ) -> dict[str, Any]:
        """Calcula una huella estable de los payloads fiscales del lote."""
        stmt = (
            select(LoteComprobanteGrupo.id, LoteComprobanteGrupo.payload_json)
            .join(LoteComprobante, LoteComprobante.id == LoteComprobanteGrupo.lote_id)
            .where(
                LoteComprobanteGrupo.lote_id == lote_id,
                LoteComprobante.empresa_id == empresa_id,
                LoteComprobanteGrupo.payload_json.is_not(None),
            )
            .order_by(LoteComprobanteGrupo.id)
        )
        if estados is not None:
            stmt = stmt.where(LoteComprobanteGrupo.estado.in_(estados))
        if grupo_ids:
            stmt = stmt.where(LoteComprobanteGrupo.id.in_(grupo_ids))

        rows = (await self.db.execute(stmt)).all()
        grupos = [
            {
                "grupo_id": grupo_id,
                "payload_hash": hashlib.sha256(
                    json.dumps(
                        payload_json or {},
                        sort_keys=True,
                        separators=(",", ":"),
                        default=str,
                    ).encode("utf-8")
                ).hexdigest(),
            }
            for grupo_id, payload_json in rows
        ]
        encoded = json.dumps(
            grupos,
            sort_keys=True,
            separators=(",", ":"),
        ).encode("utf-8")
        return {
            "grupo_ids": [grupo["grupo_id"] for grupo in grupos],
            "grupos_hash": hashlib.sha256(encoded).hexdigest(),
        }

    async def obtener_lote(self, lote_id: int, empresa_id: int) -> LoteComprobante:
        """Obtiene un lote con grupos y filas."""
        result = await self.db.execute(
            select(LoteComprobante)
            .options(
                selectinload(LoteComprobante.grupos),
                selectinload(LoteComprobante.filas),
            )
            .where(
                LoteComprobante.id == lote_id,
                LoteComprobante.empresa_id == empresa_id,
            )
        )
        lote = result.scalar_one_or_none()
        if lote is None:
            raise LoteComprobanteError("No se encontró el lote solicitado")
        return lote

    async def obtener_seguimiento_lote(
        self,
        lote_id: int,
        empresa_id: int,
    ) -> LoteComprobanteSeguimientoResponse:
        """Proyecta el estado mínimo de un lote aislado por emisor."""
        result = await self.db.execute(
            select(
                LoteComprobante.id,
                LoteComprobante.estado,
                LoteComprobante.modo_procesamiento,
                LoteComprobante.procesamiento_async,
                LoteComprobante.total_filas,
                LoteComprobante.total_grupos,
                LoteComprobante.grupos_validos,
                LoteComprobante.grupos_con_error,
                LoteComprobante.grupos_emitidos,
                LoteComprobante.grupos_fallidos,
                LoteComprobante.grupos_reconciliados_externos,
                LoteComprobante.grupos_descartados,
                LoteComprobante.mensaje_resumen,
                LoteComprobante.started_at,
                LoteComprobante.finished_at,
                LoteComprobante.updated_at,
            ).where(
                LoteComprobante.id == lote_id,
                LoteComprobante.empresa_id == empresa_id,
            )
        )
        seguimiento = result.mappings().one_or_none()
        if seguimiento is None:
            raise LoteComprobanteError("No se encontró el lote solicitado")
        return LoteComprobanteSeguimientoResponse.model_validate(dict(seguimiento))

    async def obtener_lote_resumen(
        self, lote_id: int, empresa_id: int
    ) -> LoteComprobante:
        """Obtiene un lote sin cargar relaciones pesadas."""
        result = await self.db.execute(
            select(LoteComprobante).where(
                LoteComprobante.id == lote_id,
                LoteComprobante.empresa_id == empresa_id,
            )
        )
        lote = result.scalar_one_or_none()
        if lote is None:
            raise LoteComprobanteError("No se encontró el lote solicitado")
        return lote

    async def obtener_grupos_lote_paginados(
        self,
        lote_id: int,
        empresa_id: int,
        page: int,
        per_page: int,
        estado: str | None = None,
    ) -> tuple[list[LoteComprobanteGrupo], int]:
        """Obtiene una página de grupos del lote validando el emisor activo."""
        await self.obtener_lote_resumen(lote_id, empresa_id)
        filtros = [LoteComprobanteGrupo.lote_id == lote_id]
        if estado:
            filtros.append(LoteComprobanteGrupo.estado == estado)

        total_result = await self.db.execute(
            select(func.count()).select_from(LoteComprobanteGrupo).where(*filtros)
        )
        total = int(total_result.scalar_one() or 0)
        offset = max(page - 1, 0) * per_page
        grupos_result = await self.db.execute(
            select(LoteComprobanteGrupo)
            .options(selectinload(LoteComprobanteGrupo.filas))
            .where(*filtros)
            .order_by(LoteComprobanteGrupo.orden, LoteComprobanteGrupo.id)
            .offset(offset)
            .limit(per_page)
        )
        return list(grupos_result.scalars().all()), total

    async def obtener_resumen_operativo_lote(
        self, lote_id: int, empresa_id: int
    ) -> dict[str, Any]:
        """Calcula metadatos completos del lote sin traer todas sus filas."""
        await self.obtener_lote_resumen(lote_id, empresa_id)
        result = await self.db.execute(
            select(
                LoteComprobanteGrupo.payload_json,
                LoteComprobanteGrupo.punto_venta_numero,
                LoteComprobanteGrupo.total_estimado,
            ).where(
                LoteComprobanteGrupo.lote_id == lote_id,
                LoteComprobanteGrupo.estado == "validado",
            )
        )
        filas = list(result.all())
        fechas, puntos_venta, token = self._crear_token_confirmacion_desde_payloads(
            [(payload, punto) for payload, punto, _ in filas]
        )
        fallidos_result = await self.db.execute(
            select(
                LoteComprobanteGrupo.payload_json,
                LoteComprobanteGrupo.punto_venta_numero,
            ).where(
                LoteComprobanteGrupo.lote_id == lote_id,
                LoteComprobanteGrupo.estado == "fallido",
            )
        )
        (
            fechas_fallidos,
            puntos_fallidos,
            token_fallidos,
        ) = self._crear_token_confirmacion_desde_payloads(list(fallidos_result.all()))
        duplicados = await self.obtener_confirmacion_duplicado_logico_grupos(
            lote_id=lote_id,
            empresa_id=empresa_id,
            estados={"validado"},
        )
        return {
            "confirmacion_fecha_fiscal": token,
            "mensaje_confirmacion_fecha_fiscal": (
                self._crear_mensaje_confirmacion_fecha_fiscal(fechas, puntos_venta)
            ),
            "confirmacion_reintento_fallidos": token_fallidos,
            "mensaje_confirmacion_reintento_fallidos": (
                self._crear_mensaje_confirmacion_fecha_fiscal(
                    fechas_fallidos,
                    puntos_fallidos,
                )
            ),
            "confirmacion_duplicado_logico": duplicados[
                "confirmacion_duplicado_logico"
            ],
            "mensaje_confirmacion_duplicado_logico": duplicados[
                "mensaje_confirmacion_duplicado_logico"
            ],
            "cantidad_duplicados_logicos": duplicados["cantidad_duplicados_logicos"],
            "fechas_emision_validas": fechas,
            "puntos_venta_validos": puntos_venta,
            "totales_listos_para_emitir": self._calcular_totales_payloads(filas),
        }

    async def obtener_confirmacion_fiscal_grupos(
        self,
        lote_id: int,
        empresa_id: int,
        estados: set[str],
        grupo_ids: list[int] | None = None,
    ) -> dict[str, str | list[int]]:
        """Calcula el token fiscal exacto para grupos seleccionados."""
        await self.obtener_lote_resumen(lote_id, empresa_id)
        filtros = [
            LoteComprobanteGrupo.lote_id == lote_id,
            LoteComprobanteGrupo.estado.in_(estados),
        ]
        if grupo_ids:
            filtros.append(LoteComprobanteGrupo.id.in_(grupo_ids))

        result = await self.db.execute(
            select(
                LoteComprobanteGrupo.id,
                LoteComprobanteGrupo.payload_json,
                LoteComprobanteGrupo.punto_venta_numero,
            ).where(*filtros)
        )
        filas = list(result.all())
        if not filas:
            raise LoteComprobanteError(
                "No hay comprobantes seleccionados para resolver"
            )

        ids_encontrados = {int(grupo_id) for grupo_id, _, _ in filas}
        if grupo_ids and ids_encontrados != set(grupo_ids):
            raise LoteComprobanteError(
                "Solo pueden resolverse comprobantes pendientes del lote activo"
            )

        fechas, puntos_venta, token = self._crear_token_confirmacion_desde_payloads(
            [(payload, punto) for _, payload, punto in filas]
        )
        if not fechas or not puntos_venta:
            raise LoteComprobanteError(
                "Los comprobantes seleccionados no tienen fecha fiscal o punto de venta definido"
            )
        return {
            "confirmacion_fecha_fiscal": token,
            "mensaje_confirmacion_fecha_fiscal": (
                self._crear_mensaje_confirmacion_fecha_fiscal(fechas, puntos_venta)
            ),
            "ids_grupos": sorted(ids_encontrados),
        }

    async def obtener_confirmacion_duplicado_logico_grupos(
        self,
        lote_id: int,
        empresa_id: int,
        estados: set[str],
        grupo_ids: list[int] | None = None,
    ) -> dict[str, Any]:
        """Calcula una confirmación adicional para duplicados lógicos probables."""
        await self.obtener_lote_resumen(lote_id, empresa_id)
        filtros = [
            LoteComprobanteGrupo.lote_id == lote_id,
            LoteComprobanteGrupo.estado.in_(estados),
        ]
        if grupo_ids:
            filtros.append(LoteComprobanteGrupo.id.in_(grupo_ids))

        result = await self.db.execute(
            select(
                LoteComprobanteGrupo.id,
                LoteComprobanteGrupo.comprobante_ref,
                LoteComprobanteGrupo.payload_json,
                LoteComprobanteGrupo.punto_venta_numero,
            ).where(*filtros)
        )
        filas = list(result.all())
        if not filas:
            return {
                "confirmacion_duplicado_logico": "",
                "mensaje_confirmacion_duplicado_logico": "",
                "cantidad_duplicados_logicos": 0,
                "ids_grupos": [],
            }

        idempotencia = IdempotenciaFiscalService(self.db)
        huellas: dict[int, str] = {}
        refs: dict[int, str] = {}
        duplicados_ids: set[int] = set()
        for grupo_id, comprobante_ref, payload, punto_venta_numero in filas:
            try:
                request = EmitirComprobanteRequest.model_validate(payload or {})
                totales = self.facturacion_service._calcular_totales(request.items)
                huella = idempotencia.calcular_huella_logica(
                    request=request,
                    punto_venta_numero=int(punto_venta_numero or 0),
                    total=totales["total"],
                )
                huellas[int(grupo_id)] = huella
                refs[int(grupo_id)] = str(comprobante_ref)
                punto_venta = await self.facturacion_service._obtener_punto_venta(
                    request.punto_venta_id,
                    empresa_id,
                )
                if punto_venta and await idempotencia.buscar_duplicado_logico(
                    request=request,
                    punto_venta=punto_venta,
                    total=totales["total"],
                ):
                    duplicados_ids.add(int(grupo_id))
            except DATABASE_TEMPORARILY_UNAVAILABLE_ERRORS:
                raise
            except Exception:  # pragma: no cover - defensivo, no bloquea validación
                continue

        grupos_por_huella: dict[str, list[int]] = defaultdict(list)
        for grupo_id, huella in huellas.items():
            grupos_por_huella[huella].append(grupo_id)
        for ids in grupos_por_huella.values():
            if len(ids) > 1:
                duplicados_ids.update(ids)

        if not duplicados_ids:
            return {
                "confirmacion_duplicado_logico": "",
                "mensaje_confirmacion_duplicado_logico": "",
                "cantidad_duplicados_logicos": 0,
                "ids_grupos": [],
            }

        material = "|".join(
            sorted(
                huellas[grupo_id] for grupo_id in duplicados_ids if grupo_id in huellas
            )
        )
        token_hash = hashlib.sha256(material.encode("utf-8")).hexdigest()
        cantidad = len(duplicados_ids)
        refs_label = ", ".join(refs[grupo_id] for grupo_id in sorted(duplicados_ids))
        return {
            "confirmacion_duplicado_logico": (
                f"duplicados_logicos={token_hash};cantidad={cantidad}"
            ),
            "mensaje_confirmacion_duplicado_logico": (
                "Se detectaron comprobantes probablemente duplicados "
                f"({refs_label}). Confirmá explícitamente si corresponde "
                "solicitar CAE de todos modos."
            ),
            "cantidad_duplicados_logicos": cantidad,
            "ids_grupos": sorted(duplicados_ids),
        }

    async def _agregar_advertencia_duplicado_logico(
        self,
        grupo_ids: list[int],
        mensaje: str,
    ) -> None:
        """Agrega una advertencia visible sin cambiar el estado emitible."""
        if not grupo_ids or not mensaje:
            return

        result = await self.db.execute(
            select(LoteComprobanteGrupo)
            .options(selectinload(LoteComprobanteGrupo.filas))
            .where(LoteComprobanteGrupo.id.in_(grupo_ids))
        )
        for grupo in result.scalars().all():
            mensajes_grupo = list(grupo.mensajes_json or [])
            if mensaje not in mensajes_grupo:
                mensajes_grupo.append(mensaje)
            grupo.mensajes_json = mensajes_grupo
            for fila in grupo.filas:
                mensajes_fila = list(fila.mensajes_json or [])
                if mensaje not in mensajes_fila:
                    mensajes_fila.append(mensaje)
                fila.mensajes_json = mensajes_fila

    @staticmethod
    def _crear_token_confirmacion_desde_payloads(
        filas: list[tuple[dict[str, Any] | None, int | None]]
    ) -> tuple[list[str], list[int], str]:
        """Construye fechas, puntos y token fiscal desde payloads de grupos."""
        fechas = sorted(
            {
                str((payload or {}).get("fecha_emision"))[:10]
                for payload, _ in filas
                if (payload or {}).get("fecha_emision")
            }
        )
        puntos_venta = sorted({int(punto) for _, punto in filas if punto is not None})
        token = (
            f"fechas={','.join(fechas)};"
            f"puntos_venta={','.join(str(punto) for punto in puntos_venta)}"
        )
        return fechas, puntos_venta, token

    async def recuperar_lote_interrumpido_pre_arca(
        self,
        *,
        lote_id: int,
        empresa_id: int,
        operacion_id: int,
        estado_reanudable: str,
        estados_claim: set[str],
        mensaje_seguro: str,
    ) -> bool:
        """Revierte lote y operación en una transacción solo antes de ARCA."""
        try:
            await self.db.rollback()
            idempotencia = IdempotenciaFiscalService(self.db)
            operacion_marcada = (
                await idempotencia.marcar_operacion_interrumpida_pre_arca(
                    operacion_id,
                    commit=False,
                )
            )
            if not operacion_marcada:
                await self.db.rollback()
                return False
            result = await self.db.execute(
                update(LoteComprobante)
                .where(
                    LoteComprobante.id == lote_id,
                    LoteComprobante.empresa_id == empresa_id,
                    LoteComprobante.estado.in_(estados_claim),
                )
                .values(
                    estado=estado_reanudable,
                    started_at=None,
                    finished_at=None,
                    mensaje_resumen=mensaje_seguro,
                )
            )
            if result.rowcount != 1:
                await self.db.rollback()
                return False
            await self.db.commit()
            return True
        except Exception as recovery_exc:
            logger.error(
                "event=pre_arca_lote_recovery_failed tipo_error=%s",
                type(recovery_exc).__name__,
            )
            try:
                await self.db.rollback()
            except Exception as rollback_exc:
                logger.error(
                    "event=pre_arca_lote_recovery_rollback_failed tipo_error=%s",
                    type(rollback_exc).__name__,
                )
            return False

    async def recuperar_lote_worker_interrumpido_pre_arca(
        self,
        *,
        lote_id: int,
        empresa_id: int,
    ) -> bool:
        """Devuelve a cola un lote del worker solo si su operación es reanudable."""
        try:
            await self.db.rollback()
            metadata = (
                await self.db.execute(
                    select(LoteComprobante.metadata_json).where(
                        LoteComprobante.id == lote_id,
                        LoteComprobante.empresa_id == empresa_id,
                    )
                )
            ).scalar_one_or_none()
            operacion_id = (metadata or {}).get("operacion_idempotente_id")
            if not isinstance(operacion_id, int):
                await self.db.rollback()
                return False
        except Exception as recovery_exc:
            logger.error(
                "event=pre_arca_worker_lookup_failed tipo_error=%s",
                type(recovery_exc).__name__,
            )
            try:
                await self.db.rollback()
            except Exception as rollback_exc:
                logger.error(
                    "event=pre_arca_worker_lookup_rollback_failed tipo_error=%s",
                    type(rollback_exc).__name__,
                )
            return False

        try:
            sin_intentos = ~exists(
                select(IntentoEmisionFiscal.id).where(
                    IntentoEmisionFiscal.operacion_id == operacion_id
                )
            )
            result = await self.db.execute(
                update(LoteComprobante)
                .where(
                    LoteComprobante.id == lote_id,
                    LoteComprobante.empresa_id == empresa_id,
                    LoteComprobante.estado.in_({"procesando", "en_cola"}),
                    sin_intentos,
                )
                .values(
                    estado="en_cola",
                    started_at=None,
                    finished_at=None,
                    mensaje_resumen=(
                        "El lote volvió a la cola para un reintento seguro."
                    ),
                )
            )
            if result.rowcount != 1:
                await self.db.rollback()
                return False
            await self.db.commit()
            return True
        except Exception as recovery_exc:
            logger.error(
                "event=pre_arca_worker_recovery_failed tipo_error=%s",
                type(recovery_exc).__name__,
            )
            try:
                await self.db.rollback()
            except Exception as rollback_exc:
                logger.error(
                    "event=pre_arca_worker_recovery_rollback_failed tipo_error=%s",
                    type(rollback_exc).__name__,
                )
            return False

    async def recuperar_reintento_interrumpido_pre_arca(
        self,
        *,
        lote_id: int,
        grupo_id: int,
        operacion_id: int,
        mensajes_previos: list[str] | None,
    ) -> bool:
        """Restaura únicamente el grupo reclamado y su operación antes de ARCA."""
        try:
            await self.db.rollback()
            idempotencia = IdempotenciaFiscalService(self.db)
            operacion_marcada = (
                await idempotencia.marcar_operacion_interrumpida_pre_arca(
                    operacion_id,
                    commit=False,
                )
            )
            if not operacion_marcada:
                await self.db.rollback()
                return False
            result = await self.db.execute(
                update(LoteComprobanteGrupo)
                .where(
                    LoteComprobanteGrupo.id == grupo_id,
                    LoteComprobanteGrupo.lote_id == lote_id,
                    LoteComprobanteGrupo.estado.in_({"reintentando", "fallido"}),
                )
                .values(
                    estado="fallido",
                    mensajes_json=(
                        mensajes_previos
                        or ["El reintento se interrumpió antes de solicitar CAE."]
                    ),
                )
            )
            if result.rowcount != 1:
                await self.db.rollback()
                return False
            await self.db.commit()
            return True
        except Exception as recovery_exc:
            logger.error(
                "event=pre_arca_retry_recovery_failed tipo_error=%s",
                type(recovery_exc).__name__,
            )
            try:
                await self.db.rollback()
            except Exception as rollback_exc:
                logger.error(
                    "event=pre_arca_retry_recovery_rollback_failed tipo_error=%s",
                    type(rollback_exc).__name__,
                )
            return False

    async def reintentar_grupos_fallidos(
        self,
        lote_id: int,
        empresa_id: int,
        usuario_id: int | None,
        grupo_ids: list[int] | None = None,
        operacion_id: int | None = None,
        confirmacion_duplicado_logico: bool = False,
        fase_solicitud_arca: FaseSolicitudArca | None = None,
    ) -> LoteComprobante:
        """Reemite grupos fallidos ya seleccionados y confirmados."""
        fase_solicitud_arca = fase_solicitud_arca or FaseSolicitudArca()
        lote = await self.obtener_lote_resumen(lote_id, empresa_id)
        self._validar_lote_resoluble(lote)
        grupos = await self._obtener_grupos_para_resolver(
            lote_id=lote_id,
            estados={"fallido"},
            grupo_ids=grupo_ids,
        )
        if not grupos:
            raise LoteComprobanteError(
                "El lote no tiene comprobantes fallidos para reintentar"
            )

        for grupo in grupos:
            mensajes_previos = list(grupo.mensajes_json or [])
            try:
                grupo_reclamado = await self._reclamar_grupo_para_reintento(
                    lote_id=lote_id,
                    grupo_id=grupo.id,
                )
            except DATABASE_TEMPORARILY_UNAVAILABLE_ERRORS:
                recuperada = False
                if operacion_id is not None and not fase_solicitud_arca.iniciada:
                    recuperada = await self.recuperar_reintento_interrumpido_pre_arca(
                        lote_id=lote_id,
                        grupo_id=grupo.id,
                        operacion_id=operacion_id,
                        mensajes_previos=mensajes_previos,
                    )
                fase_solicitud_arca.registrar_recuperacion_pre_arca(recuperada)
                raise
            if grupo_reclamado is None:
                continue

            try:
                request = EmitirComprobanteRequest.model_validate(
                    grupo_reclamado.payload_json or {}
                )
                request = request.model_copy(
                    update={
                        "confirmacion_duplicado_logico": (confirmacion_duplicado_logico)
                    }
                )
            except Exception as exc:  # pragma: no cover - defensa operacional
                try:
                    await self._marcar_reintento_fallido(
                        lote_id=lote_id,
                        empresa_id=empresa_id,
                        usuario_id=usuario_id,
                        grupo=grupo_reclamado,
                        exc=exc,
                    )
                except DATABASE_TEMPORARILY_UNAVAILABLE_ERRORS:
                    recuperada = False
                    if operacion_id is not None and not fase_solicitud_arca.iniciada:
                        recuperada = (
                            await self.recuperar_reintento_interrumpido_pre_arca(
                                lote_id=lote_id,
                                grupo_id=grupo_reclamado.id,
                                operacion_id=operacion_id,
                                mensajes_previos=mensajes_previos,
                            )
                        )
                    fase_solicitud_arca.registrar_recuperacion_pre_arca(recuperada)
                    raise
                continue

            lock = await self.facturacion_service._get_number_lock(
                request.empresa_id,
                request.punto_venta_id,
                request.tipo_comprobante,
            )
            async with lock:
                try:
                    resultado = (
                        await self.facturacion_service._emitir_comprobante_locked(
                            request,
                            commit=False,
                            operacion_id=operacion_id,
                            usuario_id=usuario_id,
                            lote_id=lote_id,
                            grupo_id=grupo_reclamado.id,
                            fase_solicitud_arca=fase_solicitud_arca,
                        )
                    )
                    await self._aplicar_resultado_emision_grupo(
                        grupo_reclamado,
                        resultado,
                    )
                    self._registrar_evento_lote(
                        lote_id=lote_id,
                        accion="reintentar_fallido",
                        usuario_id=usuario_id,
                        grupo_id=grupo_reclamado.id,
                        metadata_json={
                            "resultado": (
                                "autorizado" if resultado.exito else "fallido"
                            ),
                            "numero": resultado.numero,
                            "cae": resultado.cae,
                        },
                    )
                    await self.db.flush()
                    lote = await self.obtener_lote_resumen(lote_id, empresa_id)
                    await self._actualizar_estado_lote(lote)
                    await self.db.commit()
                except DATABASE_TEMPORARILY_UNAVAILABLE_ERRORS:
                    recuperada = False
                    if operacion_id is not None and not fase_solicitud_arca.iniciada:
                        recuperada = (
                            await self.recuperar_reintento_interrumpido_pre_arca(
                                lote_id=lote_id,
                                grupo_id=grupo_reclamado.id,
                                operacion_id=operacion_id,
                                mensajes_previos=mensajes_previos,
                            )
                        )
                    fase_solicitud_arca.registrar_recuperacion_pre_arca(recuperada)
                    raise
                except Exception as exc:  # pragma: no cover - defensa operacional
                    try:
                        await self._marcar_reintento_fallido(
                            lote_id=lote_id,
                            empresa_id=empresa_id,
                            usuario_id=usuario_id,
                            grupo=grupo_reclamado,
                            exc=exc,
                        )
                    except DATABASE_TEMPORARILY_UNAVAILABLE_ERRORS:
                        recuperada = False
                        if (
                            operacion_id is not None
                            and not fase_solicitud_arca.iniciada
                        ):
                            recuperada = (
                                await self.recuperar_reintento_interrumpido_pre_arca(
                                    lote_id=lote_id,
                                    grupo_id=grupo_reclamado.id,
                                    operacion_id=operacion_id,
                                    mensajes_previos=mensajes_previos,
                                )
                            )
                        fase_solicitud_arca.registrar_recuperacion_pre_arca(recuperada)
                        raise

        return await self.obtener_lote_resumen(lote_id, empresa_id)

    async def descartar_grupos(
        self,
        lote_id: int,
        empresa_id: int,
        usuario_id: int | None,
        grupo_ids: list[int],
        motivo: str,
    ) -> LoteComprobante:
        """Marca grupos pendientes como descartados por decisión operativa."""
        motivo = motivo.strip()
        if not motivo:
            raise LoteComprobanteError(
                "Debes indicar un motivo para descartar comprobantes"
            )

        lote = await self.obtener_lote_resumen(lote_id, empresa_id)
        self._validar_lote_resoluble(lote)
        grupos = await self._obtener_grupos_para_resolver(
            lote_id=lote_id,
            estados=self.ESTADOS_RESOLUBLES,
            grupo_ids=grupo_ids,
        )
        if len(grupos) != len(set(grupo_ids)):
            raise LoteComprobanteError(
                "Solo pueden descartarse comprobantes pendientes no emitidos"
            )

        for grupo in grupos:
            estado_anterior = grupo.estado
            grupo_reclamado = await self._reclamar_grupo_para_accion_resolutiva(
                lote_id=lote_id,
                grupo_id=grupo.id,
                estados=self.ESTADOS_RESOLUBLES,
                estado_claim="descartando",
                mensaje_error=(
                    "Solo pueden descartarse comprobantes pendientes no emitidos"
                ),
            )
            grupo_reclamado.estado = "descartado"
            grupo_reclamado.mensajes_json = [f"Descartado por usuario: {motivo}"]
            await self._marcar_filas(
                grupo_reclamado,
                "descartado",
                grupo_reclamado.mensajes_json,
            )
            self._registrar_evento_lote(
                lote_id=lote_id,
                accion="descartar_grupo",
                usuario_id=usuario_id,
                grupo_id=grupo_reclamado.id,
                motivo=motivo,
                metadata_json={"estado_anterior": estado_anterior},
            )

        await self.db.flush()
        lote = await self.obtener_lote_resumen(lote_id, empresa_id)
        await self._actualizar_estado_lote(lote)
        await self.db.commit()
        return await self.obtener_lote_resumen(lote_id, empresa_id)

    async def reconciliar_emitidos_externos(
        self,
        lote_id: int,
        empresa: Empresa,
        usuario_id: int | None,
        comprobantes: list[dict[str, Any]],
        consultar_comprobante,
    ) -> LoteComprobante:
        """Reconcila grupos emitidos fuera de FactuFlow verificando ARCA."""
        lote = await self.obtener_lote_resumen(lote_id, empresa.id)
        self._validar_lote_resoluble(lote, permitir_reconciliacion_tecnica=True)
        if not comprobantes:
            raise LoteComprobanteError("Debes indicar comprobantes para reconciliar")
        self._validar_reconciliacion_sin_numeros_duplicados(comprobantes)

        try:
            for item in comprobantes:
                motivo = str(item.get("motivo") or "").strip()
                if not motivo:
                    raise LoteComprobanteError(
                        "Debes indicar un motivo para registrar comprobantes externos"
                    )

                grupo = await self._reclamar_grupo_para_accion_resolutiva(
                    lote_id=lote_id,
                    grupo_id=int(item["grupo_id"]),
                    estados=self.ESTADOS_RECONCILIABLES,
                    estado_claim="reconciliando",
                    mensaje_error=(
                        "Solo pueden reconciliarse comprobantes pendientes del lote activo"
                    ),
                )
                try:
                    request = EmitirComprobanteRequest.model_validate(
                        grupo.payload_json or {}
                    )
                except Exception as exc:
                    raise LoteComprobanteError(
                        "El comprobante seleccionado no tiene datos fiscales completos para reconciliar"
                    ) from exc
                self._validar_datos_reconciliacion_externa(grupo, request, item)
                try:
                    consulta_arca = await consultar_comprobante(
                        punto_venta=int(item["punto_venta_numero"]),
                        tipo_cbte=int(item["tipo_comprobante"]),
                        numero=int(item["numero"]),
                    )
                except (ArcaServiceError, ArcaValidationError, ValueError) as exc:
                    raise LoteComprobanteError(
                        "No se pudo verificar el comprobante externo contra ARCA. "
                        f"Detalle: {exc}"
                    ) from exc
                self._validar_consulta_arca_externa(
                    empresa=empresa,
                    grupo=grupo,
                    request=request,
                    item=item,
                    consulta_arca=consulta_arca,
                )
                comprobante = await self._crear_o_vincular_comprobante_externo(
                    empresa=empresa,
                    request=request,
                    punto_venta_numero=int(item["punto_venta_numero"]),
                    numero=int(item["numero"]),
                    cae=str(consulta_arca.cae),
                    cae_vencimiento=str(consulta_arca.cae_vencimiento),
                )

                grupo.estado = "autorizado_externo"
                grupo.cae = comprobante.cae
                grupo.numero_asignado = comprobante.numero
                grupo.comprobante_id = comprobante.id
                grupo.mensajes_json = [
                    "Comprobante emitido fuera de FactuFlow y verificado contra ARCA.",
                    f"CAE {comprobante.cae}",
                    f"Motivo: {motivo}",
                ]
                await self._marcar_filas(
                    grupo, "autorizado_externo", grupo.mensajes_json
                )
                self._registrar_evento_lote(
                    lote_id=lote_id,
                    accion="reconciliar_externo",
                    usuario_id=usuario_id,
                    grupo_id=grupo.id,
                    motivo=motivo,
                    metadata_json={
                        "tipo_comprobante": item["tipo_comprobante"],
                        "punto_venta_numero": item["punto_venta_numero"],
                        "numero": item["numero"],
                        "cae": comprobante.cae,
                    },
                )
                await self.db.flush()
        except IntegrityError as exc:
            await self.db.rollback()
            raise LoteComprobanteError(
                "El comprobante externo ya está vinculado a otro grupo del lote"
            ) from exc
        except Exception:
            await self.db.rollback()
            raise

        lote = await self.obtener_lote_resumen(lote_id, empresa.id)
        await self._actualizar_estado_lote(lote)
        await self.db.commit()
        return await self.obtener_lote_resumen(lote_id, empresa.id)

    async def eliminar_lote_sin_emision(
        self,
        lote_id: int,
        empresa_id: int,
        usuario_id: int | None,
        motivo: str,
    ) -> None:
        """Elimina físicamente un lote sin evidencia fiscal emitida."""
        motivo = motivo.strip()
        if not motivo:
            raise LoteComprobanteError("Debes indicar un motivo para eliminar el lote")
        lote = await self.obtener_lote_resumen(lote_id, empresa_id)
        if lote.estado in {"en_cola", "procesando", "requiere_reconciliacion"}:
            raise LoteComprobanteError("No se puede eliminar un lote activo o incierto")
        if await self._lote_tiene_emision_o_incertidumbre(lote_id):
            raise LoteComprobanteError(
                "No se puede eliminar un lote con comprobantes emitidos o inciertos"
            )

        self._registrar_evento_lote(
            lote_id=lote_id,
            accion="eliminar_lote",
            usuario_id=usuario_id,
            motivo=motivo,
            metadata_json={
                "lote_id_original": lote.id,
                "nombre_archivo": lote.nombre_archivo,
                "archivo_hash": lote.archivo_hash,
                "empresa_id": lote.empresa_id,
                "estado": lote.estado,
                "total_grupos": lote.total_grupos,
            },
        )
        await self.db.flush()
        await self.db.execute(
            update(LoteComprobanteEvento)
            .where(LoteComprobanteEvento.lote_id == lote_id)
            .values(lote_id=None, grupo_id=None)
        )
        await self.db.delete(lote)
        await self.db.commit()

    async def compactar_lote(
        self,
        lote_id: int,
        empresa_id: int,
        usuario_id: int | None,
        commit: bool = True,
    ) -> LoteComprobante:
        """Elimina filas detalladas de un lote cerrado conservando grupos y resumen."""
        lote = await self.obtener_lote_resumen(lote_id, empresa_id)
        if lote.estado not in {
            "completado",
            "cerrado_reconciliado",
            "cerrado_con_descartes",
        }:
            raise LoteComprobanteError("Solo se pueden compactar lotes cerrados")
        if lote.compactado_at is not None:
            return lote

        result = await self.db.execute(
            delete(LoteComprobanteFila).where(LoteComprobanteFila.lote_id == lote_id)
        )
        lote.compactado_at = datetime.utcnow()
        self._registrar_evento_lote(
            lote_id=lote_id,
            accion="compactar_lote",
            usuario_id=usuario_id,
            motivo="Compactación para ahorro de almacenamiento",
            metadata_json={"filas_eliminadas": result.rowcount or 0},
        )
        if commit:
            await self.db.commit()
            await self.db.refresh(lote)
        else:
            await self.db.flush()
        return lote

    def _validar_lote_resoluble(
        self,
        lote: LoteComprobante,
        permitir_reconciliacion_tecnica: bool = False,
    ) -> None:
        """Bloquea acciones manuales sobre lotes activos, cerrados o inciertos."""
        if lote.estado in {"en_cola", "procesando"}:
            raise LoteComprobanteError("El lote está en proceso y no puede resolverse")
        if (
            lote.estado == "requiere_reconciliacion"
            and not permitir_reconciliacion_tecnica
        ):
            raise LoteComprobanteError(
                "El lote requiere reconciliación técnica antes de resolver pendientes"
            )
        if lote.estado in {
            "completado",
            "cerrado_reconciliado",
            "cerrado_con_descartes",
        }:
            raise LoteComprobanteError("El lote ya está cerrado")

    async def _obtener_grupos_para_resolver(
        self,
        lote_id: int,
        estados: set[str],
        grupo_ids: list[int] | None,
    ) -> list[LoteComprobanteGrupo]:
        """Obtiene grupos de un lote que pueden recibir una acción resolutiva."""
        filtros = [
            LoteComprobanteGrupo.lote_id == lote_id,
            LoteComprobanteGrupo.estado.in_(estados),
        ]
        if grupo_ids:
            filtros.append(LoteComprobanteGrupo.id.in_(grupo_ids))
        result = await self.db.execute(
            select(LoteComprobanteGrupo)
            .options(selectinload(LoteComprobanteGrupo.filas))
            .where(*filtros)
            .order_by(LoteComprobanteGrupo.orden, LoteComprobanteGrupo.id)
        )
        grupos = list(result.scalars().all())
        if grupo_ids and {grupo.id for grupo in grupos} != set(grupo_ids):
            raise LoteComprobanteError(
                "La selección contiene comprobantes que no están pendientes"
            )
        return grupos

    async def _obtener_grupo_para_resolver(
        self,
        lote_id: int,
        grupo_id: int,
        estados: set[str],
    ) -> LoteComprobanteGrupo:
        """Obtiene un único grupo resoluble por id."""
        grupos = await self._obtener_grupos_para_resolver(
            lote_id=lote_id,
            estados=estados,
            grupo_ids=[grupo_id],
        )
        if not grupos:
            raise LoteComprobanteError("No se encontró el comprobante pendiente")
        return grupos[0]

    async def _reclamar_grupo_para_reintento(
        self,
        lote_id: int,
        grupo_id: int,
    ) -> LoteComprobanteGrupo | None:
        """Toma un grupo fallido antes de solicitar CAE para evitar doble emisión."""
        result = await self.db.execute(
            update(LoteComprobanteGrupo)
            .where(
                LoteComprobanteGrupo.id == grupo_id,
                LoteComprobanteGrupo.lote_id == lote_id,
                LoteComprobanteGrupo.estado == "fallido",
            )
            .values(
                estado="reintentando",
                mensajes_json=["Reintento de emisión en curso."],
            )
        )
        if result.rowcount != 1:
            return None

        await self.db.flush()
        await self.db.commit()
        return await self._obtener_grupo_para_resolver(
            lote_id=lote_id,
            grupo_id=grupo_id,
            estados={"reintentando"},
        )

    async def _reclamar_grupo_para_accion_resolutiva(
        self,
        lote_id: int,
        grupo_id: int,
        estados: set[str],
        estado_claim: str,
        mensaje_error: str,
    ) -> LoteComprobanteGrupo:
        """Toma un grupo con transición condicional antes de resolverlo."""
        result = await self.db.execute(
            update(LoteComprobanteGrupo)
            .where(
                LoteComprobanteGrupo.id == grupo_id,
                LoteComprobanteGrupo.lote_id == lote_id,
                LoteComprobanteGrupo.estado.in_(estados),
            )
            .values(estado=estado_claim)
        )
        if result.rowcount != 1:
            raise LoteComprobanteError(mensaje_error)

        await self.db.flush()
        return await self._obtener_grupo_para_resolver(
            lote_id=lote_id,
            grupo_id=grupo_id,
            estados={estado_claim},
        )

    async def _marcar_reintento_fallido(
        self,
        lote_id: int,
        empresa_id: int,
        usuario_id: int | None,
        grupo: LoteComprobanteGrupo,
        exc: Exception,
    ) -> None:
        """Persiste un reintento fallido y actualiza el resumen del lote."""
        logger.exception("Error reintentando grupo fallido %s", grupo.id)
        grupo.estado = "fallido"
        grupo.mensajes_json = [str(exc)]
        await self._marcar_filas(grupo, "fallido", grupo.mensajes_json)
        self._registrar_evento_lote(
            lote_id=lote_id,
            accion="reintentar_fallido",
            usuario_id=usuario_id,
            grupo_id=grupo.id,
            metadata_json={"resultado": "fallido", "error": str(exc)},
        )
        await self.db.flush()
        lote = await self.obtener_lote_resumen(lote_id, empresa_id)
        await self._actualizar_estado_lote(lote)
        await self.db.commit()

    @staticmethod
    def _validar_reconciliacion_sin_numeros_duplicados(
        comprobantes: list[dict[str, Any]],
    ) -> None:
        """Evita cerrar más de un grupo contra el mismo comprobante externo."""
        claves_vistas: set[tuple[int, int, int]] = set()
        for item in comprobantes:
            try:
                clave = (
                    int(item["tipo_comprobante"]),
                    int(item["punto_venta_numero"]),
                    int(item["numero"]),
                )
            except (KeyError, TypeError, ValueError) as exc:
                raise LoteComprobanteError(
                    "Los datos de comprobante externo no son válidos"
                ) from exc

            if clave in claves_vistas:
                raise LoteComprobanteError(
                    "No podés reconciliar dos grupos con el mismo comprobante externo"
                )
            claves_vistas.add(clave)

    def _validar_datos_reconciliacion_externa(
        self,
        grupo: LoteComprobanteGrupo,
        request: EmitirComprobanteRequest,
        item: dict[str, Any],
    ) -> None:
        """Confirma que el usuario está reconciliando el grupo correcto."""
        fecha = self._parse_date(item.get("fecha_emision"))
        total = self._parse_decimal(item.get("total"))
        if total is None:
            raise LoteComprobanteError("El total informado no es válido")
        total = self._redondear_centavos(total)
        if int(item["tipo_comprobante"]) != request.tipo_comprobante:
            raise LoteComprobanteError("El tipo informado no coincide con el lote")
        if int(item["punto_venta_numero"]) != int(grupo.punto_venta_numero or 0):
            raise LoteComprobanteError(
                "El punto de venta informado no coincide con el lote"
            )
        if fecha != request.fecha_emision:
            raise LoteComprobanteError("La fecha informada no coincide con el lote")
        if total != self._redondear_centavos(Decimal(str(grupo.total_estimado))):
            raise LoteComprobanteError("El total informado no coincide con el lote")

    def _validar_consulta_arca_externa(
        self,
        empresa: Empresa,
        grupo: LoteComprobanteGrupo,
        request: EmitirComprobanteRequest,
        item: dict[str, Any],
        consulta_arca,
    ) -> None:
        """Valida que ARCA confirme exactamente el comprobante externo."""
        fecha_esperada = self._parse_date(item.get("fecha_emision"))
        total_esperado = self._parse_decimal(item.get("total"))
        if total_esperado is None:
            raise LoteComprobanteError("El total informado no es válido")
        total_esperado = self._redondear_centavos(total_esperado)
        fecha_arca = self._parse_arca_yyyymmdd(str(consulta_arca.fecha_cbte))
        total_arca = self._redondear_centavos(Decimal(str(consulta_arca.imp_total)))
        cae_informado = str(item.get("cae") or "").strip()
        try:
            tipo_doc_arca = int(consulta_arca.tipo_doc)
            tipo_doc_esperado = int(request.tipo_documento)
        except (TypeError, ValueError) as exc:
            raise LoteComprobanteError(
                "ARCA no devolvió un tipo de documento de receptor válido"
            ) from exc
        nro_doc_arca = clean_cuit(str(consulta_arca.nro_doc or ""))
        nro_doc_esperado = clean_cuit(str(request.numero_documento or ""))

        if consulta_arca.resultado != "A":
            raise LoteComprobanteError("ARCA no informa el comprobante como autorizado")
        if not str(consulta_arca.cae or "").strip():
            raise LoteComprobanteError("ARCA no devolvió un CAE válido")
        if not str(consulta_arca.cae_vencimiento or "").strip():
            raise LoteComprobanteError("ARCA no devolvió vencimiento de CAE válido")
        if clean_cuit(str(consulta_arca.cuit_emisor)) != clean_cuit(empresa.cuit):
            raise LoteComprobanteError(
                "El comprobante consultado no pertenece al emisor activo"
            )
        if int(consulta_arca.tipo_cbte) != int(item["tipo_comprobante"]):
            raise LoteComprobanteError("ARCA devolvió otro tipo de comprobante")
        if int(consulta_arca.punto_venta) != int(item["punto_venta_numero"]):
            raise LoteComprobanteError("ARCA devolvió otro punto de venta")
        if int(consulta_arca.numero) != int(item["numero"]):
            raise LoteComprobanteError("ARCA devolvió otro número de comprobante")
        if tipo_doc_arca != tipo_doc_esperado:
            raise LoteComprobanteError(
                "El tipo de documento del receptor informado por ARCA no coincide con el lote"
            )
        if nro_doc_arca != nro_doc_esperado:
            raise LoteComprobanteError(
                "El número de documento del receptor informado por ARCA no coincide con el lote"
            )
        if fecha_arca != fecha_esperada:
            raise LoteComprobanteError(
                "La fecha fiscal de ARCA no coincide con el lote"
            )
        if total_arca != total_esperado:
            raise LoteComprobanteError(
                "El total autorizado en ARCA no coincide con el lote"
            )
        if cae_informado and cae_informado != str(consulta_arca.cae):
            raise LoteComprobanteError("El CAE informado no coincide con ARCA")
        if total_esperado != self._redondear_centavos(
            Decimal(str(grupo.total_estimado))
        ):
            raise LoteComprobanteError("El total informado no coincide con el grupo")

    async def _crear_o_vincular_comprobante_externo(
        self,
        empresa: Empresa,
        request: EmitirComprobanteRequest,
        punto_venta_numero: int,
        numero: int,
        cae: str,
        cae_vencimiento: str,
    ) -> Comprobante:
        """Crea o vincula el comprobante local emitido externamente."""
        punto_venta = await self._obtener_punto_venta_por_numero(
            empresa_id=empresa.id,
            numero=punto_venta_numero,
        )
        if punto_venta is None:
            raise LoteComprobanteError(
                "El punto de venta externo no está cargado para el emisor activo"
            )
        totales = self.facturacion_service._calcular_totales(request.items)
        total_request = self._redondear_centavos(totales["total"])

        result = await self.db.execute(
            select(Comprobante).where(
                Comprobante.empresa_id == empresa.id,
                Comprobante.punto_venta_id == punto_venta.id,
                Comprobante.tipo_comprobante == request.tipo_comprobante,
                Comprobante.numero == numero,
            )
        )
        existente = result.scalar_one_or_none()
        if existente is not None:
            grupo_vinculado_id = await self.db.scalar(
                select(LoteComprobanteGrupo.id)
                .where(LoteComprobanteGrupo.comprobante_id == existente.id)
                .limit(1)
            )
            if grupo_vinculado_id is not None:
                raise LoteComprobanteError(
                    "El comprobante externo ya está vinculado a otro grupo del lote"
                )
            if (
                existente.estado != "autorizado"
                or existente.cae != cae
                or existente.fecha_emision != request.fecha_emision
                or self._redondear_centavos(Decimal(str(existente.total)))
                != total_request
            ):
                raise LoteComprobanteError(
                    "Ya existe un comprobante local con ese número y datos distintos"
                )
            return existente

        resultado_arca = ResultadoArcaExterno(
            cae=cae,
            cae_vencimiento=cae_vencimiento,
        )
        return await self.facturacion_service._guardar_comprobante(
            request,
            numero,
            totales,
            resultado_arca,
            punto_venta,
            origen_emision="arca_web",
            commit=False,
        )

    async def _obtener_punto_venta_por_numero(
        self, empresa_id: int, numero: int
    ) -> PuntoVenta | None:
        """Busca un punto de venta del emisor por número."""
        result = await self.db.execute(
            select(PuntoVenta).where(
                PuntoVenta.empresa_id == empresa_id,
                PuntoVenta.numero == numero,
            )
        )
        return result.scalar_one_or_none()

    async def _lote_tiene_emision_o_incertidumbre(self, lote_id: int) -> bool:
        """Indica si un lote ya tiene evidencia fiscal que impide borrado físico."""
        result = await self.db.execute(
            select(func.count())
            .select_from(LoteComprobanteGrupo)
            .where(
                LoteComprobanteGrupo.lote_id == lote_id,
                (
                    LoteComprobanteGrupo.estado.in_(
                        {
                            "autorizado",
                            "autorizado_externo",
                            "requiere_reconciliacion",
                            "reintentando",
                        }
                    )
                    | LoteComprobanteGrupo.comprobante_id.is_not(None)
                    | LoteComprobanteGrupo.cae.is_not(None)
                ),
            )
        )
        return int(result.scalar_one() or 0) > 0

    def _registrar_evento_lote(
        self,
        lote_id: int,
        accion: str,
        usuario_id: int | None,
        motivo: str | None = None,
        grupo_id: int | None = None,
        metadata_json: dict[str, Any] | None = None,
    ) -> None:
        """Agrega un evento auditable del lote en la sesión actual."""
        self.db.add(
            LoteComprobanteEvento(
                lote_id=lote_id,
                grupo_id=grupo_id,
                usuario_id=usuario_id,
                accion=accion,
                motivo=motivo,
                metadata_json=metadata_json,
            )
        )

    @staticmethod
    def _parse_arca_yyyymmdd(value: str) -> date:
        """Convierte una fecha ARCA `YYYYMMDD` a `date`."""
        return date(int(value[0:4]), int(value[4:6]), int(value[6:8]))

    def _crear_mensaje_confirmacion_fecha_fiscal(
        self, fechas: list[str], puntos_venta: list[int]
    ) -> str:
        """Construye el texto de confirmación fiscal irreversible."""
        if not fechas:
            return "No hay comprobantes pendientes con fecha fiscal para confirmar."
        fechas_label = ", ".join(
            date.fromisoformat(fecha[:10]).strftime("%d/%m/%y") for fecha in fechas
        )
        fecha_texto = (
            f"fecha {fechas_label}" if len(fechas) == 1 else f"fechas {fechas_label}"
        )
        punto_texto = ""
        if puntos_venta:
            punto_label = ", ".join(f"{punto:04d}" for punto in puntos_venta)
            punto_texto = (
                f" para el punto de venta {punto_label}"
                if len(puntos_venta) == 1
                else f" para los puntos de venta {punto_label}"
            )
        return (
            f"Está seguro que quiere emitir comprobantes con {fecha_texto}{punto_texto}? "
            "Recuerde que luego no podrá emitir comprobantes con fecha anterior para ese mismo punto de venta."
        )

    def _calcular_totales_payloads(
        self, filas: list[tuple[dict[str, Any] | None, int | None, Decimal]]
    ) -> dict[str, Decimal | int]:
        """Calcula totales fiscales desde los payloads ya validados."""
        totales: dict[str, Decimal | int] = {
            "comprobantes": 0,
            "neto": Decimal("0"),
            "iva21": Decimal("0"),
            "iva105": Decimal("0"),
            "total": Decimal("0"),
            "valores_invalidos": 0,
        }
        for payload, _, total_estimado in filas:
            payload = payload or {}
            items = payload.get("items") if isinstance(payload, dict) else None
            if not isinstance(items, list):
                totales["valores_invalidos"] = int(totales["valores_invalidos"]) + 1
                continue

            neto_grupo = Decimal("0")
            iva21_grupo = Decimal("0")
            iva105_grupo = Decimal("0")
            for item in items:
                if not isinstance(item, dict):
                    totales["valores_invalidos"] = int(totales["valores_invalidos"]) + 1
                    continue
                cantidad = self._parse_decimal(item.get("cantidad"), Decimal("0"))
                precio = self._parse_decimal(item.get("precio_unitario"), Decimal("0"))
                descuento = self._parse_decimal(
                    item.get("descuento_porcentaje"), Decimal("0")
                )
                iva = self._parse_decimal(item.get("iva_porcentaje"), Decimal("0"))
                if None in {cantidad, precio, descuento, iva}:
                    totales["valores_invalidos"] = int(totales["valores_invalidos"]) + 1
                    continue
                bruto = cantidad * precio
                neto_item = bruto - bruto * (descuento / Decimal("100"))
                neto_grupo += neto_item
                if iva == Decimal("21"):
                    iva21_grupo += neto_item * Decimal("0.21")
                elif iva == Decimal("10.5"):
                    iva105_grupo += neto_item * Decimal("0.105")

            totales["comprobantes"] = int(totales["comprobantes"]) + 1
            totales["neto"] = self._round_money(Decimal(totales["neto"]) + neto_grupo)
            totales["iva21"] = self._round_money(
                Decimal(totales["iva21"]) + iva21_grupo
            )
            totales["iva105"] = self._round_money(
                Decimal(totales["iva105"]) + iva105_grupo
            )
            total_grupo = total_estimado or self._round_money(
                neto_grupo + iva21_grupo + iva105_grupo
            )
            totales["total"] = self._round_money(
                Decimal(totales["total"]) + total_grupo
            )
        return totales

    @staticmethod
    def _round_money(value: Decimal) -> Decimal:
        """Redondea importes a centavos."""
        return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    async def procesar_lote(
        self,
        lote_id: int,
        empresa_id: int,
        reanudar: bool = False,
        operacion_id: int | None = None,
        usuario_id: int | None = None,
        confirmacion_duplicado_logico: bool = False,
        fase_solicitud_arca: FaseSolicitudArca | None = None,
    ) -> LoteComprobante:
        """Procesa un lote válido, en forma sincrónica o desde background."""
        fase_solicitud_arca = fase_solicitud_arca or FaseSolicitudArca()
        lote = await self.obtener_lote_resumen(lote_id, empresa_id)
        if operacion_id is None:
            operacion_id = (lote.metadata_json or {}).get("operacion_idempotente_id")
        if not confirmacion_duplicado_logico:
            confirmacion_duplicado_logico = bool(
                (lote.metadata_json or {}).get("confirmacion_duplicado_logico")
            )
        metadata_actual = lote.metadata_json or {}
        if operacion_id is not None and (
            metadata_actual.get("operacion_idempotente_id") != operacion_id
            or bool(metadata_actual.get("confirmacion_duplicado_logico"))
            != confirmacion_duplicado_logico
        ):
            metadata = dict(lote.metadata_json or {})
            metadata["operacion_idempotente_id"] = operacion_id
            metadata["confirmacion_duplicado_logico"] = confirmacion_duplicado_logico
            lote.metadata_json = metadata
            await self.db.flush()

        if lote.estado == "procesando":
            if reanudar:
                stale_before = datetime.utcnow() - timedelta(
                    minutes=settings.batch_processing_stale_minutes
                )
                if lote.updated_at < stale_before:
                    return await self.bloquear_lote_procesando_stale(
                        lote_id,
                        empresa_id,
                        motivo=(
                            "El worker detectó que el lote seguía procesando después "
                            "de la ventana segura y lo bloqueó para reconciliación "
                            "antes de cualquier nuevo CAE."
                        ),
                    )
            return lote
        if lote.estado in self.ESTADOS_TERMINALES:
            if reanudar and operacion_id is not None:
                await self._guardar_respuesta_operacion_background(
                    lote,
                    operacion_id,
                )
            return lote
        if lote.estado not in self.ESTADOS_PROCESABLES:
            raise LoteComprobanteError(
                "El lote debe estar validado o en cola antes de emitir"
            )
        if not (lote.metadata_json or {}).get("opciones_concepto"):
            raise LoteComprobanteError(
                "Este lote fue validado antes de confirmar el concepto fiscal. Revalidá el archivo eligiendo Productos, Servicios o Definido por archivo antes de emitir."
            )
        if not (lote.metadata_json or {}).get("opciones_descripcion_item"):
            raise LoteComprobanteError(
                "Este lote fue validado antes de confirmar la descripción facturada. Revalidá el archivo eligiendo descripción desde archivo o una descripción fija antes de emitir."
            )

        procesamiento_async = (
            lote.procesamiento_async
            or lote.modo_procesamiento == "background"
            or lote.total_grupos > settings.batch_sync_limit
        )
        modo_procesamiento = "background" if procesamiento_async else "sincronico"
        await self._tomar_lote_para_procesamiento(
            lote_id=lote_id,
            empresa_id=empresa_id,
            procesamiento_async=procesamiento_async,
            modo_procesamiento=modo_procesamiento,
            reanudar=reanudar,
        )
        await self.db.commit()
        lote = await self.obtener_lote_resumen(lote_id, empresa_id)
        logger.info(
            "Procesando lote %s de empresa %s en modo %s",
            lote.id,
            empresa_id,
            lote.modo_procesamiento,
        )

        config_batch = await self._resolver_configuracion_batch_arca(lote, empresa_id)
        self._registrar_metadata_batch_arca(lote, config_batch)
        await self.db.commit()

        grupos = await self._obtener_grupos_emitibles(lote_id)
        pendientes: list[GrupoPendienteEmision] = []
        for grupo in grupos:
            try:
                payload = grupo.payload_json or {}
                request = EmitirComprobanteRequest.model_validate(payload)
                request = request.model_copy(
                    update={
                        "confirmacion_duplicado_logico": (confirmacion_duplicado_logico)
                    }
                )
                if reanudar and await self._reconciliar_grupo_autorizado_existente(
                    grupo, request
                ):
                    await self.db.flush()
                    await self._actualizar_progreso_lote(lote_id)
                    await self.db.commit()
                    continue

                pendientes.append(GrupoPendienteEmision(grupo=grupo, request=request))
            except DATABASE_TEMPORARILY_UNAVAILABLE_ERRORS:
                raise
            except Exception as exc:  # pragma: no cover - fallback defensivo
                logger.exception("Error procesando grupo %s", grupo.comprobante_ref)
                grupo.estado = "fallido"
                grupo.mensajes_json = [str(exc)]
                await self._marcar_filas(grupo, "fallido", grupo.mensajes_json)
                await self.db.flush()
                await self._actualizar_progreso_lote(lote_id)
                await self.db.commit()

        for sublote in self._iterar_sublotes_emision(pendientes, config_batch):
            if self._sublote_requiere_emision_unitaria(sublote, config_batch):
                for pendiente in sublote:
                    try:
                        resultado = await self.facturacion_service.emitir_comprobante(
                            pendiente.request,
                            operacion_id=operacion_id,
                            usuario_id=usuario_id,
                            lote_id=lote_id,
                            grupo_id=pendiente.grupo.id,
                            fase_solicitud_arca=fase_solicitud_arca,
                        )
                        await self._aplicar_resultado_emision_grupo(
                            pendiente.grupo,
                            resultado,
                        )
                    except DATABASE_TEMPORARILY_UNAVAILABLE_ERRORS:
                        raise
                    except Exception as exc:  # pragma: no cover - fallback defensivo
                        logger.exception(
                            "Error procesando grupo %s",
                            pendiente.grupo.comprobante_ref,
                        )
                        pendiente.grupo.estado = "fallido"
                        pendiente.grupo.mensajes_json = [str(exc)]
                        await self._marcar_filas(
                            pendiente.grupo,
                            "fallido",
                            pendiente.grupo.mensajes_json,
                        )

                    await self.db.flush()
                    await self._actualizar_progreso_lote(lote_id)
                    await self.db.commit()
                continue

            try:
                resultados = await self.facturacion_service.emitir_comprobantes_lote(
                    [pendiente.request for pendiente in sublote],
                    max_registros=config_batch.chunk_size,
                    contextos=[
                        {
                            "operacion_id": operacion_id,
                            "usuario_id": usuario_id,
                            "lote_id": lote_id,
                            "grupo_id": pendiente.grupo.id,
                        }
                        for pendiente in sublote
                    ],
                    fase_solicitud_arca=fase_solicitud_arca,
                )
            except DATABASE_TEMPORARILY_UNAVAILABLE_ERRORS:
                raise
            except Exception as exc:  # pragma: no cover - fallback defensivo
                logger.exception("Error procesando sublote ARCA")
                resultados = [
                    EmitirComprobanteResponse(
                        exito=False,
                        tipo_comprobante=pendiente.request.tipo_comprobante,
                        punto_venta=0,
                        numero=0,
                        fecha=pendiente.request.fecha_emision,
                        total=Decimal("0"),
                        mensaje="Error inesperado",
                        errores=[str(exc)],
                    )
                    for pendiente in sublote
                ]

            if len(resultados) < len(sublote):
                resultados.extend(
                    [
                        EmitirComprobanteResponse(
                            exito=False,
                            tipo_comprobante=pendiente.request.tipo_comprobante,
                            punto_venta=0,
                            numero=0,
                            fecha=pendiente.request.fecha_emision,
                            total=Decimal("0"),
                            mensaje="Error inesperado",
                            errores=[
                                "FactuFlow no recibió respuesta para este comprobante dentro del sublote."
                            ],
                            requiere_reconciliacion=True,
                            categoria_error="arca_batch_respuesta_incompleta",
                        )
                        for pendiente in sublote[len(resultados) :]
                    ]
                )

            for pendiente, resultado in zip(sublote, resultados):
                await self._aplicar_resultado_emision_grupo(
                    pendiente.grupo,
                    resultado,
                )
                await self.db.flush()
                await self._actualizar_progreso_lote(lote_id)
                await self.db.commit()

        lote = await self.obtener_lote_resumen(lote_id, empresa_id)
        lote.finished_at = datetime.utcnow()
        lote.estado = "cargado"
        await self._actualizar_estado_lote(lote)
        self._aplicar_aviso_batch_arca(lote)
        await self.db.commit()
        await self.db.refresh(lote)
        if reanudar and operacion_id is not None:
            await self._guardar_respuesta_operacion_background(
                lote,
                operacion_id,
            )
        logger.info(
            "Lote %s finalizado con estado %s: emitidos=%s fallidos=%s",
            lote.id,
            lote.estado,
            lote.grupos_emitidos,
            lote.grupos_fallidos,
        )
        return lote

    async def _guardar_respuesta_operacion_background(
        self,
        lote: LoteComprobante,
        operacion_id: int,
    ) -> None:
        """Actualiza la operación idempotente de un lote procesado por worker."""
        result = await self.db.execute(
            select(OperacionIdempotente).where(
                OperacionIdempotente.id == operacion_id,
                OperacionIdempotente.empresa_id == lote.empresa_id,
            )
        )
        operacion = result.scalar_one_or_none()
        if operacion is None:
            return

        respuesta = LoteProcesamientoResponse(
            lote=LoteComprobanteResponse.model_validate(lote),
            mensaje=lote.mensaje_resumen or "Lote procesado",
            en_progreso=lote.estado in {"en_cola", "procesando"},
        )
        estado = "finalizado"
        if lote.estado == "requiere_reconciliacion":
            estado = "requiere_reconciliacion"
        elif respuesta.en_progreso:
            estado = "en_proceso"

        await IdempotenciaFiscalService(self.db).guardar_respuesta_operacion(
            operacion,
            response_json=respuesta,
            estado=estado,
        )

    async def _tomar_lote_para_procesamiento(
        self,
        lote_id: int,
        empresa_id: int,
        procesamiento_async: bool,
        modo_procesamiento: str,
        reanudar: bool = False,
    ) -> None:
        """Marca un lote como procesando con transición atómica."""
        estados_tomables = self.ESTADOS_PROCESABLES
        estado_condition = LoteComprobante.estado.in_(estados_tomables)

        result = await self.db.execute(
            update(LoteComprobante)
            .where(
                LoteComprobante.id == lote_id,
                LoteComprobante.empresa_id == empresa_id,
                estado_condition,
            )
            .values(
                estado="procesando",
                procesamiento_async=procesamiento_async,
                modo_procesamiento=modo_procesamiento,
                started_at=datetime.utcnow(),
                mensaje_resumen="Procesando comprobantes...",
            )
        )
        if result.rowcount != 1:
            raise LoteComprobanteError(
                "El lote ya está siendo procesado o fue procesado previamente."
            )

    async def _actualizar_progreso_lote(self, lote_id: int) -> None:
        """Persiste los contadores parciales mientras un lote se procesa."""
        result = await self.db.execute(
            select(LoteComprobante).where(LoteComprobante.id == lote_id)
        )
        lote = result.scalar_one()
        await self._actualizar_estado_lote(lote)

    async def _resolver_configuracion_batch_arca(
        self,
        lote: LoteComprobante,
        empresa_id: int,
    ) -> ConfiguracionBatchArca:
        """Resuelve el tamaño de sublote ARCA que se usará en el procesamiento."""
        if not settings.arca_fecaesolicitar_batch_enabled:
            return ConfiguracionBatchArca(
                habilitado=False,
                reg_x_req=None,
                chunk_size=1,
            )

        try:
            reg_x_req = (
                await self.facturacion_service.obtener_registros_maximos_por_request(
                    empresa_id
                )
            )
            if reg_x_req < 1:
                raise LoteComprobanteError("ARCA devolvió RegXReq menor a 1")

            limite_configurado = settings.arca_fecaesolicitar_batch_max_registros
            chunk_size = reg_x_req
            if limite_configurado > 0:
                chunk_size = min(chunk_size, limite_configurado)

            return ConfiguracionBatchArca(
                habilitado=chunk_size > 1,
                reg_x_req=reg_x_req,
                chunk_size=max(1, chunk_size),
            )
        except DATABASE_TEMPORARILY_UNAVAILABLE_ERRORS:
            raise
        except Exception as exc:
            logger.warning(
                "No se pudo consultar RegXReq para lote %s; se degrada a modo unitario. error=%s",
                lote.id,
                str(exc),
            )
            return ConfiguracionBatchArca(
                habilitado=False,
                reg_x_req=None,
                chunk_size=1,
                fallback_unitario=True,
                fallback_motivo=str(exc),
            )

    @staticmethod
    def _registrar_metadata_batch_arca(
        lote: LoteComprobante,
        config: ConfiguracionBatchArca,
    ) -> None:
        """Guarda metadata operativa sobre el modo de emisión ARCA usado."""
        metadata = dict(lote.metadata_json or {})
        if config.fallback_unitario:
            modo = "unitario_fallback"
        elif config.habilitado:
            modo = "batch"
        else:
            modo = "unitario"

        metadata["arca_batch"] = {
            "modo": modo,
            "reg_x_req": config.reg_x_req,
            "chunk_size": config.chunk_size,
            "fallback_unitario": config.fallback_unitario,
            "fallback_motivo": config.fallback_motivo,
        }
        lote.metadata_json = metadata

    @staticmethod
    def _aplicar_aviso_batch_arca(lote: LoteComprobante) -> None:
        """Agrega al resumen un aviso si el lote debió degradar a modo unitario."""
        arca_batch = (lote.metadata_json or {}).get("arca_batch") or {}
        if not arca_batch.get("fallback_unitario"):
            return

        aviso = (
            "ARCA no informó la capacidad máxima por request; FactuFlow emitió "
            "en modo unitario para no bloquear la operación."
        )
        if not lote.mensaje_resumen:
            lote.mensaje_resumen = aviso
        elif aviso not in lote.mensaje_resumen:
            lote.mensaje_resumen = f"{lote.mensaje_resumen} {aviso}"

    @staticmethod
    def _iterar_sublotes_emision(
        pendientes: list[GrupoPendienteEmision],
        config: ConfiguracionBatchArca,
    ) -> list[list[GrupoPendienteEmision]]:
        """Divide grupos pendientes por punto, tipo y tamaño efectivo."""
        grupos_por_clave: dict[tuple[int, int], list[GrupoPendienteEmision]] = {}
        orden_claves: list[tuple[int, int]] = []
        for pendiente in pendientes:
            clave = (
                pendiente.request.punto_venta_id,
                pendiente.request.tipo_comprobante,
            )
            if clave not in grupos_por_clave:
                grupos_por_clave[clave] = []
                orden_claves.append(clave)
            grupos_por_clave[clave].append(pendiente)

        sublotes: list[list[GrupoPendienteEmision]] = []
        for clave in orden_claves:
            grupo_clave = grupos_por_clave[clave]
            primer = grupo_clave[0].request
            chunk_size = config.chunk_size
            if (
                not config.habilitado
                or config.fallback_unitario
                or primer.tipo_comprobante
                in FacturacionService.TIPOS_COMPROBANTE_FCE_MIPYME
            ):
                chunk_size = 1

            for index in range(0, len(grupo_clave), chunk_size):
                sublotes.append(grupo_clave[index : index + chunk_size])

        return sublotes

    @staticmethod
    def _sublote_requiere_emision_unitaria(
        sublote: list[GrupoPendienteEmision],
        config: ConfiguracionBatchArca,
    ) -> bool:
        """Indica si el sublote debe usar el flujo unitario existente."""
        if len(sublote) <= 1:
            return True
        if not config.habilitado or config.fallback_unitario:
            return True
        tipo_comprobante = sublote[0].request.tipo_comprobante
        return tipo_comprobante in FacturacionService.TIPOS_COMPROBANTE_FCE_MIPYME

    async def _aplicar_resultado_emision_grupo(
        self,
        grupo: LoteComprobanteGrupo,
        resultado: EmitirComprobanteResponse,
    ) -> None:
        """Actualiza grupo y filas según la respuesta de emisión."""
        if resultado.exito:
            grupo.estado = "autorizado"
            grupo.cae = resultado.cae
            grupo.numero_asignado = resultado.numero
            grupo.comprobante_id = resultado.comprobante_id
            grupo.mensajes_json = [
                resultado.mensaje,
                f"CAE {resultado.cae}",
            ]
            await self._marcar_filas(grupo, "autorizado", grupo.mensajes_json)
        elif resultado.requiere_reconciliacion:
            grupo.estado = "requiere_reconciliacion"
            grupo.cae = resultado.cae
            grupo.numero_asignado = resultado.numero
            grupo.mensajes_json = resultado.errores or [resultado.mensaje]
            await self._marcar_filas(
                grupo,
                "requiere_reconciliacion",
                grupo.mensajes_json,
            )
        else:
            grupo.estado = "fallido"
            grupo.mensajes_json = resultado.errores or [resultado.mensaje]
            await self._marcar_filas(grupo, "fallido", grupo.mensajes_json)

    @staticmethod
    def _sanitizar_valor_excel_observado(value: Any) -> Any:
        """Evita que Excel interprete texto observado como formula."""
        if not isinstance(value, str) or value == "":
            return value

        if value[0] in {"=", "+", "-", "@", "\t", "\r"}:
            return f"'{value}"

        stripped = value.lstrip()
        if stripped and stripped[0] in {"=", "+", "-", "@"}:
            return f"'{value}"

        return value

    async def generar_archivo_observado(self, lote_id: int, empresa_id: int) -> bytes:
        """Genera un archivo observado con resultados por fila."""
        lote = await self.obtener_lote(lote_id, empresa_id)
        if lote.compactado_at is not None:
            raise LoteComprobanteError(
                "El lote fue compactado y ya no conserva el detalle de filas para generar el observado."
            )
        filas_por_numero = {fila.fila_excel: fila for fila in lote.filas}
        workbook = Workbook()
        hoja = workbook.active
        hoja.title = "Resultados"
        salida_headers = self.TEMPLATE_COLUMNS + [
            "resultado_estado",
            "resultado_mensajes",
        ]
        hoja.append(salida_headers)
        for cell in hoja[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")

        for fila_numero in sorted(filas_por_numero):
            fila = filas_por_numero[fila_numero]
            row = [
                self._sanitizar_valor_excel_observado(fila.datos_json.get(col))
                for col in self.TEMPLATE_COLUMNS
            ]
            row.extend(
                [
                    self._sanitizar_valor_excel_observado(fila.estado),
                    self._sanitizar_valor_excel_observado(
                        " | ".join(fila.mensajes_json or [])
                    ),
                ]
            )
            hoja.append(row)
            if fila.estado in {"autorizado", "autorizado_externo"}:
                fill_color = "E2F0D9"
            elif fila.estado == "descartado":
                fill_color = "E7E6E6"
            elif fila.estado == "requiere_reconciliacion":
                fill_color = "FFF2CC"
            else:
                fill_color = "FDE9D9"
            for cell in hoja[hoja.max_row]:
                cell.fill = PatternFill("solid", fgColor=fill_color)

        stream = BytesIO()
        workbook.save(stream)
        return stream.getvalue()

    async def _leer_excel_para_lote(
        self,
        file_bytes: bytes,
        empresa: Empresa,
        formato_version_id: int | None,
    ) -> dict[str, Any]:
        """Lee una plantilla oficial o aplica un formato configurable."""
        formatos_service = FormatosImportacionService(self.db)

        if formato_version_id is not None:
            try:
                version = await formatos_service.obtener_version(
                    formato_version_id,
                    empresa.id,
                )
                importacion = await formatos_service.importar_con_version(
                    file_bytes,
                    empresa,
                    version,
                )
            except FormatoImportacionError as exc:
                raise LoteComprobanteError(str(exc)) from exc

            return self._serializar_importacion_configurable(importacion)

        try:
            headers = formatos_service.leer_headers(file_bytes)
        except FormatoImportacionError as exc:
            raise LoteComprobanteError(str(exc)) from exc

        if formatos_service.es_plantilla_oficial(headers, self.TEMPLATE_COLUMNS):
            return {
                "filas": self._leer_excel(file_bytes),
                "headers_detectados": headers,
                "mapeo_usado": {
                    "tipo": "plantilla_oficial",
                    "campos": {
                        column: {"origen": "header", "header": column}
                        for column in self.TEMPLATE_COLUMNS
                    },
                },
                "formato_importacion_id": None,
                "formato_importacion_version_id": None,
                "formato_nombre": "Plantilla oficial FactuFlow",
            }

        raise LoteComprobanteError(
            "El archivo no coincide con la plantilla oficial. Elegí y confirmá un formato de importación antes de validar el lote."
        )

    def _serializar_importacion_configurable(
        self, importacion: ImportacionNormalizada
    ) -> dict[str, Any]:
        """Adapta la importación configurable a metadatos persistibles."""
        return {
            "filas": importacion.filas,
            "headers_detectados": importacion.headers_detectados,
            "mapeo_usado": importacion.mapeo_usado,
            "formato_importacion_id": importacion.formato.id,
            "formato_importacion_version_id": importacion.version.id,
            "formato_nombre": importacion.formato.nombre,
        }

    def _leer_excel(self, file_bytes: bytes) -> list[dict[str, Any]]:
        """Lee y normaliza las filas del Excel."""
        try:
            workbook = load_workbook(
                BytesIO(file_bytes), data_only=True, read_only=True
            )
        except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
            raise LoteComprobanteError(
                "No se pudo leer el archivo Excel. Verificá que sea un .xlsx válido generado desde la plantilla."
            ) from exc
        sheet = (
            workbook["Comprobantes"]
            if "Comprobantes" in workbook.sheetnames
            else workbook.active
        )
        headers = [self._normalize_header(cell.value) for cell in sheet[1]]

        missing = [col for col in self.TEMPLATE_COLUMNS if col not in headers]
        if missing:
            raise LoteComprobanteError(
                f"Faltan columnas obligatorias en la plantilla: {', '.join(missing)}"
            )

        filas: list[dict[str, Any]] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if all(cell in (None, "") for cell in row):
                continue
            row_dict = {
                headers[idx]: self._normalize_cell_value(value)
                for idx, value in enumerate(row)
                if idx < len(headers) and headers[idx]
            }
            filas.append(row_dict)
            if len(filas) > settings.batch_max_rows:
                raise LoteComprobanteError(
                    f"El archivo supera el máximo permitido de {settings.batch_max_rows} filas"
                )

        if not filas:
            raise LoteComprobanteError("La plantilla no contiene filas para procesar")
        return filas

    def _validar_empresa_del_lote(
        self, filas_excel: list[dict[str, Any]], empresa: Empresa
    ) -> None:
        """Verifica que todo el lote pertenezca a la empresa activa."""
        cuit_esperado = clean_cuit(empresa.cuit)
        cuit_detectados = {
            clean_cuit(row.get("empresa_cuit", ""))
            for row in filas_excel
            if clean_cuit(row.get("empresa_cuit", ""))
        }

        if cuit_detectados != {cuit_esperado}:
            raise LoteComprobanteError(
                "El archivo mezcla empresas o no coincide con la empresa activa. Revisa la columna empresa_cuit y vuelve a subir el lote."
            )

    async def _preparar_idempotencia(self, empresa_id: int, archivo_hash: str) -> None:
        """Evita duplicados y libera reintentos seguros sin CAE emitido."""
        result = await self.db.execute(
            select(LoteComprobante).where(
                LoteComprobante.empresa_id == empresa_id,
                LoteComprobante.archivo_hash == archivo_hash,
            )
        )
        lote_existente = result.scalar_one_or_none()
        if lote_existente is None:
            return

        if self._lote_permite_reintento(lote_existente):
            hash_anterior = lote_existente.archivo_hash
            lote_existente.archivo_hash = hashlib.sha256(
                (
                    f"{hash_anterior}:reintento:{lote_existente.id}:"
                    f"{datetime.utcnow().isoformat()}"
                ).encode("utf-8")
            ).hexdigest()
            metadata = dict(lote_existente.metadata_json or {})
            metadata["reemplazado_por_reintento"] = {
                "archivo_hash_original": hash_anterior,
                "fecha": datetime.utcnow().isoformat(),
                "motivo": "El lote anterior no emitió comprobantes fiscales",
            }
            lote_existente.metadata_json = metadata
            await self.db.flush()
            return

        raise LoteComprobanteError(
            "Ese archivo ya fue cargado previamente. Revisá el lote existente antes de volver a subirlo."
        )

    def _lote_permite_reintento(self, lote: LoteComprobante) -> bool:
        """Indica si un lote previo puede reemplazarse por una nueva validación."""
        if lote.grupos_emitidos and lote.grupos_emitidos > 0:
            return False
        return lote.estado in {"con_errores", "fallido"}

    async def _obtener_puntos_venta(self, empresa_id: int) -> dict[int, PuntoVenta]:
        """Obtiene los puntos de venta activos de la empresa indexados por número."""
        result = await self.db.execute(
            select(PuntoVenta).where(
                PuntoVenta.empresa_id == empresa_id,
                PuntoVenta.activo.is_(True),
                PuntoVenta.es_webservice.is_(True),
                PuntoVenta.bloqueado.is_(False),
                PuntoVenta.fecha_baja.is_(None),
            )
        )
        puntos = result.scalars().all()
        return {punto.numero: punto for punto in puntos}

    async def _obtener_certificado_activo(self, empresa_id: int) -> Certificado | None:
        """Obtiene el certificado activo para el ambiente actual."""
        ambiente = (
            ArcaAmbiente.PRODUCCION.value
            if settings.arca_env.lower() == ArcaAmbiente.PRODUCCION.value
            else ArcaAmbiente.HOMOLOGACION.value
        )
        result = await self.db.execute(
            select(Certificado)
            .where(
                Certificado.empresa_id == empresa_id,
                Certificado.activo.is_(True),
                Certificado.ambiente == ambiente,
            )
            .order_by(Certificado.fecha_vencimiento.desc(), Certificado.id.desc())
            .limit(1)
        )
        return result.scalar_one_or_none()

    def _validar_grupo(
        self,
        comprobante_ref: str,
        rows: list[tuple[int, dict[str, Any]]],
        empresa: Empresa,
        puntos_venta: dict[int, PuntoVenta],
    ) -> dict[str, Any]:
        """Valida un grupo de filas que forman un comprobante."""
        mensajes: list[str] = []
        header = rows[0][1]

        for _, row in rows[1:]:
            for field in self.HEADER_FIELDS:
                if (
                    str(row.get(field, "")).strip()
                    != str(header.get(field, "")).strip()
                ):
                    mensajes.append(
                        f"El campo {field} no coincide en todas las filas del comprobante {comprobante_ref}"
                    )
                    break

        empresa_cuit = clean_cuit(header.get("empresa_cuit", ""))
        if empresa_cuit != empresa.cuit:
            mensajes.append(
                f"El comprobante {comprobante_ref} no pertenece a la empresa activa ({empresa.cuit})"
            )

        punto_venta_numero = self._parse_int(header.get("punto_venta_numero"))
        if punto_venta_numero is None or punto_venta_numero not in puntos_venta:
            mensajes.append(
                f"El punto de venta del comprobante {comprobante_ref} no existe o no está activo"
            )

        tipo_comprobante = self._parse_int(header.get("tipo_comprobante"))
        if tipo_comprobante not in {1, 2, 3, 6, 7, 8, 11, 12, 13}:
            mensajes.append(
                f"El tipo de comprobante del grupo {comprobante_ref} no está soportado"
            )
        else:
            tipo_error = self._validar_tipo_comprobante_emisor(
                empresa, tipo_comprobante
            )
            if tipo_error:
                mensajes.append(f"{comprobante_ref}: {tipo_error}")

        concepto = self._parse_int(header.get("concepto"))
        if concepto not in {1, 2, 3}:
            mensajes.append(
                f"El concepto del comprobante {comprobante_ref} debe ser 1, 2 o 3"
            )

        fecha_emision = self._parse_date(header.get("fecha_emision"))
        if not fecha_emision:
            mensajes.append(
                f"El comprobante {comprobante_ref} requiere fecha de emisión explícita"
            )
        elif concepto in {1, 2, 3}:
            try:
                self.facturacion_service._validar_fecha_emision_arca(
                    fecha_emision, concepto
                )
            except ValidationError as exc:
                mensajes.append(f"{comprobante_ref}: {exc}")

        tipo_documento = self._parse_tipo_documento(
            header.get("cliente_tipo_documento")
        )
        numero_documento = clean_cuit(header.get("cliente_numero_documento", ""))
        condicion_iva = self._parse_condicion_iva(header.get("cliente_condicion_iva"))

        fecha_servicio_desde = self._parse_date(header.get("fecha_servicio_desde"))
        fecha_servicio_hasta = self._parse_date(header.get("fecha_servicio_hasta"))
        fecha_vto_pago = self._parse_date(header.get("fecha_vto_pago"))
        if concepto in {2, 3}:
            if (
                not fecha_servicio_desde
                or not fecha_servicio_hasta
                or not fecha_vto_pago
            ):
                mensajes.append(
                    f"El comprobante {comprobante_ref} requiere fechas de servicio y vencimiento"
                )

        comprobantes_asociados: list[ComprobanteAsociadoCreate] = []
        if tipo_comprobante in {2, 3, 7, 8, 12, 13}:
            asociado_tipo = self._parse_int(header.get("asociado_tipo_comprobante"))
            asociado_pv = self._parse_int(header.get("asociado_punto_venta"))
            asociado_numero = self._parse_int(header.get("asociado_numero"))
            asociado_fecha = self._parse_date(header.get("asociado_fecha"))
            asociado_cuit = clean_cuit(header.get("asociado_cuit", ""))
            if not asociado_tipo or not asociado_pv or not asociado_numero:
                mensajes.append(
                    f"La nota de crédito/débito {comprobante_ref} requiere comprobante asociado: tipo, punto de venta y número"
                )
            else:
                comprobantes_asociados.append(
                    ComprobanteAsociadoCreate(
                        tipo_comprobante=asociado_tipo,
                        punto_venta=asociado_pv,
                        numero=asociado_numero,
                        fecha=asociado_fecha,
                        cuit=asociado_cuit or None,
                    )
                )

        items: list[ItemComprobanteCreate] = []
        for index, (_, row) in enumerate(rows):
            descripcion = str(row.get("item_descripcion", "")).strip()
            if not descripcion:
                mensajes.append(
                    f"Todas las filas del comprobante {comprobante_ref} deben informar item_descripcion"
                )
                continue

            cantidad = self._parse_decimal(row.get("item_cantidad"))
            precio_unitario = self._parse_decimal(row.get("item_precio_unitario"))
            descuento = self._parse_decimal(
                row.get("item_descuento_porcentaje"), Decimal("0")
            )
            iva = self._parse_decimal(row.get("item_iva_porcentaje"), Decimal("0"))
            if cantidad is None or cantidad <= 0:
                mensajes.append(
                    f"La cantidad del ítem '{descripcion}' en {comprobante_ref} debe ser mayor a cero"
                )
                continue
            if precio_unitario is None or precio_unitario < 0:
                mensajes.append(
                    f"El precio unitario del ítem '{descripcion}' en {comprobante_ref} es inválido"
                )
                continue
            if iva is None or iva not in self.ALICUOTAS_IVA_PERMITIDAS:
                mensajes.append(
                    f"La alícuota de IVA del ítem '{descripcion}' en {comprobante_ref} debe ser 0, 10.5, 21 o 27"
                )
                continue
            if (
                tipo_comprobante in FacturacionService.TIPOS_COMPROBANTE_C
                and iva != Decimal("0")
            ):
                mensajes.append(
                    f"Los comprobantes tipo C no pueden incluir IVA en el ítem '{descripcion}' de {comprobante_ref}"
                )
                continue

            items.append(
                ItemComprobanteCreate(
                    codigo=str(row.get("item_codigo", "")).strip() or None,
                    descripcion=descripcion,
                    cantidad=cantidad,
                    unidad=str(row.get("item_unidad", "")).strip() or "unidad",
                    precio_unitario=precio_unitario,
                    descuento_porcentaje=descuento or Decimal("0"),
                    iva_porcentaje=iva or Decimal("0"),
                    orden=index,
                )
            )

        if not items:
            mensajes.append(f"El comprobante {comprobante_ref} no tiene ítems válidos")
        elif not mensajes:
            total_archivo = self._total_informado_por_archivo(rows)
            if total_archivo is not None:
                total_calculado = self.facturacion_service._calcular_totales(items)[
                    "total"
                ]
                diferencia = abs(total_calculado - total_archivo)
                if diferencia > Decimal("0.01"):
                    mensajes.append(
                        f"El total calculado para {comprobante_ref} "
                        f"({total_calculado:.2f}) no coincide con el total "
                        f"informado por el archivo ({total_archivo:.2f}). "
                        "Revisá si la columna usada como precio unitario ya "
                        "incluye IVA."
                    )

        if tipo_documento is None:
            tipo_documento_raw = str(header.get("cliente_tipo_documento", "")).strip()
            if not tipo_documento_raw and not numero_documento:
                tipo_documento = 99
            else:
                mensajes.append(
                    f"El tipo de documento del receptor en {comprobante_ref} es inválido"
                )

        if (
            tipo_documento == 80
            and numero_documento
            and not validate_cuit(numero_documento)
        ):
            mensajes.append(f"El CUIT del receptor en {comprobante_ref} no es válido")

        if tipo_comprobante in {1, 2, 3} and tipo_documento != 80:
            mensajes.append(
                f"Los comprobantes tipo A requieren que el receptor tenga CUIT en {comprobante_ref}"
            )

        if condicion_iva is None:
            condicion_raw = str(header.get("cliente_condicion_iva", "")).strip()
            if not condicion_raw and tipo_documento == 99:
                condicion_iva = "CF"
            else:
                mensajes.append(
                    f"La condición IVA del receptor en {comprobante_ref} no es válida"
                )
        elif tipo_documento == 80 and condicion_iva == "CF":
            mensajes.append(
                f"El receptor en {comprobante_ref} tiene CUIT pero figura como consumidor final. Configurá una condición IVA del receptor o dejá el documento vacío cuando la normativa lo permita."
            )

        payload: EmitirComprobanteRequest | None = None
        if not mensajes:
            try:
                payload = self.facturacion_service.normalizar_receptor(
                    EmitirComprobanteRequest(
                        empresa_id=empresa.id,
                        punto_venta_id=puntos_venta[punto_venta_numero].id,
                        tipo_comprobante=tipo_comprobante,
                        concepto=concepto,
                        fecha_emision=fecha_emision,
                        confirmacion_fecha_fiscal=True,
                        tipo_documento=tipo_documento,
                        numero_documento=numero_documento,
                        razon_social=str(
                            header.get("cliente_razon_social", "")
                        ).strip(),
                        condicion_iva=condicion_iva,
                        domicilio=str(header.get("cliente_domicilio", "")).strip()
                        or None,
                        fecha_servicio_desde=fecha_servicio_desde,
                        fecha_servicio_hasta=fecha_servicio_hasta,
                        fecha_vto_pago=fecha_vto_pago,
                        comprobantes_asociados=comprobantes_asociados,
                        observaciones=str(header.get("observaciones", "")).strip()
                        or None,
                        moneda="PES",
                        cotizacion=Decimal("1"),
                        guardar_cliente=False,
                        items=items,
                    )
                )
            except ValidationError as exc:
                mensajes.append(f"{comprobante_ref}: {exc}")

        if mensajes:
            fechas_payload = {
                "concepto": concepto,
                "fecha_emision": fecha_emision.isoformat() if fecha_emision else None,
                "fecha_servicio_desde": (
                    fecha_servicio_desde.isoformat() if fecha_servicio_desde else None
                ),
                "fecha_servicio_hasta": (
                    fecha_servicio_hasta.isoformat() if fecha_servicio_hasta else None
                ),
                "fecha_vto_pago": fecha_vto_pago.isoformat()
                if fecha_vto_pago
                else None,
            }
            return {
                "estado": "con_error",
                "mensajes": sorted(set(mensajes)),
                "payload": fechas_payload,
                "tipo_comprobante": tipo_comprobante,
                "concepto": concepto,
                "punto_venta_numero": punto_venta_numero,
                "cliente_documento": numero_documento,
                "cliente_razon_social": header.get("cliente_razon_social", ""),
                "fecha_emision": fecha_emision,
                "total_estimado": Decimal("0"),
            }

        assert payload is not None
        totales = self.facturacion_service._calcular_totales(items)
        return {
            "estado": "validado",
            "mensajes": ["Validado correctamente. Listo para emitir."],
            "payload": payload.model_dump(mode="json"),
            "tipo_comprobante": tipo_comprobante,
            "concepto": concepto,
            "punto_venta_numero": punto_venta_numero,
            "cliente_documento": payload.numero_documento,
            "cliente_razon_social": payload.razon_social,
            "fecha_emision": payload.fecha_emision,
            "total_estimado": totales["total"],
        }

    def _total_informado_por_archivo(
        self, rows: list[tuple[int, dict[str, Any]]]
    ) -> Decimal | None:
        """Suma el total informado por el archivo externo, si existe."""
        total = Decimal("0")
        encontrado = False
        for _, row in rows:
            valor = row.get("importe_total")
            if valor in (None, ""):
                continue
            importe = self._parse_decimal(valor)
            if importe is None:
                continue
            total += importe
            encontrado = True
        return total if encontrado else None

    async def _clasificar_grupos_validos_stale(
        self,
        lote: LoteComprobante,
    ) -> ClasificacionGruposStale:
        """Separa grupos pendientes intactos de grupos con evidencia fiscal."""
        grupos = await self._obtener_grupos_emitibles(lote.id)
        if not grupos:
            return ClasificacionGruposStale(
                intactos=(),
                con_evidencia_fiscal=(),
                total_validos=0,
            )

        grupo_ids = [grupo.id for grupo in grupos]
        intentos_por_grupo = dict(
            (
                await self.db.execute(
                    select(IntentoEmisionFiscal.grupo_id, func.count())
                    .where(
                        IntentoEmisionFiscal.lote_id == lote.id,
                        IntentoEmisionFiscal.grupo_id.in_(grupo_ids),
                    )
                    .group_by(IntentoEmisionFiscal.grupo_id)
                )
            ).all()
        )
        claves_comprobantes_locales = (
            await self._obtener_claves_comprobantes_locales_candidatos(grupos)
        )
        intactos: list[LoteComprobanteGrupo] = []
        con_evidencia_fiscal: list[LoteComprobanteGrupo] = []
        for grupo in grupos:
            cantidad_intentos = int(intentos_por_grupo.get(grupo.id, 0) or 0)
            clave_local = self._clave_evidencia_local_grupo(grupo)
            tiene_comprobante_local = (
                clave_local is not None and clave_local in claves_comprobantes_locales
            )
            if (
                self._grupo_validado_stale_intacto(grupo, cantidad_intentos)
                and not tiene_comprobante_local
            ):
                intactos.append(grupo)
            else:
                con_evidencia_fiscal.append(grupo)

        return ClasificacionGruposStale(
            intactos=tuple(intactos),
            con_evidencia_fiscal=tuple(con_evidencia_fiscal),
            total_validos=len(grupos),
        )

    @staticmethod
    def _grupo_validado_stale_intacto(
        grupo: LoteComprobanteGrupo,
        cantidad_intentos: int,
    ) -> bool:
        """Indica si un grupo validado no tiene señales de haber pedido CAE."""
        return (
            cantidad_intentos == 0
            and not grupo.cae
            and grupo.numero_asignado is None
            and grupo.comprobante_id is None
        )

    async def _obtener_claves_comprobantes_locales_candidatos(
        self,
        grupos: list[LoteComprobanteGrupo],
    ) -> set[tuple[int, int, int, date, Decimal, int, str, str]]:
        """Busca comprobantes locales que impiden tratar un grupo como intacto."""
        claves_grupos = {
            clave
            for grupo in grupos
            if (clave := self._clave_evidencia_local_grupo(grupo)) is not None
        }
        if not claves_grupos:
            return set()

        empresa_ids = {clave[0] for clave in claves_grupos}
        punto_venta_ids = {clave[1] for clave in claves_grupos}
        tipos = {clave[2] for clave in claves_grupos}
        fechas = {clave[3] for clave in claves_grupos}
        comprobantes = list(
            (
                await self.db.execute(
                    select(Comprobante).where(
                        Comprobante.empresa_id.in_(empresa_ids),
                        Comprobante.punto_venta_id.in_(punto_venta_ids),
                        Comprobante.tipo_comprobante.in_(tipos),
                        Comprobante.fecha_emision.in_(fechas),
                        Comprobante.estado == "autorizado",
                        Comprobante.cae.is_not(None),
                    )
                )
            )
            .scalars()
            .all()
        )
        return {
            clave
            for comprobante in comprobantes
            if (clave := self._clave_evidencia_local_comprobante(comprobante))
            in claves_grupos
        }

    def _clave_evidencia_local_grupo(
        self,
        grupo: LoteComprobanteGrupo,
    ) -> tuple[int, int, int, date, Decimal, int, str, str] | None:
        """Construye una clave fiscal del grupo para detectar evidencia local."""
        try:
            request = EmitirComprobanteRequest.model_validate(grupo.payload_json or {})
        except Exception:
            return None
        total = self.facturacion_service._calcular_totales(request.items)["total"]
        return (
            request.empresa_id,
            request.punto_venta_id,
            request.tipo_comprobante,
            request.fecha_emision,
            self._round_money(Decimal(str(total))),
            int(request.tipo_documento or 0),
            str(request.numero_documento or ""),
            self._normalizar_texto_snapshot(request.razon_social),
        )

    def _clave_evidencia_local_comprobante(
        self,
        comprobante: Comprobante,
    ) -> tuple[int, int, int, date, Decimal, int, str, str] | None:
        """Construye una clave fiscal de un comprobante local autorizado."""
        if comprobante.fecha_emision is None:
            return None
        return (
            comprobante.empresa_id,
            comprobante.punto_venta_id,
            comprobante.tipo_comprobante,
            comprobante.fecha_emision,
            self._round_money(Decimal(str(comprobante.total))),
            int(comprobante.receptor_tipo_documento or 0),
            str(comprobante.receptor_numero_documento or ""),
            self._normalizar_texto_snapshot(comprobante.receptor_razon_social),
        )

    async def _preflight_reanudar_grupos_intactos_stale(
        self,
        lote: LoteComprobante,
        grupos: tuple[LoteComprobanteGrupo, ...],
    ) -> tuple[bool, list[dict[str, Any]], str | None]:
        """Valida que los grupos intactos puedan volver a cola sin riesgo fiscal."""
        combos: dict[tuple[int, int, int], EmitirComprobanteRequest] = {}
        for grupo in grupos:
            if not grupo.payload_json:
                return (
                    False,
                    [],
                    f"El grupo {grupo.comprobante_ref} no conserva payload fiscal.",
                )
            try:
                request = EmitirComprobanteRequest.model_validate(grupo.payload_json)
            except Exception as exc:
                logger.warning(
                    "No se pudo validar payload fiscal del grupo stale %s",
                    grupo.comprobante_ref,
                    exc_info=True,
                )
                return False, [], str(exc)
            if request.empresa_id != lote.empresa_id:
                return (
                    False,
                    [],
                    f"El grupo {grupo.comprobante_ref} pertenece a otra empresa.",
                )
            combos[
                (
                    request.empresa_id,
                    request.punto_venta_id,
                    request.tipo_comprobante,
                )
            ] = request

        checks: list[dict[str, Any]] = []
        for empresa_id, punto_venta_id, tipo_comprobante in sorted(combos):
            try:
                checks.append(
                    await self.facturacion_service.verificar_numeracion_alineada_para_emision(
                        empresa_id=empresa_id,
                        punto_venta_id=punto_venta_id,
                        tipo_comprobante=tipo_comprobante,
                    )
                )
            except DATABASE_TEMPORARILY_UNAVAILABLE_ERRORS:
                raise
            except Exception as exc:
                logger.warning(
                    "No se pudo verificar numeración ARCA/local para lote stale %s",
                    lote.id,
                    exc_info=True,
                )
                return False, checks, str(exc)

        return True, checks, None

    async def _marcar_grupos_validos_stale_para_reconciliacion(
        self,
        lote: LoteComprobante,
        grupos: tuple[LoteComprobanteGrupo, ...]
        | list[LoteComprobanteGrupo]
        | None = None,
    ) -> int:
        """Marca grupos válidos de un lote stale como fiscalmente inciertos."""
        mensaje = (
            "El lote quedó en estado incierto durante el procesamiento. "
            "No reintentes este comprobante: primero hay que reconciliar "
            "contra ARCA para confirmar si fue autorizado."
        )
        if grupos is None:
            grupos = (
                (
                    await self.db.execute(
                        select(LoteComprobanteGrupo).where(
                            LoteComprobanteGrupo.lote_id == lote.id,
                            LoteComprobanteGrupo.estado == "validado",
                        )
                    )
                )
                .scalars()
                .all()
            )
        grupos = list(grupos)
        for grupo in grupos:
            grupo.estado = "requiere_reconciliacion"
            grupo.mensajes_json = [mensaje]
        if grupos:
            grupo_ids = [grupo.id for grupo in grupos]
            await self.db.execute(
                update(LoteComprobanteFila)
                .where(LoteComprobanteFila.grupo_id.in_(grupo_ids))
                .values(
                    estado="requiere_reconciliacion",
                    mensajes_json=[mensaje],
                )
            )
            await self.db.flush()
        return len(grupos)

    async def _actualizar_contadores_lote(
        self, lote: LoteComprobante
    ) -> dict[str, int]:
        """Recalcula contadores del lote desde sus grupos."""
        conteos = dict(
            (
                await self.db.execute(
                    select(LoteComprobanteGrupo.estado, func.count())
                    .where(LoteComprobanteGrupo.lote_id == lote.id)
                    .group_by(LoteComprobanteGrupo.estado)
                )
            ).all()
        )
        lote.total_grupos = sum(conteos.values())
        lote.grupos_validos = conteos.get("validado", 0)
        lote.grupos_con_error = conteos.get("con_error", 0)
        lote.grupos_emitidos = conteos.get("autorizado", 0)
        lote.grupos_fallidos = conteos.get("fallido", 0)
        lote.grupos_reconciliados_externos = conteos.get("autorizado_externo", 0)
        lote.grupos_descartados = conteos.get("descartado", 0)
        return conteos

    async def _actualizar_estado_lote(self, lote: LoteComprobante) -> None:
        """Recalcula contadores y estado del lote."""
        conteos = await self._actualizar_contadores_lote(lote)
        grupos_reconciliacion = conteos.get("requiere_reconciliacion", 0)
        grupos_reconciliacion += conteos.get("reintentando", 0)
        grupos_resueltos = (
            lote.grupos_emitidos
            + lote.grupos_reconciliados_externos
            + lote.grupos_descartados
        )

        if lote.estado == "procesando":
            total_emitible = (
                lote.grupos_emitidos
                + lote.grupos_fallidos
                + grupos_reconciliacion
                + lote.grupos_validos
            )
            procesados = lote.grupos_emitidos + lote.grupos_fallidos
            procesados += grupos_reconciliacion
            if total_emitible:
                lote.mensaje_resumen = (
                    f"Procesando comprobante {procesados} de {total_emitible}..."
                )
            else:
                lote.mensaje_resumen = "Procesando comprobantes..."
            return
        if lote.estado == "en_cola":
            lote.mensaje_resumen = "El lote está en cola para procesamiento."
            return

        if grupos_reconciliacion:
            lote.estado = "requiere_reconciliacion"
            lote.mensaje_resumen = (
                "ARCA pudo haber autorizado comprobantes que no quedaron guardados. "
                "No reintentes este lote hasta reconciliarlo."
            )
        elif lote.total_grupos and lote.grupos_emitidos == lote.total_grupos:
            lote.estado = "completado"
            lote.mensaje_resumen = (
                "Todos los comprobantes del lote fueron emitidos por FactuFlow."
            )
        elif lote.total_grupos and grupos_resueltos == lote.total_grupos:
            if lote.grupos_descartados:
                lote.estado = "cerrado_con_descartes"
                lote.mensaje_resumen = "El lote quedó cerrado con comprobantes descartados por decisión operativa."
            else:
                lote.estado = "cerrado_reconciliado"
                lote.mensaje_resumen = "El lote quedó cerrado con comprobantes emitidos fuera de FactuFlow y verificados contra ARCA."
        elif (lote.grupos_emitidos or lote.grupos_reconciliados_externos) and (
            lote.grupos_validos or lote.grupos_fallidos or lote.grupos_con_error
        ):
            lote.estado = "autorizado_parcial"
            lote.mensaje_resumen = (
                "El lote tiene comprobantes emitidos y pendientes por resolver."
            )
        elif lote.grupos_con_error:
            lote.estado = "con_errores"
            lote.mensaje_resumen = (
                "El lote tiene comprobantes con errores de validación."
            )
        elif lote.grupos_validos:
            lote.estado = "validado"
            lote.mensaje_resumen = "El lote se validó correctamente y puede emitirse."
        elif lote.grupos_fallidos:
            lote.estado = "fallido"
            lote.mensaje_resumen = "No se pudo emitir ningún comprobante del lote."
        else:
            lote.estado = "cargado"
            lote.mensaje_resumen = "El lote fue cargado."

    def _validar_tipo_comprobante_emisor(
        self, empresa: Empresa, tipo_comprobante: int
    ) -> str | None:
        """Valida compatibilidad básica entre emisor y tipo de comprobante."""
        condicion = str(empresa.condicion_iva or "").strip().upper()
        tipos_a = {1, 2, 3}
        tipos_c = {11, 12, 13}

        if condicion == "RI" and tipo_comprobante in tipos_c:
            return (
                "El emisor Responsable Inscripto no puede emitir comprobantes "
                "tipo C. Usá un formato particular con Factura A/B según corresponda."
            )
        if condicion in {"MONOTRIBUTO", "EXENTO"} and tipo_comprobante in tipos_a:
            return "El emisor no Responsable Inscripto no puede emitir comprobantes tipo A."
        return None

    async def _obtener_grupos_lote(self, lote_id: int) -> list[LoteComprobanteGrupo]:
        result = await self.db.execute(
            select(LoteComprobanteGrupo)
            .options(selectinload(LoteComprobanteGrupo.filas))
            .where(LoteComprobanteGrupo.lote_id == lote_id)
            .order_by(LoteComprobanteGrupo.orden)
            .execution_options(populate_existing=True)
        )
        return list(result.scalars().all())

    async def _obtener_grupos_emitibles(
        self, lote_id: int
    ) -> list[LoteComprobanteGrupo]:
        result = await self.db.execute(
            select(LoteComprobanteGrupo)
            .options(selectinload(LoteComprobanteGrupo.filas))
            .where(
                LoteComprobanteGrupo.lote_id == lote_id,
                LoteComprobanteGrupo.estado == "validado",
            )
            .order_by(LoteComprobanteGrupo.orden)
            .execution_options(populate_existing=True)
        )
        return list(result.scalars().all())

    async def _reconciliar_grupos_autorizados_existentes(self, lote_id: int) -> int:
        """Vincula comprobantes locales ya autorizados sin solicitar CAE."""
        reconciliados = 0
        for grupo in await self._obtener_grupos_emitibles(lote_id):
            try:
                payload = grupo.payload_json or {}
                request = EmitirComprobanteRequest.model_validate(payload)
                if await self._reconciliar_grupo_autorizado_existente(grupo, request):
                    reconciliados += 1
            except DATABASE_TEMPORARILY_UNAVAILABLE_ERRORS:
                raise
            except Exception:  # pragma: no cover - fallback defensivo
                logger.exception(
                    "No se pudo reconciliar localmente el grupo %s",
                    grupo.comprobante_ref,
                )
        return reconciliados

    async def _lote_tiene_intentos_fiscales_inciertos(self, lote_id: int) -> bool:
        """Indica si un lote tiene intentos fiscales que requieren reconciliación."""
        intento_incierto = (
            await self.db.execute(
                select(IntentoEmisionFiscal.id)
                .where(
                    IntentoEmisionFiscal.lote_id == lote_id,
                    IntentoEmisionFiscal.estado.in_(
                        {"en_proceso", "requiere_reconciliacion"}
                    ),
                )
                .limit(1)
            )
        ).scalar_one_or_none()
        return intento_incierto is not None

    async def _lote_tiene_evidencia_fiscal_local_coherente(self, lote_id: int) -> bool:
        """Verifica grupos autorizados antes de cerrar un lote stale localmente."""
        grupos_autorizados = list(
            (
                await self.db.execute(
                    select(LoteComprobanteGrupo)
                    .options(selectinload(LoteComprobanteGrupo.filas))
                    .where(
                        LoteComprobanteGrupo.lote_id == lote_id,
                        LoteComprobanteGrupo.estado == "autorizado",
                    )
                    .order_by(LoteComprobanteGrupo.orden)
                )
            )
            .scalars()
            .all()
        )
        for grupo in grupos_autorizados:
            try:
                request = EmitirComprobanteRequest.model_validate(
                    grupo.payload_json or {}
                )
            except Exception:
                logger.exception(
                    "No se pudo validar payload local del grupo autorizado %s",
                    grupo.comprobante_ref,
                )
                return False
            if not await self._reconciliar_grupo_autorizado_existente(grupo, request):
                return False
        return True

    async def _reconciliar_grupo_autorizado_existente(
        self, grupo: LoteComprobanteGrupo, request: EmitirComprobanteRequest
    ) -> bool:
        """Vincula solo comprobantes respaldados por un intento fiscal del grupo."""
        totales = self.facturacion_service._calcular_totales(request.items)
        result_intentos = await self.db.execute(
            select(IntentoEmisionFiscal)
            .where(
                IntentoEmisionFiscal.lote_id == grupo.lote_id,
                IntentoEmisionFiscal.grupo_id == grupo.id,
            )
            .order_by(IntentoEmisionFiscal.id)
        )
        intentos_grupo = list(result_intentos.scalars().all())
        if any(
            intento.estado in {"en_proceso", "requiere_reconciliacion"}
            for intento in intentos_grupo
        ):
            return False
        intentos_autorizados = [
            intento for intento in intentos_grupo if intento.estado == "autorizado"
        ]
        if len(intentos_autorizados) != 1:
            return False

        intento = intentos_autorizados[0]
        if (
            intento.comprobante_id is None
            or intento.numero_planificado is None
            or not intento.cae
        ):
            return False

        comprobante = (
            await self.db.execute(
                select(Comprobante)
                .options(
                    selectinload(Comprobante.punto_venta),
                    selectinload(Comprobante.items),
                )
                .where(Comprobante.id == intento.comprobante_id)
            )
        ).scalar_one_or_none()
        if comprobante is None:
            return False

        comprobante_vinculado = (
            await self.db.execute(
                select(LoteComprobanteGrupo.id)
                .where(
                    LoteComprobanteGrupo.comprobante_id == comprobante.id,
                    LoteComprobanteGrupo.id != grupo.id,
                )
                .limit(1)
            )
        ).scalar_one_or_none()
        if comprobante_vinculado is not None:
            return False

        if not self._intento_local_coincide_con_grupo(
            intento=intento,
            comprobante=comprobante,
            request=request,
            total=self._redondear_centavos(Decimal(str(totales["total"]))),
        ):
            return False

        grupo.estado = "autorizado"
        grupo.cae = comprobante.cae
        grupo.numero_asignado = comprobante.numero
        grupo.comprobante_id = comprobante.id
        grupo.mensajes_json = [
            "Comprobante ya guardado y respaldado por intento fiscal del lote.",
            f"CAE {comprobante.cae}",
        ]
        await self._marcar_filas(grupo, "autorizado", grupo.mensajes_json)
        return True

    def _intento_local_coincide_con_grupo(
        self,
        *,
        intento: IntentoEmisionFiscal,
        comprobante: Comprobante,
        request: EmitirComprobanteRequest,
        total: Decimal,
    ) -> bool:
        """Valida que intento, comprobante y payload del grupo sean el mismo CAE."""
        total_intento = self._redondear_centavos(Decimal(str(intento.total)))
        total_comprobante = self._redondear_centavos(Decimal(str(comprobante.total)))
        idempotencia = IdempotenciaFiscalService(self.db)
        payload = request.model_dump(mode="json")
        payload_hash = idempotencia.calcular_payload_hash(
            idempotencia.payload_sin_confirmacion_duplicado(payload)
        )
        huella_request = idempotencia.calcular_huella_logica(
            request=request,
            punto_venta_numero=intento.punto_venta_numero,
            total=total,
        )
        return all(
            [
                intento.payload_hash == payload_hash,
                intento.huella_logica == huella_request,
                self._comprobante_snapshot_coincide_con_request(comprobante, request),
                intento.empresa_id == request.empresa_id,
                intento.punto_venta_id == request.punto_venta_id,
                intento.tipo_comprobante == request.tipo_comprobante,
                intento.fecha_emision == request.fecha_emision,
                total_intento == total,
                intento.receptor_tipo_documento == request.tipo_documento,
                intento.receptor_numero_documento == request.numero_documento,
                intento.receptor_razon_social == request.razon_social,
                comprobante.empresa_id == request.empresa_id,
                comprobante.punto_venta_id == request.punto_venta_id,
                comprobante.tipo_comprobante == request.tipo_comprobante,
                comprobante.numero == intento.numero_planificado,
                comprobante.fecha_emision == request.fecha_emision,
                total_comprobante == total,
                comprobante.estado == "autorizado",
                comprobante.cae == intento.cae,
                bool(comprobante.cae),
                bool(comprobante.cae_vencimiento),
                comprobante.cae_vencimiento == intento.cae_vencimiento,
                comprobante.receptor_tipo_documento == request.tipo_documento,
                comprobante.receptor_numero_documento == request.numero_documento,
                comprobante.receptor_razon_social == request.razon_social,
            ]
        )

    def _comprobante_snapshot_coincide_con_request(
        self,
        comprobante: Comprobante,
        request: EmitirComprobanteRequest,
    ) -> bool:
        """Compara el snapshot fiscal local guardado contra el request original."""
        if comprobante.concepto != request.concepto:
            return False
        if comprobante.fecha_servicio_desde != request.fecha_servicio_desde:
            return False
        if comprobante.fecha_servicio_hasta != request.fecha_servicio_hasta:
            return False
        if comprobante.fecha_vto_pago != request.fecha_vto_pago:
            return False
        if comprobante.fecha_vencimiento != request.fecha_vto_pago:
            return False
        if self._normalizar_texto_snapshot(comprobante.moneda).upper() != (
            self._normalizar_texto_snapshot(request.moneda).upper()
        ):
            return False
        if self._decimal_snapshot(comprobante.cotizacion) != self._decimal_snapshot(
            request.cotizacion
        ):
            return False
        if self._normalizar_texto_snapshot(
            comprobante.receptor_condicion_iva
        ) != self._normalizar_texto_snapshot(request.condicion_iva):
            return False
        if self._normalizar_texto_snapshot(
            comprobante.receptor_domicilio
        ) != self._normalizar_texto_snapshot(request.domicilio):
            return False

        items_comprobante = sorted(comprobante.items, key=lambda item: item.orden)
        items_request = sorted(request.items, key=lambda item: item.orden)
        if len(items_comprobante) != len(items_request):
            return False
        for item_comprobante, item_request in zip(items_comprobante, items_request):
            if item_comprobante.orden != item_request.orden:
                return False
            if self._normalizar_texto_snapshot(
                item_comprobante.codigo
            ) != self._normalizar_texto_snapshot(item_request.codigo):
                return False
            if self._normalizar_texto_snapshot(
                item_comprobante.descripcion
            ) != self._normalizar_texto_snapshot(item_request.descripcion):
                return False
            if self._normalizar_texto_snapshot(
                item_comprobante.unidad
            ) != self._normalizar_texto_snapshot(item_request.unidad):
                return False
            if self._decimal_snapshot(item_comprobante.cantidad) != (
                self._decimal_snapshot(item_request.cantidad)
            ):
                return False
            if self._decimal_snapshot(item_comprobante.precio_unitario) != (
                self._decimal_snapshot(item_request.precio_unitario)
            ):
                return False
            if self._decimal_snapshot(item_comprobante.descuento_porcentaje) != (
                self._decimal_snapshot(item_request.descuento_porcentaje)
            ):
                return False
            if self._decimal_snapshot(item_comprobante.iva_porcentaje) != (
                self._decimal_snapshot(item_request.iva_porcentaje)
            ):
                return False
        return True

    @staticmethod
    def _normalizar_texto_snapshot(value: Any) -> str:
        """Normaliza textos persistidos para comparar snapshots locales."""
        return str(value or "").strip()

    @staticmethod
    def _decimal_snapshot(value: Any) -> Decimal:
        """Normaliza decimales persistidos para comparar snapshots locales."""
        if value is None:
            return Decimal("0")
        return Decimal(str(value))

    async def _marcar_filas(
        self, grupo: LoteComprobanteGrupo, estado: str, mensajes: list[str]
    ) -> None:
        for fila in grupo.filas:
            fila.estado = estado
            fila.mensajes_json = mensajes

    def _normalize_header(self, value: Any) -> str:
        return str(value or "").strip().lower()

    def _normalize_cell_value(self, value: Any) -> Any:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        if isinstance(value, Decimal):
            return float(value)
        return value

    def _parse_int(self, value: Any) -> int | None:
        try:
            if value in (None, ""):
                return None
            return int(value)
        except (TypeError, ValueError):
            return None

    def _parse_decimal(
        self, value: Any, default: Decimal | None = None
    ) -> Decimal | None:
        if value in (None, ""):
            return default
        if isinstance(value, Decimal):
            return value
        if isinstance(value, (int, float)):
            return Decimal(str(value))
        text = str(value).strip().replace("$", "").replace(" ", "")
        if "," in text and "." in text:
            text = text.replace(".", "").replace(",", ".")
        elif "," in text:
            text = text.replace(",", ".")
        try:
            return Decimal(text)
        except (InvalidOperation, ValueError):
            return None

    @staticmethod
    def _redondear_centavos(value: Decimal) -> Decimal:
        """Redondea importes fiscales a centavos para comparaciones exactas."""
        return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    def _parse_date(self, value: Any) -> date | None:
        if value in (None, ""):
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, (int, float)) and 1 <= value <= 60000:
            try:
                return from_excel(value).date()
            except (TypeError, ValueError):
                return None
        text = str(value).strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        try:
            return datetime.fromisoformat(text).date()
        except ValueError:
            return None

    def _parse_tipo_documento(self, value: Any) -> int | None:
        normalized = str(value or "").strip().upper()
        return self.TIPO_DOCUMENTO_MAP.get(normalized)

    def _parse_condicion_iva(self, value: Any) -> str | None:
        normalized = str(value or "").strip().upper()
        return self.CONDICION_IVA_MAP.get(normalized)
