"""Servicios para formatos configurables de importación de Excel."""

from __future__ import annotations

import re
import unicodedata
from copy import deepcopy
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.datetime import from_excel
from openpyxl.utils.cell import column_index_from_string, get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import or_, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.arca.utils import clean_cuit, validate_cuit
from app.models.empresa import Empresa
from app.models.formato_importacion import (
    FormatoImportacion,
    FormatoImportacionCampo,
    FormatoImportacionRegla,
    FormatoImportacionVersion,
)


class FormatoImportacionError(Exception):
    """Error funcional al detectar o aplicar un formato de importación."""


@dataclass
class CandidatoFormato:
    """Resultado de detección de un formato contra encabezados de Excel."""

    formato_id: int | None
    formato_version_id: int | None
    nombre: str
    alcance: str
    version: int | None
    score: float
    confianza: str
    columnas_detectadas: list[str]
    columnas_faltantes: list[str]
    mensajes: list[str]


@dataclass
class ImportacionNormalizada:
    """Archivo externo convertido al contrato interno de lotes."""

    filas: list[dict[str, Any]]
    headers_detectados: list[str]
    mapeo_usado: dict[str, Any]
    formato: FormatoImportacion
    version: FormatoImportacionVersion


@dataclass
class ExcelColumnaAnalizada:
    """Columna detectada en un Excel de ejemplo."""

    indice: int
    letra: str
    encabezado: str


@dataclass
class ExcelAnalisis:
    """Encabezados detectados para iniciar una plantilla visual."""

    hoja: str
    fila_encabezado: int
    columnas: list[ExcelColumnaAnalizada]


@dataclass
class CompatibilidadMensaje:
    """Mensaje de compatibilidad de plantilla, perfil y emisor."""

    codigo: str
    campo: str | None
    severidad: str
    mensaje: str


@dataclass
class CompatibilidadPlantilla:
    """Resultado de compatibilidad para una plantilla visual."""

    estado: str
    faltantes: list[CompatibilidadMensaje]
    omitibles: list[CompatibilidadMensaje]
    advertencias: list[CompatibilidadMensaje]
    conflictos: list[CompatibilidadMensaje]


FORMATO_BANCARIO_CONFIG: dict[str, Any] = {
    "tipo": "extracto_bancario_creditos",
    "plantilla_sistema_protegida": True,
    "header_row": 1,
    "modo_agrupacion": "fila",
    "plantilla": {
        "nombre_publico": "Extracto bancario - créditos IVA exento",
        "columnas": [
            {
                "campo_destino": "fecha_origen",
                "etiqueta": "Fecha",
                "origen": "header",
                "requerido": False,
                "transformacion": "fecha",
                "ejemplo": "2026-05-31",
            },
            {
                "campo_destino": "importe_total",
                "etiqueta": "Créditos",
                "origen": "header",
                "requerido": True,
                "transformacion": "decimal",
                "ejemplo": "10000.00",
            },
            {
                "campo_destino": "cliente_razon_social",
                "etiqueta": "Leyendas Adicionales1",
                "origen": "header",
                "requerido": False,
                "transformacion": "texto",
                "ejemplo": "Consumidor Final",
            },
            {
                "campo_destino": "cliente_numero_documento",
                "etiqueta": "Leyendas Adicionales2",
                "origen": "header",
                "requerido": False,
                "transformacion": "documento",
                "ejemplo": "30123456789",
            },
            {
                "campo_destino": "punto_venta_numero",
                "etiqueta": "Pto Vta",
                "origen": "header",
                "requerido": True,
                "transformacion": "entero",
                "ejemplo": "1",
            },
        ],
    },
    "campos": {
        "fecha_origen": {
            "origen": "header",
            "encabezados": ["Fecha"],
            "transformacion": "fecha",
            "requerido": False,
        },
        "importe_total": {
            "origen": "header",
            "encabezados": [
                "Creditos",
                "Créditos",
                "Credito",
                "Crédito",
                "Importe Credito",
                "Importe Crédito",
                "Importe acreditado",
            ],
            "transformacion": "decimal",
            "requerido": True,
        },
        "cliente_razon_social": {
            "origen": "header",
            "encabezados": [
                "Leyendas Adicionales1",
                "Leyendas Adicionales 1",
                "Nombre",
                "Cliente",
                "Razon Social",
                "Razón Social",
            ],
            "transformacion": "texto",
            "requerido": False,
            "default": "",
        },
        "cliente_numero_documento": {
            "origen": "header",
            "encabezados": [
                "Leyendas Adicionales2",
                "Leyendas Adicionales 2",
                "CUIT",
                "Documento",
                "Numero Documento",
                "Número Documento",
            ],
            "transformacion": "documento",
            "requerido": False,
            "default": "",
        },
        "punto_venta_numero": {
            "origen": "header",
            "encabezados": [
                "Pto Vta",
                "Pto. Vta.",
                "Punto Venta",
                "Punto de Venta",
                "PV",
            ],
            "transformacion": "entero",
            "requerido": True,
        },
        "tipo_comprobante": {"origen": "constante", "valor": 11},
        "cliente_condicion_iva": {
            "origen": "constante",
            "valor": "Consumidor Final",
        },
        "item_cantidad": {"origen": "constante", "valor": 1},
        "item_unidad": {"origen": "constante", "valor": "unidad"},
        "item_iva_porcentaje": {"origen": "constante", "valor": 0},
        "item_descuento_porcentaje": {"origen": "constante", "valor": 0},
        "guardar_cliente": {"origen": "constante", "valor": False},
    },
}

CATALOGO_CAMPOS_PLANTILLA: list[dict[str, Any]] = [
    {
        "codigo": "empresa_cuit",
        "etiqueta": "CUIT del emisor",
        "grupo": "Emisor",
        "descripcion": "Identifica al emisor del comprobante.",
        "requerido_base": True,
        "transformaciones": ["documento"],
        "origenes": ["header", "columna", "constante", "empresa"],
    },
    {
        "codigo": "punto_venta_numero",
        "etiqueta": "Punto de venta",
        "grupo": "Comprobante",
        "descripcion": "Punto de venta ARCA usado para emitir.",
        "requerido_base": True,
        "transformaciones": ["entero"],
        "origenes": ["header", "columna", "constante"],
    },
    {
        "codigo": "tipo_comprobante",
        "etiqueta": "Tipo de comprobante",
        "grupo": "Comprobante",
        "descripcion": "Factura A/B/C, nota de crédito o nota de débito.",
        "requerido_base": True,
        "transformaciones": ["entero"],
        "origenes": ["header", "columna", "constante"],
    },
    {
        "codigo": "concepto",
        "etiqueta": "Concepto ARCA",
        "grupo": "Comprobante",
        "descripcion": "Productos, servicios o productos y servicios.",
        "requerido_base": False,
        "transformaciones": ["entero"],
        "origenes": ["header", "columna", "constante"],
    },
    {
        "codigo": "fecha_emision",
        "etiqueta": "Fecha de emisión",
        "grupo": "Fechas",
        "descripcion": "Fecha fiscal explícita del comprobante.",
        "requerido_base": True,
        "transformaciones": ["fecha"],
        "origenes": ["header", "columna"],
    },
    {
        "codigo": "fecha_servicio_desde",
        "etiqueta": "Servicio desde",
        "grupo": "Fechas",
        "descripcion": "Inicio del período facturado para servicios.",
        "requerido_base": False,
        "transformaciones": ["fecha"],
        "origenes": ["header", "columna", "constante"],
    },
    {
        "codigo": "fecha_servicio_hasta",
        "etiqueta": "Servicio hasta",
        "grupo": "Fechas",
        "descripcion": "Fin del período facturado para servicios.",
        "requerido_base": False,
        "transformaciones": ["fecha"],
        "origenes": ["header", "columna", "constante"],
    },
    {
        "codigo": "fecha_vto_pago",
        "etiqueta": "Vencimiento de pago",
        "grupo": "Fechas",
        "descripcion": "Vencimiento obligatorio cuando se facturan servicios.",
        "requerido_base": False,
        "transformaciones": ["fecha"],
        "origenes": ["header", "columna", "constante"],
    },
    {
        "codigo": "cliente_tipo_documento",
        "etiqueta": "Tipo doc. receptor",
        "grupo": "Receptor",
        "descripcion": "Tipo de documento del receptor, por ejemplo CUIT o DNI.",
        "requerido_base": False,
        "transformaciones": ["texto", "entero"],
        "origenes": ["header", "columna", "constante"],
    },
    {
        "codigo": "cliente_numero_documento",
        "etiqueta": "Documento receptor",
        "grupo": "Receptor",
        "descripcion": "Número de documento o CUIT del receptor.",
        "requerido_base": False,
        "transformaciones": ["documento"],
        "origenes": ["header", "columna", "constante"],
    },
    {
        "codigo": "cliente_razon_social",
        "etiqueta": "Razón social receptor",
        "grupo": "Receptor",
        "descripcion": "Nombre fiscal o comercial del receptor.",
        "requerido_base": False,
        "transformaciones": ["texto"],
        "origenes": ["header", "columna", "constante"],
    },
    {
        "codigo": "cliente_condicion_iva",
        "etiqueta": "Condición IVA receptor",
        "grupo": "Receptor",
        "descripcion": "Condición fiscal del receptor.",
        "requerido_base": False,
        "transformaciones": ["texto"],
        "origenes": ["header", "columna", "constante"],
    },
    {
        "codigo": "cliente_domicilio",
        "etiqueta": "Domicilio receptor",
        "grupo": "Receptor",
        "descripcion": "Domicilio fiscal o declarado del receptor.",
        "requerido_base": False,
        "transformaciones": ["texto"],
        "origenes": ["header", "columna", "constante"],
    },
    {
        "codigo": "item_codigo",
        "etiqueta": "Código ítem",
        "grupo": "Ítems",
        "descripcion": "Código interno o SKU del producto/servicio.",
        "requerido_base": False,
        "transformaciones": ["texto"],
        "origenes": ["header", "columna", "constante"],
    },
    {
        "codigo": "item_descripcion",
        "etiqueta": "Descripción ítem",
        "grupo": "Ítems",
        "descripcion": "Descripción facturada en el comprobante.",
        "requerido_base": True,
        "transformaciones": ["texto"],
        "origenes": ["header", "columna", "constante"],
    },
    {
        "codigo": "item_cantidad",
        "etiqueta": "Cantidad",
        "grupo": "Ítems",
        "descripcion": "Cantidad del producto o servicio.",
        "requerido_base": True,
        "transformaciones": ["decimal"],
        "origenes": ["header", "columna", "constante"],
    },
    {
        "codigo": "item_unidad",
        "etiqueta": "Unidad",
        "grupo": "Ítems",
        "descripcion": "Unidad de medida visual del ítem.",
        "requerido_base": False,
        "transformaciones": ["texto"],
        "origenes": ["header", "columna", "constante"],
    },
    {
        "codigo": "item_precio_unitario",
        "etiqueta": "Precio unitario",
        "grupo": "Ítems",
        "descripcion": "Precio unitario sin IVA cuando corresponde.",
        "requerido_base": True,
        "transformaciones": ["decimal"],
        "origenes": ["header", "columna", "constante"],
    },
    {
        "codigo": "item_descuento_porcentaje",
        "etiqueta": "Descuento %",
        "grupo": "Ítems",
        "descripcion": "Descuento porcentual del ítem.",
        "requerido_base": False,
        "transformaciones": ["decimal"],
        "origenes": ["header", "columna", "constante"],
    },
    {
        "codigo": "item_iva_porcentaje",
        "etiqueta": "IVA %",
        "grupo": "Ítems",
        "descripcion": "Alícuota de IVA del ítem.",
        "requerido_base": True,
        "transformaciones": ["decimal"],
        "origenes": ["header", "columna", "constante"],
    },
    {
        "codigo": "importe_total",
        "etiqueta": "Total de control",
        "grupo": "Totales",
        "descripcion": "Total esperado para validar el importe del comprobante.",
        "requerido_base": False,
        "transformaciones": ["decimal"],
        "origenes": ["header", "columna", "constante"],
    },
    {
        "codigo": "observaciones",
        "etiqueta": "Observaciones",
        "grupo": "Comprobante",
        "descripcion": "Texto interno o visible según la emisión.",
        "requerido_base": False,
        "transformaciones": ["texto"],
        "origenes": ["header", "columna", "constante"],
    },
    {
        "codigo": "asociado_tipo_comprobante",
        "etiqueta": "Tipo comp. asociado",
        "grupo": "Comprobantes asociados",
        "descripcion": "Tipo del comprobante asociado para notas.",
        "requerido_base": False,
        "transformaciones": ["entero"],
        "origenes": ["header", "columna", "constante"],
    },
    {
        "codigo": "asociado_punto_venta",
        "etiqueta": "PV asociado",
        "grupo": "Comprobantes asociados",
        "descripcion": "Punto de venta del comprobante asociado para notas.",
        "requerido_base": False,
        "transformaciones": ["entero"],
        "origenes": ["header", "columna", "constante"],
    },
    {
        "codigo": "asociado_numero",
        "etiqueta": "Número asociado",
        "grupo": "Comprobantes asociados",
        "descripcion": "Número del comprobante asociado para notas.",
        "requerido_base": False,
        "transformaciones": ["entero"],
        "origenes": ["header", "columna", "constante"],
    },
    {
        "codigo": "asociado_fecha",
        "etiqueta": "Fecha asociado",
        "grupo": "Comprobantes asociados",
        "descripcion": "Fecha del comprobante asociado si está disponible.",
        "requerido_base": False,
        "transformaciones": ["fecha"],
        "origenes": ["header", "columna", "constante"],
    },
    {
        "codigo": "asociado_cuit",
        "etiqueta": "CUIT asociado",
        "grupo": "Comprobantes asociados",
        "descripcion": "CUIT informado en el comprobante asociado.",
        "requerido_base": False,
        "transformaciones": ["documento"],
        "origenes": ["header", "columna", "constante"],
    },
]

TIPOS_A = {1, 2, 3}
TIPOS_B = {6, 7, 8}
TIPOS_C = {11, 12, 13}
TIPOS_COMPROBANTE_SOPORTADOS = TIPOS_A | TIPOS_B | TIPOS_C
TIPOS_NOTA_DEBITO = {2, 7, 12}
TIPOS_NOTA_CREDITO = {3, 8, 13}
TIPOS_NOTA = TIPOS_NOTA_DEBITO | TIPOS_NOTA_CREDITO
CONCEPTOS_SERVICIOS = {2, 3}
CONCEPTOS_PRODUCTOS = {1}
ALICUOTAS_IVA_PERMITIDAS = {
    Decimal("0"),
    Decimal("10.5"),
    Decimal("21"),
    Decimal("27"),
}
CAMPOS_ORIGEN_EMPRESA = {"empresa_cuit"}
CAMPOS_CONSTANTE_REQUERIDA_NO_VACIA = {
    "tipo_comprobante",
    "punto_venta_numero",
    "item_descripcion",
    "item_precio_unitario",
    "item_iva_porcentaje",
}
CAMPOS_FISCALES_REQUERIDOS_GUARDADO = (
    "tipo_comprobante",
    "item_iva_porcentaje",
)
CAMPOS_VISUALES_REQUERIDOS_GUARDADO = ("item_precio_unitario",)
CAMPOS_FISCALES_DESEMPATE = (
    "tipo_comprobante",
    "punto_venta_numero",
    "concepto",
    "cliente_condicion_iva",
    "item_cantidad",
    "item_unidad",
    "item_iva_porcentaje",
)
ORIGENES_PERMITIDOS_POR_CAMPO = {
    str(item["codigo"]): set(item.get("origenes", []))
    for item in CATALOGO_CAMPOS_PLANTILLA
}
TRANSFORMACIONES_PERMITIDAS_POR_CAMPO = {
    str(item["codigo"]): set(item.get("transformaciones", []))
    for item in CATALOGO_CAMPOS_PLANTILLA
}
EXCEL_FORMULA_PREFIXES = ("=", "+", "-", "@")
EXCEL_MAX_COLUMN_INDEX = 16383


class FormatosImportacionService:
    """Gestiona formatos reutilizables y aplica mapeos de Excel externos."""

    FORMATO_BANCARIO_NOMBRE = "Extracto bancario - creditos IVA exento"
    CONSUMIDOR_FINAL_IDENTIFICACION_MINIMA = Decimal("10000000")

    def __init__(self, db: AsyncSession):
        self.db = db

    def catalogo_campos(self) -> list[dict[str, Any]]:
        """Devuelve los campos disponibles para construir plantillas."""
        return deepcopy(CATALOGO_CAMPOS_PLANTILLA)

    async def asegurar_formatos_base(self) -> None:
        """Crea formatos globales base si la base aun no los tiene."""
        result = await self.db.execute(
            select(FormatoImportacion)
            .options(selectinload(FormatoImportacion.versiones))
            .where(
                FormatoImportacion.nombre == self.FORMATO_BANCARIO_NOMBRE,
                FormatoImportacion.alcance == "global",
            )
        )
        existente = result.scalar_one_or_none()
        if existente is not None:
            await self._actualizar_formato_base_si_corresponde(existente)
            return

        formato = FormatoImportacion(
            nombre=self.FORMATO_BANCARIO_NOMBRE,
            descripcion=(
                "Formato global para extractos donde Creditos es el importe, "
                "Leyendas Adicionales1 el receptor, Leyendas Adicionales2 el "
                "documento y Pto Vta el punto de venta."
            ),
            alcance="global",
            activo=True,
        )
        self.db.add(formato)
        await self.db.flush()
        await self._crear_version_formato_base(formato, version_numero=1)
        await self.db.commit()

    async def _actualizar_formato_base_si_corresponde(
        self, formato: FormatoImportacion
    ) -> None:
        """Versiona el formato global si todavia usa una regla fiscal antigua."""
        version_vigente = self._version_vigente(formato)
        if version_vigente is None:
            await self._crear_version_formato_base(formato, version_numero=1)
            await self.db.commit()
            return

        campos = version_vigente.configuracion_json.get("campos", {})
        concepto = campos.get("concepto")
        item_descripcion = campos.get("item_descripcion")
        tiene_metadatos_visuales = version_vigente.configuracion_json.get(
            "plantilla_sistema_protegida"
        ) and version_vigente.configuracion_json.get("plantilla")
        if concepto is None and item_descripcion is None and tiene_metadatos_visuales:
            return

        if concepto is None and item_descripcion is None:
            version_vigente.configuracion_json = deepcopy(FORMATO_BANCARIO_CONFIG)
            version_vigente.headers_firma_json = self._construir_firma_headers(
                FORMATO_BANCARIO_CONFIG
            )
            await self.db.commit()
            return

        for version in formato.versiones:
            if version.estado == "vigente":
                version.estado = "reemplazada"
        proxima_version = max(version.version for version in formato.versiones) + 1
        await self._crear_version_formato_base(formato, version_numero=proxima_version)
        await self.db.commit()

    async def _crear_version_formato_base(
        self, formato: FormatoImportacion, version_numero: int
    ) -> FormatoImportacionVersion:
        """Crea una version vigente del formato bancario global."""
        version = FormatoImportacionVersion(
            formato_id=formato.id,
            version=version_numero,
            estado="vigente",
            configuracion_json=FORMATO_BANCARIO_CONFIG,
            headers_firma_json={
                "requeridos": ["Creditos", "Pto Vta"],
                "opcionales": [
                    "Fecha",
                    "Leyendas Adicionales1",
                    "Leyendas Adicionales2",
                ],
            },
        )
        self.db.add(version)
        await self.db.flush()
        self._agregar_campos_desde_config(version.id, FORMATO_BANCARIO_CONFIG)
        self.db.add(
            FormatoImportacionRegla(
                version_id=version.id,
                nombre="Cada fila genera un comprobante",
                tipo="agrupacion",
                configuracion_json={"modo": "fila"},
                orden=1,
                activo=True,
            )
        )
        return version

    async def listar_formatos(self, empresa_id: int) -> list[FormatoImportacion]:
        """Lista formatos globales y formatos particulares del emisor."""
        await self.asegurar_formatos_base()
        result = await self.db.execute(
            select(FormatoImportacion)
            .options(selectinload(FormatoImportacion.versiones))
            .where(
                FormatoImportacion.activo.is_(True),
                or_(
                    FormatoImportacion.alcance == "global",
                    FormatoImportacion.empresa_id == empresa_id,
                ),
            )
            .order_by(FormatoImportacion.alcance, FormatoImportacion.nombre)
        )
        return list(result.scalars().all())

    async def crear_formato(
        self,
        empresa_id: int,
        nombre: str,
        descripcion: str | None,
        configuracion: dict[str, Any],
        alcance: str = "emisor",
    ) -> FormatoImportacion:
        """Crea una plantilla con versión inicial."""
        self._validar_alcance(alcance)
        configuracion = self._normalizar_configuracion_usuario(configuracion)
        self._validar_configuracion_formato(
            configuracion,
            exigir_constantes_requeridas=True,
        )
        formato = FormatoImportacion(
            nombre=nombre,
            descripcion=descripcion,
            alcance=alcance,
            empresa_id=empresa_id if alcance == "emisor" else None,
            activo=True,
        )
        self.db.add(formato)
        await self.db.flush()
        version = FormatoImportacionVersion(
            formato_id=formato.id,
            version=1,
            estado="vigente",
            configuracion_json=configuracion,
            headers_firma_json=self._construir_firma_headers(configuracion),
        )
        self.db.add(version)
        await self.db.flush()
        self._agregar_campos_desde_config(version.id, configuracion)
        await self.db.commit()
        await self.db.refresh(formato)
        return formato

    async def obtener_formato(
        self, formato_id: int, empresa_id: int
    ) -> FormatoImportacion:
        """Obtiene una plantilla accesible desde el emisor activo."""
        await self.asegurar_formatos_base()
        result = await self.db.execute(
            select(FormatoImportacion)
            .options(
                selectinload(FormatoImportacion.versiones).selectinload(
                    FormatoImportacionVersion.campos
                ),
                selectinload(FormatoImportacion.versiones).selectinload(
                    FormatoImportacionVersion.reglas
                ),
            )
            .where(FormatoImportacion.id == formato_id)
            .execution_options(populate_existing=True)
        )
        formato = result.scalar_one_or_none()
        if formato is None or not formato.activo:
            raise FormatoImportacionError("No se encontró la plantilla seleccionada")
        if formato.alcance != "global" and formato.empresa_id != empresa_id:
            raise FormatoImportacionError(
                "La plantilla seleccionada no pertenece al emisor activo"
            )
        return formato

    async def actualizar_formato(
        self,
        formato_id: int,
        empresa_id: int,
        nombre: str | None = None,
        descripcion: str | None = None,
        actualizar_descripcion: bool = False,
        configuracion: dict[str, Any] | None = None,
        alcance: str | None = None,
    ) -> FormatoImportacion:
        """Actualiza datos y crea una nueva versión vigente si cambia el mapeo."""
        formato = await self.obtener_formato(formato_id, empresa_id)
        self._validar_no_protegido(formato)
        if alcance is not None:
            self._validar_alcance(alcance)
            if alcance != formato.alcance:
                raise FormatoImportacionError(
                    "No se puede cambiar el alcance de una plantilla existente. "
                    "Cloná la plantilla con el nuevo alcance."
                )
            formato.alcance = alcance
            formato.empresa_id = empresa_id if alcance == "emisor" else None
        if nombre is not None:
            formato.nombre = nombre
        if actualizar_descripcion:
            formato.descripcion = descripcion

        if configuracion is not None:
            configuracion = self._normalizar_configuracion_usuario(configuracion)
            self._validar_configuracion_formato(
                configuracion,
                exigir_constantes_requeridas=True,
            )
            version_vigente = self._version_vigente(formato)
            if (
                version_vigente is None
                or version_vigente.configuracion_json != configuracion
            ):
                for version in formato.versiones:
                    if version.estado == "vigente":
                        version.estado = "reemplazada"
                proxima_version = (
                    max((version.version for version in formato.versiones), default=0)
                    + 1
                )
                version = FormatoImportacionVersion(
                    formato_id=formato.id,
                    version=proxima_version,
                    estado="vigente",
                    configuracion_json=configuracion,
                    headers_firma_json=self._construir_firma_headers(configuracion),
                )
                self.db.add(version)
                await self.db.flush()
                self._agregar_campos_desde_config(version.id, configuracion)

        await self.db.commit()
        return await self.obtener_formato(formato.id, empresa_id)

    async def desactivar_formato(self, formato_id: int, empresa_id: int) -> None:
        """Desactiva una plantilla sin borrar versiones usadas por lotes."""
        formato = await self.obtener_formato(formato_id, empresa_id)
        self._validar_no_protegido(formato)
        formato.activo = False
        await self.db.commit()

    async def clonar_formato(
        self,
        formato_id: int,
        empresa_id: int,
        nombre: str | None = None,
        alcance: str = "emisor",
    ) -> FormatoImportacion:
        """Clona una plantilla accesible y la deja editable."""
        formato = await self.obtener_formato(formato_id, empresa_id)
        version = self._version_vigente(formato)
        if version is None:
            raise FormatoImportacionError("La plantilla no tiene una versión vigente")
        configuracion = self._normalizar_configuracion_usuario(
            deepcopy(version.configuracion_json)
        )
        configuracion["plantilla_clonada_de"] = {
            "formato_id": formato.id,
            "version_id": version.id,
        }
        return await self.crear_formato(
            empresa_id=empresa_id,
            nombre=nombre or f"Copia de {formato.nombre}",
            descripcion=formato.descripcion,
            configuracion=configuracion,
            alcance=alcance,
        )

    async def obtener_version(
        self, version_id: int, empresa_id: int, exigir_vigente: bool = True
    ) -> FormatoImportacionVersion:
        """Obtiene una versión accesible para el emisor activo."""
        await self.asegurar_formatos_base()
        result = await self.db.execute(
            select(FormatoImportacionVersion)
            .options(
                selectinload(FormatoImportacionVersion.formato),
                selectinload(FormatoImportacionVersion.campos),
                selectinload(FormatoImportacionVersion.reglas),
            )
            .where(FormatoImportacionVersion.id == version_id)
        )
        version = result.scalar_one_or_none()
        if version is None or not version.formato.activo:
            raise FormatoImportacionError("No se encontró el formato seleccionado")
        if exigir_vigente and version.estado != "vigente":
            raise FormatoImportacionError(
                "La versión de la plantilla ya no está vigente. "
                "Actualizá el perfil o elegí una plantilla vigente."
            )
        if (
            version.formato.alcance != "global"
            and version.formato.empresa_id != empresa_id
        ):
            raise FormatoImportacionError(
                "El formato seleccionado no pertenece al emisor activo"
            )
        return version

    async def detectar_formato(
        self,
        file_bytes: bytes,
        empresa_id: int,
        template_columns: list[str],
    ) -> tuple[list[str], list[CandidatoFormato]]:
        """Detecta candidatos por encabezados y cantidad de columnas."""
        headers = self.leer_headers(file_bytes)
        candidatos: list[CandidatoFormato] = []
        if self.es_plantilla_oficial(headers, template_columns):
            candidatos.append(
                CandidatoFormato(
                    formato_id=None,
                    formato_version_id=None,
                    nombre="Plantilla oficial FactuFlow",
                    alcance="sistema",
                    version=None,
                    score=1.0,
                    confianza="alta",
                    columnas_detectadas=template_columns,
                    columnas_faltantes=[],
                    mensajes=["El archivo coincide con la plantilla oficial."],
                )
            )

        formatos = await self.listar_formatos(empresa_id)
        for formato in formatos:
            version = self._version_vigente(formato)
            if version is None:
                continue
            candidatos.append(self._evaluar_formato(headers, formato, version))

        candidatos.sort(key=lambda item: item.score, reverse=True)
        return headers, candidatos

    def analizar_excel(self, file_bytes: bytes) -> ExcelAnalisis:
        """Analiza encabezados de un Excel para iniciar una plantilla."""
        sheet, headers = self._leer_sheet_y_headers(file_bytes)
        columnas = [
            ExcelColumnaAnalizada(
                indice=index,
                letra=get_column_letter(index + 1),
                encabezado=header,
            )
            for index, header in enumerate(headers)
            if str(header or "").strip()
        ]
        return ExcelAnalisis(
            hoja=sheet.title,
            fila_encabezado=1,
            columnas=columnas,
        )

    async def evaluar_compatibilidad(
        self,
        configuracion: dict[str, Any],
        empresa_id: int,
        perfil_configuracion: dict[str, Any] | None = None,
    ) -> CompatibilidadPlantilla:
        """Evalúa compatibilidad entre plantilla, perfil de carga y emisor."""
        empresa = await self._obtener_empresa(empresa_id)
        configuracion = self._normalizar_configuracion_usuario(configuracion)
        self._validar_configuracion_formato(configuracion)
        perfil_configuracion = perfil_configuracion or {}
        campos = configuracion.get("campos", {})
        resolvedores = {
            campo
            for campo, detalle in campos.items()
            if self._campo_resuelto_por_plantilla(campo, detalle)
        }
        if "importe_total" in resolvedores:
            # El importador usa importe_total como fallback para precio unitario.
            resolvedores.add("item_precio_unitario")
        faltantes: list[CompatibilidadMensaje] = []
        omitibles: list[CompatibilidadMensaje] = []
        advertencias: list[CompatibilidadMensaje] = []
        conflictos: list[CompatibilidadMensaje] = []

        def agregar(
            bucket: list[CompatibilidadMensaje],
            codigo: str,
            campo: str | None,
            severidad: str,
            mensaje: str,
        ) -> None:
            bucket.append(
                CompatibilidadMensaje(
                    codigo=codigo,
                    campo=campo,
                    severidad=severidad,
                    mensaje=mensaje,
                )
            )

        def resuelto(campo: str) -> bool:
            return campo in resolvedores or self._campo_resuelto_por_perfil(
                campo, perfil_configuracion
            )

        def exige_desde_archivo(campo: str) -> bool:
            return self._perfil_exige_campo_desde_archivo(campo, perfil_configuracion)

        campos_requeridos_base = [
            "tipo_comprobante",
            "item_precio_unitario",
            "item_iva_porcentaje",
        ]
        for campo in campos_requeridos_base:
            if not resuelto(campo):
                agregar(
                    faltantes,
                    f"falta_{campo}",
                    campo,
                    "error",
                    f"Falta {self._etiqueta_campo(campo)}.",
                )

        if not resuelto("fecha_emision"):
            if exige_desde_archivo("fecha_emision"):
                agregar(
                    conflictos,
                    "perfil_fecha_archivo_sin_columna",
                    "fecha_emision",
                    "error",
                    "El perfil espera fecha de emisión desde archivo, pero la "
                    "plantilla no tiene esa columna.",
                )
            else:
                agregar(
                    advertencias,
                    "fecha_emision_manual",
                    "fecha_emision",
                    "warning",
                    "Falta fecha de emisión en la plantilla; deberá quedar "
                    "definida antes de validar el lote.",
                )

        if not resuelto("punto_venta_numero"):
            if exige_desde_archivo("punto_venta_numero"):
                agregar(
                    conflictos,
                    "perfil_punto_venta_archivo_sin_columna",
                    "punto_venta_numero",
                    "error",
                    "El perfil espera punto de venta desde archivo, pero la "
                    "plantilla no tiene esa columna.",
                )
            else:
                agregar(
                    advertencias,
                    "punto_venta_manual",
                    "punto_venta_numero",
                    "warning",
                    "Falta punto de venta en la plantilla; se completará desde "
                    "el perfil o en Emisión masiva.",
                )

        if not resuelto("item_descripcion"):
            if exige_desde_archivo("item_descripcion"):
                agregar(
                    conflictos,
                    "perfil_descripcion_archivo_sin_columna",
                    "item_descripcion",
                    "error",
                    "El perfil espera descripción desde archivo, pero la "
                    "plantilla no tiene esa columna.",
                )
            elif self._campo_resuelto_por_perfil(
                "item_descripcion", perfil_configuracion
            ):
                agregar(
                    omitibles,
                    "descripcion_fija_perfil",
                    "item_descripcion",
                    "info",
                    "No hace falta descripción porque el perfil fija un texto "
                    "para todos los ítems.",
                )
            else:
                agregar(
                    advertencias,
                    "descripcion_manual",
                    "item_descripcion",
                    "warning",
                    "Falta descripción; deberá completarse antes de validar "
                    "si la plantilla no la trae.",
                )

        if not resuelto("item_cantidad"):
            if (
                self._concepto_principal(configuracion, perfil_configuracion)
                in CONCEPTOS_SERVICIOS
            ):
                agregar(
                    omitibles,
                    "cantidad_omitible_servicios",
                    "item_cantidad",
                    "info",
                    "No hace falta cantidad para servicios si se fija en 1 "
                    "desde la plantilla o el perfil de emisión.",
                )
            else:
                agregar(
                    advertencias,
                    "cantidad_manual",
                    "item_cantidad",
                    "warning",
                    "Falta cantidad; para productos conviene traerla desde el "
                    "archivo o fijarla explícitamente.",
                )

        concepto = self._concepto_principal(configuracion, perfil_configuracion)
        if concepto in CONCEPTOS_SERVICIOS:
            for campo in (
                "fecha_servicio_desde",
                "fecha_servicio_hasta",
                "fecha_vto_pago",
            ):
                if not resuelto(campo):
                    if exige_desde_archivo(campo):
                        agregar(
                            conflictos,
                            f"perfil_{campo}_archivo_sin_columna",
                            campo,
                            "error",
                            f"El perfil espera {self._etiqueta_campo(campo)} "
                            "desde archivo, pero la plantilla no tiene esa "
                            "columna.",
                        )
                    else:
                        agregar(
                            advertencias,
                            f"{campo}_servicio_manual",
                            campo,
                            "warning",
                            f"Para servicios debe quedar definida la columna "
                            f"{self._etiqueta_campo(campo)} o una regla "
                            "equivalente antes de validar.",
                        )
        elif concepto in CONCEPTOS_PRODUCTOS:
            for campo in (
                "fecha_servicio_desde",
                "fecha_servicio_hasta",
                "fecha_vto_pago",
            ):
                if campo not in resolvedores:
                    agregar(
                        omitibles,
                        f"{campo}_omitible_productos",
                        campo,
                        "info",
                        f"No hace falta {self._etiqueta_campo(campo)} para "
                        "productos.",
                    )

        tipo = self._tipo_comprobante_configurado(configuracion)
        if tipo in TIPOS_C and empresa.condicion_iva == "RI":
            agregar(
                conflictos,
                "ri_no_emite_c",
                "tipo_comprobante",
                "error",
                "Responsable Inscripto no es compatible con comprobantes C.",
            )
        if tipo in TIPOS_A and empresa.condicion_iva in {"Monotributo", "Exento"}:
            agregar(
                conflictos,
                "monotributo_exento_no_emite_a",
                "tipo_comprobante",
                "error",
                "Monotributo o Exento no es compatible con comprobantes A.",
            )
        if tipo in TIPOS_A | TIPOS_B and empresa.condicion_iva == "RI":
            agregar(
                advertencias,
                "ri_tipo_a_b",
                "tipo_comprobante",
                "warning",
                "Responsable Inscripto puede requerir elegir A o B según el "
                "receptor; revisa si la plantilla debe traer tipo de "
                "comprobante.",
            )
        if tipo in TIPOS_C and empresa.condicion_iva in {"Monotributo", "Exento"}:
            agregar(
                omitibles,
                "monotributo_tipo_c",
                "tipo_comprobante",
                "info",
                "Para Monotributo o Exento el tipo C puede quedar fijo en la "
                "plantilla.",
            )

        if tipo in TIPOS_NOTA:
            for campo in (
                "asociado_tipo_comprobante",
                "asociado_punto_venta",
                "asociado_numero",
            ):
                if not resuelto(campo):
                    agregar(
                        faltantes,
                        f"nota_falta_{campo}",
                        campo,
                        "error",
                        "Las notas de crédito y débito requieren comprobante "
                        "asociado: tipo, punto de venta y número.",
                    )

        if not conflictos and not faltantes:
            estado = "compatible" if not advertencias else "advertencia"
        else:
            estado = "incompatible"

        return CompatibilidadPlantilla(
            estado=estado,
            faltantes=faltantes,
            omitibles=omitibles,
            advertencias=advertencias,
            conflictos=conflictos,
        )

    async def generar_excel_formato(
        self,
        formato_id: int,
        empresa_id: int,
    ) -> tuple[bytes, str]:
        """Genera el archivo XLSX visual de una plantilla."""
        formato = await self.obtener_formato(formato_id, empresa_id)
        version = self._version_vigente(formato)
        if version is None:
            raise FormatoImportacionError("La plantilla no tiene versión vigente")
        columnas = self._columnas_generacion(version.configuracion_json)
        if not columnas:
            raise FormatoImportacionError("La plantilla no tiene columnas descargables")

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Comprobantes"
        instrucciones = workbook.create_sheet("Instrucciones")
        metadata = workbook.create_sheet("_factuflow")
        metadata.sheet_state = "hidden"

        header_fill = PatternFill("solid", fgColor="1D4ED8")
        header_font = Font(color="FFFFFF", bold=True)
        required_fill = PatternFill("solid", fgColor="DBEAFE")
        optional_fill = PatternFill("solid", fgColor="F8FAFC")

        columnas_usadas: set[int] = set()
        siguiente_columna = 1

        for columna in columnas:
            indice_fijo = columna.get("indice_columna")
            if indice_fijo is not None:
                column_index = int(indice_fijo) + 1
            else:
                while siguiente_columna in columnas_usadas:
                    siguiente_columna += 1
                column_index = siguiente_columna
                siguiente_columna += 1

            if column_index in columnas_usadas:
                raise FormatoImportacionError(
                    "La plantilla tiene más de un campo asignado a la misma columna"
                )
            columnas_usadas.add(column_index)

            etiqueta = str(columna["etiqueta"])
            cell = sheet.cell(
                row=1,
                column=column_index,
                value=self._valor_excel_seguro(etiqueta),
            )
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            sheet.cell(row=2, column=column_index).fill = (
                required_fill if columna.get("requerido") else optional_fill
            )
            sheet.column_dimensions[get_column_letter(column_index)].width = max(
                16, min(len(etiqueta) + 4, 42)
            )
            if columna.get("campo_destino") == "tipo_comprobante":
                self._agregar_validacion_lista(
                    sheet,
                    column_index,
                    ["1", "2", "3", "6", "7", "8", "11", "12", "13"],
                )
            if columna.get("campo_destino") == "concepto":
                self._agregar_validacion_lista(sheet, column_index, ["1", "2", "3"])
            if columna.get("campo_destino") == "item_iva_porcentaje":
                self._agregar_validacion_lista(
                    sheet,
                    column_index,
                    ["0", "10.5", "21", "27"],
                )

        sheet.freeze_panes = "A2"

        instrucciones["A1"] = "Plantilla de carga masiva"
        instrucciones["A1"].font = Font(bold=True, size=14)
        instrucciones["A3"] = "Uso"
        instrucciones["A3"].font = Font(bold=True)
        instrucciones["A4"] = (
            "Carga datos reales desde la fila 2 de la hoja Comprobantes. "
            "Los ejemplos están solo en esta hoja y no completan valores "
            "fiscales por defecto."
        )
        instrucciones["A6"] = "Campos de la plantilla"
        instrucciones["A6"].font = Font(bold=True)
        instrucciones["A7"] = "Columna"
        instrucciones["B7"] = "Campo FactuFlow"
        instrucciones["C7"] = "Requerido"
        instrucciones["D7"] = "Ejemplo"
        for cell in instrucciones[7]:
            cell.font = Font(bold=True)
        fila = 8
        for columna in columnas:
            instrucciones.cell(
                row=fila,
                column=1,
                value=self._valor_excel_seguro(columna["etiqueta"]),
            )
            instrucciones.cell(
                row=fila,
                column=2,
                value=self._valor_excel_seguro(
                    self._etiqueta_campo(columna["campo_destino"])
                ),
            )
            instrucciones.cell(
                row=fila,
                column=3,
                value="Sí" if columna.get("requerido") else "No",
            )
            instrucciones.cell(
                row=fila,
                column=4,
                value=self._valor_excel_seguro(columna.get("ejemplo", "")),
            )
            fila += 1
        instrucciones.column_dimensions["A"].width = 32
        instrucciones.column_dimensions["B"].width = 40
        instrucciones.column_dimensions["C"].width = 14
        instrucciones.column_dimensions["D"].width = 28

        metadata["A1"] = "factuflow_template_metadata"
        metadata["A2"] = "formato_id"
        metadata["B2"] = formato.id
        metadata["A3"] = "formato_version_id"
        metadata["B3"] = version.id
        metadata["A4"] = "alcance"
        metadata["B4"] = formato.alcance
        metadata["A5"] = "nombre"
        metadata["B5"] = self._valor_excel_seguro(formato.nombre)
        metadata["A6"] = "advertencia"
        metadata[
            "B6"
        ] = "Metadatos no fiscales. FactuFlow no confía en esta hoja para emitir."

        output = BytesIO()
        workbook.save(output)
        filename = self._nombre_archivo_seguro(formato.nombre)
        return output.getvalue(), filename

    async def importar_con_version(
        self,
        file_bytes: bytes,
        empresa: Empresa,
        version: FormatoImportacionVersion,
    ) -> ImportacionNormalizada:
        """Convierte un Excel externo al formato interno de lote."""
        sheet, headers = self._leer_sheet_y_headers(file_bytes, version)
        mapeo = self._resolver_mapeo(headers, version.configuracion_json)
        faltantes = [
            campo
            for campo, detalle in mapeo["campos"].items()
            if detalle.get("requerido") and not detalle.get("encontrado")
        ]
        if faltantes:
            raise FormatoImportacionError(
                "El archivo no tiene columnas requeridas para este formato: "
                + ", ".join(faltantes)
            )

        header_row = int(version.configuracion_json.get("header_row", 1))
        filas: list[dict[str, Any]] = []
        for fila_excel, row in enumerate(
            sheet.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            if all(cell in (None, "") for cell in row):
                continue
            valores = self._extraer_valores_configurados(row, mapeo)
            filas.append(self._armar_fila_canonica(valores, empresa, fila_excel))

        if not filas:
            raise FormatoImportacionError(
                "La plantilla no contiene filas para procesar"
            )

        return ImportacionNormalizada(
            filas=filas,
            headers_detectados=headers,
            mapeo_usado=mapeo,
            formato=version.formato,
            version=version,
        )

    def leer_headers(self, file_bytes: bytes) -> list[str]:
        """Lee los encabezados de la hoja más probable del Excel."""
        sheet, headers = self._leer_sheet_y_headers(file_bytes)
        return headers

    def es_plantilla_oficial(
        self, headers: list[str], template_columns: list[str]
    ) -> bool:
        """Indica si los encabezados coinciden con la plantilla oficial."""
        normalized = {self.normalizar_etiqueta(header) for header in headers}
        return all(
            self.normalizar_etiqueta(col) in normalized for col in template_columns
        )

    @classmethod
    def normalizar_etiqueta(cls, value: Any) -> str:
        """Normaliza encabezados para tolerar acentos, espacios y símbolos."""
        text = cls.reparar_texto(value)
        decomposed = unicodedata.normalize("NFKD", text)
        ascii_text = decomposed.encode("ascii", "ignore").decode("ascii")
        return re.sub(r"[^a-z0-9]+", "", ascii_text.lower())

    @classmethod
    def reparar_texto(cls, value: Any) -> str:
        """Corrige texto comúnmente mojibakeado en encabezados de Excel."""
        text = str(value or "").strip()
        if "Ã" not in text and "Â" not in text:
            return text
        try:
            return text.encode("latin1").decode("utf-8")
        except UnicodeError:
            return text

    async def _obtener_empresa(self, empresa_id: int) -> Empresa:
        """Obtiene el emisor usado para evaluar compatibilidad."""
        result = await self.db.execute(select(Empresa).where(Empresa.id == empresa_id))
        empresa = result.scalar_one_or_none()
        if empresa is None:
            raise FormatoImportacionError("No se encontró el emisor activo")
        return empresa

    def _validar_alcance(self, alcance: str) -> None:
        """Valida el alcance admitido para plantillas."""
        if alcance not in {"global", "emisor"}:
            raise FormatoImportacionError("El alcance de la plantilla es inválido")

    def _validar_no_protegido(self, formato: FormatoImportacion) -> None:
        """Evita editar o desactivar plantillas internas del sistema."""
        version = self._version_vigente(formato)
        if version and version.configuracion_json.get("plantilla_sistema_protegida"):
            raise FormatoImportacionError(
                "Esta plantilla es interna del sistema. Podés clonarla, pero no "
                "editarla directamente."
            )

    def _normalizar_configuracion_usuario(
        self, configuracion: dict[str, Any]
    ) -> dict[str, Any]:
        """Normaliza una configuración editable sin metadatos protegidos."""
        normalizada = deepcopy(configuracion)
        normalizada.pop("plantilla_sistema_protegida", None)
        plantilla = normalizada.get("plantilla")
        if isinstance(plantilla, dict) and isinstance(plantilla.get("columnas"), list):
            campos = dict(normalizada.get("campos") or {})
            for orden, columna in enumerate(plantilla["columnas"], start=1):
                if not isinstance(columna, dict):
                    continue
                orden_configurado = self._parse_int_laxo(columna.get("orden"))
                columna["orden"] = (
                    orden_configurado
                    if orden_configurado is not None and orden_configurado > 0
                    else orden
                )
                campo = str(columna.get("campo_destino") or "").strip()
                if not campo:
                    continue
                origen = str(columna.get("origen") or "header")
                detalle_existente = campos.get(campo)
                if not isinstance(detalle_existente, dict):
                    detalle_existente = {}
                detalle: dict[str, Any] = {
                    "origen": origen,
                    "requerido": bool(columna.get("requerido", False)),
                }
                transformacion = columna.get("transformacion")
                if transformacion:
                    detalle["transformacion"] = transformacion
                default = columna.get("default")
                if default not in (None, ""):
                    detalle["default"] = default
                if origen == "header":
                    etiqueta = str(columna.get("etiqueta") or campo).strip()
                    encabezados = [etiqueta]
                    if detalle_existente.get("origen", "header") == "header":
                        encabezados.extend(detalle_existente.get("encabezados") or [])
                    detalle["encabezados"] = self._deduplicar_textos(encabezados)
                elif origen == "columna":
                    if columna.get("letra_columna"):
                        detalle["letra_columna"] = columna.get("letra_columna")
                    if columna.get("indice_columna") is not None:
                        detalle["indice_columna"] = columna.get("indice_columna")
                elif origen == "constante":
                    detalle["valor"] = columna.get("valor")
                campos[campo] = detalle
            normalizada["campos"] = campos
        return normalizada

    def _deduplicar_textos(self, valores: list[Any]) -> list[str]:
        """Conserva textos únicos respetando el orden original."""
        vistos: set[str] = set()
        resultado: list[str] = []
        for valor in valores:
            texto = str(valor or "").strip()
            if not texto:
                continue
            clave = texto.casefold()
            if clave in vistos:
                continue
            vistos.add(clave)
            resultado.append(texto)
        return resultado

    def _campo_resuelto_por_plantilla(
        self, campo: str, detalle: dict[str, Any]
    ) -> bool:
        """Indica si una configuración aporta el campo al procesamiento."""
        origen = detalle.get("origen", "header")
        if origen == "constante":
            return not self._valor_constante_vacio(detalle.get("valor"))
        if origen == "empresa":
            return campo in CAMPOS_ORIGEN_EMPRESA
        if origen == "header":
            return bool(detalle.get("encabezados"))
        if origen == "columna":
            return detalle.get("indice_columna") is not None or bool(
                detalle.get("letra_columna")
            )
        return False

    def _valor_constante_vacio(self, valor: Any) -> bool:
        """Indica si un valor fijo no aporta un dato usable."""
        if valor is None:
            return True
        if isinstance(valor, str) and not valor.strip():
            return True
        return False

    def _validar_constante_fiscal(self, campo: str, valor: Any) -> None:
        """Valida valores fijos fiscales contra dominios aceptados por lotes."""
        if self._valor_constante_vacio(valor):
            return
        if campo == "tipo_comprobante":
            tipo = self._parse_int_fiscal(valor)
            if tipo not in TIPOS_COMPROBANTE_SOPORTADOS:
                raise FormatoImportacionError(
                    "El tipo de comprobante fijo debe ser 1, 2, 3, 6, 7, "
                    "8, 11, 12 o 13"
                )
        elif campo == "concepto":
            concepto = self._parse_int_fiscal(valor)
            if concepto not in CONCEPTOS_PRODUCTOS | CONCEPTOS_SERVICIOS:
                raise FormatoImportacionError("El concepto fijo debe ser 1, 2 o 3")
        elif campo == "item_iva_porcentaje":
            iva = self._parse_decimal(valor)
            if iva not in ALICUOTAS_IVA_PERMITIDAS:
                raise FormatoImportacionError(
                    "La alícuota de IVA fija debe ser 0, 10.5, 21 o 27"
                )

    def _campo_resuelto_por_perfil(
        self, campo: str, perfil_configuracion: dict[str, Any]
    ) -> bool:
        """Indica si un perfil evita pedir un campo desde la plantilla."""
        if campo == "punto_venta_numero":
            punto_venta = perfil_configuracion.get("punto_venta") or {}
            return punto_venta.get("modo") == "fijo" and bool(punto_venta.get("numero"))
        if campo == "concepto":
            return perfil_configuracion.get("concepto_modo") in {
                "productos",
                "servicios",
            }
        if campo == "item_descripcion":
            return perfil_configuracion.get("descripcion_item_modo") == "fija" and bool(
                perfil_configuracion.get("descripcion_item_fija")
            )
        if campo == "fecha_emision":
            fecha_emision = perfil_configuracion.get("fecha_emision") or {}
            return fecha_emision.get("modo") in {
                "ultimo_dia_mes_anterior",
                "personalizada",
            }
        if campo in {"fecha_servicio_desde", "fecha_servicio_hasta"}:
            periodo = perfil_configuracion.get("periodo_servicio") or {}
            return periodo.get("modo") in {
                "mes_anterior_completo",
                "mes_actual_completo",
                "personalizado",
            }
        if campo == "fecha_vto_pago":
            vencimiento = perfil_configuracion.get("fecha_vto_pago") or {}
            return vencimiento.get("modo") in {
                "mismo_dia_emision",
                "emision_mas_dias",
                "personalizada",
            }
        return False

    def _perfil_exige_campo_desde_archivo(
        self, campo: str, perfil_configuracion: dict[str, Any]
    ) -> bool:
        """Indica si el perfil declara que el campo debe venir del Excel."""
        if campo == "punto_venta_numero":
            punto_venta = perfil_configuracion.get("punto_venta") or {}
            return punto_venta.get("modo") == "archivo"
        if campo == "concepto":
            return perfil_configuracion.get("concepto_modo") == "archivo"
        if campo == "item_descripcion":
            return perfil_configuracion.get("descripcion_item_modo") == "archivo"
        if campo == "fecha_emision":
            fecha_emision = perfil_configuracion.get("fecha_emision") or {}
            return fecha_emision.get("modo") == "archivo"
        if campo in {"fecha_servicio_desde", "fecha_servicio_hasta"}:
            periodo = perfil_configuracion.get("periodo_servicio") or {}
            return periodo.get("modo") == "archivo"
        if campo == "fecha_vto_pago":
            vencimiento = perfil_configuracion.get("fecha_vto_pago") or {}
            return vencimiento.get("modo") == "archivo"
        return False

    def _concepto_principal(
        self,
        configuracion: dict[str, Any],
        perfil_configuracion: dict[str, Any],
    ) -> int | None:
        """Resuelve el concepto fiscal fijo si la plantilla o perfil lo indican."""
        campo = configuracion.get("campos", {}).get("concepto") or {}
        if campo.get("origen") == "constante":
            return self._parse_int_fiscal(campo.get("valor"))
        modo = perfil_configuracion.get("concepto_modo")
        if modo == "productos":
            return 1
        if modo == "servicios":
            return 2
        return None

    def _tipo_comprobante_configurado(
        self, configuracion: dict[str, Any]
    ) -> int | None:
        """Devuelve el tipo fijo de comprobante si está configurado."""
        campo = configuracion.get("campos", {}).get("tipo_comprobante") or {}
        if campo.get("origen") == "constante":
            return self._parse_int_fiscal(campo.get("valor"))
        return None

    def _parse_int_fiscal(self, value: Any) -> int | None:
        """Convierte constantes fiscales enteras sin truncar fracciones."""
        if isinstance(value, bool):
            return None
        if isinstance(value, str):
            text = value.strip()
            if not re.fullmatch(r"[+-]?\d+", text):
                return None
            return int(text)
        try:
            decimal_value = Decimal(str(value))
        except (InvalidOperation, TypeError, ValueError):
            return None
        if not decimal_value.is_finite():
            return None
        if decimal_value != decimal_value.to_integral_value():
            return None
        return int(decimal_value)

    def _parse_int_laxo(self, value: Any) -> int | None:
        """Convierte valores simples a entero sin romper compatibilidad."""
        try:
            return int(Decimal(str(value)))
        except (InvalidOperation, TypeError, ValueError):
            return None

    def _etiqueta_campo(self, campo: str | None) -> str:
        """Devuelve la etiqueta pública de un campo."""
        if not campo:
            return "campo"
        for item in CATALOGO_CAMPOS_PLANTILLA:
            if item["codigo"] == campo:
                return str(item["etiqueta"])
        return campo.replace("_", " ")

    def _columnas_generacion(
        self, configuracion: dict[str, Any]
    ) -> list[dict[str, Any]]:
        """Resuelve columnas visibles que tendrá el Excel generado."""
        plantilla = configuracion.get("plantilla") or {}
        columnas_visuales = (
            plantilla.get("columnas") if isinstance(plantilla, dict) else None
        )
        columnas: list[dict[str, Any]] = []
        if isinstance(columnas_visuales, list):
            for orden, columna in enumerate(columnas_visuales):
                if not isinstance(columna, dict):
                    continue
                origen = columna.get("origen", "header")
                if origen not in {"header", "columna"}:
                    continue
                campo = str(columna.get("campo_destino") or "").strip()
                if not campo:
                    continue
                indice_columna = (
                    self._resolver_indice_columna(columna)
                    if origen == "columna"
                    else None
                )
                if origen == "columna" and indice_columna is None:
                    raise FormatoImportacionError(
                        f"El campo {campo} debe declarar letra o índice de columna"
                    )
                columnas.append(
                    {
                        "campo_destino": campo,
                        "etiqueta": str(
                            columna.get("etiqueta") or self._etiqueta_campo(campo)
                        ).strip(),
                        "origen": origen,
                        "indice_columna": indice_columna,
                        "requerido": bool(columna.get("requerido", False)),
                        "ejemplo": columna.get("ejemplo", ""),
                        "orden": self._orden_visual_columna(columna, orden),
                    }
                )
            return sorted(columnas, key=lambda item: item["orden"])

        for orden, (campo, detalle) in enumerate(
            configuracion.get("campos", {}).items()
        ):
            if detalle.get("origen", "header") not in {"header", "columna"}:
                continue
            encabezados = detalle.get("encabezados") or []
            indice_columna = (
                self._resolver_indice_columna(detalle)
                if detalle.get("origen") == "columna"
                else None
            )
            if detalle.get("origen") == "columna" and indice_columna is None:
                raise FormatoImportacionError(
                    f"El campo {campo} debe declarar letra o índice de columna"
                )
            columnas.append(
                {
                    "campo_destino": campo,
                    "etiqueta": encabezados[0]
                    if encabezados
                    else self._etiqueta_campo(campo),
                    "origen": detalle.get("origen", "header"),
                    "indice_columna": indice_columna,
                    "requerido": bool(detalle.get("requerido", False)),
                    "ejemplo": detalle.get("ejemplo", ""),
                    "orden": orden,
                }
            )
        return columnas

    def _agregar_validacion_lista(
        self,
        sheet,
        column_index: int,
        valores: list[str],
    ) -> None:
        """Agrega una validación de lista simple a una columna."""
        formula = '"' + ",".join(valores) + '"'
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        sheet.add_data_validation(validation)
        letter = get_column_letter(column_index)
        validation.add(f"{letter}2:{letter}5000")

    def _orden_visual_columna(self, columna: dict[str, Any], fallback: int) -> int:
        """Obtiene un orden de columna tolerante para metadatos legacy."""
        orden = self._parse_int_laxo(columna.get("orden"))
        if orden is None or orden <= 0:
            return fallback
        return orden

    def _valor_excel_seguro(self, value: Any) -> Any:
        """Neutraliza textos configurables que Excel interpretaría como fórmula."""
        if not isinstance(value, str) or value == "":
            return value
        stripped = value.lstrip()
        if stripped.startswith(EXCEL_FORMULA_PREFIXES):
            return f"'{value}"
        return value

    def _nombre_archivo_seguro(self, nombre: str) -> str:
        """Construye un nombre de archivo seguro para descargar."""
        base = self.normalizar_etiqueta(nombre) or "plantilla"
        return f"plantilla-{base}.xlsx"

    def _agregar_campos_desde_config(
        self, version_id: int, configuracion: dict[str, Any]
    ) -> None:
        """Persiste una copia relacional del mapeo para inspección futura."""
        for campo_destino, config in configuracion.get("campos", {}).items():
            encabezados = config.get("encabezados") or []
            self.db.add(
                FormatoImportacionCampo(
                    version_id=version_id,
                    campo_destino=campo_destino,
                    origen_tipo=config.get("origen", "header"),
                    encabezado=encabezados[0] if encabezados else None,
                    alias_json=encabezados or None,
                    letra_columna=config.get("letra_columna"),
                    indice_columna=config.get("indice_columna"),
                    valor_constante_json=config.get("valor"),
                    requerido=bool(config.get("requerido", False)),
                    transformacion=config.get("transformacion"),
                    valor_default_json=config.get("default"),
                )
            )

    def _validar_configuracion_formato(
        self,
        configuracion: dict[str, Any],
        exigir_constantes_requeridas: bool = False,
    ) -> None:
        """Valida la forma mínima del mapeo configurable antes de persistirlo."""
        if not isinstance(configuracion, dict):
            raise FormatoImportacionError("La configuración del formato es inválida")

        plantilla = configuracion.get("plantilla")
        if isinstance(plantilla, dict) and isinstance(plantilla.get("columnas"), list):
            campos_visuales: set[str] = set()
            for columna in plantilla["columnas"]:
                if not isinstance(columna, dict):
                    continue
                campo = str(columna.get("campo_destino") or "").strip()
                if not campo:
                    continue
                if campo in campos_visuales:
                    raise FormatoImportacionError(
                        f"La plantilla repite el campo {campo}"
                    )
                campos_visuales.add(campo)

        campos = configuracion.get("campos")
        if not isinstance(campos, dict) or not campos:
            raise FormatoImportacionError(
                "La configuración del formato debe incluir campos"
            )
        origenes_validos = {"header", "columna", "constante", "empresa"}
        for campo_destino, config in campos.items():
            if not isinstance(config, dict):
                raise FormatoImportacionError(
                    f"La configuración del campo {campo_destino} debe ser un objeto"
                )

            origen = config.get("origen", "header")
            if origen not in origenes_validos:
                raise FormatoImportacionError(
                    f"El campo {campo_destino} usa un origen no soportado"
                )

            if origen == "empresa" and campo_destino not in CAMPOS_ORIGEN_EMPRESA:
                raise FormatoImportacionError(
                    f"El campo {campo_destino} no puede usar datos del emisor"
                )

            constante_requerida = bool(config.get("requerido")) or (
                campo_destino in CAMPOS_CONSTANTE_REQUERIDA_NO_VACIA
            )
            if (
                exigir_constantes_requeridas
                and origen == "constante"
                and constante_requerida
                and self._valor_constante_vacio(config.get("valor"))
            ):
                raise FormatoImportacionError(
                    f"El campo {campo_destino} debe informar un valor constante"
                )
            if origen == "constante":
                self._validar_constante_fiscal(campo_destino, config.get("valor"))

            origenes_permitidos = ORIGENES_PERMITIDOS_POR_CAMPO.get(str(campo_destino))
            if origenes_permitidos is not None and origen not in origenes_permitidos:
                raise FormatoImportacionError(
                    f"El campo {campo_destino} no admite origen {origen}"
                )

            transformacion = config.get("transformacion")
            transformaciones_permitidas = TRANSFORMACIONES_PERMITIDAS_POR_CAMPO.get(
                str(campo_destino)
            )
            if (
                transformacion not in (None, "")
                and transformaciones_permitidas is not None
                and transformacion not in transformaciones_permitidas
            ):
                raise FormatoImportacionError(
                    f"El campo {campo_destino} no admite transformación "
                    f"{transformacion}"
                )

            if origen == "header":
                encabezados = config.get("encabezados")
                if not isinstance(encabezados, list) or not any(
                    str(item).strip() for item in encabezados
                ):
                    raise FormatoImportacionError(
                        f"El campo {campo_destino} debe declarar encabezados"
                    )
            elif origen == "columna":
                if self._resolver_indice_columna(config) is None:
                    raise FormatoImportacionError(
                        f"El campo {campo_destino} debe declarar letra o índice "
                        "de columna"
                    )

        tipo_constante = self._parse_int_fiscal(
            (campos.get("tipo_comprobante") or {}).get("valor")
            if isinstance(campos.get("tipo_comprobante"), dict)
            and (campos.get("tipo_comprobante") or {}).get("origen") == "constante"
            else None
        )
        iva_constante = (
            self._parse_decimal((campos.get("item_iva_porcentaje") or {}).get("valor"))
            if isinstance(campos.get("item_iva_porcentaje"), dict)
            and (campos.get("item_iva_porcentaje") or {}).get("origen") == "constante"
            else None
        )
        if (
            tipo_constante in TIPOS_C
            and iva_constante is not None
            and iva_constante != Decimal("0")
        ):
            raise FormatoImportacionError("Los comprobantes tipo C deben usar IVA 0")

        if exigir_constantes_requeridas:
            campos_requeridos = list(CAMPOS_FISCALES_REQUERIDOS_GUARDADO)
            if configuracion.get("tipo") == "plantilla_visual":
                campos_requeridos.extend(CAMPOS_VISUALES_REQUERIDOS_GUARDADO)
            for campo_requerido in campos_requeridos:
                if not self._campo_requerido_resuelto_por_plantilla(
                    campo_requerido,
                    campos,
                ):
                    raise FormatoImportacionError(
                        "La plantilla debe definir "
                        f"{self._etiqueta_campo(campo_requerido)}."
                    )

    def _campo_requerido_resuelto_por_plantilla(
        self,
        campo: str,
        campos: dict[str, Any],
    ) -> bool:
        """Indica si una plantilla satisface un campo requerido de guardado."""
        detalle = campos.get(campo)
        if isinstance(detalle, dict) and self._campo_resuelto_por_plantilla(
            campo,
            detalle,
        ):
            return True
        if campo == "item_precio_unitario":
            detalle_total = campos.get("importe_total")
            return isinstance(
                detalle_total, dict
            ) and self._campo_resuelto_por_plantilla(
                "importe_total",
                detalle_total,
            )
        return False

    def _construir_firma_headers(self, configuracion: dict[str, Any]) -> dict[str, Any]:
        """Construye una firma simple de encabezados esperados."""
        requeridos: list[str] = []
        opcionales: list[str] = []
        for config in configuracion.get("campos", {}).values():
            encabezados = config.get("encabezados") or []
            if not encabezados:
                continue
            if config.get("requerido"):
                requeridos.append(encabezados[0])
            else:
                opcionales.append(encabezados[0])
        return {"requeridos": requeridos, "opcionales": opcionales}

    def _version_vigente(
        self, formato: FormatoImportacion
    ) -> FormatoImportacionVersion | None:
        versiones = sorted(
            [version for version in formato.versiones if version.estado == "vigente"],
            key=lambda version: version.version,
            reverse=True,
        )
        return versiones[0] if versiones else None

    async def hay_ambiguedad_fiscal_entre_candidatos(
        self,
        candidatos: list[CandidatoFormato],
        empresa_id: int,
    ) -> bool:
        """Indica si candidatos igualmente buenos difieren en constantes fiscales."""
        candidatos_altos = [
            candidato
            for candidato in candidatos
            if candidato.confianza == "alta" and candidato.score >= 0.85
        ]
        if len(candidatos_altos) < 2:
            return False

        mejor_score = max(candidato.score for candidato in candidatos_altos)
        empatados = [
            candidato
            for candidato in candidatos_altos
            if abs(candidato.score - mejor_score) < 0.0001
        ]
        if len(empatados) < 2:
            return False

        firmas: set[tuple[tuple[str, Any, Any], ...]] = set()
        for candidato in empatados:
            if candidato.formato_version_id is None:
                firmas.add((("plantilla_oficial", None, None),))
                continue
            try:
                version = await self.obtener_version(
                    candidato.formato_version_id,
                    empresa_id,
                )
            except FormatoImportacionError:
                continue
            firmas.add(self._firma_fiscal_no_visible(version.configuracion_json))
        return len(firmas) > 1

    def _firma_fiscal_no_visible(
        self,
        configuracion: dict[str, Any],
    ) -> tuple[tuple[str, Any, Any], ...]:
        """Resume constantes fiscales que no se distinguen por encabezados."""
        campos = configuracion.get("campos") or {}
        firma: list[tuple[str, Any, Any]] = []
        for campo in CAMPOS_FISCALES_DESEMPATE:
            detalle = campos.get(campo) if isinstance(campos, dict) else None
            if not isinstance(detalle, dict):
                firma.append((campo, "faltante", None))
                continue
            origen = detalle.get("origen", "header")
            if origen in {"header", "columna"}:
                firma.append((campo, "visible", None))
            else:
                firma.append((campo, origen, detalle.get("valor")))
        return tuple(firma)

    def _evaluar_formato(
        self,
        headers: list[str],
        formato: FormatoImportacion,
        version: FormatoImportacionVersion,
    ) -> CandidatoFormato:
        mapeo = self._resolver_mapeo(headers, version.configuracion_json)
        campos_header = [
            (campo, detalle)
            for campo, detalle in mapeo["campos"].items()
            if detalle.get("origen") in {"header", "columna"}
        ]
        required = [item for item in campos_header if item[1].get("requerido")]
        optional = [item for item in campos_header if not item[1].get("requerido")]
        matched_required = [item for item in required if item[1].get("encontrado")]
        matched_optional = [item for item in optional if item[1].get("encontrado")]
        missing_required = [
            item[0] for item in required if not item[1].get("encontrado")
        ]
        weight = max((len(required) * 2) + len(optional), 1)
        score = ((len(matched_required) * 2) + len(matched_optional)) / weight
        if missing_required:
            confianza = "baja"
        elif score >= 0.85:
            confianza = "alta"
        elif score >= 0.6:
            confianza = "media"
        else:
            confianza = "baja"

        mensajes = [
            "Coincide con las columnas requeridas."
            if not missing_required
            else "Faltan columnas requeridas."
        ]
        return CandidatoFormato(
            formato_id=formato.id,
            formato_version_id=version.id,
            nombre=formato.nombre,
            alcance=formato.alcance,
            version=version.version,
            score=round(score, 4),
            confianza=confianza,
            columnas_detectadas=[
                item[1]["header"] for item in campos_header if item[1].get("header")
            ],
            columnas_faltantes=missing_required,
            mensajes=mensajes,
        )

    def _leer_sheet_y_headers(
        self,
        file_bytes: bytes,
        version: FormatoImportacionVersion | None = None,
    ):
        try:
            workbook = load_workbook(
                BytesIO(file_bytes), data_only=True, read_only=True
            )
        except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
            raise FormatoImportacionError(
                "No se pudo leer el archivo Excel. Verificá que sea un .xlsx válido."
            ) from exc
        sheet_name = None
        header_row = 1
        if version is not None:
            sheet_name = version.configuracion_json.get("sheet_name")
            header_row = int(version.configuracion_json.get("header_row", 1))

        if sheet_name and sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        elif "Comprobantes" in workbook.sheetnames:
            sheet = workbook["Comprobantes"]
        else:
            sheet = workbook.active

        headers = [self.reparar_texto(cell.value) for cell in sheet[header_row]]
        if not any(headers):
            raise FormatoImportacionError(
                "No se detectaron encabezados en la primera fila del Excel"
            )
        return sheet, headers

    def _resolver_mapeo(
        self, headers: list[str], configuracion: dict[str, Any]
    ) -> dict[str, Any]:
        normalizados = {
            self.normalizar_etiqueta(header): {"header": header, "index": index}
            for index, header in enumerate(headers)
            if self.normalizar_etiqueta(header)
        }
        campos: dict[str, Any] = {}
        for campo_destino, config in configuracion.get("campos", {}).items():
            origen = config.get("origen", "header")
            detalle = {
                "origen": origen,
                "requerido": bool(config.get("requerido", False)),
                "transformacion": config.get("transformacion"),
                "default": config.get("default"),
                "valor": config.get("valor"),
                "encontrado": origen == "constante"
                or (origen == "empresa" and campo_destino in CAMPOS_ORIGEN_EMPRESA),
                "header": None,
                "index": None,
            }
            if origen == "header":
                for alias in config.get("encabezados", []):
                    match = normalizados.get(self.normalizar_etiqueta(alias))
                    if match:
                        detalle.update(
                            {
                                "encontrado": True,
                                "header": match["header"],
                                "index": match["index"],
                            }
                        )
                        break
            elif origen == "columna":
                index = self._resolver_indice_columna(config)
                detalle.update({"encontrado": index is not None, "index": index})

            campos[campo_destino] = detalle
        return {
            "tipo": configuracion.get("tipo"),
            "modo_agrupacion": configuracion.get("modo_agrupacion", "fila"),
            "campos": campos,
        }

    def _resolver_indice_columna(self, config: dict[str, Any]) -> int | None:
        try:
            if config.get("indice_columna") is not None:
                index = int(config["indice_columna"])
                if index < 0 or index > EXCEL_MAX_COLUMN_INDEX:
                    raise ValueError
                return index
            if config.get("letra_columna"):
                index = column_index_from_string(str(config["letra_columna"])) - 1
                if index > EXCEL_MAX_COLUMN_INDEX:
                    raise ValueError
                return index
            return None
        except (TypeError, ValueError) as exc:
            raise FormatoImportacionError(
                "La configuración de columna del formato es inválida"
            ) from exc

    def _extraer_valores_configurados(
        self, row: tuple[Any, ...], mapeo: dict[str, Any]
    ) -> dict[str, Any]:
        valores: dict[str, Any] = {}
        for campo, detalle in mapeo["campos"].items():
            if detalle["origen"] == "constante":
                valores[campo] = detalle.get("valor")
                continue
            if detalle["origen"] == "empresa":
                continue

            value = detalle.get("default", "")
            index = detalle.get("index")
            if index is not None and index < len(row):
                value = row[index]
            valores[campo] = self._aplicar_transformacion(
                value, detalle.get("transformacion")
            )
        return valores

    def _armar_fila_canonica(
        self,
        valores: dict[str, Any],
        empresa: Empresa,
        fila_excel: int,
    ) -> dict[str, Any]:
        documento = clean_cuit(valores.get("cliente_numero_documento", ""))
        importe_total = self._parse_decimal(valores.get("importe_total"))
        precio_unitario = self._parse_decimal(valores.get("item_precio_unitario"))
        total_receptor = importe_total or precio_unitario or Decimal("0")
        condicion_iva = str(
            valores.get("cliente_condicion_iva", "Consumidor Final") or ""
        ).strip()
        es_consumidor_final = condicion_iva.upper() in {"CF", "CONSUMIDOR FINAL"}
        if (
            es_consumidor_final
            and total_receptor < self.CONSUMIDOR_FINAL_IDENTIFICACION_MINIMA
        ):
            documento = ""
        empresa_cuit = (
            clean_cuit(valores.get("empresa_cuit", ""))
            if "empresa_cuit" in valores
            else clean_cuit(empresa.cuit)
        )
        item_precio_unitario = valores.get(
            "item_precio_unitario", valores.get("importe_total", "")
        )
        return {
            "comprobante_ref": f"FILA-{fila_excel:05d}",
            "empresa_cuit": empresa_cuit,
            "punto_venta_numero": valores.get("punto_venta_numero", ""),
            "tipo_comprobante": valores.get("tipo_comprobante", ""),
            "concepto": valores.get("concepto", ""),
            "fecha_emision": valores.get("fecha_emision", ""),
            "fecha_origen": valores.get("fecha_origen", ""),
            "importe_total": valores.get("importe_total", ""),
            "cliente_tipo_documento": self._inferir_tipo_documento(documento),
            "cliente_numero_documento": documento,
            "cliente_razon_social": valores.get("cliente_razon_social", ""),
            "cliente_condicion_iva": condicion_iva or "Consumidor Final",
            "cliente_domicilio": valores.get("cliente_domicilio", ""),
            "fecha_servicio_desde": valores.get("fecha_servicio_desde", ""),
            "fecha_servicio_hasta": valores.get("fecha_servicio_hasta", ""),
            "fecha_vto_pago": valores.get("fecha_vto_pago", ""),
            "item_codigo": valores.get("item_codigo", ""),
            "item_descripcion": valores.get("item_descripcion", ""),
            "item_cantidad": valores.get("item_cantidad", 1),
            "item_unidad": valores.get("item_unidad", "unidad"),
            "item_precio_unitario": item_precio_unitario,
            "item_descuento_porcentaje": valores.get("item_descuento_porcentaje", 0),
            "item_iva_porcentaje": valores.get("item_iva_porcentaje", ""),
            "observaciones": valores.get("observaciones", ""),
            "asociado_tipo_comprobante": valores.get("asociado_tipo_comprobante", ""),
            "asociado_punto_venta": valores.get("asociado_punto_venta", ""),
            "asociado_numero": valores.get("asociado_numero", ""),
            "asociado_fecha": valores.get("asociado_fecha", ""),
            "asociado_cuit": valores.get("asociado_cuit", ""),
        }

    def _aplicar_transformacion(self, value: Any, transformacion: str | None) -> Any:
        if value in (None, ""):
            return ""
        if transformacion == "decimal":
            parsed = self._parse_decimal(value)
            return str(parsed) if parsed is not None else ""
        if transformacion == "entero":
            try:
                return int(Decimal(str(self._parse_decimal(value) or value)))
            except (InvalidOperation, ValueError):
                return value
        if transformacion == "documento":
            return clean_cuit(value)
        if transformacion == "fecha":
            return self._parse_date(value) or value
        if transformacion == "texto":
            return str(value).strip()
        return value

    def _parse_decimal(self, value: Any) -> Decimal | None:
        if isinstance(value, Decimal):
            return value
        if isinstance(value, (int, float)):
            return Decimal(str(value))
        text = str(value).strip()
        if not text:
            return None
        text = text.replace("$", "").replace(" ", "")
        if "," in text and "." in text:
            text = text.replace(".", "").replace(",", ".")
        elif "," in text:
            text = text.replace(",", ".")
        try:
            return Decimal(text)
        except (InvalidOperation, ValueError):
            return None

    def _parse_date(self, value: Any) -> str | None:
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        if isinstance(value, (int, float)) and 1 <= value <= 60000:
            try:
                return from_excel(value).date().isoformat()
            except (TypeError, ValueError):
                return None
        text = str(value).strip()
        if not text:
            return None
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(text, fmt).date().isoformat()
            except ValueError:
                continue
        return None

    def _inferir_tipo_documento(self, documento: str) -> str:
        if not documento:
            return ""
        if len(documento) == 11 and validate_cuit(documento):
            return "CUIT"
        if len(documento) in {7, 8}:
            return "DNI"
        return "CUIT" if len(documento) == 11 else ""
